import json
from django.conf import settings
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from unidecode import unidecode
import logging

from MonEcole_app.models.country_structure import (
    Pays, PAYS_AFRIQUE_EST,
    AdministrativeStructureInstance, PedagogicStructureInstance,
    StructurePedagogique as PedagogicStructureType,
    AdministrativeStructureType,
    Regime,
    Cycle, Etablissement, GestionnaireEtablissement,
    RepartitionType, RepartitionInstance, RepartitionHierarchie,
    RepartitionConfigCycle, RepartitionConfigEtabAnnee,
    CoursAnnee, EtablissementAnnee, EtablissementAnneeClasse,
    Session, Mention,
    EvaluationType,
)
from MonEcole_app.models.classe import Classe
from MonEcole_app.models.annee import Annee

# Models that may not exist yet in MonEcole — safe imports with fallbacks
try:
    from MonEcole_app.models.country_structure import Section
except ImportError:
    Section = None
try:
    from MonEcole_app.models.country_structure import Programme
except ImportError:
    Programme = None
try:
    from MonEcole_app.models.country_structure import Domaine
except ImportError:
    Domaine = None
try:
    from MonEcole_app.models.country_structure import Cours
except ImportError:
    try:
        from MonEcole_app.models.enseignmnts.matiere import Cours
    except ImportError:
        Cours = None
try:
    from MonEcole_app.models.country_structure import TypeSubdivision
except ImportError:
    TypeSubdivision = None

# AdminUser compatibility — MonEcole uses Django auth, not AdminUser
class _DummyAdminUser:
    """Compatibility stub — MonEcole doesn't use AdminUser table directly."""
    DoesNotExist = Exception
    objects = None
try:
    from MonEcole_app.models.hub_auth import AdminUser
except ImportError:
    AdminUser = _DummyAdminUser

# Forms compatibility (MonEcole doesn't have these forms)
class PaysForm:
    pass

from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

def get_user_scope(request):
    """MonEcole stub — scope is implicit via tenant middleware (subdomain)."""
    return None, None


def get_all_user_scopes(request):
    """MonEcole stub — no RBAC multi-structure filtering needed."""
    return []


def _get_tenant_etab(request):
    """
    Get the current tenant Etablissement from the TenantMiddleware.
    In MonEcole, AdminUser does NOT have an 'etablissement' FK.
    The establishment is resolved by TenantMiddleware from the subdomain.
    Returns (Etablissement, None) or (None, JsonResponse_error).
    """
    etab_id = getattr(request, 'id_etablissement', None) or request.session.get('id_etablissement')
    id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
    if not etab_id:
        return None, JsonResponse({'success': False, 'error': 'Établissement non résolu (tenant).'}, status=403)
    # Scope by both id_etablissement and pays for multi-tenant safety
    filters = {'id_etablissement': etab_id}
    if id_pays:
        filters['pays_id'] = id_pays
    etab = Etablissement.objects.select_related('pays').filter(**filters).first()
    if not etab:
        return None, JsonResponse({'success': False, 'error': f'Établissement {etab_id} introuvable.'}, status=404)
    return etab, None



def get_country_context_logic(request):
    """MonEcole stub — context is built in dashboard_views.py."""
    return {}


def parametres_generaux_view(request):
    """Not used in MonEcole."""
    from django.http import Http404
    raise Http404


def calendrier_scolaire_view(request):
    """Not used in MonEcole."""
    from django.http import Http404
    raise Http404


import sys


@require_http_methods(["GET"])
def get_pays_data(request):
    """API pour récupérer les données d'un pays."""
    sys.stderr.write(f"DEBUG STDERR: get_pays_data reached for path: {request.path}\n")
    logger = logging.getLogger(__name__)
    try:
        id_pays = request.GET.get('id_pays')
        nom = request.GET.get('nom')
        logger.debug(f"DEBUG: get_pays_data called with id_pays={id_pays}, nom={nom}")
        
        pays = None
        if id_pays:
            pays = Pays.objects.filter(id_pays=id_pays).first()
        elif nom:
            pays = Pays.objects.filter(nom__iexact=nom).first()
            
        if not pays:
            logger.debug("DEBUG: Pays not found")
            return JsonResponse({
                'exists': False,
                'pays': {'nom': nom, 'id_pays': None, 'nLevelsPedagogiques': 0, 'nLevelsAdministratifs': 0},
                'structures_pedagogiques': [],
                'structures_administratives': [],
            })
            
        # Filtrage par scope utilisateur (multi-structures)
        scope_codes = get_all_user_scopes(request)
        
        # Build structure lists from TYPES tables (source of truth for level definitions)
        # Then overlay instance counts for each level
        structures_ped = []
        ped_types = PedagogicStructureType.objects.filter(pays=pays).order_by('ordre')
        for t in ped_types:
            qs = PedagogicStructureInstance.objects.filter(pays=pays, ordre=t.ordre)
            if scope_codes:
                from django.db.models import Q
                scope_q = Q()
                for sc in scope_codes:
                    scope_q |= Q(code__startswith=sc + '-') | Q(code=sc)
                qs = qs.filter(scope_q)
            count = qs.count()
            structures_ped.append({
                'id_structure': t.id_structure,
                'typeStruct': 'PD',
                'nom': t.nom,
                'code': t.code,
                'ordre': t.ordre,
                'nInstances': count
            })

        structures_admin = []
        admin_types = AdministrativeStructureType.objects.filter(pays=pays).order_by('ordre')
        for t in admin_types:
            qs = AdministrativeStructureInstance.objects.filter(pays=pays, ordre=t.ordre)
            count = qs.count()
            structures_admin.append({
                'id_structure': t.id_structure,
                'typeStruct': 'AD',
                'nom': t.nom,
                'code': t.code,
                'ordre': t.ordre,
                'nInstances': count
            })
        
        return JsonResponse({
            'exists': True,
            'pays': {
                'id_pays': pays.id_pays,
                'nom': pays.nom,
                'sigle': pays.sigle,
                'nLevelsPedagogiques': pays.nLevelsPedagogiques,
                'nLevelsAdministratifs': pays.nLevelsAdministratifs,
            },
            'structures_pedagogiques': structures_ped,
            'structures_administratives': structures_admin,
        })
    except Pays.DoesNotExist:
        return JsonResponse({
            'exists': False,
            'pays': {'nom': nom, 'id_pays': None, 'nLevelsPedagogiques': 0, 'nLevelsAdministratifs': 0},
            'structures_pedagogiques': [],
            'structures_administratives': [],
        })
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logger.error(f"CRITICAL ERROR in get_pays_data: {str(e)}\n{tb}")
        return JsonResponse({
            'success': False,
            'error': f"Erreur interne du serveur: {str(e)}",
            'traceback': tb if settings.DEBUG else None
        }, status=500)


@require_http_methods(["POST"])
def save_pays(request):
    """API pour créer ou mettre à jour un pays."""
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays')
        nom = data.get('nom')
        sigle_value = data.get('sigle', '').upper()  # sigle reste un champ de données, pas un identifiant
        
        # Conversion robuste vers entier
        try:
            nLevelsPedagogiques = int(data.get('nLevelsPedagogiques') or 0)
            nLevelsAdministratifs = int(data.get('nLevelsAdministratifs') or 0)
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Les nombres de niveaux doivent être des entiers.'}, status=400)
        
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        
        # Lookup par id_pays
        pays = Pays.objects.filter(id_pays=id_pays).first() if id_pays else None
        if pays:
            pays.nom = nom
            if sigle_value:
                pays.sigle = sigle_value
            pays.nLevelsPedagogiques = nLevelsPedagogiques
            pays.nLevelsAdministratifs = nLevelsAdministratifs
            pays.save()
            created = False
        else:
            pays = Pays.objects.create(
                sigle=sigle_value or 'XX',
                nom=nom,
                nLevelsPedagogiques=nLevelsPedagogiques,
                nLevelsAdministratifs=nLevelsAdministratifs
            )
            created = True
        return JsonResponse({
            'success': True, 
            'pays': {
                'id_pays': pays.id_pays,
                'nom': pays.nom,
                'nLevelsPedagogiques': pays.nLevelsPedagogiques, 
                'nLevelsAdministratifs': pays.nLevelsAdministratifs
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def add_structure_pedagogique(request):
    try:
        data = json.loads(request.body)
        pays = get_object_or_404(Pays, id_pays=data.get('id_pays'))
        existing_ordres = pays.pedag_instances.values('ordre').distinct().count()
        if existing_ordres >= pays.nLevelsPedagogiques:
            return JsonResponse({'error': 'Limite atteinte'}, status=400)
        
        nom = data.get('nom')
        ordre = existing_ordres + 1
        
        PedagogicStructureInstance.objects.create(pays=pays, nom=nom, ordre=ordre, code='TEMP')
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def add_structure_administrative(request):
    try:
        data = json.loads(request.body)
        pays = get_object_or_404(Pays, id_pays=data.get('id_pays'))
        existing_ordres = pays.admin_instances.values('ordre').distinct().count()
        if existing_ordres >= pays.nLevelsAdministratifs:
            return JsonResponse({'error': 'Limite atteinte'}, status=400)
        
        nom = data.get('nom')
        ordre = existing_ordres + 1
        
        AdministrativeStructureInstance.objects.create(pays=pays, nom=nom, ordre=ordre, code='TEMP')
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_structure_pedagogique(request):
    try:
        data = json.loads(request.body)
        structure = get_object_or_404(PedagogicStructureInstance, id_structure=data.get('id_structure'))
        structure.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_structure_administrative(request):
    try:
        data = json.loads(request.body)
        structure = get_object_or_404(AdministrativeStructureInstance, id_structure=data.get('id_structure'))
        structure.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def update_structure_pedagogique(request):
    try:
        data = json.loads(request.body)
        structure = get_object_or_404(PedagogicStructureInstance, id_structure=data.get('id_structure'))
        
        nom = data.get('nom')
        if nom:
            structure.nom = nom
            structure.save()
            
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def update_structure_administrative(request):
    try:
        data = json.loads(request.body)
        structure = get_object_or_404(AdministrativeStructureInstance, id_structure=data.get('id_structure'))
        
        nom = data.get('nom')
        if nom:
            structure.nom = nom
            structure.save()
            
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def adjust_levels(request):
    try:
        data = json.loads(request.body)
        pays = get_object_or_404(Pays, id_pays=data.get('id_pays'))
        adj_type = data.get('type')
        
        if adj_type == 'PD' or adj_type == 'all':
            pays.nLevelsPedagogiques = PedagogicStructureType.objects.filter(pays=pays).count()
        if adj_type == 'AD' or adj_type == 'all':
            pays.nLevelsAdministratifs = AdministrativeStructureType.objects.filter(pays=pays).count()
            
        pays.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def generate_code_api(request):
    # This might be deprecated or unused now, but keeping for safety for now or returning empty
    return JsonResponse({'code': ''})


@require_http_methods(["POST"])
def save_structures(request):
    """API pour sauvegarder (remplacer) la liste des types de structures."""
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays')
        struct_type = data.get('type')  # 'PD' or 'AD'
        structures_list = data.get('structures', [])
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        if struct_type == 'PD':
            Model = PedagogicStructureType
            pays.nLevelsPedagogiques = len(structures_list)
        elif struct_type == 'AD':
            Model = AdministrativeStructureType
            pays.nLevelsAdministratifs = len(structures_list)
        else:
            return JsonResponse({'success': False, 'error': 'Type invalide'}, status=400)
        
        pays.save()
        
        # Delete existing types for this pays
        Model.objects.filter(pays=pays).delete()
        
        # Create new types from the list
        for idx, s in enumerate(structures_list, start=1):
            nom = s.get('nom', f'Niveau {idx}')
            code = s.get('code', '')
            if not code:
                # Auto-generate code from nom (unidecode imported at top)
                code = unidecode(nom).upper().replace(' ', '_')[:20]
            Model.objects.create(
                pays=pays,
                ordre=idx,
                nom=nom,
                code=code
            )
        
        return JsonResponse({'success': True})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logging.getLogger(__name__).error(f"Error in save_structures: {str(e)}\n{tb}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_http_methods(["GET"])
def get_structures_list(request):
    """
    API pour récupérer la liste des structures concrètes filtrées.
    Supporte le filtrage par parent_code (hiérarchique).
    """
    try:
        id_pays = request.GET.get('id_pays')
        type_struct = request.GET.get('type')
        ordre = request.GET.get('ordre')
        parent_code = request.GET.get('parent_code') # Optionnel: filtrer par hiérarchie
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        Model = PedagogicStructureInstance if type_struct == 'PD' else AdministrativeStructureInstance
        qs = Model.objects.filter(
            pays=pays,
            ordre=ordre
        )
        
        if parent_code:
            # Filtrer les descendants. Le code d'un enfant commence par "parent_code-"
            qs = qs.filter(code__startswith=f"{parent_code}-")

        # Filtrage par scope utilisateur (RBAC multi-structures)
        scope_codes = get_all_user_scopes(request)
        if scope_codes and type_struct == 'PD':
            from django.db.models import Q
            scope_q = Q()
            for sc in scope_codes:
                scope_q |= Q(code__startswith=sc + '-') | Q(code=sc)
            qs = qs.filter(scope_q)

        qs = qs.order_by('nom')
        
        results = []
        for s in qs:
            # Count ecoles
            if type_struct == 'PD':
                n_ecoles = s.etablissements.count()
            else:
                # Count ecoles in PD structures that belong to this AD structure
                n_ecoles = Etablissement.objects.filter(
                    pays=pays,
                    structure_pedagogique__administrative_parent__code__startswith=s.code
                ).count()
            
            item = {
                'id_structure': s.id_structure,
                'nom': s.nom,
                'code': s.code,
                'latitude': s.latitude,
                'longitude': s.longitude,
                'n_ecoles': n_ecoles,
                'n_eleves': 0,
            }
            if type_struct == 'PD' and hasattr(s, 'administrative_parent') and s.administrative_parent:
                item['administrative_parent'] = {
                    'id': s.administrative_parent.id_structure,
                    'nom': s.administrative_parent.nom
                }
            results.append(item)
            
        return JsonResponse({'success': True, 'structures': results})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_potential_parents(request):
    """
    API pour récupérer les parents potentiels (niveau N-1).
    """
    try:
        id_pays = request.GET.get('id_pays')
        type_struct = request.GET.get('type')
        ordre_enfant = int(request.GET.get('ordre'))
        
        if ordre_enfant <= 1:
            return JsonResponse({'success': True, 'parents': []})
            
        pays = get_object_or_404(Pays, id_pays=id_pays)
        ordre_parent = ordre_enfant - 1
        
        # NOTE: A pedagogical structure MIGHT look into administrative parents or vice versa?
        # User implies hierarchy within same system usually, but let's stick to strict same type for now?
        # "fils d'un elemnt du second ordre" implies recursive structure.
        # Assuming Same Type for now.
        
        Model = PedagogicStructureInstance if type_struct == 'PD' else AdministrativeStructureInstance
        qs = Model.objects.filter(
            pays=pays,
            ordre=ordre_parent
        ).order_by('nom')
        
        filter_code = request.GET.get('filter_code')
        if filter_code:
             qs = qs.filter(code__startswith=f"{filter_code}-")
        
        # Filtrage par scope utilisateur (RBAC multi-structures)
        scope_codes = get_all_user_scopes(request)
        if scope_codes and type_struct == 'PD':
            from django.db.models import Q
            scope_q = Q()
            for sc in scope_codes:
                scope_q |= Q(code__startswith=sc + '-') | Q(code=sc)
                # Ajouter les ancêtres de chaque scope
                for i in range(1, sc.count('-') + 1):
                    scope_q |= Q(code=sc.rsplit('-', i)[0])
            qs = qs.filter(scope_q)
        
        qs = qs.select_related('pays') # Optimization
        parents_data = list(qs.values('id_structure', 'nom', 'code'))
        
        # Resolve Breadcrumbs (Hierarchy names)
        # Collect all ancestor IDs from all codes
        all_ancestor_ids = set()
        for p in parents_data:
            if p['code']:
                # Code format: "GP_ID-P_ID-MY_ID"
                # We want ancestors, so all except the last one (which is self, or effectively the node itself)
                # Actually, 'code' for an object INCLUDES itself as the last element.
                # Ancestors are the prefix parts.
                parts = p['code'].split('-')
                if len(parts) > 1:
                    all_ancestor_ids.update(parts[:-1]) # All parts except the last one (self)
                    
        # Fetch Ancestor Names
        ancestor_map = {}
        if all_ancestor_ids:
             Model = PedagogicStructureInstance if type_struct == 'PD' else AdministrativeStructureInstance
             ancestors = Model.objects.filter(id_structure__in=all_ancestor_ids).values('id_structure', 'nom')
             for a in ancestors:
                 ancestor_map[str(a['id_structure'])] = a['nom']
                 
        # Build Strings
        for p in parents_data:
            breadcrumb = ""
            if p['code']:
                parts = p['code'].split('-')
                # Ancestors only
                ancestor_names = []
                for ancestor_id in parts[:-1]: # Skip self
                    name = ancestor_map.get(str(ancestor_id))
                    if name:
                        ancestor_names.append(name)
                
                if ancestor_names:
                    breadcrumb = " > ".join(ancestor_names)
            
            p['breadcrumb'] = breadcrumb
            p['full_display'] = f"{breadcrumb} > {p['nom']}" if breadcrumb else p['nom']

        return JsonResponse({'success': True, 'parents': parents_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_structure_instance(request):
    """
    API pour créer ou mettre à jour une structure concrète.
    Gère la logique de code hiérarchique : code = parent_code + '-' + self_id
    """
    try:
        data = json.loads(request.body)
        id_structure = data.get('id_structure') # If update
        id_pays = data.get('id_pays')
        type_struct = data.get('type')
        ordre = int(data.get('ordre'))
        nom = data.get('nom')
        parent_id = data.get('parent_id')
        lat = float(data.get('latitude') or 0.0)
        lng = float(data.get('longitude') or 0.0)
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        # LEVEL ENFORCEMENT disabled in MonEcole (uses tenant middleware)
        
        Model = PedagogicStructureInstance if type_struct == 'PD' else AdministrativeStructureInstance
        
        if id_structure:
            # UPDATE
            s = get_object_or_404(Model, id_structure=id_structure)
            s.nom = nom
            s.latitude = lat
            s.longitude = lng
            
            # Geographic link for PD
            if type_struct == 'PD' and data.get('administrative_parent_id'):
                try:
                    s.administrative_parent = AdministrativeStructureInstance.objects.get(id_structure=data.get('administrative_parent_id'))
                except:
                    pass
            s.save()
        else:
            # CREATE
            parent = None
            if ordre > 1:
                if not parent_id:
                    return JsonResponse({'success': False, 'error': "Un parent est requis pour ce niveau."}, status=400)
                parent = get_object_or_404(Model, id_structure=parent_id)
            
            # 1. Create parameters
            create_params = {
                'pays': pays,
                'nom': nom,
                'ordre': ordre,
                'latitude': lat,
                'longitude': lng,
                'code': 'TEMP'
            }
            if type_struct == 'PD' and data.get('administrative_parent_id'):
                try:
                    create_params['administrative_parent'] = AdministrativeStructureInstance.objects.get(id_structure=data.get('administrative_parent_id'))
                except:
                    pass
                    
            s = Model.objects.create(**create_params)
            
            # 2. Calculate Code
            if ordre == 1:
                new_code = str(s.id_structure)
            else:
                new_code = f"{parent.code}-{s.id_structure}"
            
            s.code = new_code
            s.save()
            
        return JsonResponse({'success': True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_structure_instance(request):
    try:
        data = json.loads(request.body)
        id_structure = data.get('id_structure')
        type_struct = data.get('type', 'PD')
        
        Model = PedagogicStructureInstance if type_struct == 'PD' else AdministrativeStructureInstance
        s = get_object_or_404(Model, id_structure=id_structure)
        
        # Check if children exist (code starts with "s.code-")
        has_children = Model.objects.filter(code__startswith=f"{s.code}-").exists()
        if has_children:
             return JsonResponse({'success': False, 'error': "Impossible de supprimer : cette structure a des enfants."}, status=400)
             
        s.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# --- EXCEL IMPORT / EXPORT ---
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from django.http import HttpResponse

@require_http_methods(["GET"])
def download_template(request):
    """
    Génère un modèle Excel pour le niveau actuel.
    Pour ordre > 1: le parent est sélectionné via la modal (parent_code passé en paramètre).
    Le template est simple: NOM DE LA STRUCTURE uniquement, parent_code dans le titre de feuille.
    """
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'La librairie Excel (openpyxl) n\'est pas installée sur le serveur.'}, status=500)

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        id_pays = request.GET.get('id_pays')
        type_struct = request.GET.get('type')
        ordre = int(request.GET.get('ordre'))
        nom_structure = request.GET.get('nom_structure', 'Structure')
        parent_code = request.GET.get('parent_code', '')
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        Model = PedagogicStructureInstance if type_struct == 'PD' else AdministrativeStructureInstance
        
        wb = openpyxl.Workbook()
        
        # --- Styles ---
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        example_font = Font(italic=True, color="888888")
        instruction_font = Font(italic=True, color="1F4E79", size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Résoudre le nom du parent si ordre > 1
        parent_name = ''
        if ordre > 1 and parent_code:
            parent_obj = Model.objects.filter(pays=pays, code=parent_code).first()
            parent_name = parent_obj.nom if parent_obj else parent_code
        
        ws = wb.active
        # Le titre de la feuille = code parent (utilisé lors de l'import legacy)
        ws.title = parent_code if parent_code else "Import"
        
        # Header
        ws['A1'] = "NOM DE LA STRUCTURE"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = header_alignment
        ws['A1'].border = thin_border
        ws.column_dimensions['A'].width = 45
        
        # Example row
        ws['A2'] = "Ex: Mon Entité"
        ws['A2'].font = example_font
        
        # Instructions sheet
        ws_inst = wb.create_sheet(title="Instructions")
        ws_inst['A1'] = f"📋 Modèle d'import — {nom_structure}"
        ws_inst['A1'].font = Font(bold=True, size=14, color="1F4E79")
        ws_inst['A3'] = "Instructions :"
        ws_inst['A3'].font = Font(bold=True, size=11)
        ws_inst['A4'] = "1. Remplissez la colonne 'NOM DE LA STRUCTURE' dans la feuille d'import."
        ws_inst['A5'] = "2. Supprimez la ligne d'exemple (ligne 2) avant d'importer."
        ws_inst['A6'] = f"3. Pays : {pays.nom} (ID {pays.id_pays}) — Type : {'Pédagogique' if type_struct == 'PD' else 'Administrative'}"
        ws_inst['A7'] = f"4. Niveau : {ordre} ({nom_structure})"
        
        if ordre > 1 and parent_name:
            ws_inst['A8'] = f"5. Parent sélectionné : {parent_name} (code: {parent_code})"
            ws_inst['A9'] = f"6. Toutes les structures seront créées sous ce parent."
            for r in range(3, 10):
                ws_inst[f'A{r}'].font = instruction_font
        else:
            for r in range(3, 8):
                ws_inst[f'A{r}'].font = instruction_font
        ws_inst.column_dimensions['A'].width = 80
        
        # Filename
        suffix = f"_{parent_name}" if parent_name else ""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Modele_{nom_structure}{suffix}_{pays.id_pays}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def import_excel(request):
    """
    Traite l'upload du fichier Excel.
    Supporte 3 modes de résolution du parent :
      1. parent_id dans POST (sélection via modal — prioritaire)
      2. Colonne PARENT dans le fichier (nom par ligne)
      3. Code parent dans le titre de la feuille (legacy)
    """
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'success': False, 'error': 'La librairie Excel (openpyxl) n\'est pas installée sur le serveur.'}, status=500)

    try:
        if 'file' not in request.FILES:
             return JsonResponse({'success': False, 'error': 'Aucun fichier fourni.'}, status=400)
             
        file = request.FILES['file']
        id_pays = request.POST.get('id_pays')
        type_struct = request.POST.get('type')
        ordre = int(request.POST.get('ordre'))
        parent_id_from_modal = request.POST.get('parent_id', '')  # From modal
        
        if not file.name.endswith('.xlsx'):
             return JsonResponse({'success': False, 'error': 'Format invalide. Utilisez .xlsx'}, status=400)
             
        pays = get_object_or_404(Pays, id_pays=id_pays)
        Model = PedagogicStructureInstance if type_struct == "PD" else AdministrativeStructureInstance
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        rows = list(ws.rows)
        if not rows:
             return JsonResponse({'success': False, 'error': 'Fichier vide.'}, status=400)
        
        # --- Détecter les headers ---
        header_row = [str(cell.value).upper().strip() if cell.value else '' for cell in rows[0]]
        
        # Chercher la colonne NOM DE LA STRUCTURE
        idx_nom = None
        for h_idx, h_val in enumerate(header_row):
            if 'NOM DE LA STRUCTURE' in h_val:
                idx_nom = h_idx
                break
        
        if idx_nom is None:
            return JsonResponse({'success': False, 'error': "En-tête 'NOM DE LA STRUCTURE' manquante dans la première ligne."}, status=400)
        
        # --- Résoudre le parent ---
        # MODE 1: parent_id from modal (prioritaire)
        modal_parent = None
        if ordre > 1 and parent_id_from_modal:
            try:
                modal_parent = Model.objects.get(id_structure=int(parent_id_from_modal), pays=pays)
                if modal_parent.ordre != ordre - 1:
                    return JsonResponse({'success': False, 'error': f"Incohérence: Le parent sélectionné '{modal_parent.nom}' est au niveau {modal_parent.ordre}, attendu {ordre - 1}."}, status=400)
            except Model.DoesNotExist:
                return JsonResponse({'success': False, 'error': f"Parent introuvable (id={parent_id_from_modal})."}, status=400)
        
        # MODE 2: No modal parent → try sheet title (legacy)
        sheet_parent = None
        if ordre > 1 and not modal_parent:
            sheet_parent_code = str(ws.title).strip()
            if sheet_parent_code and sheet_parent_code != 'Import':
                sheet_parent = Model.objects.filter(pays=pays, code=sheet_parent_code).first()
                if sheet_parent and sheet_parent.ordre != ordre - 1:
                    sheet_parent = None  # Wrong level, ignore
        
        # Determine the parent to use for all rows
        fixed_parent = modal_parent or sheet_parent  # Will be None for ordre 1
        
        if ordre > 1 and not fixed_parent:
            return JsonResponse({'success': False, 'error': "Aucun parent n'a pu être déterminé. Sélectionnez un parent via le dialogue."}, status=400)
        
        success_count = 0
        errors = []
        
        # Parcourir à partir de la 2ème ligne
        for i, row in enumerate(rows[1:], start=2):
            # Lire le nom de la structure
            if idx_nom >= len(row):
                continue
            val_nom = row[idx_nom].value
            if not val_nom or str(val_nom).strip().startswith('Ex:'):
                continue
            nom_clean = str(val_nom).strip()
            if not nom_clean:
                continue
            
            # Créer la structure
            try:
                s = Model.objects.create(
                    pays=pays,
                    nom=nom_clean,
                    ordre=ordre,
                    latitude=0.0,
                    longitude=0.0,
                    code='TEMP_IMPORT'
                )
                
                # Calculer le code hiérarchique
                if ordre == 1:
                    new_code = str(s.id_structure)
                else:
                    new_code = f"{fixed_parent.code}-{s.id_structure}"
                
                s.code = new_code
                s.save()
                success_count += 1
                
            except Exception as e:
                errors.append(f"Ligne {i}: Erreur création ({str(e)})")
        
        return JsonResponse({
            'success': True, 
            'count': success_count, 
            'errors': errors[:20]
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def export_structures_excel(request):
    """Exporte les structures filtrées en Excel."""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'openpyxl non disponible'}, status=500)
    
    try:
        id_pays = request.GET.get('id_pays')
        type_str = request.GET.get('type')
        ordre = request.GET.get('ordre')
        parent_code = request.GET.get('parent_code')
        nom_type = request.GET.get('nom_type', 'Structures')

        pays = get_object_or_404(Pays, id_pays=id_pays)
        Model = PedagogicStructureInstance if type_str == "PD" else AdministrativeStructureInstance
        qs = Model.objects.filter(pays=pays, ordre=ordre)
        if parent_code:
            qs = qs.filter(code__startswith=f"{parent_code}-")
        qs = qs.order_by('nom')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = nom_type[:30]
        
        headers = ['ID', 'Nom', 'Code Hiérarchique', 'Latitude', 'Longitude']
        ws.append(headers)
        
        for s in qs:
            ws.append([s.id_structure, s.nom, s.code, s.latitude, s.longitude])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Export_{nom_type}_{pays.id_pays}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def export_structures_pdf(request):
    """Exporte les structures filtrées en PDF."""
    try:
        id_pays = request.GET.get('id_pays')
        type_str = request.GET.get('type')
        ordre = request.GET.get('ordre')
        parent_code = request.GET.get('parent_code')
        nom_type = request.GET.get('nom_type', 'Structures')

        pays = get_object_or_404(Pays, id_pays=id_pays)
        Model = PedagogicStructureInstance if type_str == "PD" else AdministrativeStructureInstance
        qs = Model.objects.filter(pays=pays, ordre=ordre)
        if parent_code:
            qs = qs.filter(code__startswith=f"{parent_code}-")
        qs = qs.order_by('nom')

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Titre
        elements.append(Paragraph(f"Liste des {nom_type} - {pays.nom}", styles['Title']))
        elements.append(Spacer(1, 12))

        # Table data
        data = [['ID', 'Nom', 'Code', 'Lat', 'Long']]
        for s in qs:
            data.append([str(s.id_structure), s.nom, s.code, str(s.latitude), str(s.longitude)])

        t = Table(data, colWidths=[40, 200, 100, 70, 70])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(t)

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        filename = f"Export_{nom_type}_{pays.id_pays}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# --- REGIME MANAGEMENT ---

@require_http_methods(["GET"])
def regimes_view(request):
    """Vue pour la gestion des régimes."""
    context = get_country_context_logic(request)
    context['active_page'] = 'structure'
    return render(request, 'structure_app/regimes.html', context)


@require_http_methods(["GET"])
def get_regimes_data(request):
    """API pour récupérer les régimes d'un pays."""
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
             return JsonResponse({'success': False, 'error': 'id_pays manquant'})
             
        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'regimes': []})

        regimes = pays.regimes.all().order_by('regime')
        
        # Filtrage par scope utilisateur pour les comptages (multi-structures)
        scope_codes = get_all_user_scopes(request)
        
        results = []
        for r in regimes:
            ecoles_qs = Etablissement.objects.filter(pays=pays, id_regime=r.id_regime)
            if scope_codes:
                from django.db.models import Q
                scope_q = Q()
                for sc in scope_codes:
                    scope_q |= Q(structure_pedagogique__code__startswith=sc)
                ecoles_qs = ecoles_qs.filter(scope_q)
            n_ecoles = ecoles_qs.count()
            results.append({
                'id_regime': r.id_regime,
                'regime': r.regime,
                'nEtablissements': n_ecoles, # Legacy name compat? Or use n_ecoles
                'n_ecoles': n_ecoles,
                'n_eleves': 0 
            })
        return JsonResponse({'success': True, 'regimes': results})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# --- ETABLISSEMENT MANAGEMENT ---

@require_http_methods(["GET"])
def get_ped_structures_for_modal(request):
    """API pour récupérer les structures pédagogiques (tous niveaux) pour le dropdown modal."""
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        # Récupérer le dernier niveau pédagogique (les établissements sont rattachés à ce niveau)
        max_ordre = pays.nLevelsPedagogiques
        
        # Récupérer le nom du type de structure pour ce niveau (pour le label dynamique)
        parent_type_nom = "Structure d'Appartenance"
        try:
            ped_type = PedagogicStructureType.objects.get(pays=pays, ordre=max_ordre)
            parent_type_nom = ped_type.nom.upper()
        except PedagogicStructureType.DoesNotExist:
            pass
        
        qs = PedagogicStructureInstance.objects.filter(
            pays=pays,
            ordre=max_ordre
        ).select_related('administrative_parent').order_by('nom')
        
        # Filtrage par scope utilisateur (RBAC)
        scope_codes = get_all_user_scopes(request)
        if scope_codes:
            from django.db.models import Q
            scope_q = Q()
            for sc in scope_codes:
                scope_q |= Q(code__startswith=sc + '-') | Q(code=sc)
                # Include ancestors
                for i in range(1, sc.count('-') + 1):
                    scope_q |= Q(code=sc.rsplit('-', i)[0])
            qs = qs.filter(scope_q)
        
        results = []
        for s in qs:
            item = {
                'id_structure': s.id_structure,
                'nom': s.nom,
                'code': s.code,
            }
            if s.administrative_parent:
                item['admin_parent'] = s.administrative_parent.nom
            results.append(item)
        
        return JsonResponse({'success': True, 'structures': results, 'parent_type_nom': parent_type_nom})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def generate_fiche_synoptique(request):
    """Génère la fiche synoptique PDF d'un établissement."""
    try:
        etab_id = request.GET.get('id')
        if not etab_id:
            return JsonResponse({'success': False, 'error': 'ID manquant'})
        
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
        etab_filters = {'id_etablissement': etab_id}
        if id_pays:
            etab_filters['pays_id'] = id_pays
        etab = Etablissement.objects.select_related(
            'structure_pedagogique', 
            'structure_pedagogique__administrative_parent',
            'gestionnaire',
            'pays'
        ).filter(**etab_filters).first()
        if not etab:
            return JsonResponse({'success': False, 'error': 'Établissement introuvable'}, status=404)
        
        # Get regime name
        regime_name = ''
        if etab.id_regime:
            regime_obj = Regime.objects.filter(id_regime=etab.id_regime).first()
            regime_name = regime_obj.regime if regime_obj else ''
        
        # Get administrative hierarchy
        division = ''
        sous_division = ''
        province = ''
        if etab.structure_pedagogique and etab.structure_pedagogique.administrative_parent:
            adm = etab.structure_pedagogique.administrative_parent
            sous_division = adm.nom
            # Try to find parent administrative structure
            if adm.code:
                parts = adm.code.split('-')
                if len(parts) > 1:
                    parent_id = parts[-2] if len(parts) >= 2 else parts[0]
                    try:
                        parent_adm = AdministrativeStructureInstance.objects.get(id_structure=parent_id)
                        division = parent_adm.nom
                        # Try grandparent for province
                        if len(parts) > 2:
                            gp_id = parts[0]
                            try:
                                gp_adm = AdministrativeStructureInstance.objects.get(id_structure=gp_id)
                                province = gp_adm.nom
                            except:
                                pass
                    except:
                        pass
        
        # Get gestionnaire info
        gest_nom = ''
        if etab.gestionnaire:
            gest_nom = f"{etab.gestionnaire.nom} {etab.gestionnaire.postnom}"
        
        # Build PDF
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm, cm
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        
        buffer = BytesIO()
        w, h = A4
        c = canvas.Canvas(buffer, pagesize=A4)
        
        # Margins
        left = 25 * mm
        right = w - 20 * mm
        top = h - 15 * mm
        content_width = right - left
        
        y = top
        
        # ========== HEADER ==========
        # Title
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.HexColor('#1a5276'))
        c.drawCentredString(w / 2, y, "FICHE SYNOPTIQUE DE RENSEIGNEMENT")
        y -= 16
        c.drawCentredString(w / 2, y, "DE L'ECOLE / CENTRE")
        y -= 8
        
        # Underline
        c.setStrokeColor(colors.HexColor('#1a5276'))
        c.setLineWidth(1.5)
        c.line(left + 30 * mm, y, right - 30 * mm, y)
        y -= 20
        
        # ========== I. IDENTITE DE L'ECOLE ==========
        def draw_section_title(canvas_obj, ypos, number, title):
            canvas_obj.setFont("Helvetica-Bold", 11)
            canvas_obj.setFillColor(colors.HexColor('#1a5276'))
            canvas_obj.drawString(left, ypos, f"{number}.     {title}")
            # Underline
            tw = canvas_obj.stringWidth(f"{number}.     {title}", "Helvetica-Bold", 11)
            canvas_obj.setStrokeColor(colors.HexColor('#1a5276'))
            canvas_obj.setLineWidth(0.8)
            canvas_obj.line(left, ypos - 2, left + tw, ypos - 2)
            return ypos - 18
        
        def draw_field(canvas_obj, ypos, label, value, x_start=None, field_width=None):
            x = x_start if x_start else left
            fw = field_width if field_width else content_width
            canvas_obj.setFont("Helvetica", 8.5)
            canvas_obj.setFillColor(colors.black)
            canvas_obj.drawString(x + 2, ypos, f"-  {label} :")
            label_w = canvas_obj.stringWidth(f"-  {label} : ", "Helvetica", 8.5)
            # Value
            canvas_obj.setFont("Helvetica-Bold", 8.5)
            val = str(value) if value else ''
            canvas_obj.drawString(x + 2 + label_w, ypos, val)
            # Dotted underline
            val_w = canvas_obj.stringWidth(val, "Helvetica-Bold", 8.5)
            dot_start = x + 2 + label_w + val_w + 2
            if dot_start < x + fw - 5:
                canvas_obj.setFont("Helvetica", 7)
                canvas_obj.setFillColor(colors.HexColor('#aaaaaa'))
                dots = '.' * int((x + fw - dot_start) / 2.5)
                canvas_obj.drawString(dot_start, ypos, dots)
                canvas_obj.setFillColor(colors.black)
            return ypos - 14
        
        def draw_field_pair(canvas_obj, ypos, label1, val1, label2, val2):
            half = content_width / 2
            draw_field(canvas_obj, ypos, label1, val1, left, half)
            draw_field(canvas_obj, ypos, label2, val2, left + half, half)
            return ypos - 14
        
        def draw_checkbox_regime(canvas_obj, ypos, regime_val):
            canvas_obj.setFont("Helvetica", 8.5)
            canvas_obj.setFillColor(colors.black)
            canvas_obj.drawString(left + 2, ypos, "-  Régime de Gestion :")
            
            regimes = [
                ('Conventionnée', 'Conventionn'),
                ('Non Conventionnée', 'Non Conventionn'),
                ('Privée Agréée', 'Priv'),
            ]
            x_pos = left + 80
            for label, match_key in regimes:
                # Draw checkbox
                box_size = 7
                canvas_obj.setStrokeColor(colors.black)
                canvas_obj.setLineWidth(0.5)
                canvas_obj.rect(x_pos, ypos - 1, box_size, box_size)
                # Check if matched
                if regime_val and match_key.lower() in regime_val.lower():
                    canvas_obj.setFont("Helvetica-Bold", 8)
                    canvas_obj.setFillColor(colors.HexColor('#1a5276'))
                    canvas_obj.drawString(x_pos + 1.5, ypos, "✓")
                    canvas_obj.setFillColor(colors.black)
                canvas_obj.setFont("Helvetica-Bold", 8.5)
                canvas_obj.drawString(x_pos + box_size + 3, ypos, label)
                x_pos += canvas_obj.stringWidth(label, "Helvetica-Bold", 8.5) + box_size + 18
            return ypos - 16
        
        y = draw_section_title(c, y, "I", "IDENTITE DE L'ECOLE")
        
        y = draw_field(c, y, "Nom de l'Ecole/Centre", etab.nom)
        y = draw_checkbox_regime(c, y, regime_name)
        
        # Code, Matricule, DINACOPE on one line
        third = content_width / 3
        c.setFont("Helvetica", 8.5)
        c.setFillColor(colors.black)
        
        draw_field(c, y, "CODE DE L'ECOLE", etab.code_ecole or etab.code, left, third)
        draw_field(c, y, "Matricule", etab.matricule, left + third, third)
        draw_field(c, y, "N°DINACOPE", etab.no_dinacope, left + 2 * third, third)
        y -= 14
        
        y = draw_field_pair(c, y, "Année de Création", '', "Année d'Agrément", '')
        y = draw_field(c, y, "Référence d'Agrément (Arrêté)", etab.reference_agrement)
        y = draw_field(c, y, "Nom du Gestionnaire/Promoteur", gest_nom)
        y = draw_field(c, y, "Adresse physique de l'Ecole/Centre", etab.adresse)
        
        y = draw_field_pair(c, y, "Division", division, "Sous-Division", sous_division)
        y = draw_field_pair(c, y, "Avenu/Village", '', "Quartier/Groupement", '')
        y = draw_field_pair(c, y, "Commune/Territoire", '', "Ville", '')
        y = draw_field_pair(c, y, "Province", province, "Pays", etab.pays.nom if etab.pays else '')
        
        # Telephone
        half = content_width / 2
        draw_field(c, y, "N° Téléphone", etab.telephone or '', left, half)
        draw_field(c, y, "N° Téléphone 2", '', left + half, half)
        y -= 14
        
        draw_field(c, y, "B.P", etab.boite_postale or '', left, half)
        draw_field(c, y, "Adresse E-mail", etab.email or '', left + half, half)
        y -= 14
        
        # Personnel
        third = content_width / 3
        draw_field(c, y, "Nombre du Personnel Pédagogique", '', left, third)
        draw_field(c, y, "Administratif", '', left + third, third)
        draw_field(c, y, "D'appoint", '', left + 2 * third, third)
        y -= 18
        
        # ========== II. STRUCTURE DE L'ECOLE ==========
        y = draw_section_title(c, y, "II", "STRUCTURE DE L'ECOLE")
        
        # Fetch dynamic data: latest year config for this school
        latest_etab_annee = EtablissementAnnee.objects.filter(
            etablissement=etab, id_pays=etab.pays_id
        ).select_related('annee').order_by('-annee__annee').first()
        
        activated_classes = []
        annee_label = ''
        if latest_etab_annee:
            annee_label = latest_etab_annee.annee.annee
            activated_classes = EtablissementAnneeClasse.objects.filter(
                etablissement_annee=latest_etab_annee
            ).select_related('classe', 'classe__cycle', 'section').order_by(
                'classe__cycle__ordre', 'classe__ordre'
            )
        
        # Group by cycle
        from collections import OrderedDict
        cycles_data = OrderedDict()
        for eac in activated_classes:
            cycle = eac.classe.cycle
            if cycle.pk not in cycles_data:
                cycles_data[cycle.pk] = {
                    'nom': cycle.nom,
                    'ordre': cycle.ordre,
                    'hasSections': cycle.hasSections,
                    'classes': OrderedDict()
                }
            cls = eac.classe
            if cls.id_classe not in cycles_data[cycle.pk]['classes']:
                cycles_data[cycle.pk]['classes'][cls.id_classe] = {
                    'nom': cls.nom,
                    'sections': []
                }
            if eac.section:
                cycles_data[cycle.pk]['classes'][cls.id_classe]['sections'].append(
                    f"{eac.section.code} ({eac.section.nom})"
                )
        
        if annee_label:
            c.setFont("Helvetica-Bold", 8)
            c.setFillColor(colors.HexColor('#2563eb'))
            c.drawString(left + 5, y, f"Configuration pour l'année : {annee_label}")
            c.setFillColor(colors.black)
            y -= 14
        
        if not cycles_data:
            c.setFont("Helvetica-Oblique", 8.5)
            c.setFillColor(colors.HexColor('#888888'))
            c.drawString(left + 10, y, "Aucune classe configurée pour cet établissement.")
            c.setFillColor(colors.black)
            y -= 14
        else:
            for cycle_id, cycle_info in cycles_data.items():
                # Check for page break
                if y < 80:
                    c.showPage()
                    y = h - 25 * mm
                
                # Cycle header
                c.setFont("Helvetica-Bold", 9)
                c.setFillColor(colors.HexColor('#1a5276'))
                c.drawString(left + 5, y, f"❖ {cycle_info['nom'].upper()}")
                nb_classes = len(cycle_info['classes'])
                c.setFont("Helvetica", 8)
                c.drawString(left + 5 + c.stringWidth(f"❖ {cycle_info['nom'].upper()}", "Helvetica-Bold", 9) + 5, y, 
                           f"({nb_classes} classe{'s' if nb_classes > 1 else ''})")
                c.setFillColor(colors.black)
                y -= 14
                
                for cls_id, cls_info in cycle_info['classes'].items():
                    if y < 60:
                        c.showPage()
                        y = h - 25 * mm
                    
                    c.setFont("Helvetica", 8)
                    c.drawString(left + 15, y, "○")
                    c.setFont("Helvetica-Bold", 8.5)
                    c.drawString(left + 25, y, cls_info['nom'])
                    
                    if cls_info['sections']:
                        # Show sections on same line or next lines
                        sections_text = ", ".join(cls_info['sections'])
                        c.setFont("Helvetica", 7.5)
                        c.setFillColor(colors.HexColor('#555555'))
                        
                        # Calculate available width
                        cls_width = c.stringWidth(cls_info['nom'], "Helvetica-Bold", 8.5)
                        avail = content_width - 30 - cls_width - 10
                        sec_label = f"  Sections: {sections_text}"
                        
                        if c.stringWidth(sec_label, "Helvetica", 7.5) <= avail:
                            c.drawString(left + 25 + cls_width + 5, y, sec_label)
                            y -= 13
                        else:
                            y -= 12
                            # Wrap on multiple lines if needed
                            words = sections_text.split(", ")
                            line = "Sections: "
                            max_w = content_width - 40
                            for i, word in enumerate(words):
                                test_line = line + word + (", " if i < len(words) - 1 else "")
                                if c.stringWidth(test_line, "Helvetica", 7.5) > max_w and line != "Sections: ":
                                    c.drawString(left + 35, y, line.rstrip(", "))
                                    y -= 10
                                    if y < 60:
                                        c.showPage()
                                        y = h - 25 * mm
                                    line = word + (", " if i < len(words) - 1 else "")
                                else:
                                    line = test_line
                            if line:
                                c.drawString(left + 35, y, line.rstrip(", "))
                                y -= 10
                        c.setFillColor(colors.black)
                    else:
                        y -= 13
                
                y -= 4  # gap between cycles
        
        y -= 5
        
        # ========== III. DIRECTEUR ==========
        y = draw_section_title(c, y, "III", "DIRECTEUR (TRICE) / PREFET DES ETUDES")
        
        y = draw_field(c, y, "Noms", '')
        draw_field(c, y, "Nationalité", '', left, half)
        draw_field(c, y, "Tél", '', left + half, half)
        y -= 14
        y = draw_field(c, y, "Adresse E-mail", '')
        
        y -= 25
        
        # Signature
        c.setFont("Helvetica-Oblique", 9)
        c.setFillColor(colors.HexColor('#1a5276'))
        c.drawRightString(right, y, "(Signature et sceau du responsable)")
        
        y -= 30
        
        # Footnotes
        c.setFont("Helvetica", 6.5)
        c.setFillColor(colors.HexColor('#666666'))
        c.setStrokeColor(colors.HexColor('#666666'))
        c.setLineWidth(0.5)
        c.line(left, y + 5, left + 60 * mm, y + 5)
        c.drawString(left, y - 3, "¹ En annexe les attributions du personnel (Enseignants, Administratif)")
        y -= 10
        c.drawString(left, y - 3, "² En annexe la liste des effectifs par niveau, section et classe")
        
        c.showPage()
        c.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        safe_name = etab.nom.replace(' ', '_').replace("'", "")[:40]
        response['Content-Disposition'] = f'inline; filename="Fiche_Synoptique_{safe_name}.pdf"'
        return response
        
    except Etablissement.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Établissement introuvable'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@require_http_methods(["GET"])
def get_etablissements(request):
    """API pour récupérer les établissements avec filtres optionnels."""
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays: return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        qs = Etablissement.objects.filter(pays=pays).select_related(
            'structure_pedagogique', 'structure_pedagogique__administrative_parent', 'gestionnaire'
        ).order_by('nom')
        
        # Filtrage par scope utilisateur (RBAC multi-structures)
        scope_codes = get_all_user_scopes(request)
        if scope_codes:
            from django.db.models import Q
            scope_q = Q()
            for sc in scope_codes:
                scope_q |= Q(structure_pedagogique__code__startswith=sc)
            qs = qs.filter(scope_q)
        
        # Optional filters
        regime_id = request.GET.get('regime_id')
        if regime_id:
            qs = qs.filter(id_regime=regime_id)
        
        struct_ped_id = request.GET.get('structure_ped_id')
        if struct_ped_id:
            qs = qs.filter(structure_pedagogique_id=struct_ped_id)
        
        struct_adm_id = request.GET.get('structure_adm_id')
        if struct_adm_id:
            qs = qs.filter(structure_pedagogique__administrative_parent_id=struct_adm_id)
        
        gestionnaire_id = request.GET.get('gestionnaire_id')
        if gestionnaire_id:
            qs = qs.filter(gestionnaire_id=gestionnaire_id)
            
        parent_code = request.GET.get('parent_code')
        if parent_code:
            qs = qs.filter(structure_pedagogique__code__startswith=parent_code)
        
        results = []
        for e in qs:
            results.append({
                'id_etablissement': e.id_etablissement,
                'nom': e.nom,
                'code': e.code,
                'code_ecole': e.code_ecole or '',
                'matricule': e.matricule or '',
                'no_dinacope': e.no_dinacope or '',
                'reference_agrement': e.reference_agrement or '',
                'adresse': e.adresse or '',
                'email': e.email or '',
                'telephone': e.telephone or '',
                'boite_postale': e.boite_postale or '',
                'admin_email': e.admin_email or '',
                'admin_telephone': e.admin_telephone or '',
                'admin_email_verified': e.admin_email_verified,
                'admin_phone_verified': e.admin_phone_verified,
                'db_server': e.db_server or '',
                'db_name': e.db_name or '',
                'db_user': e.db_user or '',
                'db_password': e.db_password or '',
                'url': e.url or '',
                'regime_id': e.id_regime,
                'regime': Regime.objects.filter(id_regime=e.id_regime).values_list('regime', flat=True).first() or '-',
                'parent_ped_id': e.structure_pedagogique_id,
                'parent_ped': e.structure_pedagogique.nom if e.structure_pedagogique else '-',
                'parent_ped_code': e.structure_pedagogique.code if e.structure_pedagogique else '',
                'parent_adm_id': e.structure_administrative.id_structure if e.structure_administrative else None,
                'parent_adm': e.structure_administrative.nom if e.structure_administrative else '-',
                'gestionnaire_id': e.gestionnaire_id if e.gestionnaire else None,
                'gestionnaire_nom': f"{e.gestionnaire.nom} {e.gestionnaire.postnom}" if e.gestionnaire else '',
                'gestionnaire_prenom': e.gestionnaire.nom if e.gestionnaire else '',
                'gestionnaire_postnom': e.gestionnaire.postnom if e.gestionnaire else '',
                'gestionnaire_email': e.gestionnaire.email if e.gestionnaire else '',
                'gestionnaire_tel': e.gestionnaire.telephone if e.gestionnaire else '',
                'created_at': e.created_at.strftime('%d/%m/%Y') if e.created_at else '',
                'updated_at': e.updated_at.strftime('%d/%m/%Y') if e.updated_at else '',
            })
            
        return JsonResponse({'success': True, 'etablissements': results, 'count': len(results)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# already imported at top


@require_http_methods(["GET"])
def get_gestionnaires(request):
    """Retourne tous les gestionnaires, avec le nombre d'établissements associés (optionnellement filtrés par régime)."""
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        # Base queryset des établissements pour compter
        etab_qs = Etablissement.objects.filter(pays=pays)
        
        regime_id = request.GET.get('regime_id')
        if regime_id:
            etab_qs = etab_qs.filter(id_regime=regime_id)
        
        # Retourner TOUS les gestionnaires (ce sont des types de gestion, pas des personnes)
        gestionnaires = GestionnaireEtablissement.objects.all().order_by('nom')
        
        results = [{
            'id_gestionnaire': g.id_gestionnaire,
            'nom': g.nom,
            'postnom': g.postnom or '',
            'label': g.nom.strip(),
            'n_ecoles': etab_qs.filter(gestionnaire_id=g.id_gestionnaire).count(),
        } for g in gestionnaires]
        
        return JsonResponse({'success': True, 'gestionnaires': results})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def test_db_connection(request):
    """Teste la connexion à une base de données distante."""
    try:
        data = json.loads(request.body)
        db_server = (data.get('db_server') or '').strip()
        db_name = (data.get('db_name') or '').strip()
        db_user = (data.get('db_user') or '').strip()
        db_password = (data.get('db_password') or '').strip()
        
        if not db_server or not db_name or not db_user:
            return JsonResponse({'success': False, 'error': 'Serveur, nom de la base et utilisateur sont requis.'})
        
        import pymysql
        import socket
        
        # Resolve the server to check if it's the local machine
        connect_host = db_server
        try:
            server_ip = socket.gethostbyname(db_server)
            local_ip = socket.gethostbyname(socket.gethostname())
            # If the target is this same server, use localhost for MySQL auth
            if server_ip == local_ip or server_ip == '127.0.0.1' or db_server in ('localhost', '127.0.0.1'):
                connect_host = '127.0.0.1'
        except socket.gaierror:
            pass  # Keep original host if DNS fails
        
        # Try connection with resolved host, fallback to localhost
        last_error = None
        for host in [connect_host, '127.0.0.1'] if connect_host != '127.0.0.1' else ['127.0.0.1']:
            try:
                connection = pymysql.connect(
                    host=host,
                    database=db_name,
                    user=db_user,
                    password=db_password,
                    port=3306,
                    connect_timeout=10
                )
                connection.close()
                return JsonResponse({'success': True, 'message': 'Connexion réussie !'})
            except Exception as e:
                last_error = e
                continue
        
        raise last_error
    except Exception as e:
        error_msg = str(e)
        # Simplify common error messages
        if 'Access denied' in error_msg:
            error_msg = "Accès refusé — vérifiez l'utilisateur et le mot de passe."
        elif 'Unknown database' in error_msg:
            error_msg = "Base de données introuvable — vérifiez le nom."
        elif 'timed out' in error_msg or 'Can\'t connect' in error_msg:
            error_msg = "Impossible de se connecter au serveur — vérifiez l'adresse."
        return JsonResponse({'success': False, 'error': error_msg})


@require_http_methods(["POST"])
def save_etablissement(request):
    """
    Crée ou met à jour un établissement.
    Code: id_etab - id_regime - id_gestionnaire - PD - codePed
    """
    import re
    
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays')
        pays = get_object_or_404(Pays, id_pays=id_pays)
        id_etablissement = data.get('id_etablissement')
        
        # Check if updating existing
        etab = None
        if id_etablissement:
            etab = Etablissement.objects.filter(id_etablissement=id_etablissement, pays=pays).first()
        
        # 1. Gestionnaire (optionnel - par ID)
        gestionnaire = None
        id_gestionnaire = data.get('id_gestionnaire')
        if id_gestionnaire:
            try:
                gestionnaire = GestionnaireEtablissement.objects.get(id_gestionnaire=int(id_gestionnaire))
            except (GestionnaireEtablissement.DoesNotExist, ValueError):
                pass  # Si non trouvé, on laisse None
        
        # 2. Admin System Fields (optionnels)
        admin_email = (data.get('admin_email') or '').strip()
        admin_telephone = (data.get('admin_telephone') or '').strip()
        
        # Valider le format téléphone seulement s'il est fourni
        if admin_telephone:
            phone_pattern = r'^\+[0-9]{10,15}$'
            if not re.match(phone_pattern, admin_telephone):
                return JsonResponse({
                    'success': False, 
                    'error': "Format téléphone invalide. Utilisez le format international (+XXX...)."
                }, status=400)
            
        # 3. Parents & Regime
        id_regime = data.get('id_regime')
        id_parent_ped = data.get('id_parent_ped')
        
        if not id_parent_ped:
            return JsonResponse({'success': False, 'error': "La structure pédagogique parente est obligatoire."}, status=400)

        try:
            id_parent_ped = int(id_parent_ped)
            if id_regime:
                id_regime = int(id_regime)
        except (ValueError, TypeError):
             return JsonResponse({'success': False, 'error': "IDs invalides."}, status=400)

        try:
            parent_ped = PedagogicStructureInstance.objects.get(id_structure=id_parent_ped)
        except PedagogicStructureInstance.DoesNotExist:
             return JsonResponse({'success': False, 'error': "Structure pédagogique introuvable."}, status=400)
        
        if etab:
            # UPDATE existing
            etab.nom = data.get('nom')
            etab.id_regime = id_regime  # simple int, peut être None
            etab.structure_pedagogique = parent_ped
            etab.gestionnaire = gestionnaire
            etab.code_ecole = data.get('code_ecole', '')
            etab.matricule = data.get('matricule', '')
            etab.no_dinacope = data.get('no_dinacope', '')
            etab.reference_agrement = data.get('reference_agrement', '')
            etab.adresse = data.get('adresse', '')
            etab.email = data.get('email', '')
            etab.telephone = data.get('telephone', '')
            etab.boite_postale = data.get('boite_postale', '')
            etab.admin_email = admin_email
            etab.admin_telephone = admin_telephone
            etab.url = (data.get('url') or '').strip() or None
            etab.db_server = (data.get('db_server') or '').strip() or None
            etab.db_name = (data.get('db_name') or '').strip() or None
            etab.db_user = (data.get('db_user') or '').strip() or None
            etab.db_password = (data.get('db_password') or '').strip() or None
            
            # Code is automatically regenerated by model.save() on every save
            etab.save()
            
            return JsonResponse({'success': True, 'code': etab.code, 'updated': True})
        else:
            # CREATE new — le code est auto-généré par le model.save()
            etab = Etablissement.objects.create(
                pays=pays,
                nom=data.get('nom'),
                id_regime=id_regime,
                structure_pedagogique=parent_ped,
                gestionnaire=gestionnaire,
                code_ecole=data.get('code_ecole', ''),
                matricule=data.get('matricule', ''),
                no_dinacope=data.get('no_dinacope', ''),
                reference_agrement=data.get('reference_agrement', ''),
                adresse=data.get('adresse', ''),
                email=data.get('email', ''),
                telephone=data.get('telephone', ''),
                boite_postale=data.get('boite_postale', ''),
                admin_email=admin_email,
                admin_telephone=admin_telephone,
                admin_email_verified=False,
                admin_phone_verified=False,
                url=(data.get('url') or '').strip() or None,
                db_server=(data.get('db_server') or '').strip() or None,
                db_name=(data.get('db_name') or '').strip() or None,
                db_user=(data.get('db_user') or '').strip() or None,
                db_password=(data.get('db_password') or '').strip() or None,
            )
            
            return JsonResponse({'success': True, 'code': etab.code, 'created': True})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@require_http_methods(["POST"])
def save_regime(request):
    """API pour créer ou modifier un régime."""
    try:
        data = json.loads(request.body)
        id_regime = data.get('id_regime')
        id_pays = data.get('id_pays')
        nom_regime = data.get('regime')
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        if id_regime:
            regime_obj = get_object_or_404(Regime, id_regime=id_regime)
            regime_obj.regime = nom_regime
            regime_obj.save()
        else:
            Regime.objects.create(pays=pays, regime=nom_regime)
            
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_regime(request):
    """API pour supprimer un régime."""
    try:
        data = json.loads(request.body)
        id_regime = data.get('id_regime')
        regime_obj = get_object_or_404(Regime, id_regime=id_regime)
        regime_obj.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# --- NIVEAUX ET PROGRAMMES ---

@require_http_methods(["GET"])
def etablissements_view(request):
    """Vue pour la gestion des Établissements, Cycles et Niveaux d'étude."""
    context = get_country_context_logic(request)
    context['active_page'] = 'niveaux'
    return render(request, 'structure_app/etablissements.html', context)


from django.shortcuts import render, get_object_or_404, redirect

@require_http_methods(["GET"])
def programmes_view(request):
    """Vue pour la gestion des Programmes (Catalogue et Config Annuelle)."""
    context = get_country_context_logic(request)
    context['active_page'] = 'programmes'
    return render(request, 'structure_app/programmes.html', context)


@require_http_methods(["GET"])
def get_cycles_data(request):
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'cycles': []})
            
        cycles = pays.cycles.all().order_by('ordre')
        
        
        
        results = []
        for c in cycles:
            # Classes belong directly to cycles — cycle FK already scopes by country
            all_classes_qs = Classe.objects.filter(cycle_id=c.id).order_by('ordre')
            all_classes = [{'id_classe': c2.id_classe, 'nom': c2.classe, 'ordre': c2.ordre} for c2 in all_classes_qs]
            # Respect the duree limit: only return up to 'duree' classes per cycle
            if c.duree and c.duree > 0:
                classes = list(all_classes[:c.duree])
            else:
                classes = list(all_classes)
            
            cycle_data = {
                'id_cycle': c.id_cycle,
                'nom': c.nom,
                'ordre': c.ordre,
                'duree': c.duree,
                'hasSections': bool(c.hasSections),
                'coursUniformes': bool(c.coursUniformes),
                'labelSection_id': c.labelSection_id if c.labelSection_id else None,
                'labelSection_nom': c.labelSection.nom if c.labelSection else 'Sections',
                'classes': classes,
                'total_classes': len(all_classes)
            }
            
            # For cycles with sections, include sections filtered by type
            if c.hasSections:
                # Filter sections by the cycle's labelSection type + pays
                sec_filters = {'id_pays': pays.id_pays}
                if c.labelSection_id:
                    sec_filters['type_subdivision_id'] = c.labelSection_id
                cycle_sections = Section.objects.filter(**sec_filters).values('id_section', 'nom', 'code')
                
                sections_with_classes = []
                for section in cycle_sections:
                    sections_with_classes.append({
                        'id_section': section['id_section'],
                        'nom': section['nom'],
                        'code': section['code'],
                        'classes': classes  # All cycle classes are available in each section
                    })
                cycle_data['sections'] = sections_with_classes
            else:
                cycle_data['sections'] = []
            
            results.append(cycle_data)
        return JsonResponse({'success': True, 'cycles': results})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logging.getLogger(__name__).error(f"Error in get_cycles_data: {str(e)}\n{tb}")
        return JsonResponse({
            'success': False, 
            'error': f"Erreur lors du chargement des cycles: {str(e)}",
            'traceback': tb if settings.DEBUG else None
        }, status=500)


@require_http_methods(["POST"])
def save_cycle(request):
    try:
        data = json.loads(request.body)
        id_cycle = data.get('id_cycle')
        id_pays = data.get('id_pays')
        nom = data.get('nom')
        ordre = data.get('ordre', 1)
        duree = data.get('duree', 1)
        hasSections = data.get('hasSections', False)
        coursUniformes = data.get('coursUniformes', True)
        labelSection_id = data.get('labelSection_id')
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        if id_cycle:
            cycle = get_object_or_404(Cycle, id_cycle=id_cycle, pays_id=pays.id_pays)
            cycle.nom = nom
            cycle.ordre = ordre
            cycle.duree = duree
            cycle.hasSections = hasSections
            cycle.coursUniformes = coursUniformes
            cycle.labelSection_id = int(labelSection_id) if labelSection_id else None
            cycle.save()
        else:
            Cycle.objects.create(
                pays=pays, nom=nom, ordre=ordre, duree=duree,
                hasSections=hasSections, coursUniformes=coursUniformes,
                labelSection_id=int(labelSection_id) if labelSection_id else None
            )
            
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_cycle(request):
    try:
        data = json.loads(request.body)
        id_cycle = data.get('id_cycle')
        id_pays = data.get('id_pays')
        pays = get_object_or_404(Pays, id_pays=id_pays)
        get_object_or_404(Cycle, id_cycle=id_cycle, pays_id=pays.id_pays).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_section(request):
    """API pour créer ou modifier une section/filière."""
    try:
        data = json.loads(request.body)
        id_section = data.get('id_section')
        nom = data.get('nom')
        code = data.get('code')
        type_subdivision_id = data.get('type_subdivision_id')
        
        id_pays = data.get('id_pays')
        if id_section:
            s = get_object_or_404(Section, id_section=id_section)
            s.nom = nom
            s.code = code
            if type_subdivision_id:
                s.type_subdivision_id = int(type_subdivision_id)
            s.save()
        else:
            Section.objects.create(
                nom=nom, code=code,
                type_subdivision_id=int(type_subdivision_id) if type_subdivision_id else None,
                id_pays=id_pays
            )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_http_methods(["POST"])
def delete_section(request):
    try:
        data = json.loads(request.body)
        id_section = data.get('id_section')
        get_object_or_404(Section, id_section=id_section).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def download_section_template(request):
    """Télécharge un modèle Excel pour importer des sections/filières."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    label = request.GET.get('label', 'Sections')
    type_id = request.GET.get('type_id', '')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = label

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Code', 'Nom']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Example rows
    singular = label.rstrip('s')
    examples = [
        ('EX01', f'Exemple {singular} 1'),
        ('EX02', f'Exemple {singular} 2'),
    ]
    for row_idx, (code, nom) in enumerate(examples, 2):
        ws.cell(row=row_idx, column=1, value=code).font = Font(color="999999", italic=True)
        ws.cell(row=row_idx, column=2, value=nom).font = Font(color="999999", italic=True)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Modele_{label}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@require_http_methods(["POST"])
def import_sections_excel(request):
    """Importe des sections/filières depuis un fichier Excel."""
    import openpyxl
    try:
        file = request.FILES.get('file')
        type_subdivision_id = request.POST.get('type_subdivision_id')

        if not file:
            return JsonResponse({'success': False, 'error': 'Aucun fichier fourni.'})

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        count = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            code = str(row[0]).strip() if row[0] else ''
            nom = str(row[1]).strip() if row[1] else ''

            if not code or not nom:
                continue
            # Skip example rows
            if code.startswith('EX0') or nom.startswith('Exemple'):
                continue

            try:
                id_pays = request.POST.get('id_pays')
                Section.objects.update_or_create(
                    code=code, id_pays=id_pays,
                    defaults={
                        'nom': nom,
                        'type_subdivision_id': int(type_subdivision_id) if type_subdivision_id else None
                    }
                )
                count += 1
            except Exception as e:
                errors.append(f"Ligne {row_idx}: {str(e)}")

        result = {'success': True, 'count': count}
        if errors:
            result['warnings'] = errors
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ============================================================
# TypeSubdivision — CRUD for dynamic section labels
# ============================================================

def get_type_subdivisions(request):
    """Liste les appellations de subdivisions pour un pays donné."""
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'types': []})
        types = list(TypeSubdivision.objects.filter(pays=pays).values('id_type', 'nom'))
        return JsonResponse({'success': True, 'types': types})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_type_subdivision(request):
    """Créer ou modifier une appellation de subdivision."""
    try:
        data = json.loads(request.body)
        id_type = data.get('id_type')
        nom = data.get('nom', '').strip()
        id_pays = data.get('id_pays')
        if not nom or not id_pays:
            return JsonResponse({'success': False, 'error': 'Nom et id_pays requis.'}, status=400)
        pays = get_object_or_404(Pays, id_pays=id_pays)
        if id_type:
            t = get_object_or_404(TypeSubdivision, id_type=id_type, pays=pays)
            t.nom = nom
            t.save()
        else:
            # Auto-assign id_type per-country
            from django.db.models import Max
            max_id = TypeSubdivision.objects.filter(pays=pays).aggregate(m=Max('id_type'))['m'] or 0
            t = TypeSubdivision.objects.create(pays=pays, nom=nom, id_type=max_id + 1)
        return JsonResponse({'success': True, 'id_type': t.id_type, 'nom': t.nom})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_type_subdivision(request):
    """Supprimer une appellation de subdivision."""
    try:
        data = json.loads(request.body)
        id_type = data.get('id_type')
        id_pays = data.get('id_pays')
        pays = get_object_or_404(Pays, id_pays=id_pays) if id_pays else None
        filters = {'id_type': id_type}
        if pays:
            filters['pays'] = pays
        get_object_or_404(TypeSubdivision, **filters).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_http_methods(["POST"])
def save_classe(request):
    """API pour créer ou modifier une classe."""
    try:
        data = json.loads(request.body)
        id_classe = data.get('id_classe')
        id_cycle = data.get('id_cycle')
        nom = data.get('nom')
        ordre = data.get('ordre', 1)
        
        cycle = get_object_or_404(Cycle, id_cycle=id_cycle, pays_id=pays.id_pays)

        if id_classe:
            c = get_object_or_404(Classe, id_classe=id_classe)
            c.nom = nom
            c.ordre = ordre
            c.save()
        else:
            # Vérifier que le nombre de classes ne dépasse pas la durée du cycle
            current_count = cycle.classes.count()
            if cycle.duree and current_count >= cycle.duree:
                return JsonResponse({
                    'success': False, 
                    'error': f'Le cycle "{cycle.nom}" a déjà atteint son maximum de {cycle.duree} classe(s). '
                             f'Modifiez la durée du cycle pour en ajouter davantage.'
                }, status=400)
            Classe.objects.create(cycle=cycle, nom=nom, ordre=ordre, id_pays=pays.id_pays)
            
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_classe(request):
    try:
        data = json.loads(request.body)
        id_classe = data.get('id_classe')
        get_object_or_404(Classe, id_classe=id_classe).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_programmes_data(request):
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'programmes': []})
            
        programmes = pays.programmes.all().order_by('nom')
        return JsonResponse({'success': True, 'programmes': list(programmes.values('id_programme', 'nom'))})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logging.getLogger(__name__).error(f"Error in get_programmes_data: {str(e)}\n{tb}")
        return JsonResponse({
            'success': False, 
            'error': f"Erreur lors du chargement des programmes: {str(e)}",
            'traceback': tb if settings.DEBUG else None
        }, status=500)


@require_http_methods(["POST"])
def save_programme(request):
    try:
        data = json.loads(request.body)
        id_programme = data.get('id_programme')
        id_pays = data.get('id_pays')
        nom = data.get('nom')
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        if id_programme:
            p = get_object_or_404(Programme, id_programme=id_programme)
            p.nom = nom
            p.save()
        else:
            Programme.objects.create(pays=pays, nom=nom)
            
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_programme(request):
    try:
        data = json.loads(request.body)
        id_programme = data.get('id_programme')
        get_object_or_404(Programme, id_programme=id_programme).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# --- CRUD DOMAINES ---
# ============================================================

@require_http_methods(["GET"])
def get_domaines_data(request):
    """Liste les domaines pour un pays donné."""
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'domaines': []})
        domaines = list(Domaine.objects.filter(pays=pays).values(
            'id_domaine', 'nom', 'code'
        ))
        return JsonResponse({'success': True, 'domaines': domaines})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_domaine(request):
    """Crée ou modifie un domaine."""
    try:
        data = json.loads(request.body)
        id_domaine = data.get('id_domaine')
        nom = data.get('nom', '').strip()
        code = data.get('code', '').strip()
        sigle = data.get('sigle', '').strip()

        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        if not nom:
            return JsonResponse({'success': False, 'error': 'Le nom du domaine est requis.'}, status=400)

        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': False, 'error': 'Pays introuvable.'}, status=400)

        if id_domaine:
            d = get_object_or_404(Domaine, id_domaine=id_domaine)
            d.nom = nom
            d.code = code
            d.save()
        else:
            if Domaine.objects.filter(pays=pays, nom=nom).exists():
                return JsonResponse({'success': False, 'error': f'Le domaine "{nom}" existe déjà.'}, status=400)
            Domaine.objects.create(nom=nom, code=code, pays=pays)

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_domaine(request):
    """Supprime un domaine s'il n'est pas référencé."""
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        domaine = get_object_or_404(Domaine, id_domaine=data.get('id_domaine'))

        # Vérifier s'il est utilisé dans des cours ou cours_annee
        cours_count = Cours.objects.filter(domaine=domaine, id_pays=id_pays).count()
        ca_count = CoursAnnee.objects.filter(domaine=domaine, id_pays=id_pays).count()
        if cours_count > 0 or ca_count > 0:
            return JsonResponse({
                'success': False,
                'error': f'Ce domaine est utilisé par {cours_count} cours et {ca_count} config(s) annuelle(s). Impossible de le supprimer.'
            }, status=400)

        domaine.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def import_domaines_excel(request):
    """Import de domaines depuis un fichier Excel."""
    try:
        import openpyxl
        file = request.FILES.get('file')
        id_pays = request.POST.get('id_pays', '')

        if not file or not id_pays:
            return JsonResponse({'success': False, 'error': 'Fichier et id_pays requis.'}, status=400)

        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': False, 'error': 'Pays introuvable.'}, status=400)

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        created = 0
        errors = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nom = str(row[0]).strip() if row[0] else ''
            code = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if not nom:
                continue
            try:
                _, was_created = Domaine.objects.get_or_create(
                    pays=pays, nom=nom, defaults={'code': code}
                )
                if was_created:
                    created += 1
            except Exception as e:
                errors.append(f"Ligne {i}: {str(e)}")

        return JsonResponse({'success': True, 'count': created, 'errors': errors})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# --- CRUD COURS (Catalogue) ---
# ============================================================

@require_http_methods(["GET"])
def get_cours_data(request):
    """Liste les cours, filtré par pays (id_pays) et optionnellement par classe+section.
    Pour les cycles à sections, chaque classe est dupliquée par section dans le dropdown.
    """
    try:
        id_pays = request.GET.get('id_pays')
        id_classe = request.GET.get('id_classe')
        id_section = request.GET.get('id_section')  # NEW: optional section filter

        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})

        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'cours': [], 'classes': []})

        # Récupérer les sections globales (pour les cycles à sections)
        all_sections = list(Section.objects.filter(pays_id=id_pays).order_by('nom').values('id_section', 'nom', 'code'))

        # Récupérer les classes activées pour cet établissement (via EAC)
        id_etablissement = getattr(request, 'id_etablissement', None) or request.session.get('id_etablissement')
        activated_classe_ids = set()
        if id_etablissement:
            annee_active = Annee.objects.filter(pays_id=pays.id_pays, isOpen=True).order_by('-annee').first()
            if annee_active:
                etab_annee = EtablissementAnnee.objects.filter(
                    etablissement_id=id_etablissement, annee=annee_active, id_pays=pays.id_pays
                ).first()
                if etab_annee:
                    activated_classe_ids = set(
                        EtablissementAnneeClasse.objects.filter(
                            etablissement_annee=etab_annee
                        ).values_list('classe_id', flat=True)
                    )

        # Récupérer les cycles qui ont au moins une classe activée
        if activated_classe_ids:
            activated_cycle_ids = set(
                Classe.objects.filter(id__in=activated_classe_ids).values_list('cycle_id', flat=True)
            )
            cycles = Cycle.objects.filter(id__in=activated_cycle_ids).order_by('ordre')
        else:
            cycles = pays.cycles.all().order_by('ordre')

        classes_data = []
        for cycle in cycles:
            if activated_classe_ids:
                cycle_classes = Classe.objects.filter(
                    id__in=activated_classe_ids, cycle_id=cycle.id
                ).order_by('ordre')
            else:
                cycle_classes = Classe.objects.filter(cycle_id=cycle.id).order_by('ordre')
            if cycle.duree and cycle.duree > 0:
                cycle_classes = cycle_classes[:cycle.duree]
            
            if cycle.hasSections and all_sections:
                # Cycle à sections: dupliquer chaque classe par section
                for section in all_sections:
                    for cls in cycle_classes:
                        classes_data.append({
                            'id_classe': cls.id_classe,
                            'id_section': section['id_section'],
                            'nom': f"{cls.nom} — {section['nom']}",
                            'nom_classe': cls.nom,
                            'nom_section': section['nom'],
                            'code_section': section['code'],
                            'cycle_nom': f"{cycle.nom} / {section['nom']}",
                            'cycle_id': cycle.id,
                            'has_sections': True,
                        })
            else:
                # Cycle normal: une entrée par classe
                for cls in cycle_classes:
                    classes_data.append({
                        'id_classe': cls.id_classe,
                        'id_section': None,
                        'nom': cls.nom,
                        'nom_classe': cls.nom,
                        'nom_section': None,
                        'code_section': None,
                        'cycle_nom': cycle.nom,
                        'cycle_id': cycle.id,
                        'has_sections': False,
                    })

        # Récupérer les cours avec le domaine et section FK
        cours_qs = Cours.objects.filter(classe__cycle__pays=pays)
        if id_classe:
            cours_qs = cours_qs.filter(classe__id_classe=id_classe)
            # Filtrer par section si spécifié
            if id_section:
                cours_qs = cours_qs.filter(section_id=id_section)
            else:
                # Si pas de section spécifiée, ne montrer que les cours sans section
                cours_qs = cours_qs.filter(section_id__isnull=True)

        # Build lookup maps for domaine/section (IntegerFields on Cours)
        _domaine_map = {}
        if Domaine:
            _domaine_map = {d['id_domaine']: d['nom'] for d in Domaine.objects.filter(pays=pays).values('id_domaine', 'nom')}
        _section_map = {}
        if Section:
            _section_map = {s['id_section']: s['nom'] for s in Section.objects.filter(id_pays=pays.id_pays).values('id_section', 'nom')}

        cours_data = []
        for c in cours_qs:
            cours_data.append({
                'id_cours': c.id_cours,
                'cours': c.cours,
                'code_cours': c.code_cours,
                'domaine_id': c.domaine_id,
                'domaine_nom': _domaine_map.get(c.domaine_id, ''),
                'id_classe_id': c.classe_id,
                'section_id': c.section_id,
                'section_nom': _section_map.get(c.section_id, ''),
            })

        # Récupérer les domaines du pays
        domaines = list(Domaine.objects.filter(pays=pays).values('id_domaine', 'nom', 'code'))

        return JsonResponse({
            'success': True,
            'cours': cours_data,
            'classes': classes_data,
            'domaines': domaines,
            'sections': all_sections,
        })
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logging.getLogger(__name__).error(f"Error in get_cours_data: {str(e)}\n{tb}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_cours(request):
    """Crée ou modifie un cours."""
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        id_cours = data.get('id_cours')
        code_cours = data.get('code_cours', '').strip()
        nom_cours = data.get('cours', '').strip()
        domaine_id = data.get('domaine_id')  # FK maintenant
        id_classe = data.get('id_classe')
        id_section = data.get('id_section')  # NEW: optional section FK

        if not code_cours or not nom_cours or not id_classe:
            return JsonResponse({'success': False, 'error': 'Code, nom du cours et classe sont requis.'}, status=400)

        classe = get_object_or_404(Classe, id_classe=id_classe)
        if not id_pays:
            id_pays = classe.id_pays
        domaine = None
        if domaine_id:
            domaine = get_object_or_404(Domaine, id_domaine=domaine_id)
        section = None
        if id_section:
            section = get_object_or_404(Section, id_section=id_section)

        if id_cours:
            c = get_object_or_404(Cours, id_cours=id_cours)
            c.cours = nom_cours
            c.code_cours = code_cours
            c.domaine = domaine
            c.classe = classe
            c.section = section
            c.save()
        else:
            if Cours.objects.filter(classe=classe, section=section, code_cours=code_cours, id_pays=id_pays).exists():
                section_label = f" (section {section.nom})" if section else ""
                return JsonResponse({'success': False, 'error': f'Le code "{code_cours}" existe déjà pour cette classe{section_label}.'}, status=400)
            Cours.objects.create(
                cours=nom_cours,
                code_cours=code_cours,
                domaine=domaine,
                classe=classe,
                section=section,
                id_pays=id_pays
            )

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_cours(request):
    """Supprime un cours."""
    try:
        data = json.loads(request.body)
        id_cours = data.get('id_cours')
        get_object_or_404(Cours, id_cours=id_cours).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def download_cours_template(request):
    """Génère un modèle Excel pour importer des cours dans une classe."""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'openpyxl non installé.'}, status=500)
    
    try:
        id_classe = request.GET.get('id_classe')
        id_section = request.GET.get('id_section')
        id_pays = request.GET.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        classe = get_object_or_404(Classe, id_classe=id_classe)
        if not id_pays:
            id_pays = classe.id_pays
        
        wb = openpyxl.Workbook()
        ws = wb.active
        section_label = ''
        if id_section:
            sec = Section.objects.filter(id_section=id_section).first()
            section_label = f"_{sec.nom}" if sec else ''
        ws.title = f"Cours_{classe.nom}{section_label}"[:31]
        
        # Build domaine map for reverse lookup
        domaine_map = {}
        if Domaine:
            pays_obj = Pays.objects.filter(id_pays=id_pays).first()
            if pays_obj:
                domaine_map = {d['id_domaine']: d['nom'] for d in Domaine.objects.filter(pays=pays_obj).values('id_domaine', 'nom')}
        
        # En-têtes
        headers = ['CODE_COURS', 'COURS', 'DOMAINE']
        ws.append(headers)
        
        # Pré-remplir avec les cours existants de cette classe/section
        existing_cours = Cours.objects.filter(classe__id_classe=id_classe, id_pays=id_pays).order_by('code_cours')
        if id_section:
            existing_cours = existing_cours.filter(section_id=id_section)
        else:
            existing_cours = existing_cours.filter(section_id__isnull=True)
        
        if existing_cours.exists():
            for c in existing_cours:
                dom_name = domaine_map.get(c.domaine_id, '') if c.domaine_id else ''
                ws.append([c.code_cours or '', c.cours or '', dom_name])
        else:
            # Lignes d'exemple si pas de cours existants
            ws.append(['MATH01', 'Mathématiques', 'Sciences'])
            ws.append(['FR01', 'Français', 'Langues'])
        
        # Instructions
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1, value="Instructions:")
        ws.cell(row=last_row + 1, column=1, value="1. Modifiez ou ajoutez des cours à partir de la ligne 2.")
        ws.cell(row=last_row + 2, column=1, value="2. Le CODE_COURS est la clé unique — les cours existants seront mis à jour.")
        ws.cell(row=last_row + 3, column=1, value="3. Le DOMAINE est optionnel.")
        
        # Largeurs de colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        
        # Style en-têtes
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Modele_Cours_{classe.nom.replace(' ', '_')}{section_label.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def import_cours_excel(request):
    """Importe des cours depuis un fichier Excel."""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'success': False, 'error': 'openpyxl non installé.'}, status=500)
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'Aucun fichier fourni.'}, status=400)
        
        file = request.FILES['file']
        id_classe = request.POST.get('id_classe')
        id_section = request.POST.get('id_section')  # NEW: optional section
        
        if not file.name.endswith('.xlsx'):
            return JsonResponse({'success': False, 'error': 'Format invalide. Utilisez .xlsx'}, status=400)
        
        classe = get_object_or_404(Classe, id_classe=id_classe)
        section = Section.objects.filter(id_section=id_section).first() if id_section else None
        
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.rows)
        
        if not rows:
            return JsonResponse({'success': False, 'error': 'Fichier vide.'}, status=400)
        
        header_row = [str(cell.value).upper().strip() if cell.value else '' for cell in rows[0]]
        
        required = ['CODE_COURS', 'COURS']
        for req in required:
            if req not in header_row:
                return JsonResponse({'success': False, 'error': f"En-tête '{req}' manquante."}, status=400)
        
        idx_code = header_row.index('CODE_COURS')
        idx_cours = header_row.index('COURS')
        idx_domaine = header_row.index('DOMAINE') if 'DOMAINE' in header_row else None
        
        success_count = 0
        errors = []
        
        for i, row in enumerate(rows[1:], start=2):
            val_code = row[idx_code].value if idx_code < len(row) else None
            val_cours = row[idx_cours].value if idx_cours < len(row) else None
            val_domaine = row[idx_domaine].value if idx_domaine is not None and idx_domaine < len(row) else ''
            
            if not val_code or not val_cours:
                continue
            if str(val_code).strip() in ('MATH01', 'FR01'):
                continue  # Ignorer les lignes d'exemple
            
            code = str(val_code).strip()
            nom = str(val_cours).strip()
            domaine = str(val_domaine).strip() if val_domaine else ''
            
            try:
                obj, created = Cours.objects.update_or_create(
                    classe=classe,
                    section=section,
                    code_cours=code,
                    id_pays=classe.id_pays,
                    defaults={'cours': nom, 'domaine': Domaine.objects.filter(pays=classe.cycle.pays, nom=domaine).first() if domaine else None}
                )
                success_count += 1
            except Exception as e:
                errors.append(f"Ligne {i}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'count': success_count,
            'errors': errors[:10]
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# --- CRUD COURS_ANNEE (Configuration Annuelle des Cours) ---

@require_http_methods(["GET"])
def get_cours_annee_data(request):
    """Liste la configuration annuelle des cours, filtrée par classe (+section) et année.
    Retourne TOUS les cours du catalogue avec un flag is_configured pour checkbox behavior.
    """
    try:
        id_classe = request.GET.get('id_classe')
        id_annee = request.GET.get('id_annee')
        id_section = request.GET.get('id_section')  # NEW: optional section filter
        id_pays = request.GET.get('id_pays')

        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})

        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'cours_annee': [], 'annees': []})

        annees = list(Annee.objects.filter(pays=pays).order_by('-annee').values('id_annee', 'annee', 'isOpen'))

        # Récupérer les domaines du pays
        domaines = list(Domaine.objects.filter(pays=pays).values('id_domaine', 'nom', 'code'))

        if not id_classe or not id_annee:
            return JsonResponse({'success': True, 'cours_annee': [], 'annees': annees, 'domaines': domaines})

        # Tous les cours du catalogue pour cette classe (+section)
        cours_catalogue = Cours.objects.filter(classe__id_classe=id_classe, id_pays=id_pays).order_by('cours')
        if id_section:
            cours_catalogue = cours_catalogue.filter(section_id=id_section)
        else:
            cours_catalogue = cours_catalogue.filter(section_id__isnull=True)

        # Les configs existantes pour cette année (même filtre section)
        # IMPORTANT: filter etablissement__isnull=True for national configs only
        configs_map = {}
        configs_qs = CoursAnnee.objects.filter(id_pays=id_pays,
            cours__classe__id_classe=id_classe, annee__id_annee=id_annee,
            etablissement__isnull=True
        ).select_related('cours')
        if id_section:
            configs_qs = configs_qs.filter(cours__section_id=id_section)
        else:
            configs_qs = configs_qs.filter(cours__section_id__isnull=True)
        for ca in configs_qs:
            configs_map[ca.cours_id] = ca

        # Construire la réponse : TOUS les cours avec leur état
        cours_data = []
        for c in cours_catalogue:
            ca = configs_map.get(c.pk)
            # Domaine : priorité au CoursAnnee, sinon fallback sur Cours
            dom_id = (ca.domaine_id if ca and ca.domaine_id else c.domaine_id)
            entry = {
                'id_cours': c.id_cours,
                'code_cours': c.code_cours,
                'cours': c.cours,
                'domaine_id': dom_id,
                'domaine_nom': '',
                'is_configured': ca is not None,
                'section_id': c.section_id,
                'section_nom': '',
            }
            if ca:
                entry.update({
                    'id_cours_annee': ca.id_cours_annee,
                    'maxima_exam': ca.maxima_exam,
                    'maxima_tj': ca.maxima_tj, 'maxima_periode': ca.maxima_periode,
                    'credits': ca.credits, 'heure_semaine': ca.heure_semaine,
                    'is_obligatory': ca.is_obligatory, 'ordre': ca.ordre,
                    'compte_au_nombre_echec': ca.compte_au_nombre_echec,
                    'total_considerable_trimestre': ca.total_considerable_trimestre,
                    'est_considerer_echec_lorsque_pourcentage_est': ca.est_considerer_echec_lorsque_pourcentage_est,
                    'is_second_semester': ca.is_second_semester,
                })

            cours_data.append(entry)

        return JsonResponse({
            'success': True,
            'cours_annee': cours_data,
            'annees': annees,
            'domaines': domaines
        })
    except Exception as e:
        import traceback
        logging.getLogger(__name__).error(f"Error in get_cours_annee_data: {str(e)}\n{traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_cours_annee(request):
    """Crée ou modifie une configuration annuelle de cours."""
    try:
        data = json.loads(request.body)
        id_cours_annee = data.get('id_cours_annee')
        cours_id = data.get('cours_id')
        annee_id = data.get('annee_id')

        int_fields_names = ['maxima_exam', 'maxima_tj', 'maxima_periode', 'credits',
                            'heure_semaine', 'ordre', 'est_considerer_echec_lorsque_pourcentage_est']
        bool_fields_names = ['is_obligatory', 'compte_au_nombre_echec', 'total_considerable_trimestre', 'is_second_semester']

        if id_cours_annee:
            # UPDATE: only update fields that are explicitly in the request
            ca = get_object_or_404(CoursAnnee, id_cours_annee=id_cours_annee)
            for f in int_fields_names:
                if f in data:
                    v = data[f]
                    try:
                        setattr(ca, f, int(v) if v is not None and v != '' else None)
                    except (ValueError, TypeError):
                        setattr(ca, f, None)
            for f in bool_fields_names:
                if f in data:
                    setattr(ca, f, bool(data[f]))
            # Handle domaine_id update
            if 'domaine_id' in data:
                domaine_id = data['domaine_id']
                ca.domaine_id = int(domaine_id) if domaine_id else None
            ca.save()
        else:
            # CREATE: need cours_id and annee_id
            if not cours_id or not annee_id:
                return JsonResponse({'success': False, 'error': 'Cours et année requis.'}, status=400)

            fields = {}
            for f in int_fields_names:
                v = data.get(f)
                try:
                    fields[f] = int(v) if v is not None and v != '' else None
                except (ValueError, TypeError):
                    fields[f] = None
            for f in bool_fields_names:
                fields[f] = bool(data.get(f, False))

            cours = get_object_or_404(Cours, id_cours=cours_id)
            annee = get_object_or_404(Annee, id_annee=annee_id)
            # Support optional etablissement_id for per-establishment configs
            etablissement_id = data.get('etablissement_id')
            etab = None
            if etablissement_id:
                etab = get_object_or_404(Etablissement, id_etablissement=etablissement_id)
            if CoursAnnee.objects.filter(cours=cours, annee=annee, etablissement=etab, id_pays=cours.id_pays).exists():
                return JsonResponse({'success': False, 'error': 'Ce cours est déjà configuré pour cette année.'}, status=400)
            # Default domaine from catalogue
            domaine_id = data.get('domaine_id') or (cours.domaine_id if cours.domaine_id else None)
            domaine = Domaine.objects.filter(id_domaine=domaine_id).first() if domaine_id else None
            id_pays_val = getattr(request, 'id_pays', None) or request.session.get('id_pays') or (cours.id_pays if cours else None)
            CoursAnnee.objects.create(cours=cours, annee=annee, etablissement=etab, domaine=domaine, id_pays=id_pays_val, **fields)

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_cours_annee(request):
    """Supprime une configuration annuelle de cours."""
    try:
        data = json.loads(request.body)
        get_object_or_404(CoursAnnee, id_cours_annee=data.get('id_cours_annee')).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def bulk_activate_cours_annee(request):
    """Active en masse les cours du catalogue pour une classe/année (+section)."""
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        id_classe = data.get('id_classe')
        id_annee = data.get('id_annee')
        id_section = data.get('id_section')  # NEW: optional section

        if not id_classe or not id_annee:
            return JsonResponse({'success': False, 'error': 'Classe et année requis.'}, status=400)

        cours_catalogue = Cours.objects.filter(classe__id_classe=id_classe, id_pays=id_pays)
        if id_section:
            cours_catalogue = cours_catalogue.filter(section_id=id_section)
        else:
            cours_catalogue = cours_catalogue.filter(section_id__isnull=True)
        annee = get_object_or_404(Annee, id_annee=id_annee)

        created = 0
        for c in cours_catalogue:
            _, was_created = CoursAnnee.objects.get_or_create(
                cours=c, annee=annee, defaults={'ordre': c.id_cours, 'id_pays': id_pays}
            )
            if was_created:
                created += 1

        return JsonResponse({'success': True, 'count': created})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def download_cours_annee_template(request):
    """Génère un modèle Excel pour importer la config annuelle des cours d'une classe."""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'openpyxl non installé.'}, status=500)
    try:
        id_classe = request.GET.get('id_classe')
        id_annee = request.GET.get('id_annee')
        id_pays = request.GET.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        classe = get_object_or_404(Classe, id_classe=id_classe)
        annee = get_object_or_404(Annee, id_annee=id_annee)

        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Config_{classe.nom}"[:31]

        ws.cell(row=1, column=1, value=f"CLASSE:{id_classe}|ANNEE:{id_annee}|{classe.nom}|{annee.annee}")
        ws.cell(row=1, column=1).fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")

        headers = ['CODE_COURS', 'COURS', 'DOMAINE', 'MAXIMA_EXAM', 'MAXIMA_TJ',
                    'MAXIMA_PERIODE', 'CREDITS', 'HEURE_SEMAINE',
                    'OBLIGATOIRE', 'ORDRE', 'COMPTE_ECHEC', 'TOTAL_TRIMESTRE',
                    'SEUIL_ECHEC_%', 'SEMESTRE_2_UNIQUEMENT']
        ws.append(headers)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")

        # Build domaine map for name lookup
        domaine_map = {}
        if Domaine:
            pays_obj = Pays.objects.filter(id_pays=id_pays).first()
            if pays_obj:
                domaine_map = {d['id_domaine']: d['nom'] for d in Domaine.objects.filter(pays=pays_obj).values('id_domaine', 'nom')}

        id_section = request.GET.get('id_section')
        cours_qs = Cours.objects.filter(classe__id_classe=id_classe, id_pays=id_pays).order_by('domaine_id', 'code_cours')
        if id_section:
            cours_qs = cours_qs.filter(section_id=id_section)
        else:
            cours_qs = cours_qs.filter(section_id__isnull=True)

        for c in cours_qs:
            config = CoursAnnee.objects.filter(cours=c, annee=annee, id_pays=id_pays).first()
            dom_name = domaine_map.get(c.domaine_id, '') if c.domaine_id else ''
            ws.append([
                c.code_cours, c.cours, dom_name,
                config.maxima_exam if config else '',
                config.maxima_tj if config else '',
                config.maxima_periode if config else '', config.credits if config else '',
                config.heure_semaine if config else '',
                1 if config and config.is_obligatory else 0,
                config.ordre if config else '',
                1 if config and config.compte_au_nombre_echec else 0,
                1 if config and config.total_considerable_trimestre else 0,
                config.est_considerer_echec_lorsque_pourcentage_est if config else '',
                1 if config and config.is_second_semester else 0,
            ])

        widths = [15, 35, 20, 12, 8, 8, 10, 10, 8, 12, 10, 8, 12, 12, 12, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Config_Cours_{classe.nom}_{annee.annee}.xlsx".replace(' ', '_')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def import_cours_annee_excel(request):
    """Importe la configuration annuelle des cours depuis un fichier Excel."""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'success': False, 'error': 'openpyxl non installé.'}, status=500)
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'Aucun fichier fourni.'}, status=400)

        file = request.FILES['file']
        id_classe = request.POST.get('id_classe')
        id_annee = request.POST.get('id_annee')

        if not file.name.endswith('.xlsx'):
            return JsonResponse({'success': False, 'error': 'Format invalide. Utilisez .xlsx'}, status=400)

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        meta = str(ws.cell(row=1, column=1).value or '')
        if meta.startswith('CLASSE:'):
            parts = meta.split('|')
            if not id_classe:
                id_classe = parts[0].replace('CLASSE:', '')
            if not id_annee and len(parts) > 1:
                id_annee = parts[1].replace('ANNEE:', '')

        if not id_classe or not id_annee:
            return JsonResponse({'success': False, 'error': 'Classe et année non identifiées.'}, status=400)

        id_pays = request.POST.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        classe = get_object_or_404(Classe, id_classe=id_classe)
        annee = get_object_or_404(Annee, id_annee=id_annee)
        if not id_pays:
            id_pays = classe.id_pays

        rows = list(ws.rows)
        header_row = [str(cell.value).upper().strip() if cell.value else '' for cell in rows[1]]
        if 'CODE_COURS' not in header_row:
            return JsonResponse({'success': False, 'error': "En-tête 'CODE_COURS' manquante."}, status=400)

        idx = {h: i for i, h in enumerate(header_row)}
        success_count = 0
        errors = []

        def val(row, key):
            return row[idx[key]].value if key in idx and idx[key] < len(row) else None

        def to_int(v):
            if v is None or v == '':
                return None
            try:
                return int(v)
            except (ValueError, TypeError):
                return None

        def to_bool(v):
            return str(v).strip() in ('1', 'True', 'true', 'oui', 'Oui') if v else False

        for i, row in enumerate(rows[2:], start=3):
            code_cours = val(row, 'CODE_COURS')
            if not code_cours:
                continue
            code_cours = str(code_cours).strip()
            try:
                cours = Cours.objects.get(classe=classe, code_cours=code_cours, id_pays=classe.id_pays)
            except Cours.DoesNotExist:
                errors.append(f"Ligne {i}: Code '{code_cours}' introuvable")
                continue
            try:
                CoursAnnee.objects.update_or_create(
                    cours=cours, annee=annee,
                    defaults={'id_pays': classe.id_pays,

                        'maxima_exam': to_int(val(row, 'MAXIMA_EXAM')),
                        'maxima_tj': to_int(val(row, 'MAXIMA_TJ')),
                        'maxima_periode': to_int(val(row, 'MAXIMA_PERIODE')),
                        'credits': to_int(val(row, 'CREDITS')),
                        'heure_semaine': to_int(val(row, 'HEURE_SEMAINE')),
                        'is_obligatory': to_bool(val(row, 'OBLIGATOIRE')),
                        'ordre': to_int(val(row, 'ORDRE')),
                        'compte_au_nombre_echec': to_bool(val(row, 'COMPTE_ECHEC')),
                        'total_considerable_trimestre': to_bool(val(row, 'TOTAL_TRIMESTRE')),
                        'est_considerer_echec_lorsque_pourcentage_est': to_int(val(row, 'SEUIL_ECHEC_%')),
                        'is_second_semester': to_bool(val(row, 'SEMESTRE_2_UNIQUEMENT')),
                    }
                )
                success_count += 1
            except Exception as e:
                errors.append(f"Ligne {i}: {str(e)}")

        return JsonResponse({'success': True, 'count': success_count, 'errors': errors[:10]})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_etablissement(request):
    try:
        data = json.loads(request.body)
        etab = get_object_or_404(Etablissement, id_etablissement=data.get('id_etablissement'))
        etab.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# CODE APPENDED BY AGENT

@require_http_methods(["GET"])
def download_ecole_template(request):
    """
    Génère un modèle Excel pour l'importation d'établissements.
    Colonnes (basées sur le modèle de données) :
      - nom (obligatoire)
      - code_ecole
      - matricule
      - no_dinacope
      - reference_agrement
      - admin_email (obligatoire)
    
    Le code est auto-généré : id_etab-id_regime-id_gestionnaire-PD-codePed
    Le régime n'est PAS inclus (assigné en modification).
    L'id_etablissement est auto-incrémenté.
    
    Structure du fichier :
      - Ligne 1 : Métadonnées (CODE: parent_ped_id|parent_ped_code|parent_ped_nom)
      - Ligne 2 : Info lisible
      - Ligne 3 : En-têtes des colonnes
      - Ligne 4 : Exemple
      - Ligne 5+ : Données à remplir
    """
    from openpyxl.styles import Font, Protection, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'La librairie Excel (openpyxl) n\'est pas installée.'}, status=500)

    try:
        id_pays = request.GET.get('id_pays')
        id_parent_ped = request.GET.get('id_parent_ped')
        
        if not id_pays or not id_parent_ped:
            return JsonResponse({'error': 'id_pays et structure pédagogique sont obligatoires.'}, status=400)
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        parent_ped = get_object_or_404(PedagogicStructureInstance, id_structure=id_parent_ped)
        
        # Charger les régimes du pays pour l'en-tête de la colonne ID_REGIME
        regimes = Regime.objects.filter(pays=pays).order_by('id_regime')
        regime_legend = ', '.join([f"{r.id_regime}={r.regime}" for r in regimes])
        
        # Charger les gestionnaires pour l'en-tête de la colonne ID_GESTIONNAIRE
        gestionnaires = GestionnaireEtablissement.objects.all().order_by('id_gestionnaire')
        gest_legend = ', '.join([f"{g.id_gestionnaire}={g.nom}" for g in gestionnaires])
        
        # Construire le code de contexte pour l'en-tête (sera lu à l'import)
        meta_code = f"{id_parent_ped}|{parent_ped.code}|{parent_ped.nom}"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Etablissements"
        
        # Styles
        locked_style = Protection(locked=True)
        unlocked_style = Protection(locked=False)
        
        meta_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        meta_font = Font(bold=True, color="FFFFFF", size=11)
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        required_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Colonnes du modèle
        regime_header = f"ID_REGIME ({regime_legend})" if regime_legend else "ID_REGIME"
        gest_header = f"ID_GESTIONNAIRE ({gest_legend})" if gest_legend else "ID_GESTIONNAIRE"
        headers = [
            "NOM*",
            regime_header,
            gest_header,
            "MATRICULE",
            "NO_DINACOPE",
            "REFERENCE_AGREMENT",
        ]
        
        col_widths = [35, 40, 50, 18, 18, 22]
        
        # --- Ligne 1 : Métadonnées (contexte) ---
        meta_text = f"CODE: {meta_code}"
        ws.cell(row=1, column=1, value=meta_text)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        meta_cell = ws.cell(row=1, column=1)
        meta_cell.fill = meta_fill
        meta_cell.font = meta_font
        meta_cell.alignment = Alignment(horizontal='center', vertical='center')
        meta_cell.protection = locked_style
        
        # --- Ligne 2 : Info lisible (contexte humain) ---
        info_text = f"Structure Péd.: {parent_ped.nom} (Code: {parent_ped.code}) | Régimes: {regime_legend or 'aucun'} | Gestionnaires: {gest_legend or 'aucun'}"
        ws.cell(row=2, column=1, value=info_text)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        info_cell = ws.cell(row=2, column=1)
        info_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        info_cell.font = Font(italic=True, color="1B5E20", size=9)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')
        info_cell.protection = locked_style
        
        # --- Ligne 3 : En-têtes des colonnes ---
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
            cell.protection = locked_style
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
        
        # --- Ligne 4 : Exemple ---
        first_regime_id = regimes.first().id_regime if regimes.exists() else 1
        first_gest_id = gestionnaires.first().id_gestionnaire if gestionnaires.exists() else 1
        example_row = [
            "Ex: École Primaire Saint-Joseph",
            str(first_regime_id),
            str(first_gest_id),
            "MAT456",
            "DIN789",
            "AGR001",
        ]
        for col_idx, val in enumerate(example_row, 1):
            cell = ws.cell(row=4, column=col_idx, value=val)
            cell.protection = unlocked_style
            cell.font = Font(italic=True, color="999999")
            cell.border = thin_border
        
        # --- Lignes 5 à 504 : Lignes vides (pré-formatées) ---
        for row_idx in range(5, 505):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.protection = unlocked_style
                cell.border = thin_border
                # Mettre en surbrillance la colonne obligatoire (NOM seulement)
                if col_idx == 1:
                    cell.fill = required_fill
        
        # Activer la protection de la feuille
        ws.protection.sheet = True
        ws.protection.password = 'monekole2026'
        ws.protection.enable()
        
        # Hauteur des lignes
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 30
        
        # --- Instructions dans une deuxième feuille ---
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            ["INSTRUCTIONS D'UTILISATION"],
            [""],
            ["1. Ne modifiez PAS la ligne 1 (bande verte). Elle contient le code de contexte nécessaire à l'import."],
            ["2. Seule la colonne NOM (marquée *) est obligatoire. Les autres colonnes sont optionnelles."],
            ["3. Remplissez les données à partir de la ligne 5 (la ligne 4 est un exemple)."],
            ["4. Le champ ID_ETABLISSEMENT est auto-incrémenté, il n'apparaît pas dans ce modèle."],
            ["5. La colonne ID_REGIME indique le type de régime. Référez-vous à la légende dans l'en-tête."],
            [f"   Régimes disponibles: {regime_legend or 'aucun défini'}"],
            ["6. La colonne ID_GESTIONNAIRE indique le gestionnaire. Référez-vous à la légende dans l'en-tête."],
            [f"   Gestionnaires disponibles: {gest_legend or 'aucun défini'}"],
            ["7. Le champ CODE est généré automatiquement selon la règle :"],
            ["   CODE = id_etablissement - id_regime - id_gestionnaire - PD - codePed"],
            [f"   Exemple pour ce contexte : [ID]-{first_regime_id}-[ID_GEST]-PD-{parent_ped.code}"],
            [""],
            [f"Pays : {pays.nom} (ID {pays.id_pays})"],
            [f"Structure Pédagogique parente : {parent_ped.nom}"],
        ]
        for row_data in instructions:
            ws_inst.append(row_data)
        
        ws_inst.column_dimensions['A'].width = 80
        ws_inst.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Modele_Import_Etablissements_{pays.id_pays}_{parent_ped.nom.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def import_ecoles_excel(request):
    """
    Importe des établissements depuis un fichier Excel.
    Structure attendue :
      - Ligne 1 : Métadonnées (CODE: parent_ped_id|parent_ped_code|parent_ped_nom)
      - Ligne 2 : Info lisible (ignorée)
      - Ligne 3 : En-têtes (NOM*, CODE_ECOLE, MATRICULE, NO_DINACOPE, REFERENCE_AGREMENT, ADMIN_EMAIL*)
      - Ligne 4+ : Données (ligne 4 peut être l'exemple à ignorer)
    
    Le code est auto-généré : id_etab-id_regime-id_gestionnaire-PD-codePed
    """
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'success': False, 'error': 'La librairie Excel (openpyxl) n\'est pas installée.'}, status=500)

    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'Aucun fichier fourni.'}, status=400)
            
        file = request.FILES['file']
        id_pays = request.POST.get('id_pays')
        
        # Paramètre optionnel du formulaire (peut être déduit du fichier)
        form_id_parent_ped = request.POST.get('id_parent_ped')
        
        # Fallback regime & gestionnaire depuis le formulaire modal
        form_id_regime = request.POST.get('id_regime')
        form_id_gestionnaire = request.POST.get('id_gestionnaire')
        
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant.'}, status=400)
        
        if not file.name.endswith('.xlsx'):
            return JsonResponse({'success': False, 'error': 'Format invalide. Utilisez .xlsx'}, status=400)

        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        # Résoudre le gestionnaire du formulaire (fallback global)
        form_gestionnaire = None
        if form_id_gestionnaire:
            try:
                form_gestionnaire = GestionnaireEtablissement.objects.get(id_gestionnaire=int(form_id_gestionnaire))
            except (GestionnaireEtablissement.DoesNotExist, ValueError):
                pass
        
        # Résoudre le régime du formulaire (fallback global)
        form_regime_int = None
        if form_id_regime:
            try:
                form_regime_int = int(form_id_regime)
                if not Regime.objects.filter(id_regime=form_regime_int, pays=pays).exists():
                    form_regime_int = None
            except (ValueError, TypeError):
                pass
        
        # Charger le fichier Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # --- Lire le contexte depuis la ligne 1 ---
        code_cell_value = str(ws.cell(row=1, column=1).value or "")
        id_parent_ped = form_id_parent_ped
        
        if code_cell_value.startswith("CODE:"):
            meta_str = code_cell_value.replace("CODE:", "").strip()
            meta_parts = meta_str.split("|")
            # Format: parent_ped_id|parent_ped_code|parent_ped_nom
            if len(meta_parts) >= 2:
                if not id_parent_ped:
                    id_parent_ped = meta_parts[0].strip()
        
        if not id_parent_ped:
            return JsonResponse({'success': False, 'error': 'Structure pédagogique parente non spécifiée. Utilisez le modèle téléchargé.'}, status=400)
        
        # Récupérer le parent pédagogique
        try:
            parent_ped = PedagogicStructureInstance.objects.get(id_structure=int(id_parent_ped))
        except (PedagogicStructureInstance.DoesNotExist, ValueError):
            return JsonResponse({'success': False, 'error': f'Structure pédagogique introuvable (ID={id_parent_ped}).'}, status=400)
        
        # --- Lire les en-têtes (ligne 3) ---
        header_row = []
        for cell in ws[3]:
            val = str(cell.value or "").strip().upper()
            header_row.append(val)
        
        # Mapping des colonnes attendues
        col_map = {}
        expected_cols = {
            'NOM*': 'nom',
            'NOM': 'nom',
            'MATRICULE': 'matricule',
            'NO_DINACOPE': 'no_dinacope',
            'REFERENCE_AGREMENT': 'reference_agrement',
        }
        
        for idx, h in enumerate(header_row):
            if h in expected_cols:
                col_map[expected_cols[h]] = idx
            # Détecter la colonne ID_REGIME (l'en-tête peut contenir la légende)
            elif h.startswith('ID_REGIME'):
                col_map['id_regime'] = idx
            # Détecter la colonne ID_GESTIONNAIRE (l'en-tête peut contenir la légende)
            elif h.startswith('ID_GESTIONNAIRE'):
                col_map['id_gestionnaire'] = idx
        
        if 'nom' not in col_map:
            return JsonResponse({'success': False, 'error': "Colonne 'NOM*' introuvable dans les en-têtes (ligne 3)."}, status=400)
        
        # --- Parcourir les données (ligne 4+) ---
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            # Lire le nom
            nom_val = row[col_map['nom']] if col_map['nom'] < len(row) else None
            if not nom_val:
                continue
            nom_val = str(nom_val).strip()
            if nom_val.startswith("Ex:") or not nom_val:
                continue
            
            # Lire les champs
            def get_val(key, default=''):
                if key in col_map and col_map[key] < len(row):
                    v = row[col_map[key]]
                    return str(v).strip() if v else default
                return default
            
            matricule = get_val('matricule') or None
            no_dinacope = get_val('no_dinacope') or None
            reference_agrement = get_val('reference_agrement') or None
            
            # Lire le régime (per-row Excel overrides form fallback)
            id_regime_val = form_regime_int  # Default from form
            if 'id_regime' in col_map and col_map['id_regime'] < len(row):
                raw_regime = row[col_map['id_regime']]
                if raw_regime is not None and str(raw_regime).strip():
                    try:
                        id_regime_val = int(float(str(raw_regime).strip()))
                        # Vérifier que le régime existe pour ce pays
                        if not Regime.objects.filter(id_regime=id_regime_val, pays=pays).exists():
                            errors.append(f"Ligne {row_idx}: Régime ID={id_regime_val} inexistant pour ce pays. Utilisation du régime par défaut.")
                            id_regime_val = form_regime_int
                    except (ValueError, TypeError):
                        errors.append(f"Ligne {row_idx}: Valeur ID_REGIME invalide '{raw_regime}'. Utilisation du régime par défaut.")
                        id_regime_val = form_regime_int
            
            # Lire le gestionnaire (per-row Excel overrides form fallback)
            gestionnaire = form_gestionnaire  # Default from form
            if 'id_gestionnaire' in col_map and col_map['id_gestionnaire'] < len(row):
                raw_gest = row[col_map['id_gestionnaire']]
                if raw_gest is not None and str(raw_gest).strip():
                    try:
                        id_gest_val = int(float(str(raw_gest).strip()))
                        gest_from_excel = GestionnaireEtablissement.objects.filter(id_gestionnaire=id_gest_val).first()
                        if gest_from_excel:
                            gestionnaire = gest_from_excel
                        else:
                            errors.append(f"Ligne {row_idx}: Gestionnaire ID={id_gest_val} inexistant. Utilisation du gestionnaire par défaut.")
                    except (ValueError, TypeError):
                        errors.append(f"Ligne {row_idx}: Valeur ID_GESTIONNAIRE invalide '{raw_gest}'. Utilisation du gestionnaire par défaut.")
            
            try:
                # Créer l'établissement — le code est auto-généré par le model.save()
                etab = Etablissement.objects.create(
                    pays=pays,
                    nom=nom_val,
                    id_regime=id_regime_val,
                    structure_pedagogique=parent_ped,
                    gestionnaire=gestionnaire,
                    matricule=matricule,
                    no_dinacope=no_dinacope,
                    reference_agrement=reference_agrement,
                    admin_email_verified=False,
                    admin_phone_verified=False,
                )
                
                created_count += 1
                
            except Exception as e_row:
                errors.append(f"Ligne {row_idx}: {str(e_row)}")
        
        msg = f"{created_count} établissement(s) importé(s) avec succès."
        if errors:
            msg += f" {len(errors)} erreur(s) rencontrée(s)."
            
        return JsonResponse({
            'success': True, 
            'message': msg, 
            'errors': errors[:20],
            'created_count': created_count
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)




# --- YEARLY CONFIGURATION ---

@require_http_methods(["GET"])
def configuration_annuelle_view(request):
    """Vue pour la configuration annuelle des établissements."""
    context = get_country_context_logic(request)
    context['active_page'] = 'configuration'
    return render(request, 'structure_app/configuration_annuelle.html', context)


@require_http_methods(["GET"])
def resultats_scolaires_view(request):
    """Vue pour la page Résultats Scolaires."""
    context = get_country_context_logic(request)
    context['active_page'] = 'resultats'
    # Charger le pays déterminé par le domaine et ses années
    pays = context.get('pays_from_domain')
    if pays:
        context['current_pays'] = pays
        context['annees'] = list(Annee.objects.filter(pays=pays).order_by('-annee').values('id_annee', 'annee', 'isOpen'))
    else:
        context['annees'] = []
    return render(request, 'structure_app/resultats_scolaires.html', context)


@require_http_methods(["GET"])
def get_resultats_etablissements(request):
    """
    API pour récupérer les classes configurées et leurs cours
    pour tous les établissements d'une année scolaire.
    Se base uniquement sur les données de countryStructure.
    """
    try:
        id_pays = request.GET.get('id_pays')
        id_annee = request.GET.get('id_annee')

        if not id_annee:
            return JsonResponse({'success': False, 'error': 'Année requise'})

        # Résoudre le pays : par id_pays ou par domaine
        if id_pays:
            pays = get_object_or_404(Pays, id_pays=id_pays)
        else:
            ctx = get_country_context_logic(request)
            pays = ctx.get('pays_from_domain')
            if not pays:
                return JsonResponse({'success': False, 'error': 'Pays non déterminé'})

        annee = get_object_or_404(Annee, id_annee=id_annee)

        # Établissements configurés pour cette année
        etab_annees = EtablissementAnnee.objects.filter(
            annee=annee,
            etablissement__pays=pays
        ).select_related('etablissement', 'etablissement__structure_pedagogique')

        # Filtrage RBAC
        scope_codes = get_all_user_scopes(request)
        if scope_codes:
            from django.db.models import Q
            scope_q = Q()
            for sc in scope_codes:
                scope_q |= Q(etablissement__structure_pedagogique__code__startswith=sc)
            etab_annees = etab_annees.filter(scope_q)

        results = []
        for ea in etab_annees:
            etab = ea.etablissement

            # Classes configurées + compter les cours pour chaque classe
            classes_config = ea.classes_config.select_related(
                'classe', 'classe__cycle', 'section'
            ).all()

            config_classes = []
            classes_with_cours = 0
            total_cours = 0

            for eac in classes_config:
                nb_cours = Cours.objects.filter(classe=eac.classe, id_pays=pays.id_pays).count()
                has_cours = nb_cours > 0
                if has_cours:
                    classes_with_cours += 1
                total_cours += nb_cours

                config_classes.append({
                    'name': eac.classe.nom,
                    'id_classe': eac.classe.id_classe,
                    'cycle': eac.classe.cycle.nom if eac.classe.cycle else '',
                    'section': eac.section.nom if eac.section else None,
                    'groupe': eac.groupe,
                    'nb_cours': nb_cours,
                    'has_cours': has_cours,
                })

            total_classes = len(config_classes)
            status = 'none'
            if classes_with_cours == total_classes and total_classes > 0:
                status = 'complete'
            elif classes_with_cours > 0:
                status = 'partial'

            percentage = round((classes_with_cours / total_classes) * 100) if total_classes > 0 else 0

            db_configured = bool(etab.db_server and etab.db_name and etab.db_user)

            results.append({
                'id': etab.id_etablissement,
                'nom': etab.nom,
                'code': etab.code or '',
                'structure_ped': etab.structure_pedagogique.nom if etab.structure_pedagogique else '-',
                'totalClasses': total_classes,
                'classesWithCours': classes_with_cours,
                'totalCours': total_cours,
                'status': status,
                'classes': config_classes,
                'percentage': percentage,
                'db_connected': db_configured,
            })

        return JsonResponse({
            'success': True,
            'etablissements': results,
            'count': len(results)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_annees_data(request):
    """API pour récupérer les années scolaires d'un pays."""
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'annees': []})
        
        annees = pays.annees.all().order_by('-annee')
        results = []
        for a in annees:
            nb_etabs = a.etablissements_config.count()
            # Safety: if annee is null (from old buggy saves), reconstruct it
            annee_label = a.annee
            if not annee_label and a.dateOuverture and a.dateCloture:
                annee_label = f"{a.dateOuverture.year}-{a.dateCloture.year}"
            
            results.append({
                'id_annee': a.id_annee,
                'annee': annee_label or "Année Sans Nom",
                'dateOuverture': a.dateOuverture.isoformat() if a.dateOuverture else None,
                'dateCloture': a.dateCloture.isoformat() if a.dateCloture else None,
                'isOpen': a.isOpen,
                'nb_etablissements': nb_etabs
            })
        return JsonResponse({'success': True, 'annees': results})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_annee(request):
    """API pour créer ou modifier une année scolaire."""
    try:
        data = json.loads(request.body)
        id_annee = data.get('id_annee')
        id_pays = data.get('id_pays')
        annee_str = data.get('annee')
        dateOuverture = data.get('dateOuverture')
        dateCloture = data.get('dateCloture')
        isOpen = data.get('isOpen', True)
        
        pays = get_object_or_404(Pays, id_pays=id_pays)
        
        if not annee_str or not dateOuverture or not dateCloture:
            return JsonResponse({'success': False, 'error': 'Données d\'année incomplètes.'}, status=400)

        if id_annee:
            annee_obj = get_object_or_404(Annee, id_annee=id_annee)
            annee_obj.annee = annee_str
            annee_obj.dateOuverture = dateOuverture
            annee_obj.dateCloture = dateCloture
            annee_obj.isOpen = isOpen
            annee_obj.save()
        else:
            # Check if year already exists for this country
            if Annee.objects.filter(pays_id=pays.id_pays, annee=annee_str).exists():
                return JsonResponse({'success': False, 'error': f'L\'année {annee_str} existe déjà pour ce pays.'})
            
            Annee.objects.create(
                pays_id=pays.id_pays,
                annee=annee_str,
                dateOuverture=dateOuverture,
                dateCloture=dateCloture,
                isOpen=isOpen
            )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_annee(request):
    """API pour supprimer une année scolaire."""
    try:
        data = json.loads(request.body)
        id_annee = data.get('id_annee')
        get_object_or_404(Annee, id_annee=id_annee).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_sections_list(request):
    """API pour récupérer la liste globale des sections/filières.
    Paramètres optionnels:
    - type_id: filtrer par type_subdivision_id (1=Sections, 2=Filières)
    """
    try:
        id_pays = request.GET.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        sections = Section.objects.all().order_by('type_subdivision', 'code')
        if id_pays:
            sections = sections.filter(pays_id=id_pays)
        type_id = request.GET.get('type_id')
        if type_id:
            sections = sections.filter(type_subdivision_id=type_id)
        return JsonResponse({
            'success': True,
            'sections': list(sections.values(
                'id_section', 'code', 'nom',
                'type_subdivision_id', 'type_subdivision__nom'
            ))
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_etablissement_config(request):
    """API pour récupérer la configuration d'un établissement pour une année."""
    try:
        id_etablissement = request.GET.get('id_etablissement')
        id_annee = request.GET.get('id_annee')
        
        if not id_etablissement or not id_annee:
            return JsonResponse({'success': False, 'error': 'Paramètres manquants'})
        
        # Resolve pays context for multi-tenant filtering
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays') or request.GET.get('id_pays')
        
        # Get the Etablissement and Annee (no get_object_or_404 — must return JSON, not HTML 404)
        etab_filters = {'id_etablissement': id_etablissement}
        if id_pays:
            etab_filters['pays_id'] = id_pays
        etablissement = Etablissement.objects.filter(**etab_filters).first()
        if not etablissement:
            return JsonResponse({'success': False, 'error': f'Établissement {id_etablissement} introuvable'})
        # Frontend sends annee_active.id (PK) — resolve robustly
        annee = Annee.objects.filter(pk=id_annee).first()
        if not annee and id_pays:
            annee = Annee.objects.filter(id_annee=id_annee, pays_id=id_pays).first()
        if not annee:
            return JsonResponse({'success': False, 'error': f'Année {id_annee} introuvable'})
        
        etab_annee, created = EtablissementAnnee.objects.get_or_create(
            etablissement=etablissement,
            annee=annee,
            defaults={'id_pays': etablissement.pays_id}
        )
        
        # Get activated classes with sections and groups
        activated = []
        for eac in etab_annee.classes_config.select_related('classe', 'section').all():
            activated.append({
                'id': eac.id,
                'classe_id': eac.classe_id,
                'classe_nom': str(eac.classe) if eac.classe else '-',
                'section_id': eac.section_id,
                'section_nom': str(eac.section) if eac.section else '-',
                'groupe': eac.groupe,
                'classe_par_annee_id': eac.id,
            })
        
        return JsonResponse({
            'success': True,
            'etablissement_annee_id': etab_annee.id,
            'activated_classes': activated
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_etablissement_config(request):
    """
    API pour sauvegarder la configuration complète d'un établissement pour une année.
    Reçoit une liste de classes/sections à activer.
    """
    try:
        data = json.loads(request.body)
        id_etablissement = data.get('id_etablissement')
        id_annee = data.get('id_annee')
        classes_config = data.get('classes', [])  # [{classe_id: X, section_id: Y or null}, ...]
        
        # Resolve pays context for multi-tenant filtering
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays') or data.get('id_pays')
        
        etab_filters = {'id_etablissement': id_etablissement}
        if id_pays:
            etab_filters['pays_id'] = id_pays
        etablissement = Etablissement.objects.filter(**etab_filters).first()
        if not etablissement:
            return JsonResponse({'success': False, 'error': 'Établissement introuvable'})
        # Frontend sends annee_active.id (PK) — resolve robustly
        annee = Annee.objects.filter(pk=id_annee).first()
        if not annee and id_pays:
            annee = Annee.objects.filter(id_annee=id_annee, pays_id=id_pays).first()
        if not annee:
            return JsonResponse({'success': False, 'error': 'Année introuvable'})
        
        # Get or create the link
        etab_annee, created = EtablissementAnnee.objects.get_or_create(
            etablissement=etablissement,
            annee=annee,
            defaults={'id_pays': etablissement.pays_id}
        )
        
        # Clear existing config — use raw SQL on Hub to avoid Django cross-DB
        # cascade check (eleve_inscription FK lives in Spoke, not Hub)
        from django.db import connections as db_conns
        with db_conns['countryStructure'].cursor() as hub_cur:
            hub_cur.execute(
                "DELETE FROM etablissements_annees_classes WHERE etablissement_annee_id = %s",
                [etab_annee.id]
            )
        
        # Add new config
        for item in classes_config:
            classe_id = item.get('classe_id')
            section_id = item.get('section_id')
            groupes = item.get('groupes', [])  # List of group letters, e.g. ['A', 'B', 'C'] or []
            
            classe = Classe.objects.filter(id_classe=classe_id).first()
            if not classe:
                continue  # Skip unknown classes silently
            section = None
            if section_id:
                section = Section.objects.filter(id_section=section_id).first()
            
            if groupes and len(groupes) > 0:
                # Create one record per group
                for grp in groupes:
                    EtablissementAnneeClasse.objects.create(
                        etablissement_annee=etab_annee,
                        classe=classe,
                        section=section,
                        groupe=grp
                    )
            else:
                # No groups — single record with groupe=NULL
                EtablissementAnneeClasse.objects.create(
                    etablissement_annee=etab_annee,
                    classe=classe,
                    section=section,
                    groupe=None
                )
        
        # ============================================================
        # AUTO-PROVISION RÉPARTITIONS
        # After saving classes, auto-create RepartitionConfigEtabAnnee
        # entries based on the cycles' RepartitionConfigCycle settings.
        # ============================================================
        repartitions_created = 0
        try:
            # 1. Collect distinct cycle IDs from saved classes
            activated_cycle_ids = set()
            for item in classes_config:
                classe = Classe.objects.filter(id_classe=item.get('classe_id')).select_related('cycle').first()
                if classe:
                    activated_cycle_ids.add(classe.cycle_id)
            
            if activated_cycle_ids:
                # Get pays from etablissement for tagging instances
                etab_pays = etablissement.pays
                
                # 2. Find RepartitionConfigCycle for these cycles
                cycle_configs = RepartitionConfigCycle.objects.filter(
                    cycle_id__in=activated_cycle_ids,
                    is_active=True,
                    id_pays=etab_pays.id_pays
                ).select_related('type_racine')
                
                # 3. Find all hierarchies (parent→child type relationships)
                hierarchies = {}
                for h in RepartitionHierarchie.objects.filter(is_active=True, id_pays=etab_pays.id_pays).select_related('type_parent', 'type_enfant'):
                    if h.type_parent_id not in hierarchies:
                        hierarchies[h.type_parent_id] = []
                    hierarchies[h.type_parent_id].append(h)
                
                # Track which RepartitionInstance IDs are already linked for this etab_annee
                existing_instance_ids = set(
                    RepartitionConfigEtabAnnee.objects.filter(
                        etablissement_annee=etab_annee
                    ).values_list('repartition_id', flat=True)
                )
                
                # Collect all type_racine+nombre combos (deduplicate across cycles)
                type_combos = {}  # type_racine_id → nombre_au_niveau_racine
                for cc in cycle_configs:
                    tid = cc.type_racine_id
                    if tid not in type_combos or cc.nombre_au_niveau_racine > type_combos[tid]:
                        type_combos[tid] = cc.nombre_au_niveau_racine
                
                for type_id, nombre_racine in type_combos.items():
                    rtype = RepartitionType.objects.get(pk=type_id)
                    
                    # 4. Find or create root-level instances for this type + année + pays
                    root_instances = list(
                        RepartitionInstance.objects.filter(
                            type=rtype, annee=annee, pays=etab_pays
                        ).order_by('ordre')
                    )
                    
                    # Also try without pays filter (legacy instances)
                    if not root_instances:
                        root_instances = list(
                            RepartitionInstance.objects.filter(
                                type=rtype, annee=annee
                            ).order_by('ordre')
                        )
                        # Tag existing instances with pays
                        for inst in root_instances:
                            if not inst.pays_id:
                                inst.pays = etab_pays
                                inst.save(update_fields=['pays'])
                    
                    # If not enough instances exist, create missing ones
                    while len(root_instances) < nombre_racine:
                        n = len(root_instances) + 1
                        new_inst = RepartitionInstance.objects.create(
                            type=rtype,
                            annee=annee,
                            pays=etab_pays,
                            nom=f"{rtype.nom} {n}",
                            code=f"{rtype.code}{n}",
                            ordre=n,
                            is_active=True
                        )
                        root_instances.append(new_inst)
                    
                    # 5. Create RepartitionConfigEtabAnnee for root instances
                    for inst in root_instances[:nombre_racine]:
                        if inst.id_instance not in existing_instance_ids:
                            RepartitionConfigEtabAnnee.objects.create(
                                etablissement_annee=etab_annee,
                                repartition=inst,
                                parent=None,
                                has_parent=False,
                                debut=inst.date_debut,
                                fin=inst.date_fin,
                                is_open=True,
                                is_national=True
                            )
                            existing_instance_ids.add(inst.id_instance)
                            repartitions_created += 1
                    
                    # 6. Handle child types via hierarchies (e.g., Trimestre→Période)
                    if type_id in hierarchies:
                        for hier in hierarchies[type_id]:
                            child_type = hier.type_enfant
                            nb_children = hier.nombre_enfants
                            
                            # Get root config entries to use as parents
                            root_configs = list(
                                RepartitionConfigEtabAnnee.objects.filter(
                                    etablissement_annee=etab_annee,
                                    repartition__type=rtype,
                                    has_parent=False
                                ).select_related('repartition').order_by('repartition__ordre')
                            )
                            
                            for parent_config in root_configs[:nombre_racine]:
                                # Find or create child instances
                                child_instances = list(
                                    RepartitionInstance.objects.filter(
                                        type=child_type, annee=annee, pays=etab_pays
                                    ).order_by('ordre')
                                )
                                
                                # Also try without pays filter (legacy)
                                if not child_instances:
                                    child_instances = list(
                                        RepartitionInstance.objects.filter(
                                            type=child_type, annee=annee
                                        ).order_by('ordre')
                                    )
                                    for ci in child_instances:
                                        if not ci.pays_id:
                                            ci.pays = etab_pays
                                            ci.save(update_fields=['pays'])
                                
                                # Determine which children to attach
                                parent_idx = parent_config.repartition.ordre - 1
                                start_child = parent_idx * nb_children
                                
                                # Create missing child instances if needed
                                total_needed = start_child + nb_children
                                while len(child_instances) < total_needed:
                                    n = len(child_instances) + 1
                                    new_child = RepartitionInstance.objects.create(
                                        type=child_type,
                                        annee=annee,
                                        pays=etab_pays,
                                        nom=f"{child_type.nom} {n}",
                                        code=f"{child_type.code}{n}",
                                        ordre=n,
                                        is_active=True
                                    )
                                    child_instances.append(new_child)
                                
                                # Link children to parent
                                for child_inst in child_instances[start_child:start_child + nb_children]:
                                    if child_inst.id_instance not in existing_instance_ids:
                                        RepartitionConfigEtabAnnee.objects.create(
                                            etablissement_annee=etab_annee,
                                            repartition=child_inst,
                                            parent=parent_config,
                                            has_parent=True,
                                            debut=child_inst.date_debut,
                                            fin=child_inst.date_fin,
                                            is_open=True,
                                            is_national=True
                                        )
                                        existing_instance_ids.add(child_inst.id_instance)
                                        repartitions_created += 1
        except Exception as rep_err:
            import traceback
            traceback.print_exc()
            # Non-blocking: if repartition provisioning fails, classes are still saved

        return JsonResponse({
            'success': True,
            'count': len(classes_config),
            'repartitions_created': repartitions_created
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# DASHBOARD — Student Management APIs (cross-DB to db_monecole)
# ============================================================

def _get_spoke_connection():
    """Helper: get pymysql connection to the spoke (default) database."""
    from django.db import connections
    db_settings = connections['default'].settings_dict
    import pymysql
    return pymysql.connect(
        host=db_settings.get('HOST', 'localhost') or 'localhost',
        user=db_settings['USER'],
        password=db_settings['PASSWORD'],
        port=int(db_settings.get('PORT', 3306) or 3306),
        database=db_settings['NAME'],
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=5,
        read_timeout=10,
        autocommit=False,
    )


def _resolve_eac_keys(cur, eac_id):
    """
    Résout un EAC.id (Hub) en clés métier stables.
    Retourne dict {classe_id, groupe, section_id, cycle_id, annee_id, campus_id} ou None.
    """
    cur.execute("""
        SELECT eac.classe_id, eac.groupe, eac.section_id,
               cl.cycle_id, ea.annee_id,
               (SELECT c.idCampus FROM db_monecole.campus c
                WHERE c.id_etablissement = ea.etablissement_id AND c.is_active=1 LIMIT 1) AS campus_id
        FROM countryStructure.etablissements_annees_classes eac
        JOIN countryStructure.classes cl ON cl.id = eac.classe_id
        JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
        WHERE eac.id = %s
    """, [eac_id])
    return cur.fetchone()


def _ei_classe_filter(alias='ei'):
    """
    Retourne la clause SQL WHERE pour filtrer eleve_inscription par business keys.
    Usage: f"AND {_ei_classe_filter()}" puis passer [bk['classe_id'], bk['groupe'], bk['section_id']]
    Le COLLATE assure la compatibilité cross-DB Hub/Spoke.
    """
    return f"{alias}.classe_id = %s AND {alias}.groupe COLLATE utf8mb4_general_ci <=> %s AND {alias}.section_id <=> %s"


def _resolve_eac_orm(eac_id):
    """
    Version ORM de _resolve_eac_keys.
    Résout un EAC.id (Hub) en clés métier stables via Django ORM.
    Retourne dict {classe_id, groupe, section_id} ou None.
    """
    try:
        from MonEcole_app.models.country_structure import EtablissementAnneeClasse
        eac = EtablissementAnneeClasse.objects.filter(id=eac_id).first()
        if eac:
            return {
                'classe_id': eac.classe_id,
                'groupe': eac.groupe,
                'section_id': eac.section_id,
            }
    except Exception:
        pass
    return None


def _count_eleves(cur, id_etablissement, id_pays=None):
    """Count active students for an establishment."""
    if id_pays:
        cur.execute(
            "SELECT COUNT(*) as total FROM eleve_inscription WHERE id_etablissement=%s AND id_pays=%s AND status=1",
            [id_etablissement, id_pays]
        )
    else:
        cur.execute(
            "SELECT COUNT(*) as total FROM eleve_inscription WHERE id_etablissement=%s AND status=1",
            [id_etablissement]
        )
    row = cur.fetchone()
    return row['total'] if row else 0


# ============================================================
# API — Search parents (smart deduplication)
# ============================================================
@require_http_methods(["GET"])
def search_parents(request):
    """Search parents by name or phone. Returns matches with children count."""
    try:
        q = request.GET.get('q', '').strip()
        if len(q) < 2:
            return JsonResponse({'success': True, 'parents': []})

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                like_q = f'%{q}%'
                cur.execute("""
                    SELECT p.id_parent, p.nomsPere,
                           p.telephonePere, p.emailPere,
                           p.nomsMere,
                           p.telephoneMere, p.emailMere,
                           COUNT(e.id_eleve) AS nb_enfants,
                           GROUP_CONCAT(CONCAT(e.nom, ' ', e.prenom) SEPARATOR ', ') AS enfants_noms
                    FROM parents p
                    LEFT JOIN eleve e ON e.id_parent = p.id_parent
                    WHERE p.nomsPere LIKE %s
                       OR p.nomsMere LIKE %s
                       OR p.telephonePere LIKE %s OR p.telephoneMere LIKE %s
                    GROUP BY p.id_parent
                    ORDER BY p.nomsPere, p.nomsMere
                    LIMIT 20
                """, [like_q, like_q, like_q, like_q])
                rows = cur.fetchall()

                result = []
                for r in rows:
                    result.append({
                        'id_parent': r['id_parent'],
                        'nomsPere': r.get('nomsPere') or '',
                        'telephonePere': r.get('telephonePere') or '',
                        'emailPere': r.get('emailPere') or '',
                        'nomsMere': r.get('nomsMere') or '',
                        'telephoneMere': r.get('telephoneMere') or '',
                        'emailMere': r.get('emailMere') or '',
                        'nb_enfants': r.get('nb_enfants', 0),
                        'enfants_noms': r.get('enfants_noms') or '',
                    })
                return JsonResponse({'success': True, 'parents': result})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def dashboard_add_eleve(request):
    """Add a single student: insert into eleve + eleve_inscription in db_monecole."""
    try:
        data = json.loads(request.body)
        id_etablissement = data.get('id_etablissement')
        id_annee = data.get('id_annee')
        numero_serie = data.get('numero_serie', '').strip()
        nom = data.get('nom', '').strip()
        postnom = data.get('postnom', '').strip()
        prenom = data.get('prenom', '').strip()
        date_naissance = data.get('date_naissance')
        genre = data.get('genre', '').strip()
        classe_par_annee_id = data.get('classe_par_annee_id') or data.get('classe_par_annee_id')
        telephone = data.get('telephone', '').strip()
        ref_administrative_naissance = data.get('ref_administrative_naissance', '').strip()
        ref_administrative_residence = data.get('ref_administrative_residence', '').strip()
        id_national = data.get('IDNational', '').strip()

        # Parent: either link to existing or create new
        id_parent = data.get('id_parent')  # existing parent ID
        parent_data = data.get('parent')   # {nomsPere, telephonePere, emailPere, nomsMere, telephoneMere, emailMere}

        if not all([nom, prenom, date_naissance, genre, classe_par_annee_id, id_etablissement]):
            return JsonResponse({'success': False, 'error': 'Champs obligatoires manquants.'}, status=400)

        # Combine nom + postnom for the `nom` field in db
        full_nom = f"{nom} {postnom}".strip() if postnom else nom

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Get classe_par_annee details (campus, cycle) — direct Hub query
                cur.execute("""
                    SELECT eac.id, eac.classe_id, eac.groupe, eac.section_id,
                           ea.annee_id AS id_annee_id,
                           c.idCampus AS idCampus_id,
                           cl.cycle_id AS cycle_id
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON eac.etablissement_annee_id = ea.id
                    JOIN countryStructure.classes cl ON cl.id = eac.classe_id
                    JOIN db_monecole.campus c ON c.id_etablissement = ea.etablissement_id AND c.is_active = 1
                    WHERE eac.id = %s
                """, [classe_par_annee_id])
                ca = cur.fetchone()
                if not ca:
                    return JsonResponse({'success': False, 'error': 'Classe active introuvable.'}, status=404)

                # Handle parent: create new if needed
                if not id_parent and parent_data:
                    cur.execute("""
                        INSERT INTO parents (nomsPere, telephonePere, emailPere, pere_en_vie,
                                             nomsMere, telephoneMere, emailMere, mere_en_vie, id_pays)
                        VALUES (%s, %s, %s, 1, %s, %s, %s, 1, %s)
                    """, [
                        parent_data.get('nomsPere') or None,
                        parent_data.get('telephonePere') or None,
                        parent_data.get('emailPere') or None,
                        parent_data.get('nomsMere') or None,
                        parent_data.get('telephoneMere') or None,
                        parent_data.get('emailMere') or None,
                        int(getattr(request, 'id_pays', None) or request.session.get('id_pays') or 0),
                    ])
                    id_parent = cur.lastrowid

                # Insert into eleve
                cur.execute("""
                    INSERT INTO eleve (numero_serie, nom, prenom, genre, date_naissance, id_etablissement,
                                       telephone, id_parent,
                                       ref_administrative_naissance, ref_administrative_residence, id_pays, IDNational)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    numero_serie or None, full_nom, prenom, genre, date_naissance, id_etablissement,
                    telephone or None, id_parent or None,
                    ref_administrative_naissance or None, ref_administrative_residence or None,
                    int(getattr(request, 'id_pays', None) or request.session.get('id_pays') or 0),
                    id_national or None
                ])
                id_eleve = cur.lastrowid

                # Insert into eleve_inscription
                cur.execute("SET FOREIGN_KEY_CHECKS=0")
                cur.execute("""
                    INSERT INTO eleve_inscription
                    (date_inscription, redoublement, status, isDelegue,
                     id_annee_id, idCampus_id, classe_id, groupe, section_id,
                     id_cycle_id, id_eleve_id, id_etablissement, id_pays)
                    VALUES (CURDATE(), 0, 1, 0, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    ca['id_annee_id'], ca['idCampus_id'], ca['classe_id'],
                    ca['groupe'], ca['section_id'], ca['cycle_id'],
                    id_eleve, id_etablissement,
                    int(getattr(request, 'id_pays', None) or request.session.get('id_pays') or 0)
                ])
                cur.execute("SET FOREIGN_KEY_CHECKS=1")

                conn.commit()

                total = _count_eleves(cur, id_etablissement,
                                      getattr(request, 'id_pays', None) or request.session.get('id_pays'))

            return JsonResponse({
                'success': True,
                'id_eleve': id_eleve,
                'total_eleves': total
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def dashboard_eleve_template(request):
    """Generate and return an Excel template for student import.
    If the class already has students, they are pre-filled in the rows.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    id_etablissement = request.GET.get('id_etablissement')
    id_annee = request.GET.get('id_annee')
    classe_par_annee_id = request.GET.get('classe_par_annee_id') or request.GET.get('classe_par_annee_id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves"

    # Header style
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=10)
    data_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'N° Série', 'Nom', 'Postnom', 'Prénom',
        'Année Naissance', 'Mois Naissance', 'Jour Naissance',
        'Genre (M/F)', 'Téléphone', 'ID National'
    ]

    # Resolve class name and fetch existing students
    classe_label = ''
    existing_students = []
    if classe_par_annee_id and id_etablissement:
        try:
            conn = _get_spoke_connection()
            try:
                with conn.cursor() as cur:
                    cur.execute("""
                        SELECT cl.nom as classe_nom,
                               COALESCE(s.nom, '') as section_nom,
                               COALESCE(eac.groupe, '') as groupe
                        FROM countryStructure.etablissements_annees_classes eac
                        JOIN countryStructure.classes cl ON cl.id = eac.classe_id
                        LEFT JOIN countryStructure.sections s ON s.id = eac.section_id
                        WHERE eac.id = %s
                    """, [classe_par_annee_id])
                    row = cur.fetchone()
                    if row:
                        classe_label = row['classe_nom']
                        if row['section_nom']:
                            classe_label += ' - ' + row['section_nom']
                        if row['groupe']:
                            classe_label += ' (' + row['groupe'] + ')'

                    # Fetch existing students enrolled in this class (via business keys)
                    cur.execute("""
                        SELECT eac.classe_id, eac.groupe, eac.section_id
                        FROM countryStructure.etablissements_annees_classes eac WHERE eac.id = %s
                    """, [classe_par_annee_id])
                    bk = cur.fetchone()
                    if bk:
                        cur.execute("""
                            SELECT e.numero_serie, e.nom, e.prenom, e.genre,
                                   e.date_naissance, e.telephone, e.IDNational
                            FROM eleve e
                            JOIN eleve_inscription ei ON ei.id_eleve_id = e.id_eleve
                            WHERE ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                              AND ei.id_etablissement = %s AND ei.id_pays = %s
                            ORDER BY e.nom, e.prenom
                        """, [bk['classe_id'], bk['groupe'], bk['section_id'], id_etablissement,
                              getattr(request, 'id_pays', None) or request.session.get('id_pays')])
                        existing_students = cur.fetchall()
            finally:
                conn.close()
        except Exception:
            pass

    # Add class info row above headers
    if classe_label:
        info_font = Font(name='Arial', bold=True, color="4338CA", size=9)
        ws.cell(row=1, column=1, value=f"Classe: {classe_label}").font = info_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        data_start = 2
    else:
        data_start = 1

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    # Pre-fill existing students
    if existing_students:
        for idx, stu in enumerate(existing_students):
            r = data_start + 1 + idx
            # Split nom back into nom + postnom (nom field stores "NOM POSTNOM")
            nom_parts = (stu['nom'] or '').split(' ', 1)
            nom_val = nom_parts[0] if nom_parts else ''
            postnom_val = nom_parts[1] if len(nom_parts) > 1 else ''

            dn = stu['date_naissance']
            dn_year = ''
            dn_month = ''
            dn_day = ''
            if dn:
                if hasattr(dn, 'year'):
                    dn_year = dn.year
                    dn_month = dn.month
                    dn_day = dn.day
                else:
                    dn_str = str(dn)
                    parts = dn_str.split('-')
                    if len(parts) == 3:
                        dn_year = parts[0]
                        dn_month = parts[1]
                        dn_day = parts[2]

            row_data = [
                stu.get('numero_serie') or '',
                nom_val,
                postnom_val,
                stu['prenom'] or '',
                dn_year,
                dn_month,
                dn_day,
                stu['genre'] or '',
                stu.get('telephone') or '',
                stu.get('IDNational') or '',
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = data_font
                cell.border = border

    # Store class ID in hidden _Meta sheet
    if classe_par_annee_id:
        ws_meta = wb.create_sheet("_Meta")
        ws_meta.cell(row=1, column=1, value="classe_par_annee_id")
        ws_meta.cell(row=1, column=2, value=int(classe_par_annee_id))
        ws_meta.sheet_state = 'hidden'

    # Save to response
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f"import_eleves_{classe_label.replace(' ','_')}.xlsx" if classe_label else "modele_import_eleves.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@require_http_methods(["POST"])
def dashboard_import_eleves(request):
    """Import students from uploaded Excel file into db_monecole."""
    try:
        import openpyxl
        from datetime import datetime, date as date_type

        file = request.FILES.get('file')
        id_etablissement = request.POST.get('id_etablissement')
        id_annee = request.POST.get('id_annee')

        if not file or not id_etablissement:
            return JsonResponse({'success': False, 'error': 'Fichier et établissement requis.'}, status=400)

        id_etablissement = int(id_etablissement)
        id_annee = int(id_annee) if id_annee else None
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')

        # Get class ID: from POST first, then from _Meta sheet in template
        classe_par_annee_id_str = request.POST.get('classe_par_annee_id', '') or request.POST.get('classe_par_annee_id', '')
        classe_id = int(classe_par_annee_id_str) if classe_par_annee_id_str else None

        wb = openpyxl.load_workbook(file, read_only=False)
        ws = wb.active

        # Try to read class ID from _Meta sheet if not provided in POST
        if not classe_id and '_Meta' in wb.sheetnames:
            meta_ws = wb['_Meta']
            for row_m in meta_ws.iter_rows(min_row=1, values_only=True):
                if row_m and len(row_m) >= 2 and str(row_m[0]).strip() in ('classe_par_annee_id', 'classe_par_annee_id'):
                    classe_id = int(row_m[1])
                    break

        if not classe_id:
            return JsonResponse({'success': False, 'error': 'Classe non spécifiée. Sélectionnez une classe avant d\'importer.'}, status=400)

        # Find data start row (skip info row if present)
        header_row_idx = 1
        first_cell = ws.cell(row=1, column=1).value
        if first_cell and str(first_cell).startswith('Classe:'):
            header_row_idx = 2

        # Read header row to build column mapping
        header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
        col_map = {}  # maps logical field -> column index
        has_split_date = False  # True if date is split into 3 columns (année, mois, jour)
        if header_cells and header_cells[0]:
            headers_raw = [str(h or '').strip().lower() for h in header_cells[0]]
            for idx, h in enumerate(headers_raw):
                if not h:
                    continue
                if 'série' in h or 'serie' in h or 'n°' in h or 'numero' in h or 'numéro' in h:
                    col_map['numero_serie'] = idx
                elif 'postnom' in h:
                    col_map['postnom'] = idx
                elif 'prénom' in h or 'prenom' in h:
                    if 'père' in h or 'pere' in h:
                        col_map['prenom_pere'] = idx  # will be combined into noms_pere
                    elif 'mère' in h or 'mere' in h:
                        col_map['prenom_mere'] = idx  # will be combined into noms_mere
                    else:
                        col_map['prenom'] = idx
                elif 'noms' in h and ('père' in h or 'pere' in h):
                    col_map['noms_pere'] = idx
                elif 'noms' in h and ('mère' in h or 'mere' in h):
                    col_map['noms_mere'] = idx
                elif 'nom' in h:
                    if 'père' in h or 'pere' in h:
                        col_map['nom_pere'] = idx  # will be combined into noms_pere
                    elif 'mère' in h or 'mere' in h:
                        col_map['nom_mere'] = idx  # will be combined into noms_mere
                    else:
                        col_map['nom'] = idx
                elif 'année' in h or 'annee' in h or 'year' in h:
                    if 'naissance' in h or idx < 8:
                        col_map['annee_naissance'] = idx
                        has_split_date = True
                elif 'mois' in h or 'month' in h:
                    if 'naissance' in h or idx < 8:
                        col_map['mois_naissance'] = idx
                        has_split_date = True
                elif ('jour' in h or 'day' in h) and ('naissance' in h or idx < 8):
                    col_map['jour_naissance'] = idx
                    has_split_date = True
                elif 'date' in h or 'naissance' in h:
                    col_map['date_naissance'] = idx
                elif 'genre' in h or 'sexe' in h:
                    col_map['genre'] = idx
                elif 'email' in h or 'mail' in h:
                    col_map['email_parent'] = idx
                elif 'tel' in h or 'phone' in h or 'téléphone' in h:
                    col_map['telephone'] = idx
                elif ('id' in h and 'national' in h) or h == 'idnational':
                    col_map['id_national'] = idx

        # Fallback: if no headers detected, use positional mapping (new 3-column format)
        if 'nom' not in col_map:
            col_map = {
                'numero_serie': 0, 'nom': 1, 'postnom': 2, 'prenom': 3,
                'annee_naissance': 4, 'mois_naissance': 5, 'jour_naissance': 6,
                'genre': 7, 'telephone': 8, 'id_national': 9
            }
            has_split_date = True

        print(f"[IMPORT] Column mapping: {col_map}, split_date={has_split_date}", flush=True)

        data_start = header_row_idx + 1
        rows = list(ws.iter_rows(min_row=data_start, values_only=True))
        if not rows:
            return JsonResponse({'success': False, 'error': 'Le fichier est vide.'}, status=400)

        def _cell(row, field):
            """Get cell value by field name using col_map."""
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        def normalize_date(val):
            """Convert date value to YYYY-MM-DD string. Tries multiple formats."""
            if val is None:
                return ''
            if isinstance(val, (datetime, date_type)):
                return val.strftime('%Y-%m-%d')
            s = str(val).strip()
            if not s:
                return ''
            # Handle "2010-05-15 00:00:00" format
            if ' ' in s:
                s = s.split(' ')[0]
            # If already in YYYY-MM-DD format (accept 1 or 2 digit month/day), normalize
            import re
            m_iso = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
            if m_iso:
                return f"{int(m_iso.group(1)):04d}-{int(m_iso.group(2)):02d}-{int(m_iso.group(3)):02d}"
            # Try multiple common date formats
            # Try numeric-only (e.g. "20100503" or "03052010" or just a year)
            if re.match(r'^\d{8}$', s):
                return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
            # If just a year
            if re.match(r'^\d{4}$', s):
                return f"{s}-01-01"
            date_formats = [
                '%d/%m/%Y',    # 15/05/2010
                '%d-%m-%Y',    # 15-05-2010
                '%d.%m.%Y',    # 15.05.2010
                '%m/%d/%Y',    # 05/15/2010
                '%m-%d-%Y',    # 05-15-2010
                '%Y/%m/%d',    # 2010/05/15
                '%d/%m/%y',    # 15/05/10
                '%d-%m-%y',    # 15-05-10
                '%m/%d/%y',    # 05/15/10
                '%d %m %Y',    # 15 05 2010
                '%d %B %Y',    # 15 mai 2010
                '%d %b %Y',    # 15 mai 2010
            ]
            for fmt in date_formats:
                try:
                    parsed = datetime.strptime(s, fmt)
                    return parsed.strftime('%Y-%m-%d')
                except ValueError:
                    continue
            # Last resort: return as-is and let MySQL handle it
            return s

        def normalize_genre(val):
            """Normalize genre to M or F."""
            if not val:
                return ''
            v = str(val).strip().upper()
            if v in ('M', 'MASCULIN', 'GARÇON', 'GARCON', 'BOY', 'MALE', 'H', 'HOMME'):
                return 'M'
            if v in ('F', 'FÉMININ', 'FEMININ', 'FILLE', 'GIRL', 'FEMALE', 'FEMME'):
                return 'F'
            return v

        conn = _get_spoke_connection()
        imported = 0
        updated = 0
        errors = []
        try:
            with conn.cursor() as cur:
                # Verify class exists
                cur.execute("""
                    SELECT eac.id, eac.classe_id, eac.groupe, eac.section_id,
                           c.idCampus AS idCampus_id, ea.annee_id AS id_annee_id, cl.cycle_id AS cycle_id
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON eac.etablissement_annee_id = ea.id
                    JOIN countryStructure.classes cl ON cl.id = eac.classe_id
                    JOIN db_monecole.campus c ON c.id_etablissement = ea.etablissement_id AND c.is_active = 1
                    WHERE eac.id = %s
                """,
                    [classe_id]
                )
                ca = cur.fetchone()
                if not ca:
                    return JsonResponse({'success': False, 'error': f'Classe {classe_id} introuvable.'}, status=400)

                # trimestre/periode obsolètes — remplacés par le système de répartitions

                cur.execute("SET FOREIGN_KEY_CHECKS=0")

                for i, row in enumerate(rows, data_start):
                    try:
                        if not row:
                            continue
                        # Filter out completely empty rows
                        non_empty = [c for c in row if c is not None and str(c).strip()]
                        if not non_empty:
                            continue

                        numero_serie = str(_cell(row, 'numero_serie') or '').strip()
                        nom = str(_cell(row, 'nom') or '').strip()
                        postnom = str(_cell(row, 'postnom') or '').strip()
                        prenom = str(_cell(row, 'prenom') or '').strip()

                        # Date de naissance: support 3-column split or single column
                        if has_split_date:
                            annee_v = _cell(row, 'annee_naissance')
                            mois_v = _cell(row, 'mois_naissance')
                            jour_v = _cell(row, 'jour_naissance')
                            try:
                                y = int(annee_v) if annee_v else 0
                                m = int(mois_v) if mois_v else 0
                                d = int(jour_v) if jour_v else 0
                                if y and m and d:
                                    date_naissance = f"{y:04d}-{m:02d}-{d:02d}"
                                elif y and m:
                                    date_naissance = f"{y:04d}-{m:02d}-01"
                                elif y:
                                    date_naissance = f"{y:04d}-01-01"
                                else:
                                    date_naissance = ''
                            except (ValueError, TypeError):
                                date_naissance = ''
                        else:
                            date_naissance = normalize_date(_cell(row, 'date_naissance'))

                        genre = normalize_genre(_cell(row, 'genre'))
                        # Parent names: support both single 'noms_pere' and split 'nom_pere'+'prenom_pere'
                        if 'noms_pere' in col_map:
                            noms_pere = str(_cell(row, 'noms_pere') or '').strip()
                        else:
                            _np = str(_cell(row, 'nom_pere') or '').strip()
                            _pp = str(_cell(row, 'prenom_pere') or '').strip()
                            noms_pere = f"{_np} {_pp}".strip()
                        if 'noms_mere' in col_map:
                            noms_mere = str(_cell(row, 'noms_mere') or '').strip()
                        else:
                            _nm = str(_cell(row, 'nom_mere') or '').strip()
                            _pm = str(_cell(row, 'prenom_mere') or '').strip()
                            noms_mere = f"{_nm} {_pm}".strip()
                        telephone = str(_cell(row, 'telephone') or '').strip()
                        id_national = str(_cell(row, 'id_national') or '').strip()

                        if not nom or not prenom:
                            errors.append(f"Ligne {i}: nom ou prénom manquant")
                            continue

                        full_nom = f"{nom} {postnom}".strip() if postnom else nom

                        # Check if student already exists (match by nom + prenom + etablissement)
                        cur.execute("""
                            SELECT id_eleve FROM eleve
                            WHERE nom = %s AND prenom = %s AND id_etablissement = %s AND id_pays = %s
                            LIMIT 1
                        """, [full_nom, prenom, id_etablissement, id_pays])
                        existing = cur.fetchone()

                        if existing:
                            # UPDATE existing student — only non-empty values
                            set_parts = []
                            set_vals = []
                            if numero_serie:
                                set_parts.append("numero_serie = %s"); set_vals.append(numero_serie)
                            if genre in ('M', 'F'):
                                set_parts.append("genre = %s"); set_vals.append(genre)
                            if date_naissance:
                                set_parts.append("date_naissance = %s"); set_vals.append(date_naissance)
                            if telephone:
                                set_parts.append("telephone = %s"); set_vals.append(telephone)
                            if id_national:
                                set_parts.append("IDNational = %s"); set_vals.append(id_national)
                            if set_parts:
                                set_vals.append(existing['id_eleve'])
                                cur.execute(f"UPDATE eleve SET {', '.join(set_parts)} WHERE id_eleve = %s", set_vals)

                            # Also ensure the student has an inscription for the TARGET class
                            cur.execute("""
                                SELECT id_inscription FROM eleve_inscription
                                WHERE id_eleve_id = %s
                                  AND classe_id = %s AND groupe <=> %s AND section_id <=> %s
                                  AND id_etablissement = %s AND id_pays = %s
                                LIMIT 1
                            """, [
                                existing['id_eleve'],
                                ca['classe_id'], ca['groupe'], ca['section_id'],
                                id_etablissement, id_pays
                            ])
                            existing_inscription = cur.fetchone()
                            if not existing_inscription:
                                cur.execute("""
                                    INSERT INTO eleve_inscription
                                    (date_inscription, redoublement, status, isDelegue,
                                     id_annee_id, idCampus_id, classe_id, groupe, section_id,
                                     id_cycle_id, id_eleve_id, id_etablissement, id_pays)
                                    VALUES (CURDATE(), 0, 1, 0, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """, [
                                    ca['id_annee_id'], ca['idCampus_id'], ca['classe_id'],
                                    ca['groupe'], ca['section_id'], ca['cycle_id'],
                                    existing['id_eleve'], id_etablissement, id_pays
                                ])
                                imported += 1
                            updated += 1
                        else:
                            # For INSERT, genre and date are required
                            if not date_naissance:
                                errors.append(f"Ligne {i}: date de naissance manquante pour {nom} {prenom}")
                                continue
                            if genre not in ('M', 'F'):
                                errors.append(f"Ligne {i}: genre invalide '{genre}' pour {nom} {prenom} (M ou F)")
                                continue

                            # INSERT new student (without parent data)
                            cur.execute("""
                                INSERT INTO eleve (numero_serie, nom, prenom, genre, date_naissance, id_etablissement,
                                                   telephone, id_pays, IDNational)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, [
                                numero_serie or None, full_nom, prenom, genre, date_naissance, id_etablissement,
                                telephone or None, id_pays, id_national or None
                            ])
                            id_eleve = cur.lastrowid

                            # Insert inscription only for NEW students
                            cur.execute("""
                                INSERT INTO eleve_inscription
                                (date_inscription, redoublement, status, isDelegue,
                                 id_annee_id, idCampus_id, classe_id, groupe, section_id,
                                 id_cycle_id, id_eleve_id, id_etablissement, id_pays)
                                VALUES (CURDATE(), 0, 1, 0, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, [
                                ca['id_annee_id'], ca['idCampus_id'], ca['classe_id'],
                                ca['groupe'], ca['section_id'], ca['cycle_id'],
                                id_eleve, id_etablissement, id_pays
                            ])
                            imported += 1

                    except Exception as row_err:
                        errors.append(f"Ligne {i}: {str(row_err)}")
                        print(f"[IMPORT ERROR] Ligne {i}: {row_err}", flush=True)

                cur.execute("SET FOREIGN_KEY_CHECKS=1")
                conn.commit()
                total = _count_eleves(cur, id_etablissement, id_pays)

            return JsonResponse({
                'success': True,
                'imported': imported,
                'updated': updated,
                'errors': errors[:20],
                'total_eleves': total
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# API — Parent Update Template (export students with parent columns)
# ============================================================
@require_http_methods(["GET"])
def parent_update_template(request):
    """Generate Excel with existing students + empty parent columns for bulk update."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

    id_etablissement = request.GET.get('id_etablissement')
    classe_par_annee_id = request.GET.get('classe_par_annee_id')

    if not id_etablissement or not classe_par_annee_id:
        return JsonResponse({'success': False, 'error': 'Paramètres manquants'}, status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compléter Parents"

    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    locked_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    pere_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    mere_fill = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid")
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=10)
    locked_font = Font(name='Arial', size=10, color="64748B")
    data_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'ID Élève', 'Nom', 'Postnom', 'Prénom', 'Genre',
        'Noms Père', 'Tél. Père', 'Email Père',
        'Noms Mère', 'Tél. Mère', 'Email Mère'
    ]

    classe_label = ''
    students = []
    try:
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT cl.nom as classe_nom,
                           COALESCE(s.nom, '') as section_nom,
                           COALESCE(eac.groupe, '') as groupe
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.classes cl ON cl.id = eac.classe_id
                    LEFT JOIN countryStructure.sections s ON s.id = eac.section_id
                    WHERE eac.id = %s
                """, [classe_par_annee_id])
                row = cur.fetchone()
                if row:
                    classe_label = row['classe_nom']
                    if row['section_nom']: classe_label += ' - ' + row['section_nom']
                    if row['groupe']: classe_label += ' (' + row['groupe'] + ')'

                cur.execute("""
                    SELECT eac.classe_id, eac.groupe, eac.section_id
                    FROM countryStructure.etablissements_annees_classes eac WHERE eac.id = %s
                """, [classe_par_annee_id])
                bk = cur.fetchone()
                if bk:
                    cur.execute("""
                        SELECT e.id_eleve, e.nom, e.prenom, e.genre,
                               p.nomsPere, p.telephonePere, p.emailPere,
                               p.nomsMere, p.telephoneMere, p.emailMere
                        FROM eleve e
                        JOIN eleve_inscription ei ON ei.id_eleve_id = e.id_eleve
                        LEFT JOIN parents p ON p.id_parent = e.id_parent
                        WHERE ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                          AND ei.id_etablissement = %s AND ei.id_pays = %s AND ei.status = 1
                        ORDER BY e.nom, e.prenom
                    """, [bk['classe_id'], bk['groupe'], bk['section_id'], id_etablissement,
                          getattr(request, 'id_pays', None) or request.session.get('id_pays')])
                    students = cur.fetchall()
        finally:
            conn.close()
    except Exception:
        pass

    # Info row
    info_font = Font(name='Arial', bold=True, color="4338CA", size=9)
    ws.cell(row=1, column=1, value=f"Classe: {classe_label} — Complétez les colonnes Parents (colonnes vertes)").font = info_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Pere/Mere header colors
    for col in range(6, 9):  # Père columns
        ws.cell(row=2, column=col).fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    for col in range(9, 12):  # Mère columns
        ws.cell(row=2, column=col).fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")

    # Column widths
    widths = [10, 18, 16, 16, 8, 24, 16, 22, 24, 16, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Data rows
    for idx, stu in enumerate(students):
        r = 3 + idx
        nom_parts = (stu['nom'] or '').split(' ', 1)
        nom_val = nom_parts[0] if nom_parts else ''
        postnom_val = nom_parts[1] if len(nom_parts) > 1 else ''

        locked_row = [
            stu['id_eleve'], nom_val, postnom_val, stu['prenom'] or '', stu['genre'] or ''
        ]
        parent_row = [
            stu.get('nomsPere') or '', stu.get('telephonePere') or '', stu.get('emailPere') or '',
            stu.get('nomsMere') or '', stu.get('telephoneMere') or '', stu.get('emailMere') or '',
        ]

        # Locked columns (student identity)
        for col, val in enumerate(locked_row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = locked_font
            cell.fill = locked_fill
            cell.border = border

        # Parent columns (editable)
        for col, val in enumerate(parent_row, 6):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = data_font
            cell.border = border
            if col <= 8:
                cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FDF2F8", end_color="FDF2F8", fill_type="solid")

    # Meta sheet
    ws_meta = wb.create_sheet("_Meta")
    ws_meta.cell(row=1, column=1, value="classe_par_annee_id")
    ws_meta.cell(row=1, column=2, value=int(classe_par_annee_id))
    ws_meta.cell(row=2, column=1, value="type")
    ws_meta.cell(row=2, column=2, value="parent_update")
    ws_meta.sheet_state = 'hidden'

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f"completer_parents_{classe_label.replace(' ','_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ============================================================
# API — Import Parent Updates from Excel
# ============================================================
@require_http_methods(["POST"])
def import_parent_updates(request):
    """Import parent data from Excel and create/link parents to students."""
    try:
        import openpyxl

        file = request.FILES.get('file')
        id_etablissement = request.POST.get('id_etablissement')

        if not file or not id_etablissement:
            return JsonResponse({'success': False, 'error': 'Fichier et établissement requis.'}, status=400)

        id_etablissement = int(id_etablissement)
        wb = openpyxl.load_workbook(file, read_only=False)
        ws = wb.active

        # Find data start (skip info row)
        data_start = 3  # row 1=info, row 2=headers, row 3+=data
        rows = list(ws.iter_rows(min_row=data_start, values_only=True))
        if not rows:
            return JsonResponse({'success': False, 'error': 'Le fichier est vide.'}, status=400)

        conn = _get_spoke_connection()
        updated = 0
        created_parents = 0
        errors = []
        try:
            with conn.cursor() as cur:
                for i, row in enumerate(rows, data_start):
                    try:
                        if not row or len(row) < 8:
                            continue
                        non_empty = [c for c in row if c is not None and str(c).strip()]
                        if not non_empty:
                            continue

                        id_eleve = row[0]
                        if not id_eleve:
                            continue
                        id_eleve = int(id_eleve)

                        # Parent data (new simplified columns)
                        noms_pere = str(row[5] or '').strip()
                        tel_pere = str(row[6] or '').strip()
                        email_pere = str(row[7] or '').strip()
                        noms_mere = str(row[8] or '').strip() if len(row) > 8 else ''
                        tel_mere = str(row[9] or '').strip() if len(row) > 9 else ''
                        email_mere = str(row[10] or '').strip() if len(row) > 10 else ''

                        if not noms_pere and not noms_mere:
                            continue  # No parent data provided

                        # Check if parent already exists (match by nomsPere + telephonePere)
                        existing_parent = None
                        if noms_pere and tel_pere:
                            cur.execute("""
                                SELECT id_parent FROM parents
                                WHERE nomsPere = %s AND telephonePere = %s
                                LIMIT 1
                            """, [noms_pere, tel_pere])
                            existing_parent = cur.fetchone()
                        elif noms_pere:
                            cur.execute("""
                                SELECT id_parent FROM parents
                                WHERE nomsPere = %s AND nomsMere = %s
                                LIMIT 1
                            """, [noms_pere, noms_mere])
                            existing_parent = cur.fetchone()

                        if existing_parent:
                            parent_id = existing_parent['id_parent']
                            # Update parent with any new info
                            cur.execute("""
                                UPDATE parents SET
                                    nomsPere = CASE WHEN %s != '' THEN %s ELSE nomsPere END,
                                    telephonePere = CASE WHEN %s != '' THEN %s ELSE telephonePere END,
                                    emailPere = CASE WHEN %s != '' THEN %s ELSE emailPere END,
                                    nomsMere = CASE WHEN %s != '' THEN %s ELSE nomsMere END,
                                    telephoneMere = CASE WHEN %s != '' THEN %s ELSE telephoneMere END,
                                    emailMere = CASE WHEN %s != '' THEN %s ELSE emailMere END
                                WHERE id_parent = %s
                            """, [
                                noms_pere, noms_pere,
                                tel_pere, tel_pere,
                                email_pere, email_pere,
                                noms_mere, noms_mere,
                                tel_mere, tel_mere,
                                email_mere, email_mere,
                                parent_id
                            ])
                        else:
                            # Create new parent
                            id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
                            cur.execute("""
                                INSERT INTO parents (nomsPere, telephonePere, emailPere,
                                                     nomsMere, telephoneMere, emailMere, id_pays)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """, [
                                noms_pere or '', tel_pere or '', email_pere or '',
                                noms_mere or '', tel_mere or '', email_mere or '',
                                int(id_pays)
                            ])
                            parent_id = cur.lastrowid
                            created_parents += 1

                        # Link student to parent
                        cur.execute("UPDATE eleve SET id_parent = %s WHERE id_eleve = %s", [parent_id, id_eleve])
                        updated += 1

                    except Exception as row_err:
                        errors.append(f"Ligne {i}: {str(row_err)}")

                conn.commit()

            return JsonResponse({
                'success': True,
                'updated': updated,
                'created_parents': created_parents,
                'errors': errors[:20]
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# ============================================================
# API — Campus CRUD
# ============================================================
@require_http_methods(["GET"])
def dashboard_campus_list(request):
    """List campuses filtered by id_etablissement."""
    id_etablissement = request.GET.get('id_etablissement')
    id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
    try:
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                if id_etablissement and id_pays:
                    cur.execute("SELECT * FROM campus WHERE id_etablissement=%s AND id_pays=%s ORDER BY campus", [id_etablissement, id_pays])
                elif id_etablissement:
                    cur.execute("SELECT * FROM campus WHERE id_etablissement=%s ORDER BY campus", [id_etablissement])
                else:
                    cur.execute("SELECT * FROM campus WHERE id_pays=%s ORDER BY campus", [id_pays]) if id_pays else cur.execute("SELECT * FROM campus ORDER BY campus")
                rows = cur.fetchall()
                return JsonResponse({'success': True, 'campus': [
                    {
                        'idCampus': r['idCampus'],
                        'campus': r['campus'],
                        'adresse': r['adresse'] or '',
                        'localisation': r['localisation'] or '',
                        'is_active': bool(r['is_active']),
                        'id_etablissement': r['id_etablissement'],
                    } for r in rows
                ]})
        finally:
            conn.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def dashboard_campus_create(request):
    """Create a new campus."""
    import json
    data = json.loads(request.body)
    campus_name = data.get('campus', '').strip()
    if not campus_name:
        return JsonResponse({'success': False, 'error': 'Nom du campus requis'}, status=400)
    try:
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Auto-increment id_campus per establishment
                etab_id = data.get('id_etablissement')
                id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
                cur.execute(
                    "SELECT COALESCE(MAX(id_campus), 0) + 1 AS next_id FROM campus WHERE id_etablissement = %s AND id_pays = %s",
                    [etab_id, id_pays]
                )
                next_id_campus = cur.fetchone()['next_id']
                cur.execute("""
                    INSERT INTO campus (id_campus, campus, adresse, localisation, is_active, id_etablissement, id_pays)
                    VALUES (%s, %s, %s, %s, 1, %s, %s)
                """, [
                    next_id_campus,
                    campus_name,
                    data.get('adresse', ''),
                    data.get('localisation', ''),
                    etab_id,
                    id_pays,
                ])
                conn.commit()
                return JsonResponse({'success': True, 'idCampus': cur.lastrowid, 'id_campus': next_id_campus})
        finally:
            conn.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def dashboard_campus_update(request):
    """Update a campus (name, adresse, localisation, is_active)."""
    import json
    data = json.loads(request.body)
    campus_id = data.get('idCampus')
    if not campus_id:
        return JsonResponse({'success': False, 'error': 'id_campus requis'}, status=400)
    try:
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                sets = []
                params = []
                for field in ('campus', 'adresse', 'localisation'):
                    if field in data:
                        sets.append(f"{field} = %s")
                        params.append(data[field])
                if 'is_active' in data:
                    sets.append("is_active = %s")
                    params.append(int(data['is_active']))
                if not sets:
                    return JsonResponse({'success': False, 'error': 'Aucun champ à modifier'}, status=400)
                params.append(campus_id)
                cur.execute(f"UPDATE campus SET {', '.join(sets)} WHERE idCampus = %s", params)
                conn.commit()
                return JsonResponse({'success': True})
        finally:
            conn.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def dashboard_campus_delete(request):
    """Delete a campus (only if no students are enrolled)."""
    import json
    data = json.loads(request.body)
    campus_id = data.get('idCampus')
    if not campus_id:
        return JsonResponse({'success': False, 'error': 'id_campus requis'}, status=400)
    try:
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Check if students are enrolled in this campus
                cur.execute("SELECT COUNT(*) as nb FROM eleve_inscription WHERE idCampus_id = %s", [campus_id])
                row = cur.fetchone()
                if row and row['nb'] > 0:
                    return JsonResponse({
                        'success': False,
                        'error': f"Impossible de supprimer : {row['nb']} élève(s) inscrit(s) dans ce campus."
                    }, status=400)
                cur.execute("DELETE FROM campus WHERE idCampus = %s", [campus_id])
                conn.commit()
                return JsonResponse({'success': True})
        finally:
            conn.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# API — Eleves stats (filtrable par classe)
@require_http_methods(["GET"])
def dashboard_eleves_stats(request):
    """Return eleves stats, optionally filtered by classe_par_annee_id."""
    id_etablissement = request.GET.get('id_etablissement')
    classe_par_annee_id = request.GET.get('classe_par_annee_id') or request.GET.get('classe_par_annee_id')  # optional filter

    if not id_etablissement:
        return JsonResponse({'success': False, 'error': 'id_etablissement requis'}, status=400)

    try:
        from django.db import connections
        db_settings = connections['default'].settings_dict
        import pymysql
        conn = pymysql.connect(
            host=db_settings.get('HOST', 'localhost') or 'localhost',
            user=db_settings['USER'],
            password=db_settings['PASSWORD'],
            port=int(db_settings.get('PORT', 3306) or 3306),
            database=connections['default'].settings_dict['NAME'],
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=5,
            read_timeout=10,
        )
        try:
            with conn.cursor() as cur:
                # Get campus for this establishment
                id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement=%s AND id_pays=%s AND is_active=1",
                            [id_etablissement, id_pays])
                campus_ids = [r['idCampus'] for r in cur.fetchall()]

                if not campus_ids:
                    return JsonResponse({'success': True, 'total': 0, 'garcons': 0, 'filles': 0,
                                         'age_gender': []})

                placeholders = ','.join(['%s'] * len(campus_ids))
                params = list(campus_ids)

                # Build WHERE clause
                where = f"ei.idCampus_id IN ({placeholders}) AND ei.status=1 AND ei.id_pays=%s"
                params.append(id_pays)

                id_annee = request.GET.get('id_annee')
                if id_annee:
                    where += " AND ei.id_annee_id=%s"
                    params.append(int(id_annee))

                if classe_par_annee_id:
                    # Resolve EAC.id → business keys
                    cur.execute("""
                        SELECT eac.classe_id, eac.groupe, eac.section_id
                        FROM countryStructure.etablissements_annees_classes eac WHERE eac.id = %s
                    """, [int(classe_par_annee_id)])
                    bk = cur.fetchone()
                    if bk:
                        where += " AND ei.classe_id=%s AND ei.groupe <=> %s AND ei.section_id <=> %s"
                        params.extend([bk['classe_id'], bk['groupe'], bk['section_id']])

                # Gender totals
                cur.execute(f"""
                    SELECT COUNT(*) as total,
                           SUM(CASE WHEN e.genre='M' THEN 1 ELSE 0 END) as garcons,
                           SUM(CASE WHEN e.genre='F' THEN 1 ELSE 0 END) as filles
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve=ei.id_eleve_id
                    WHERE {where}
                """, params)
                row = cur.fetchone()
                total = int(row['total'] or 0)
                garcons = int(row['garcons'] or 0)
                filles = int(row['filles'] or 0)

                # Age + gender breakdown (per year)
                cur.execute(f"""
                    SELECT TIMESTAMPDIFF(YEAR, e.date_naissance, CURDATE()) as age,
                           COUNT(*) as total,
                           SUM(CASE WHEN e.genre='M' THEN 1 ELSE 0 END) as garcons,
                           SUM(CASE WHEN e.genre='F' THEN 1 ELSE 0 END) as filles
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve=ei.id_eleve_id
                    WHERE {where}
                          AND e.date_naissance IS NOT NULL
                          AND e.date_naissance != '0000-00-00'
                    GROUP BY age ORDER BY age
                """, params)
                age_gender = [
                    {'age': int(r['age']), 'total': int(r['total']),
                     'garcons': int(r['garcons'] or 0), 'filles': int(r['filles'] or 0)}
                    for r in cur.fetchall()
                ]

                return JsonResponse({
                    'success': True,
                    'total': total,
                    'garcons': garcons,
                    'filles': filles,
                    'age_gender': age_gender,
                })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# API — List students for a class (for inline grid editing)
# ============================================================
@require_http_methods(["GET"])
def dashboard_eleves_list(request):
    """Return the list of students enrolled in a specific class with all attributes."""
    import sys
    id_etablissement = request.GET.get('id_etablissement')
    classe_par_annee_id = request.GET.get('classe_par_annee_id') or request.GET.get('classe_par_annee_id')
    search_query = request.GET.get('search', '').strip()
    print(f"[dashboard_eleves_list] CALLED: id_etablissement={id_etablissement}, classe_par_annee_id={classe_par_annee_id}, search={search_query}", file=sys.stderr, flush=True)

    if not id_etablissement:
        return JsonResponse({'success': False, 'error': 'id_etablissement requis'}, status=400)

    # Search mode: find students by name across all classes
    if search_query and not classe_par_annee_id:
        try:
            conn = _get_spoke_connection()
            try:
                with conn.cursor() as cur:
                    like_q = f'%{search_query}%'
                    id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
                    cur.execute("""
                        SELECT e.id_eleve, e.numero_serie, e.nom, e.prenom, e.genre,
                               e.date_naissance, e.telephone, e.id_parent, e.IDNational,
                               e.ref_administrative_naissance, e.ref_administrative_residence,
                               e.imageUrl,
                               p.nomsPere, p.telephonePere,
                               p.nomsMere, p.telephoneMere,
                               ei.id_inscription, ei.date_inscription, ei.redoublement, ei.isDelegue
                        FROM eleve e
                        JOIN eleve_inscription ei ON ei.id_eleve_id = e.id_eleve
                        LEFT JOIN parents p ON p.id_parent = e.id_parent
                        WHERE ei.id_etablissement = %s AND ei.id_pays = %s AND ei.status = 1
                          AND (e.nom LIKE %s OR e.prenom LIKE %s OR e.matricule LIKE %s)
                        ORDER BY e.nom, e.prenom
                        LIMIT 50
                    """, [id_etablissement, id_pays, like_q, like_q, like_q])
                    students = cur.fetchall()
                    result = []
                    for stu in students:
                        dn = stu.get('date_naissance')
                        dn_str = dn.strftime('%Y-%m-%d') if dn and hasattr(dn, 'strftime') else (str(dn) if dn else '')
                        full_nom = stu.get('nom') or ''
                        nom_parts = full_nom.split(' ', 1)
                        result.append({
                            'id_eleve': stu['id_eleve'],
                            'id_inscription': stu.get('id_inscription'),
                            'numero_serie': stu.get('numero_serie') or '',
                            'nom': nom_parts[0] if nom_parts else '',
                            'postnom': nom_parts[1] if len(nom_parts) > 1 else '',
                            'prenom': stu.get('prenom') or '',
                            'genre': stu.get('genre') or '',
                            'date_naissance': dn_str,
                            'telephone': stu.get('telephone') or '',
                            'IDNational': stu.get('IDNational') or '',
                            'id_parent': stu.get('id_parent') or 0,
                            'nom_pere': stu.get('nomsPere') or '',
                            'tel_pere': stu.get('telephonePere') or '',
                            'nom_mere': stu.get('nomsMere') or '',
                            'tel_mere': stu.get('telephoneMere') or '',
                            'imageUrl': stu.get('imageUrl') or '',
                            'ref_administrative_naissance': stu.get('ref_administrative_naissance') or '',
                            'ref_administrative_residence': stu.get('ref_administrative_residence') or '',
                            'date_inscription': str(stu['date_inscription']) if stu.get('date_inscription') else '',
                            'redoublement': stu.get('redoublement', 0),
                            'isDelegue': stu.get('isDelegue', 0),
                        })
                    return JsonResponse({'success': True, 'eleves': result, 'total': len(result)})
            finally:
                conn.close()
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    if not classe_par_annee_id:
        return JsonResponse({'success': False, 'error': 'classe_par_annee_id requis'}, status=400)

    try:
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Resolve EAC.id → business keys
                cur.execute("""
                    SELECT eac.classe_id, eac.groupe, eac.section_id
                    FROM countryStructure.etablissements_annees_classes eac WHERE eac.id = %s
                """, [classe_par_annee_id])
                bk = cur.fetchone()
                print(f"[dashboard_eleves_list] EAC lookup for id={classe_par_annee_id}: bk={bk}", file=sys.stderr, flush=True)
                if not bk:
                    return JsonResponse({'success': False, 'error': 'Classe introuvable'}, status=404)

                cur.execute("""
                    SELECT e.id_eleve, e.numero_serie, e.nom, e.prenom, e.genre,
                           e.date_naissance, e.telephone, e.imageUrl, e.id_parent,
                           e.ref_administrative_naissance, e.ref_administrative_residence, e.IDNational,
                           p.nomsPere, p.telephonePere, p.emailPere,
                           p.nomsMere, p.telephoneMere, p.emailMere,
                           ei.id_inscription, ei.date_inscription, ei.redoublement,
                           ei.status, ei.isDelegue
                    FROM eleve e
                    JOIN eleve_inscription ei ON ei.id_eleve_id = e.id_eleve
                    LEFT JOIN parents p ON p.id_parent = e.id_parent
                    WHERE ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                      AND ei.id_etablissement = %s AND ei.id_pays = %s AND ei.status = 1
                    ORDER BY e.nom, e.prenom
                """, [bk['classe_id'], bk['groupe'], bk['section_id'], id_etablissement,
                      getattr(request, 'id_pays', None) or request.session.get('id_pays')])
                students = cur.fetchall()

                result = []
                for stu in students:
                    dn = stu.get('date_naissance')
                    dn_str = ''
                    if dn:
                        dn_str = dn.strftime('%Y-%m-%d') if hasattr(dn, 'strftime') else str(dn)

                    # Split nom back into nom + postnom
                    full_nom = stu.get('nom') or ''
                    nom_parts = full_nom.split(' ', 1)
                    nom_val = nom_parts[0] if nom_parts else ''
                    postnom_val = nom_parts[1] if len(nom_parts) > 1 else ''

                    result.append({
                        'id_eleve': stu['id_eleve'],
                        'numero_serie': stu.get('numero_serie') or '',
                        'nom': nom_val,
                        'postnom': postnom_val,
                        'prenom': stu.get('prenom') or '',
                        'genre': stu.get('genre') or '',
                        'date_naissance': dn_str,
                        'id_parent': stu.get('id_parent') or 0,
                        'nom_pere': stu.get('nomsPere') or '',
                        'tel_pere': stu.get('telephonePere') or '',
                        'nom_mere': stu.get('nomsMere') or '',
                        'tel_mere': stu.get('telephoneMere') or '',
                        'telephone': stu.get('telephone') or '',
                        'imageUrl': stu.get('imageUrl') or '',
                        'id_inscription': stu.get('id_inscription'),
                        'date_inscription': str(stu['date_inscription']) if stu.get('date_inscription') else '',
                        'redoublement': stu.get('redoublement', 0),
                        'isDelegue': stu.get('isDelegue', 0),
                        'ref_administrative_naissance': stu.get('ref_administrative_naissance') or '',
                        'ref_administrative_residence': stu.get('ref_administrative_residence') or '',
                        'IDNational': stu.get('IDNational') or '',
                    })

                return JsonResponse({'success': True, 'eleves': result, 'total': len(result)})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# API — Update a student inline
# ============================================================
@require_http_methods(["POST"])
def dashboard_update_eleve(request):
    """Update one or more fields of a student in db_monecole."""
    try:
        data = json.loads(request.body)
        id_eleve = data.get('id_eleve')
        if not id_eleve:
            return JsonResponse({'success': False, 'error': 'id_eleve requis.'}, status=400)

        # Allowed fields for update
        allowed_fields = {
            'numero_serie', 'nom', 'postnom', 'prenom', 'genre',
            'date_naissance', 'telephone', 'id_parent',
            'ref_administrative_naissance', 'ref_administrative_residence',
            'IDNational'
        }

        updates = {}
        for key, val in data.items():
            if key in allowed_fields:
                updates[key] = val

        if not updates:
            return JsonResponse({'success': False, 'error': 'Aucun champ à mettre à jour.'}, status=400)

        # Handle nom + postnom combination
        if 'nom' in updates or 'postnom' in updates:
            nom = updates.pop('nom', None)
            postnom = updates.pop('postnom', None)
            # We need both to combine — fetch existing if only one is provided
            if nom is None or postnom is None:
                conn_temp = _get_spoke_connection()
                try:
                    with conn_temp.cursor() as cur_temp:
                        cur_temp.execute("SELECT nom, prenom FROM eleve WHERE id_eleve=%s", [id_eleve])
                        existing = cur_temp.fetchone()
                        if existing:
                            existing_parts = (existing.get('nom') or '').split(' ', 1)
                            if nom is None:
                                nom = existing_parts[0] if existing_parts else ''
                            if postnom is None:
                                postnom = existing_parts[1] if len(existing_parts) > 1 else ''
                finally:
                    conn_temp.close()
            full_nom = f"{nom} {postnom}".strip() if postnom else (nom or '')
            updates['nom'] = full_nom

        if not updates:
            return JsonResponse({'success': True, 'message': 'Rien à mettre à jour.'})

        set_clauses = ', '.join([f"{k}=%s" for k in updates.keys()])
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
        values = list(updates.values()) + [id_eleve, int(id_pays)]

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(f"UPDATE eleve SET {set_clauses} WHERE id_eleve=%s AND id_pays=%s", values)
                conn.commit()
            return JsonResponse({'success': True, 'message': 'Élève mis à jour.'})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# API — Upload student photo
# ============================================================
@require_http_methods(["POST"])
def dashboard_upload_photo(request):
    """Upload a student photo and save to media/Photos/EtabID_x/Image_x_y.ext"""
    try:
        photo = request.FILES.get('photo')
        id_eleve = request.POST.get('id_eleve')
        id_etablissement = request.POST.get('id_etablissement')

        if not photo or not id_eleve or not id_etablissement:
            return JsonResponse({'success': False, 'error': 'photo, id_eleve et id_etablissement requis.'}, status=400)

        import os
        from django.conf import settings

        # Create directory structure: media/Photos/EtabID_x/
        photos_dir = os.path.join(settings.MEDIA_ROOT, 'Photos', f'EtabID_{id_etablissement}')
        os.makedirs(photos_dir, exist_ok=True)

        # Also ensure Documents dir exists
        docs_dir = os.path.join(settings.MEDIA_ROOT, 'Documents', f'EtabID_{id_etablissement}')
        os.makedirs(docs_dir, exist_ok=True)

        # Determine file extension
        ext = os.path.splitext(photo.name)[1].lower() or '.jpg'
        filename = f"Image_{id_etablissement}_{id_eleve}{ext}"
        filepath = os.path.join(photos_dir, filename)

        # Remove old photos with different extensions
        for old_ext in ['.png', '.jpg', '.jpeg', '.gif', '.webp']:
            old_file = os.path.join(photos_dir, f"Image_{id_etablissement}_{id_eleve}{old_ext}")
            if os.path.exists(old_file) and old_file != filepath:
                try:
                    os.remove(old_file)
                except Exception:
                    pass

        # Save file
        with open(filepath, 'wb+') as f:
            for chunk in photo.chunks():
                f.write(chunk)

        # Build URL path
        image_url = f"/media/Photos/EtabID_{id_etablissement}/{filename}"

        # Update imageUrl in eleve table
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Ensure imageUrl column exists (defensive)
                try:
                    cur.execute("ALTER TABLE eleve ADD COLUMN imageUrl VARCHAR(255) DEFAULT NULL")
                    conn.commit()
                except Exception:
                    conn.rollback()  # Column already exists, ignore
                cur.execute("UPDATE eleve SET imageUrl=%s WHERE id_eleve=%s AND id_pays=%s",
                            [image_url, id_eleve,
                             int(getattr(request, 'id_pays', None) or request.session.get('id_pays') or 0)])
                conn.commit()
            return JsonResponse({'success': True, 'imageUrl': image_url})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# PERSONNEL / ENSEIGNANTS MANAGEMENT
# ============================================================

@require_http_methods(["GET"])
def dashboard_personnel_list(request):
    """List all personnel for the establishment, with reference table labels."""
    id_etablissement = request.GET.get('id_etablissement')
    if not id_etablissement:
        return JsonResponse({'success': False, 'error': 'id_etablissement requis'}, status=400)
    try:
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT p.*,
                           pc.categorie AS categorie_label, pc.sigle AS categorie_sigle,
                           pt.type AS type_label, pt.sigle AS type_sigle,
                           pa.poste AS poste_label,
                           d.diplome AS diplome_label, d.sigle AS diplome_sigle,
                           sp.specialite AS specialite_label, sp.sigle AS specialite_sigle,
                           v.vacation AS vacation_label, v.sigle AS vacation_sigle
                    FROM personnel p
                    LEFT JOIN personnel_categorie pc ON pc.id_personnel_category = p.id_categorie_id
                    LEFT JOIN personnel_type pt ON pt.id_type_personnel = p.id_personnel_type_id
                    LEFT JOIN personnel_posteAdministratif pa ON pa.id_posteAdministratif = p.id_posteAdministratif
                    LEFT JOIN diplome d ON d.id_diplome = p.id_diplome_id
                    LEFT JOIN specialite sp ON sp.id_specialite = p.id_specialite_id
                    LEFT JOIN vacation v ON v.id_vacation = p.id_vacation_id
                    WHERE p.id_etablissement = %s AND p.id_pays = %s
                    ORDER BY p.nom, p.postnom
                """, [id_etablissement, getattr(request, 'id_pays', None) or request.session.get('id_pays')])
                rows = cur.fetchall()

                # Load ALL reference data for dropdowns
                cur.execute("SELECT * FROM personnel_categorie ORDER BY id_personnel_category")
                categories = cur.fetchall()
                cur.execute("SELECT * FROM personnel_type ORDER BY id_type_personnel")
                types = cur.fetchall()
                cur.execute("SELECT * FROM personnel_posteAdministratif ORDER BY id_posteAdministratif")
                postes = cur.fetchall()
                cur.execute("SELECT * FROM diplome ORDER BY id_diplome")
                diplomes = cur.fetchall()
                cur.execute("SELECT * FROM specialite ORDER BY id_specialite")
                specialites = cur.fetchall()
                cur.execute("SELECT * FROM vacation ORDER BY id_vacation")
                vacations = cur.fetchall()
                cur.execute("SELECT * FROM personnelEnseignant_Taches ORDER BY id_tache")
                taches = cur.fetchall()

                result = []
                for r in rows:
                    dn = r.get('date_naissance')
                    dc = r.get('date_creation')
                    # Safely format dates (can be datetime, date, or str)
                    def safe_date(val):
                        if not val:
                            return ''
                        if isinstance(val, str):
                            return val[:10]  # Already a string, take YYYY-MM-DD part
                        try:
                            return val.strftime('%Y-%m-%d')
                        except (AttributeError, ValueError):
                            return str(val)[:10] if val else ''
                    result.append({
                        'id_personnel': r['id_personnel'],
                        'nom': r.get('nom') or '',
                        'postnom': r.get('postnom') or '',
                        'prenom': r.get('prenom') or '',
                        'matricule': r.get('matricule') or '',
                        'genre': r.get('genre') or '',
                        'date_naissance': safe_date(dn),
                        'etat_civil': r.get('etat_civil') or '',
                        'type_identite': r.get('type_identite') or '',
                        'numero_identite': r.get('numero_identite') or '',
                        'telephone': r.get('telephone') or '',
                        'email': r.get('email') or '',
                        'password': r.get('password') or '',
                        'addresse': r.get('addresse') or '',
                        'zone': r.get('zone') or '',
                        'commune': r.get('commune') or '',
                        'province': r.get('province') or '',
                        'region': r.get('region') or '',
                        'pays': r.get('pays') or '',
                        'imageUrl': r.get('imageUrl') or '',
                        'identiteUrl': r.get('identiteUrl') or '',
                        'id_categorie_id': r.get('id_categorie_id'),
                        'categorie_label': r.get('categorie_label') or '',
                        'categorie_sigle': r.get('categorie_sigle') or '',
                        'id_personnel_type_id': r.get('id_personnel_type_id'),
                        'type_label': r.get('type_label') or '',
                        'id_posteAdministratif': r.get('id_posteAdministratif'),
                        'poste_label': r.get('poste_label') or '',
                        'id_specialite_id': r.get('id_specialite_id'),
                        'specialite_label': r.get('specialite_label') or '',
                        'id_diplome_id': r.get('id_diplome_id'),
                        'diplome_label': r.get('diplome_label') or '',
                        'diplome_sigle': r.get('diplome_sigle') or '',
                        'id_vacation_id': r.get('id_vacation_id'),
                        'vacation_label': r.get('vacation_label') or '',
                        'id_tache': r.get('id_tache'),
                        'isMaitresse': r.get('isMaitresse', 0),
                        'isInstiteur': r.get('isInstiteur', 0),
                        'isDAF': r.get('isDAF', 0),
                        'isDirecteur': r.get('isDirecteur', 0),
                        'en_fonction': r.get('en_fonction', 0),
                        'isUser': r.get('isUser', 0),
                        'is_verified': r.get('is_verified', 0),
                        'email_verified': r.get('email_verified', 0),
                        'phone_verified': r.get('phone_verified', 0),
                        'date_creation': safe_date(dc),
                        'codeAnnee': r.get('codeAnnee') or '',
                    })
                # Stats
                total = len(result)
                en_fonction = sum(1 for r in result if r['en_fonction'])
                par_cat = {}
                for r in result:
                    s = r['categorie_sigle'] or 'N/A'
                    par_cat[s] = par_cat.get(s, 0) + 1
                par_genre = {}
                for r in result:
                    g = r['genre'] or 'N/A'
                    par_genre[g] = par_genre.get(g, 0) + 1

                ref = {
                    'categories': [{'id': c['id_personnel_category'], 'label': c['categorie'], 'sigle': c['sigle']} for c in categories],
                    'types': [{'id': t['id_type_personnel'], 'label': t['type'], 'sigle': t.get('sigle','')} for t in types],
                    'postes': [{'id': p['id_posteAdministratif'], 'label': p['poste']} for p in postes],
                    'diplomes': [{'id': d['id_diplome'], 'label': d['diplome'], 'sigle': d.get('sigle','')} for d in diplomes],
                    'specialites': [{'id': s['id_specialite'], 'label': s['specialite'], 'sigle': s.get('sigle','')} for s in specialites],
                    'vacations': [{'id': v['id_vacation'], 'label': v['vacation'], 'sigle': v.get('sigle','')} for v in vacations],
                    'taches': [{'id': t['id_tache'], 'label': t['tache']} for t in taches],
                }

                return JsonResponse({
                    'success': True,
                    'personnel': result,
                    'stats': {'total': total, 'en_fonction': en_fonction, 'par_categorie': par_cat, 'par_genre': par_genre},
                    'ref': ref,
                })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def dashboard_add_personnel(request):
    """Add a new personnel member. Matricule auto-generated as M_x_y.
    Plus de création dans auth_user — tout passe par personnel directement.
    Sets email_verified=0 and phone_verified=0 for new personnel.
    """
    try:
        data = json.loads(request.body)
        id_etablissement = data.get('id_etablissement')
        nom = data.get('nom', '').strip()
        postnom = data.get('postnom', '').strip()
        prenom = data.get('prenom', '').strip()
        genre = data.get('genre', '').strip()
        if not id_etablissement or not nom or not genre:
            return JsonResponse({'success': False, 'error': 'Nom et genre sont requis.'}, status=400)

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Ensure new columns exist
                for col_name in ('email_verified', 'phone_verified', 'username', 'password_hash', 'last_login'):
                    try:
                        if col_name == 'password_hash':
                            cur.execute(f"ALTER TABLE personnel ADD COLUMN {col_name} VARCHAR(255) DEFAULT ''")
                        elif col_name == 'username':
                            cur.execute(f"ALTER TABLE personnel ADD COLUMN {col_name} VARCHAR(150) DEFAULT NULL")
                        elif col_name == 'last_login':
                            cur.execute(f"ALTER TABLE personnel ADD COLUMN {col_name} DATETIME DEFAULT NULL")
                        else:
                            cur.execute(f"ALTER TABLE personnel ADD COLUMN {col_name} TINYINT(1) NOT NULL DEFAULT 0")
                        conn.commit()
                    except Exception:
                        conn.rollback()  # Column already exists

                # Générer un username unique
                import time as _time
                ts = _time.time()
                username = f"pers_{id_etablissement}_{int(ts)}"
                email_val = data.get('email') or ''

                fields = {
                    'nom': nom,
                    'postnom': postnom,
                    'prenom': prenom,
                    'matricule': f'TEMP_{id_etablissement}_{int(ts)}',
                    'genre': genre,
                    'date_naissance': data.get('date_naissance') or None,
                    'etat_civil': data.get('etat_civil', ''),
                    'type_identite': data.get('type_identite', ''),
                    'numero_identite': data.get('numero_identite') or None,
                    'telephone': data.get('telephone') or None,
                    'email': email_val or None,
                    'password': data.get('password') or None,
                    'addresse': data.get('addresse', ''),
                    'zone': data.get('zone', ''),
                    'commune': data.get('commune', ''),
                    'province': data.get('province', ''),
                    'region': data.get('region') or None,
                    'pays': data.get('pays') or None,
                    'id_categorie_id': int(data.get('id_categorie_id') or 1),
                    'id_diplome_id': int(data.get('id_diplome_id') or 1),
                    'id_personnel_type_id': int(data.get('id_personnel_type_id') or 1),
                    'id_specialite_id': int(data.get('id_specialite_id') or 1),
                    'id_vacation_id': int(data.get('id_vacation_id') or 1),
                    'id_posteAdministratif': int(data.get('id_posteAdministratif') or 0) or None,
                    'id_tache': int(data.get('id_tache') or 0) or None,
                    'isMaitresse': int(data.get('isMaitresse', 0)),
                    'isInstiteur': int(data.get('isInstiteur', 0)),
                    'isDAF': int(data.get('isDAF', 0)),
                    'isDirecteur': int(data.get('isDirecteur', 0)),
                    'isUser': 0,
                    'en_fonction': int(data.get('en_fonction', 1)),
                    'is_verified': 0,
                    'email_verified': 0,
                    'phone_verified': 0,
                    'username': username,
                    'password_hash': '',
                    'id_etablissement': int(id_etablissement),
                    'imageUrl': '',
                    'identiteUrl': '',
                    'code_secret': None,
                    'codeAnnee': data.get('codeAnnee') or None,
                    'date_creation': __import__('datetime').date.today().strftime('%Y-%m-%d'),
                    'id_pays': int(getattr(request, 'id_pays', None) or request.session.get('id_pays') or 0),
                }
                cols = ', '.join(fields.keys())
                placeholders = ', '.join(['%s'] * len(fields))
                cur.execute(f"INSERT INTO personnel ({cols}) VALUES ({placeholders})", list(fields.values()))
                new_id = cur.lastrowid

                # Auto-generate matricule: M_x_y
                matricule = f"M_{id_etablissement}_{new_id}"
                cur.execute("UPDATE personnel SET matricule=%s WHERE id_personnel=%s", [matricule, new_id])
                conn.commit()

                # Count total
                cur.execute("SELECT COUNT(*) as c FROM personnel WHERE id_etablissement=%s AND id_pays=%s",
                            [id_etablissement, getattr(request, 'id_pays', None) or request.session.get('id_pays')])
                total = cur.fetchone()['c']

                return JsonResponse({'success': True, 'id_personnel': new_id, 'matricule': matricule, 'total': total})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def dashboard_update_personnel(request):
    """Inline update a personnel field."""
    try:
        data = json.loads(request.body)
        id_personnel = data.get('id_personnel')
        field = data.get('field')
        value = data.get('value')
        if not id_personnel or not field:
            return JsonResponse({'success': False, 'error': 'Missing params'}, status=400)

        # Whitelist fields
        allowed = {
            'nom', 'postnom', 'prenom',
            'matricule', 'genre', 'date_naissance', 'etat_civil', 'type_identite', 'numero_identite',
            'telephone', 'email', 'password', 'addresse', 'zone', 'commune', 'province', 'region', 'pays',
            'id_categorie_id', 'id_diplome_id', 'id_personnel_type_id', 'id_specialite_id', 'id_vacation_id',
            'isMaitresse', 'isInstiteur', 'isDAF', 'isDirecteur', 'en_fonction',
            'id_posteAdministratif', 'id_tache', 'codeAnnee',
        }
        if field not in allowed:
            return JsonResponse({'success': False, 'error': f'Champ "{field}" non modifiable.'}, status=400)

        # For boolean/int fields, convert
        bool_fields = {'isMaitresse', 'isInstiteur', 'isDAF', 'isDirecteur', 'en_fonction'}
        fk_fields = {'id_categorie_id', 'id_diplome_id', 'id_personnel_type_id', 'id_specialite_id',
                      'id_vacation_id', 'id_posteAdministratif', 'id_tache'}
        if field in bool_fields:
            value = 1 if str(value).strip().lower() in ('1', 'true', 'oui', 'yes') else 0
        elif field in fk_fields:
            value = int(value) if value else None

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
                cur.execute(f"UPDATE personnel SET {field}=%s WHERE id_personnel=%s AND id_pays=%s", [value, id_personnel, id_pays])
                conn.commit()
            return JsonResponse({'success': True})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def dashboard_upload_personnel_photo(request):
    """Upload a photo for a personnel member. Same naming convention as student photos."""
    try:
        photo = request.FILES.get('photo')
        id_personnel = request.POST.get('id_personnel')
        id_etablissement = request.POST.get('id_etablissement')
        if not photo or not id_personnel or not id_etablissement:
            return JsonResponse({'success': False, 'error': 'Paramètres manquants.'}, status=400)

        # Create directory
        etab_dir = os.path.join(settings.MEDIA_ROOT, 'Photos', f'EtabID_{id_etablissement}')
        os.makedirs(etab_dir, exist_ok=True)

        # Name: Image_x_y.ext
        ext = os.path.splitext(photo.name)[1] or '.jpg'
        filename = f'Image_{id_etablissement}_{id_personnel}{ext}'
        filepath = os.path.join(etab_dir, filename)

        # Save file
        with open(filepath, 'wb+') as dest:
            for chunk in photo.chunks():
                dest.write(chunk)

        # Update DB
        image_url = f'/media/Photos/EtabID_{id_etablissement}/{filename}'
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
                cur.execute("UPDATE personnel SET imageUrl=%s WHERE id_personnel=%s AND id_pays=%s", [filename, id_personnel, id_pays])
                conn.commit()
            return JsonResponse({'success': True, 'imageUrl': image_url})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# PERSONNEL TEMPLATE DOWNLOAD & IMPORT
# ============================================================

@require_http_methods(["GET"])
def dashboard_personnel_template(request):
    """Download Excel template for personnel import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    id_etablissement = request.GET.get('id_etablissement')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personnel"

    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=10)
    data_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'Nom', 'Postnom', 'Prénom',
        'Année Naissance', 'Mois Naissance', 'Jour Naissance',
        'Genre (M/F)', 'État Civil', 'Type Pièce', 'N° Pièce',
        'Email', 'Téléphone'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 18

    # Store id_etablissement in hidden _Meta sheet
    if id_etablissement:
        ws_meta = wb.create_sheet("_Meta")
        ws_meta.cell(row=1, column=1, value="id_etablissement")
        ws_meta.cell(row=1, column=2, value=int(id_etablissement))
        ws_meta.sheet_state = 'hidden'

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_personnel.xlsx"'
    return response


@require_http_methods(["POST"])
def dashboard_import_personnel(request):
    """Import personnel from uploaded Excel file."""
    try:
        import openpyxl
        from datetime import datetime, date as date_type

        file = request.FILES.get('file')
        id_etablissement = request.POST.get('id_etablissement')

        if not file or not id_etablissement:
            return JsonResponse({'success': False, 'error': 'Fichier et établissement requis.'}, status=400)

        id_etablissement = int(id_etablissement)

        wb = openpyxl.load_workbook(file, read_only=False)
        ws = wb.active

        # Try to read id_etablissement from _Meta if not provided
        if '_Meta' in wb.sheetnames:
            meta_ws = wb['_Meta']
            for row_m in meta_ws.iter_rows(min_row=1, values_only=True):
                if row_m and len(row_m) >= 2 and str(row_m[0]).strip() == 'id_etablissement':
                    id_etablissement = int(row_m[1])
                    break

        # Read header row to build column mapping
        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map = {}
        has_split_date = False
        if header_cells and header_cells[0]:
            headers_raw = [str(h or '').strip().lower() for h in header_cells[0]]
            for idx, h in enumerate(headers_raw):
                if not h:
                    continue
                if 'postnom' in h:
                    col_map['postnom'] = idx
                elif 'prénom' in h or 'prenom' in h:
                    col_map['prenom'] = idx
                elif 'nom' in h:
                    col_map['nom'] = idx
                elif 'année' in h or 'annee' in h or 'year' in h:
                    col_map['annee_naissance'] = idx
                    has_split_date = True
                elif 'mois' in h or 'month' in h:
                    col_map['mois_naissance'] = idx
                    has_split_date = True
                elif 'jour' in h or 'day' in h:
                    col_map['jour_naissance'] = idx
                    has_split_date = True
                elif 'date' in h or 'naissance' in h:
                    col_map['date_naissance'] = idx
                elif 'genre' in h or 'sexe' in h:
                    col_map['genre'] = idx
                elif 'état' in h or 'etat' in h or 'civil' in h:
                    col_map['etat_civil'] = idx
                elif 'type' in h and ('pièce' in h or 'piece' in h or 'identite' in h or 'identité' in h):
                    col_map['type_identite'] = idx
                elif 'n°' in h or 'numero' in h or 'numéro' in h:
                    col_map['numero_identite'] = idx
                elif 'email' in h or 'mail' in h:
                    col_map['email'] = idx
                elif 'tel' in h or 'phone' in h or 'téléphone' in h:
                    col_map['telephone'] = idx

        data_start = 2
        inserted = 0
        updated = 0
        errors = []

        # Diagnostic logging
        print(f"[IMPORT PERSONNEL] col_map: {col_map}")
        print(f"[IMPORT PERSONNEL] has_split_date: {has_split_date}")
        print(f"[IMPORT PERSONNEL] id_etablissement: {id_etablissement}")
        # Log first data row
        first_data_rows = list(ws.iter_rows(min_row=data_start, max_row=data_start, values_only=True))
        if first_data_rows:
            print(f"[IMPORT PERSONNEL] First data row: {first_data_rows[0]}")


        def normalize_pers_date(val):
            """Flexible date parser for personnel import."""
            if val is None:
                return None
            if isinstance(val, (datetime, date_type)):
                return val.strftime('%Y-%m-%d')
            s = str(val).strip()
            if not s:
                return None
            if ' ' in s:
                s = s.split(' ')[0]
            import re
            # YYYY-M-D (accept 1 or 2 digit month/day)
            m_iso = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
            if m_iso:
                return f"{int(m_iso.group(1)):04d}-{int(m_iso.group(2)):02d}-{int(m_iso.group(3)):02d}"
            # Just a year
            if re.match(r'^\d{4}$', s):
                return f"{s}-01-01"
            for fmt in ['%d/%m/%Y','%d-%m-%Y','%d.%m.%Y','%m/%d/%Y','%Y/%m/%d','%d/%m/%y','%d-%m-%y']:
                try:
                    return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
                except ValueError:
                    continue
            return None

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Ensure email_verified and phone_verified columns exist
                for col_name in ('email_verified', 'phone_verified'):
                    try:
                        cur.execute(f"ALTER TABLE personnel ADD COLUMN {col_name} TINYINT(1) NOT NULL DEFAULT 0")
                        conn.commit()
                    except Exception:
                        conn.rollback()  # Column already exists

                for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), data_start):
                    if not row or all(v is None for v in row):
                        continue

                    def get_val(field):
                        if field in col_map:
                            idx = col_map[field]
                            return str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ''
                        return ''

                    nom = get_val('nom')
                    if not nom:
                        errors.append(f"Ligne {row_idx}: Nom vide, ignorée")
                        continue

                    postnom = get_val('postnom')
                    prenom = get_val('prenom')
                    genre = get_val('genre').upper()
                    if genre and genre not in ('M', 'F'):
                        genre = 'M' if genre.startswith('M') else 'F'

                    # Parse date — flexible
                    date_naissance = None
                    if has_split_date:
                        try:
                            annee_raw = get_val('annee_naissance')
                            mois_raw = get_val('mois_naissance')
                            jour_raw = get_val('jour_naissance')
                            annee = int(float(annee_raw)) if annee_raw else None
                            mois = int(float(mois_raw)) if mois_raw else 1
                            jour = int(float(jour_raw)) if jour_raw else 1
                            if annee:
                                date_naissance = f"{annee:04d}-{mois:02d}-{jour:02d}"
                        except (ValueError, TypeError):
                            pass
                    elif 'date_naissance' in col_map:
                        raw = row[col_map['date_naissance']] if col_map['date_naissance'] < len(row) else None
                        date_naissance = normalize_pers_date(raw)

                    etat_civil = get_val('etat_civil')
                    type_identite = get_val('type_identite')
                    numero_identite = get_val('numero_identite')
                    telephone = get_val('telephone')
                    email = get_val('email')

                    try:
                        # Check if personnel already exists (nom + prenom + etablissement + pays)
                        id_pays = int(getattr(request, 'id_pays', None) or request.session.get('id_pays') or 0)
                        cur.execute("""
                            SELECT id_personnel FROM personnel
                            WHERE nom = %s AND prenom = %s AND id_etablissement = %s AND id_pays = %s
                            LIMIT 1
                        """, [nom, prenom, id_etablissement, id_pays])
                        existing = cur.fetchone()

                        if existing:
                            # UPDATE only non-empty values
                            set_parts = []
                            set_vals = []
                            if postnom:
                                set_parts.append("postnom = %s"); set_vals.append(postnom)
                            if genre in ('M', 'F'):
                                set_parts.append("genre = %s"); set_vals.append(genre)
                            if date_naissance:
                                set_parts.append("date_naissance = %s"); set_vals.append(date_naissance)
                            if etat_civil:
                                set_parts.append("etat_civil = %s"); set_vals.append(etat_civil)
                            if type_identite:
                                set_parts.append("type_identite = %s"); set_vals.append(type_identite)
                            if numero_identite:
                                set_parts.append("numero_identite = %s"); set_vals.append(numero_identite)
                            if telephone:
                                set_parts.append("telephone = %s"); set_vals.append(telephone)
                            if email:
                                set_parts.append("email = %s"); set_vals.append(email)
                            if set_parts:
                                set_vals.extend([existing['id_personnel'], id_pays])
                                cur.execute(f"UPDATE personnel SET {', '.join(set_parts)} WHERE id_personnel = %s AND id_pays = %s", set_vals)
                            updated += 1
                        else:
                            # INSERT new personnel directly (no auth_user)
                            import time as _time
                            ts = _time.time()
                            username = f"pers_{id_etablissement}_{row_idx}_{int(ts)}"

                            # INSERT new personnel
                            fields = {
                                'nom': nom,
                                'postnom': postnom,
                                'prenom': prenom,
                                'matricule': f'TEMP_{id_etablissement}_{row_idx}_{int(ts)}',
                                'genre': genre or 'M',
                                'date_naissance': date_naissance,
                                'etat_civil': etat_civil,
                                'type_identite': type_identite,
                                'numero_identite': numero_identite or None,
                                'telephone': telephone or None,
                                'email': email or None,
                                'password': None,
                                'addresse': '',
                                'zone': '',
                                'commune': '',
                                'province': '',
                                'region': None,
                                'pays': None,
                                'id_categorie_id': 1,
                                'id_diplome_id': 1,
                                'id_personnel_type_id': 1,
                                'id_specialite_id': 1,
                                'id_vacation_id': 1,
                                'id_posteAdministratif': None,
                                'id_tache': None,
                                'isMaitresse': 0,
                                'isInstiteur': 0,
                                'isDAF': 0,
                                'isDirecteur': 0,
                                'isUser': 0,
                                'en_fonction': 1,
                                'is_verified': 0,
                                'email_verified': 0,
                                'phone_verified': 0,
                                'username': username,
                                'password_hash': '',
                                'id_etablissement': id_etablissement,
                                'imageUrl': '',
                                'identiteUrl': '',
                                'code_secret': None,
                                'codeAnnee': None,
                                'date_creation': __import__('datetime').date.today().strftime('%Y-%m-%d'),
                                'id_pays': id_pays,
                            }
                            cols = ', '.join(fields.keys())
                            placeholders = ', '.join(['%s'] * len(fields))
                            cur.execute(f"INSERT INTO personnel ({cols}) VALUES ({placeholders})", list(fields.values()))
                            new_id = cur.lastrowid
                            matricule = f"M_{id_etablissement}_{new_id}"
                            cur.execute("UPDATE personnel SET matricule=%s WHERE id_personnel=%s", [matricule, new_id])
                            inserted += 1
                    except Exception as row_err:
                        import traceback
                        print(f"[IMPORT PERSONNEL] Ligne {row_idx} erreur: {row_err}")
                        traceback.print_exc()
                        errors.append(f"Ligne {row_idx}: {str(row_err)}")

                conn.commit()
        finally:
            conn.close()

        return JsonResponse({
            'success': True,
            'inserted': inserted,
            'updated': updated,
            'errors': errors[:20],
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ============================================================
# PERSONNEL REFERENCE TABLES CRUD
# ============================================================

@require_http_methods(["POST"])
def dashboard_personnel_ref_crud(request):
    """Generic CRUD for personnel reference tables."""
    TABLE_MAP = {
        'categories': ('personnel_categorie', 'id_personnel_category', [('categorie', 'Catégorie'), ('sigle', 'Sigle')]),
        'types': ('personnel_type', 'id_type_personnel', [('type', 'Type'), ('sigle', 'Sigle')]),
        'postes': ('personnel_posteAdministratif', 'id_posteAdministratif', [('poste', 'Poste')]),
        'diplomes': ('diplome', 'id_diplome', [('diplome', 'Diplôme'), ('sigle', 'Sigle')]),
        'specialites': ('specialite', 'id_specialite', [('specialite', 'Spécialité'), ('sigle', 'Sigle')]),
        'vacations': ('vacation', 'id_vacation', [('vacation', 'Vacation'), ('sigle', 'Sigle')]),
        'taches': ('personnelEnseignant_Taches', 'id_tache', [('tache', 'Tâche')]),
    }
    try:
        data = json.loads(request.body)
        table_key = data.get('table')
        action = data.get('action')
        if table_key not in TABLE_MAP:
            return JsonResponse({'success': False, 'error': f'Table inconnue: {table_key}'}, status=400)
        table_name, pk_col, columns = TABLE_MAP[table_key]
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                if action == 'list':
                    cur.execute(f"SELECT * FROM {table_name} ORDER BY {pk_col}")
                    return JsonResponse({'success': True, 'rows': cur.fetchall()})
                elif action == 'add':
                    values = data.get('values', {})
                    col_names = [c[0] for c in columns]
                    vals = [values.get(c, '') for c in col_names]
                    if not vals[0]:
                        return JsonResponse({'success': False, 'error': 'La valeur principale est requise.'}, status=400)
                    cur.execute(f"INSERT INTO {table_name} ({', '.join(col_names)}) VALUES ({', '.join(['%s']*len(col_names))})", vals)
                    conn.commit()
                    return JsonResponse({'success': True, 'id': cur.lastrowid})
                elif action == 'update':
                    row_id = data.get('id')
                    field = data.get('field')
                    value = data.get('value')
                    if field not in [c[0] for c in columns]:
                        return JsonResponse({'success': False, 'error': f'Champ non autorisé: {field}'}, status=400)
                    cur.execute(f"UPDATE {table_name} SET {field}=%s WHERE {pk_col}=%s", [value, row_id])
                    conn.commit()
                    return JsonResponse({'success': True})
                elif action == 'delete':
                    row_id = data.get('id')
                    if not row_id:
                        return JsonResponse({'success': False, 'error': 'ID requis.'}, status=400)
                    cur.execute(f"DELETE FROM {table_name} WHERE {pk_col}=%s", [row_id])
                    conn.commit()
                    return JsonResponse({'success': True})
                else:
                    return JsonResponse({'success': False, 'error': f'Action inconnue: {action}'}, status=400)
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# ATTRIBUTION COURS
# ============================================================

@require_http_methods(["POST"])
def dashboard_attribution_cours(request):
    """Manage course-teacher attribution for a class."""
    try:
        data = json.loads(request.body)
        action = data.get('action')  # list, assign, remove
        id_etablissement = data.get('id_etablissement')
        id_classe = data.get('id_classe')

        if not id_etablissement:
            return JsonResponse({'success': False, 'error': 'Établissement requis.'}, status=400)
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                if action == 'list':
                    if not id_classe:
                        return JsonResponse({'success': False, 'error': 'Classe requise.'}, status=400)

                    # Get courses from Hub via ORM
                    # already imported at top
                    from django.db.models import Q
                    eac = EtablissementAnneeClasse.objects.select_related(
                        'classe', 'etablissement_annee', 'etablissement_annee__annee'
                    ).get(id=int(id_classe))
                    real_classe_id = eac.classe_id
                    annee_id = eac.etablissement_annee.annee_id
                    etab_id_hub = eac.etablissement_annee.etablissement_id

                    cours_annee_list = CoursAnnee.objects.filter(
                        cours__classe_id=real_classe_id, annee_id=annee_id, id_pays=id_pays
                    ).filter(
                        Q(etablissement__isnull=True) | Q(etablissement_id=etab_id_hub)
                    ).select_related('cours').order_by('cours__cours')

                    # Now get attributions from spoke for these courses
                    # attribution_cours.id_cours_id → cours_annee.id_cours_annee (NOT cours.id_cours!)
                    ca_annee_ids = [ca.id_cours_annee for ca in cours_annee_list]
                    attributions_map = {}
                    if ca_annee_ids:
                        placeholders = ','.join(['%s'] * len(ca_annee_ids))
                        cur.execute(f"""
                            SELECT ac.id_attribution, ac.id_cours_id, ac.id_personnel_id,
                                   ac.attribution_type_id, ac.date_attribution,
                                   p.nom AS pers_nom, p.postnom AS pers_postnom,
                                   p.prenom AS pers_prenom, p.matricule
                            FROM attribution_cours ac
                            LEFT JOIN personnel p ON p.id_personnel = ac.id_personnel_id
                            WHERE ac.id_cours_id IN ({placeholders})
                              AND ac.classe_id = %s AND ac.groupe <=> %s AND ac.section_id <=> %s
                              AND ac.id_etablissement = %s AND ac.id_pays = %s
                        """, ca_annee_ids + [real_classe_id, eac.groupe, eac.section_id, id_etablissement, id_pays])
                        for r in cur.fetchall():
                            attributions_map[r['id_cours_id']] = r

                    courses = []
                    for ca in cours_annee_list:
                        attr = attributions_map.get(ca.id_cours_annee, {})
                        pers_name = ''
                        if attr.get('pers_nom'):
                            pers_name = f"{attr['pers_nom']} {attr.get('pers_postnom') or ''} {attr.get('pers_prenom') or ''}".strip()
                        courses.append({
                            'id_cours_classe': ca.id_cours_annee,
                            'id_cours': ca.cours_id,
                            'cours': ca.cours.cours,
                            'code_cours': ca.cours.code_cours,
                            'maxima_exam': ca.maxima_exam,
                            'id_attribution': attr.get('id_attribution'),
                            'id_personnel': attr.get('id_personnel_id'),
                            'personnel_nom': pers_name,
                            'matricule': attr.get('matricule') or '',
                            'attribution_type_id': attr.get('attribution_type_id'),
                            'date_attribution': attr['date_attribution'].strftime('%Y-%m-%d') if attr.get('date_attribution') else '',
                        })

                    # Get personnel en_fonction for the dropdown
                    cur.execute("""
                        SELECT id_personnel, nom, postnom, prenom, matricule
                        FROM personnel
                        WHERE id_etablissement = %s AND id_pays = %s AND en_fonction = 1
                        ORDER BY nom, postnom
                    """, [id_etablissement, id_pays])
                    personnel = []
                    for p in cur.fetchall():
                        personnel.append({
                            'id': p['id_personnel'],
                            'label': f"{p['nom']} {p.get('postnom') or ''} {p.get('prenom') or ''}".strip(),
                            'matricule': p.get('matricule') or '',
                        })

                    # Get attribution types
                    cur.execute("SELECT id_attribution_type, attribution_type FROM attribution_type ORDER BY id_attribution_type")
                    attr_types = [{'id': t['id_attribution_type'], 'label': t['attribution_type']} for t in cur.fetchall()]

                    return JsonResponse({
                        'success': True,
                        'courses': courses,
                        'personnel': personnel,
                        'attribution_types': attr_types,
                    })

                elif action == 'assign':
                    id_cours_classe = data.get('id_cours_classe') or data.get('id_cours')
                    id_personnel = data.get('id_personnel')
                    id_cycle = data.get('id_cycle')
                    attr_type_id = data.get('attribution_type_id', 1)

                    if not all([id_cours_classe, id_personnel, id_classe]):
                        return JsonResponse({'success': False, 'error': 'Cours, personnel et classe requis.'}, status=400)

                    # Get current année, campus and business keys — via helper
                    bk = _resolve_eac_keys(cur, id_classe)
                    if not bk:
                        return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=400)
                    id_annee = bk['annee_id']
                    idCampus = bk['campus_id'] or 1

                    # Get cycle from Hub via cours_annee → cours → classe → cycle
                    try:
                        from MonEcole_app.models.enseignmnts.matiere import Cours_par_classe
                        ca_obj = Cours_par_classe.objects.select_related('id_cours__classe__cycle').get(id_cours_classe=int(id_cours_classe))
                        if ca_obj.id_cours and ca_obj.id_cours.classe:
                            id_cycle = ca_obj.id_cours.classe.cycle_id
                        else:
                            id_cycle = id_cycle or 1
                    except Exception:
                        id_cycle = id_cycle or 1

                    # Check if attribution already exists (by cours_annee id + business keys)
                    cur.execute("""
                        SELECT id_attribution FROM attribution_cours
                        WHERE id_cours_id=%s AND classe_id=%s AND groupe <=> %s AND section_id <=> %s AND id_etablissement=%s AND id_pays=%s
                    """, [id_cours_classe, bk['classe_id'], bk['groupe'], bk['section_id'], id_etablissement, id_pays])
                    existing = cur.fetchone()

                    if existing:
                        cur.execute("""
                            UPDATE attribution_cours
                            SET id_personnel_id=%s, attribution_type_id=%s, date_attribution=%s
                            WHERE id_attribution=%s
                        """, [id_personnel, attr_type_id, __import__('datetime').date.today().strftime('%Y-%m-%d'), existing['id_attribution']])
                    else:
                        cur.execute("""
                            INSERT INTO attribution_cours
                            (attribution_type_id, id_annee_id, idCampus_id, classe_id, groupe, section_id, id_cours_id, id_cycle_id, id_personnel_id, date_attribution, id_etablissement, id_pays)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, [attr_type_id, id_annee, idCampus, bk['classe_id'], bk['groupe'], bk['section_id'],
                              id_cours_classe, id_cycle, id_personnel,
                              __import__('datetime').date.today().strftime('%Y-%m-%d'), id_etablissement, id_pays])

                    conn.commit()
                    return JsonResponse({'success': True})

                elif action == 'remove':
                    id_attribution = data.get('id_attribution')
                    if not id_attribution:
                        return JsonResponse({'success': False, 'error': 'ID attribution requis.'}, status=400)
                    cur.execute("DELETE FROM attribution_cours WHERE id_attribution=%s AND id_etablissement=%s AND id_pays=%s",
                                [id_attribution, id_etablissement, id_pays])
                    conn.commit()
                    return JsonResponse({'success': True})

                else:
                    return JsonResponse({'success': False, 'error': 'Action invalide.'}, status=400)
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# APIs TABLES DE RÉFÉRENCE (migrées depuis db_monecole)
# Session, Mention
# ============================================================

@require_http_methods(["GET"])
def get_sessions_data(request):
    """Retourne la liste des sessions (référence commune)."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        sessions = list(Session.objects.filter(id_pays=etab.pays_id).values('id_session', 'session'))
        return JsonResponse({'success': True, 'sessions': sessions})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)






@csrf_exempt
@require_http_methods(["POST"])
def update_admin_instance(request):
    """API pour mettre à jour une instance administrative avec cascade aux enfants."""
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        data = json.loads(request.body)
        instance_id = int(data.get('id', 0))
        new_nom = (data.get('nom') or '').strip()

        if not instance_id or not new_nom:
            return JsonResponse({'success': False, 'error': 'ID et nom requis.'}, status=400)

        inst = AdministrativeStructureInstance.objects.filter(id_structure=instance_id).first()
        if not inst:
            return JsonResponse({'success': False, 'error': 'Instance introuvable.'}, status=404)

        old_nom = inst.nom
        inst.nom = new_nom
        inst.save(update_fields=['nom'])

        # Count descendants (children whose code contains this instance's id)
        old_id_str = str(instance_id)
        children_updated = 0
        all_descendants = AdministrativeStructureInstance.objects.filter(
            pays=inst.pays,
            ordre__gt=inst.ordre
        )
        for child in all_descendants:
            if child.code:
                parts = child.code.split('-')
                if old_id_str in parts:
                    children_updated += 1

        return JsonResponse({
            'success': True,
            'message': f'"{old_nom}" renommé en "{new_nom}". {children_updated} descendant(s).',
            'instance': {
                'id': inst.id_structure,
                'nom': inst.nom,
                'code': inst.code,
                'ordre': inst.ordre,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_mentions_data(request):
    """Retourne la liste des mentions (barèmes de notation)."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        mentions = list(Mention.objects.filter(id_pays=etab.pays_id).values(
            'id_mention', 'mention', 'abbreviation', 'min', 'max'
        ))
        return JsonResponse({'success': True, 'mentions': mentions})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def save_session(request):
    """Créer ou modifier une session."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        id_session = data.get('id_session')
        session_name = data.get('session')
        if id_session:
            s = get_object_or_404(Session, id_session=id_session, id_pays=etab.pays_id)
            s.session = session_name
            s.save()
        else:
            Session.objects.create(session=session_name, id_pays=etab.pays_id)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)






@csrf_exempt
@require_http_methods(["POST"])
def save_mention(request):
    """Créer ou modifier une mention."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        id_mention = data.get('id_mention')
        mention_name = data.get('mention')
        abbreviation = data.get('abbreviation')
        min_val = data.get('min')
        max_val = data.get('max')
        if id_mention:
            m = get_object_or_404(Mention, id_mention=id_mention, id_pays=etab.pays_id)
            m.mention = mention_name
            m.abbreviation = abbreviation
            m.min = min_val
            m.max = max_val
            m.save()
        else:
            Mention.objects.create(
                mention=mention_name, abbreviation=abbreviation,
                min=min_val, max=max_val, id_pays=etab.pays_id
            )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# DASHBOARD ÉTABLISSEMENT
# ============================================================

@require_http_methods(["GET"])
def dashboard_etablissement_view(request):
    """
    Dashboard spécialisé pour les administrateurs d'établissement.
    Affiche la fiche de l'école, les stats structurelles et les actions rapides.
    """
    context = get_country_context_logic(request)
    context['active_page'] = 'dashboard_etab'

    active_section = request.GET.get('section', 'dashboard')
    context['active_section'] = active_section

    user_id = request.session.get('user_id')
    if not user_id:
        from django.shortcuts import redirect
        return redirect('/login/')

    try:
        etab, err = _get_tenant_etab(request)
        if err:
            from django.shortcuts import redirect
            return redirect('/login/')
    except Exception:
        from django.shortcuts import redirect
        return redirect('/login/')

    pays = etab.pays

    # --- Établissement courant (tenant) ---
    etab_list = [{
        'id_admin': user_id,
        'id_etablissement': etab.id_etablissement,
        'nom': etab.nom,
        'is_current': True,
    }]

    # --- Année scolaire active (fallback to most recent if none open) ---
    annee_active = Annee.objects.filter(pays_id=pays.id_pays, isOpen=True).order_by('-annee').first()
    if not annee_active:
        annee_active = Annee.objects.filter(pays_id=pays.id_pays).order_by('-annee').first()
    annees_list = list(Annee.objects.filter(pays_id=pays.id_pays).order_by('-annee').values(
        'id_annee', 'annee', 'isOpen'
    ))

    # --- Stats pour l'année active ---
    stats = {
        'n_classes': 0,
        'n_cycles': 0,
        'n_cours': 0,
        'n_trimestres_ouverts': 0,
    }
    cycles_detail = []
    classes_detail = []
    trimestres_detail = []
    repartitions_notes = []
    cours_par_domaine = []

    if annee_active:
        etab_annee = EtablissementAnnee.objects.filter(
            etablissement=etab, annee=annee_active, id_pays=pays.id_pays
        ).first()

        if etab_annee:
            # Classes configurées
            classes_config = EtablissementAnneeClasse.objects.filter(
                etablissement_annee=etab_annee
            ).select_related('classe__cycle', 'section')

            stats['n_classes'] = classes_config.count()

            # Cycles distincts
            cycle_ids = set()
            cycle_counts = {}
            for cc in classes_config:
                cid = cc.classe.cycle_id
                cycle_ids.add(cid)
                cycle_counts[cid] = cycle_counts.get(cid, 0) + 1

            stats['n_cycles'] = len(cycle_ids)

            # Detail par cycle
            if cycle_ids:
                cycles_qs = Cycle.objects.filter(id__in=cycle_ids).order_by('ordre')
                for c in cycles_qs:
                    cycles_detail.append({
                        'nom': c.nom,
                        'n_classes': cycle_counts.get(c.id, 0),
                        'ordre': c.ordre,
                    })

            # Detail par classe
            classes_detail = []
            for cc in classes_config:
                classe_label = str(cc.classe) if cc.classe else '-'
                if cc.groupe:
                    classe_label += f" {cc.groupe}"
                classes_detail.append({
                    'eac_id': cc.id,
                    'classe_nom': classe_label,
                    'cycle_nom': str(cc.classe.cycle) if cc.classe and cc.classe.cycle else '-',
                    'section_nom': str(cc.section) if cc.section else '-',
                    'groupe': cc.groupe or '',
                })

            # Cours pour l'année (via les classes de l'établissement)
            classe_ids = [cc.classe_id for cc in classes_config]
            cours_annee_qs = CoursAnnee.objects.filter(
                annee=annee_active,
                cours__classe_id__in=classe_ids,
                id_pays=pays.id_pays
            ).select_related('cours')
            stats['n_cours'] = cours_annee_qs.count()

            # --- Cours par Domaine (pour le chart) ---
            # Build domaine name map (domaine_id is IntegerField, not FK)
            all_dom_ids = set()
            for ca in cours_annee_qs:
                if ca.domaine_id:
                    all_dom_ids.add(ca.domaine_id)
                if ca.cours and ca.cours.domaine_id:
                    all_dom_ids.add(ca.cours.domaine_id)
            dom_name_map = {}
            if all_dom_ids and Domaine:
                dom_name_map = {d['id_domaine']: d['nom'] for d in Domaine.objects.filter(id_domaine__in=all_dom_ids).values('id_domaine', 'nom')}

            domaine_counts = {}
            for ca in cours_annee_qs:
                dom_id = ca.domaine_id or (ca.cours.domaine_id if ca.cours else None)
                dom_name = dom_name_map.get(dom_id, 'Sans domaine') if dom_id else 'Sans domaine'
                domaine_counts[dom_name] = domaine_counts.get(dom_name, 0) + 1
            # Sort by count descending, limit to top 8
            cours_par_domaine = sorted(domaine_counts.items(), key=lambda x: x[1], reverse=True)[:8]

            # Répartitions temporelles via le nouveau système
            # Filtrer par les types de répartition liés aux cycles actifs
            active_cycle_ids = set(cc.classe.cycle_id for cc in classes_config if cc.classe and cc.classe.cycle_id)
            allowed_type_ids = set(
                RepartitionConfigCycle.objects.filter(
                    cycle_id__in=active_cycle_ids,
                    is_active=True,
                    id_pays=pays.id_pays
                ).values_list('type_racine_id', flat=True)
            ) if active_cycle_ids else set()
            # Toujours inclure le type Période (subdivision universelle)
            if allowed_type_ids:
                periode_type = RepartitionType.objects.filter(code='P', id_pays=pays.id_pays).values_list('id_type', flat=True).first()
                if periode_type:
                    allowed_type_ids.add(periode_type)

            # Determine which types are containers (parents in hierarchy) — they are NOT leaf
            parent_type_ids = set(
                RepartitionHierarchie.objects.filter(
                    is_active=True, id_pays=pays.id_pays
                ).values_list('type_parent_id', flat=True)
            )

            if etab.is_calendar_synched:
                # Synchronisé: lire directement les RepartitionInstance nationales
                ri_qs = RepartitionInstance.objects.filter(
                    annee=annee_active,
                    pays=etab.pays,
                    is_active=True
                ).select_related('type').order_by('type__nom', 'ordre')
                if allowed_type_ids:
                    ri_qs = ri_qs.filter(type_id__in=allowed_type_ids)
                repartitions_notes = []
                for ri in ri_qs:
                    repartitions_notes.append({
                        'id': ri.pk,
                        'id_instance': ri.id_instance,
                        'nom': ri.nom,
                        'code': ri.code,
                        'type': ri.type.nom if ri.type else '',
                        'type_code': ri.type.code if ri.type else '',
                        'type_id': ri.type_id,
                        'ordre': ri.ordre,
                        'is_open': ri.is_active,
                        'is_leaf': ri.type_id not in parent_type_ids,
                        'debut': str(ri.date_debut or ''),
                        'fin': str(ri.date_fin or ''),
                        'parent_instance_id': None,
                        'parent_nom': None,
                    })

                # Build parent→child mapping using hierarchy
                # For each hierarchy (e.g. Trimestre→Période, nombre_enfants=2),
                # assign each child period to its parent container by order
                hierarchies = RepartitionHierarchie.objects.filter(is_active=True, id_pays=pays.id_pays)
                for hier in hierarchies:
                    parent_items = sorted(
                        [r for r in repartitions_notes if r['type_id'] == hier.type_parent_id],
                        key=lambda x: x['ordre']
                    )
                    child_items = sorted(
                        [r for r in repartitions_notes if r['type_id'] == hier.type_enfant_id],
                        key=lambda x: x['ordre']
                    )
                    nb = hier.nombre_enfants
                    for pi, parent in enumerate(parent_items):
                        start = pi * nb
                        end = start + nb
                        for child in child_items[start:end]:
                            child['parent_instance_id'] = parent['id_instance']
                            child['parent_nom'] = parent['nom']

                stats['n_trimestres_ouverts'] = sum(1 for r in repartitions_notes if r.get('is_open'))
            else:
                # Personnalisé: lire les RepartitionConfigEtabAnnee
                rep_qs = RepartitionConfigEtabAnnee.objects.filter(
                    etablissement_annee=etab_annee
                )
                if allowed_type_ids:
                    rep_qs = rep_qs.filter(repartition__type_id__in=allowed_type_ids)
                repartitions_raw = rep_qs.select_related('repartition__type').order_by('repartition__type__nom', 'repartition__ordre').values(
                    'id', 'repartition__id_instance', 'repartition__nom', 'repartition__code',
                    'repartition__type__nom', 'repartition__type__code',
                    'repartition__ordre', 'debut', 'fin', 'is_open'
                )
                stats['n_trimestres_ouverts'] = sum(1 for r in repartitions_raw if r.get('is_open'))
                repartitions_notes = []
                for rc in repartitions_raw:
                    type_id = None
                    # Get type_id for is_leaf check
                    type_code = rc.get('repartition__type__code', '')
                    if type_code:
                        rt = RepartitionType.objects.filter(code=type_code, id_pays=pays.id_pays).values_list('id_type', flat=True).first()
                        type_id = rt
                    repartitions_notes.append({
                        'id': rc.get('id'),
                        'id_instance': rc.get('repartition__id_instance'),
                        'nom': rc.get('repartition__nom', '-'),
                        'code': rc.get('repartition__code', ''),
                        'type': rc.get('repartition__type__nom', ''),
                        'type_code': type_code,
                        'type_id': type_id,
                        'ordre': rc.get('repartition__ordre', 0),
                        'is_open': bool(rc.get('is_open')),
                        'is_leaf': type_id not in parent_type_ids if type_id else True,
                        'debut': str(rc.get('debut', '') or ''),
                        'fin': str(rc.get('fin', '') or ''),
                        'parent_instance_id': None,
                        'parent_nom': None,
                    })

                # Build parent→child mapping (same logic as synced mode)
                hierarchies = RepartitionHierarchie.objects.filter(is_active=True, id_pays=pays.id_pays)
                for hier in hierarchies:
                    parent_items = sorted(
                        [r for r in repartitions_notes if r.get('type_id') == hier.type_parent_id],
                        key=lambda x: x['ordre']
                    )
                    child_items = sorted(
                        [r for r in repartitions_notes if r.get('type_id') == hier.type_enfant_id],
                        key=lambda x: x['ordre']
                    )
                    nb = hier.nombre_enfants
                    for pi, parent in enumerate(parent_items):
                        start = pi * nb
                        end = start + nb
                        for child in child_items[start:end]:
                            child['parent_instance_id'] = parent['id_instance']
                            child['parent_nom'] = parent['nom']

            # Deduplicate by name for trimestres widget
            trim_seen = {}
            for rc in repartitions_notes:
                r_name = rc.get('nom', '-')
                debut_val = rc.get('debut')
                fin_val = rc.get('fin')
                def safe_fmt(v, fmt='%d/%m/%Y'):
                    if v is None or v == '':
                        return '-'
                    if isinstance(v, str):
                        # Try parsing YYYY-MM-DD to dd/mm/yyyy
                        parts = v.split('-')
                        if len(parts) == 3:
                            return f"{parts[2]}/{parts[1]}/{parts[0]}"
                        return v
                    try:
                        return v.strftime(fmt)
                    except Exception:
                        return str(v)
                def safe_iso(v):
                    if v is None or v == '':
                        return None
                    if isinstance(v, str):
                        return v[:10] if len(v) >= 10 else v
                    try:
                        return v.isoformat()
                    except Exception:
                        return str(v)
                if r_name not in trim_seen:
                    trim_seen[r_name] = {
                        'id': rc.get('id'),
                        'trimestre': r_name,
                        'debut': safe_fmt(debut_val),
                        'fin': safe_fmt(fin_val),
                        'debut_iso': safe_iso(debut_val),
                        'fin_iso': safe_iso(fin_val),
                        'isOpen': bool(rc.get('is_open')),
                    }
                else:
                    if rc.get('is_open'):
                        trim_seen[r_name]['isOpen'] = True
            trimestres_detail = list(trim_seen.values())

    # --- Cours par Domaine as JSON for frontend chart ---
    cours_domaine_json = json.dumps(
        [{'domaine': d, 'count': c} for d, c in cours_par_domaine] if annee_active else [],
        ensure_ascii=False
    )

    # --- Admin display name ---
    admin_email = request.session.get('user_email', '')
    admin_display_name = admin_email.split('@')[0] if admin_email else 'Administrateur'

    # --- Fiche établissement ---
    regime_nom = '-'
    if etab.id_regime:
        try:
            regime_nom = Regime.objects.get(id_regime=etab.id_regime, pays=pays).regime
        except Regime.DoesNotExist:
            pass

    etab_data = {
        'id_etablissement': etab.id_etablissement,
        'nom': etab.nom,
        'sigle': etab.sigle or '',
        'code_ecole': etab.code_ecole or '',
        'matricule': etab.matricule or '',
        'no_dinacope': etab.no_dinacope or '',
        'reference_agrement': etab.reference_agrement or '',
        'adresse': etab.adresse or '',
        'email': etab.email or '',
        'telephone': etab.telephone or '',
        'fax': etab.fax or '',
        'boite_postale': etab.boite_postale or '',
        'representant': etab.representant or '',
        'emplacement': etab.emplacement or '',
        'url': etab.url or '',
        'logo_ecole': etab.logo_ecole or '',
        'regime_nom': regime_nom,
        'id_regime': etab.id_regime,
        'structure_pedagogique': etab.structure_pedagogique.nom if etab.structure_pedagogique else '-',
        'gestionnaire': str(etab.gestionnaire) if etab.gestionnaire else '-',
        'gestionnaire_email': etab.gestionnaire.email if etab.gestionnaire and etab.gestionnaire.email else '',
        'gestionnaire_telephone': etab.gestionnaire.telephone if etab.gestionnaire and etab.gestionnaire.telephone else '',
        'code': etab.code,
        'pays_nom': pays.nom,
        'id_pays': pays.id_pays,
        'pays_id': pays.id_pays,
        'ref_administrative': etab.ref_administrative or '',
        'nom_rue': etab.nom_rue or '',
        'numero_rue': etab.numero_rue or '',
        'latitude': float(etab.latitude) if etab.latitude else None,
        'longitude': float(etab.longitude) if etab.longitude else None,
        'annee_creation': etab.annee_creation or '',
        'annee_agrement': etab.annee_agrement or '',
        'document_agrement': etab.document_agrement or '',
        'is_calendar_synched': etab.is_calendar_synched,
    }

    # --- Régimes pour le dropdown ---
    regimes = list(Regime.objects.filter(pays=pays).values('id_regime', 'regime'))

    # --- Administrative hierarchy chain (from ref_administrative, single source of truth) ---
    admin_chain = []  # [{type_nom, instance_nom, instance_id, ordre}, ...]
    admin_types = list(
        AdministrativeStructureType.objects.filter(pays=pays).order_by('ordre').values(
            'id_structure', 'code', 'nom', 'ordre'
        )
    )

    # Build chain from ref_administrative (e.g. "1-2-14-47")
    ref_admin = etab.ref_administrative or ''
    ref_parts = [p.strip() for p in ref_admin.split('-') if p.strip()]
    if ref_parts:
        all_ids = []
        for part in ref_parts:
            try:
                all_ids.append(int(part))
            except (ValueError, TypeError):
                pass

        if all_ids:
            # Fetch all referenced instances in one query — SCOPED BY PAYS
            ancestors_qs = AdministrativeStructureInstance.objects.filter(
                id_structure__in=all_ids,
                pays=pays
            ).order_by('ordre')

            # Map type_nom by ordre
            type_by_ordre = {t['ordre']: t['nom'] for t in admin_types}

            for anc in ancestors_qs:
                admin_chain.append({
                    'type_nom': type_by_ordre.get(anc.ordre, f'Niveau {anc.ordre}'),
                    'instance_nom': anc.nom,
                    'instance_id': anc.id_structure,
                    'ordre': anc.ordre,
                })

    # Fetch admin instances grouped by ordre for cascade dropdowns
    admin_instances_by_ordre = {}
    all_admin_instances = AdministrativeStructureInstance.objects.filter(
        pays=pays
    ).order_by('ordre', 'nom').values('id_structure', 'nom', 'ordre', 'code', 'latitude', 'longitude')
    for inst in all_admin_instances:
        o = inst['ordre']
        if o not in admin_instances_by_ordre:
            admin_instances_by_ordre[o] = []
        admin_instances_by_ordre[o].append({
            'id': inst['id_structure'],
            'nom': inst['nom'],
            'code': inst['code'],
            'lat': float(inst.get('latitude') or 0),
            'lng': float(inst.get('longitude') or 0),
        })

    # --- Cross-database student/teacher stats from db_monecole ---
    eleves_stats = {}
    enseignants_count = 0
    eleves_par_classe = []
    eleves_par_campus = []
    try:
        from django.db import connections
        db_settings = connections['default'].settings_dict
        import pymysql
        spoke_conn = pymysql.connect(
            host=db_settings.get('HOST', 'localhost') or 'localhost',
            user=db_settings['USER'],
            password=db_settings['PASSWORD'],
            port=int(db_settings.get('PORT', 3306) or 3306),
            database=connections['default'].settings_dict['NAME'],
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=5,
            read_timeout=10,
        )
        try:
            with spoke_conn.cursor() as cur:
                etab_id = etab.id_etablissement

                # Get campus IDs for this establishment
                cur.execute("""
                    SELECT idCampus FROM campus
                    WHERE id_etablissement = %s AND is_active = 1
                """, [etab_id])
                campus_ids = [r['idCampus'] for r in cur.fetchall()]

                annee_id = annee_active.pk if annee_active else None

                if campus_ids:
                    placeholders = ','.join(['%s'] * len(campus_ids))
                    annee_filter = ' AND ei.id_annee_id = %s' if annee_id else ''
                    base_params = campus_ids + ([annee_id] if annee_id else [])

                    # Total students + gender (filter via campus + annee)
                    cur.execute(f"""
                        SELECT
                            COUNT(*) as total,
                            SUM(CASE WHEN e.genre = 'M' THEN 1 ELSE 0 END) as garcons,
                            SUM(CASE WHEN e.genre = 'F' THEN 1 ELSE 0 END) as filles
                        FROM eleve_inscription ei
                        JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                        WHERE ei.idCampus_id IN ({placeholders}) AND ei.status = 1{annee_filter}
                    """, base_params)
                    row = cur.fetchone()
                    if row:
                        eleves_stats['total'] = int(row['total'] or 0)
                        eleves_stats['garcons'] = int(row['garcons'] or 0)
                        eleves_stats['filles'] = int(row['filles'] or 0)

                    # Age distribution — per individual year
                    cur.execute(f"""
                        SELECT
                            TIMESTAMPDIFF(YEAR, e.date_naissance, CURDATE()) as age,
                            COUNT(*) as nb
                        FROM eleve_inscription ei
                        JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                        WHERE ei.idCampus_id IN ({placeholders}) AND ei.status = 1{annee_filter}
                              AND e.date_naissance IS NOT NULL
                              AND e.date_naissance != '0000-00-00'
                        GROUP BY age
                        ORDER BY age
                    """, base_params)
                    age_rows = list(cur.fetchall())
                    eleves_stats['age_distribution'] = [
                        {'tranche': f"{int(r['age'])} ans", 'nb': int(r['nb'])} for r in age_rows
                    ]

                    # Students per class with gender
                    cur.execute(f"""
                        SELECT
                            eac.id as eac_id,
                            CONCAT(
                                cl.nom,
                                COALESCE(CONCAT(' - ', s.nom), ''),
                                COALESCE(CONCAT(' (', eac.groupe, ')'), '')
                            ) as classe_nom,
                            COUNT(DISTINCT ei.id_eleve_id) as total,
                            COUNT(DISTINCT CASE WHEN e.genre = 'M' THEN ei.id_eleve_id END) as garcons,
                            COUNT(DISTINCT CASE WHEN e.genre = 'F' THEN ei.id_eleve_id END) as filles
                        FROM eleve_inscription ei
                        JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                        JOIN countryStructure.etablissements_annees_classes eac
                          ON eac.classe_id = ei.classe_id
                          AND (eac.groupe COLLATE utf8mb4_general_ci <=> ei.groupe COLLATE utf8mb4_general_ci)
                          AND eac.section_id <=> ei.section_id
                        JOIN countryStructure.classes cl ON cl.id = eac.classe_id
                        LEFT JOIN countryStructure.sections s ON s.id = eac.section_id
                        WHERE ei.idCampus_id IN ({placeholders}) AND ei.status = 1{annee_filter}
                        GROUP BY eac.id
                        ORDER BY cl.ordre, cl.nom, eac.groupe
                    """, base_params)
                    eleves_par_classe = [
                        {'eac_id': int(r['eac_id']), 'classe_nom': r['classe_nom'], 'total': int(r['total']),
                         'garcons': int(r['garcons'] or 0), 'filles': int(r['filles'] or 0)}
                        for r in cur.fetchall()
                    ]

                    # Students per campus
                    cur.execute(f"""
                        SELECT
                            c.campus as campus_nom,
                            COUNT(*) as total
                        FROM eleve_inscription ei
                        JOIN campus c ON c.idCampus = ei.idCampus_id
                        WHERE ei.idCampus_id IN ({placeholders}) AND ei.status = 1{annee_filter}
                        GROUP BY c.idCampus, c.campus
                        ORDER BY total DESC
                    """, base_params)
                    eleves_par_campus = [
                        {'campus_nom': r['campus_nom'], 'total': int(r['total'])}
                        for r in cur.fetchall()
                    ]

                    # Teachers count
                    cur.execute(f"""
                        SELECT COUNT(DISTINCT user_id) as total
                        FROM user_enseignement
                        WHERE id_etablissement = %s
                    """, [etab_id])
                    row = cur.fetchone()
                    enseignants_count = int(row['total']) if row else 0

        finally:
            spoke_conn.close()
    except Exception:
        # Spoke DB unavailable — graceful degradation
        pass

    stats['n_eleves'] = eleves_stats.get('total', 0)
    stats['n_enseignants'] = enseignants_count

    # --- Trimestres as JSON for Gantt timeline ---
    trimestres_json = json.dumps(trimestres_detail, ensure_ascii=False, default=str)

    context.update({
        'etab': etab_data,
        'etab_list': etab_list,
        'stats': stats,
        'admin_display_name': admin_display_name,
        'annee_active': {
            'id': annee_active.pk,
            'id_annee': annee_active.id_annee,
            'annee': annee_active.annee,
            'isOpen': annee_active.isOpen,
        } if annee_active else None,
        'annees_list': annees_list,
        'annees_json': json.dumps(annees_list, ensure_ascii=False, default=str),
        'cycles_detail': cycles_detail,
        'trimestres_detail': trimestres_detail,
        'trimestres_json': trimestres_json,
        'cours_domaine_json': cours_domaine_json,
        'eleves_stats_json': json.dumps(eleves_stats, ensure_ascii=False, default=str),
        'eleves_par_classe_json': json.dumps(eleves_par_classe, ensure_ascii=False, default=str),
        'eleves_par_campus_json': json.dumps(eleves_par_campus, ensure_ascii=False, default=str),
        'regimes': json.dumps(regimes, ensure_ascii=False, default=str),
        'classes_detail': classes_detail if annee_active else [],
        'repartitions_notes_json': json.dumps(repartitions_notes if annee_active else [], ensure_ascii=False, default=str),
        'admin_chain': admin_chain,
        'admin_types': admin_types,
        'admin_types_json': json.dumps(admin_types, ensure_ascii=False, default=str),
        'admin_instances_json': json.dumps(admin_instances_by_ordre, ensure_ascii=False, default=str),
        'is_etab_admin': True,
        'is_calendar_synched': etab.is_calendar_synched,
    })

    return render(request, 'structure_app/dashboard_etablissement.html', context)


@csrf_exempt
@require_http_methods(["POST"])
def create_admin_instance(request):
    """API pour créer une nouvelle instance de structure administrative."""
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        data = json.loads(request.body)
        nom = (data.get('nom') or '').strip()
        ordre = int(data.get('ordre', 0))
        pays_id = int(data.get('pays_id', 0))
        parent_code = (data.get('parent_code') or '').strip()  # code path of parent chain

        if not nom or not ordre or not pays_id:
            return JsonResponse({'success': False, 'error': 'Données manquantes (nom, ordre, pays).'}, status=400)

        pays = Pays.objects.filter(id_pays=pays_id).first()
        if not pays:
            return JsonResponse({'success': False, 'error': 'Pays introuvable.'}, status=404)

        # --- Duplicate name check: prevent same name at same level under same parent ---
        from django.db.models.functions import Lower
        existing_qs = AdministrativeStructureInstance.objects.filter(
            pays=pays,
            ordre=ordre,
        ).annotate(nom_lower=Lower('nom')).filter(nom_lower=nom.lower())
        if parent_code:
            # Scope to same parent: children whose code starts with parent_code-
            existing_qs = existing_qs.filter(code__startswith=f"{parent_code}-")
        else:
            # First level: no parent code prefix
            pass
        if existing_qs.exists():
            return JsonResponse({
                'success': False,
                'error': f'« {nom} » existe déjà à ce niveau.'
            }, status=409)

        # Create the instance (code will be set after we get the ID)
        new_inst = AdministrativeStructureInstance.objects.create(
            nom=nom,
            ordre=ordre,
            pays=pays,
            code='',  # will update below
        )
        # Refresh from DB to get the real id_structure computed by save()
        new_inst.refresh_from_db()

        # Safety check: id_structure must be > 0
        if not new_inst.id_structure or new_inst.id_structure == 0:
            # Force assign next available id_structure
            from django.db.models import Max
            max_id = AdministrativeStructureInstance.objects.filter(pays=pays).exclude(pk=new_inst.pk).aggregate(
                m=Max('id_structure'))['m'] or 0
            new_inst.id_structure = max_id + 1

        # Build code: parent_code + "-" + id_structure  (or just id_structure if first level)
        if parent_code:
            new_inst.code = f"{parent_code}-{new_inst.id_structure}"
        else:
            new_inst.code = str(new_inst.id_structure)
        new_inst.save(update_fields=['code', 'id_structure'])

        return JsonResponse({
            'success': True,
            'instance': {
                'id': new_inst.id_structure,
                'nom': new_inst.nom,
                'code': new_inst.code,
                'ordre': new_inst.ordre,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_rue_suggestions(request):
    """Return distinct nom_rue values from all etablissements in the given pays for autocomplete."""
    try:
        pays_id = request.GET.get('pays_id')
        if not pays_id:
            return JsonResponse({'success': False, 'error': 'pays_id requis'}, status=400)

        rues = list(
            Etablissement.objects.filter(
                pays_id=pays_id,
                nom_rue__isnull=False
            ).exclude(
                nom_rue=''
            ).values_list('nom_rue', flat=True).distinct().order_by('nom_rue')
        )

        return JsonResponse({'success': True, 'rues': rues})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def update_mon_etablissement(request):
    """
    API pour mettre à jour les informations de l'établissement de l'admin courant.
    L'admin ne peut modifier QUE son propre établissement.
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        etab, err = _get_tenant_etab(request)
        if err: return err
        if not etab:
            return JsonResponse({'success': False, 'error': 'Pas un admin établissement.'}, status=403)

        data = json.loads(request.body)

        # Champs modifiables par l'admin établissement
        editable_fields = {
            'nom': 'nom',
            'sigle': 'sigle',
            'code_ecole': 'code_ecole',
            'matricule': 'matricule',
            'no_dinacope': 'no_dinacope',
            'reference_agrement': 'reference_agrement',
            'adresse': 'adresse',
            'email': 'email',
            'telephone': 'telephone',
            'fax': 'fax',
            'boite_postale': 'boite_postale',
            'representant': 'representant',
            'emplacement': 'emplacement',
            'ref_administrative': 'ref_administrative',
            'nom_rue': 'nom_rue',
            'numero_rue': 'numero_rue',
            'annee_creation': 'annee_creation',
            'annee_agrement': 'annee_agrement',
        }

        updated_fields = []
        for json_key, model_field in editable_fields.items():
            if json_key in data:
                val = data[json_key] or None
                # --- Validate ref_administrative completeness ---
                if json_key == 'ref_administrative' and val:
                    parts = [p.strip() for p in str(val).split('-') if p.strip()]
                    n_levels = etab.pays.nLevelsAdministratifs if etab.pays else 0
                    if n_levels and len(parts) != n_levels:
                        return JsonResponse({
                            'success': False,
                            'error': f'La référence administrative doit avoir {n_levels} niveaux (reçu {len(parts)}).'
                        }, status=400)
                setattr(etab, model_field, val)
                updated_fields.append(model_field)

        # Régime (id_regime)
        if 'id_regime' in data:
            etab.id_regime = int(data['id_regime']) if data['id_regime'] else None
            updated_fields.append('id_regime')

        # GPS coordinates
        if 'latitude' in data:
            from decimal import Decimal
            etab.latitude = Decimal(str(data['latitude'])) if data['latitude'] is not None else None
            updated_fields.append('latitude')
        if 'longitude' in data:
            from decimal import Decimal
            etab.longitude = Decimal(str(data['longitude'])) if data['longitude'] is not None else None
            updated_fields.append('longitude')

        if updated_fields:
            etab.save(update_fields=updated_fields)

        return JsonResponse({'success': True, 'updated': updated_fields})
    except AdminUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Utilisateur introuvable.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def toggle_calendar_synch(request):
    """
    Basculer is_calendar_synched. Quand on passe en mode personnalisé,
    auto-provisionne les RepartitionConfigEtabAnnee depuis les instances nationales.
    Retourne aussi les répartitions pour rafraîchir l'UI.
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        etab, err = _get_tenant_etab(request)
        if err: return err
        if not etab:
            return JsonResponse({'success': False, 'error': 'Pas un admin établissement.'}, status=403)

        data = json.loads(request.body)
        is_synched = bool(data.get('is_synched', True))

        etab.is_calendar_synched = is_synched
        etab.save(update_fields=['is_calendar_synched'])

        repartitions = []

        # Get allowed repartition types from active cycles
        annee_active = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        allowed_type_ids = set()
        if annee_active:
            etab_annee, _ = EtablissementAnnee.objects.get_or_create(
                etablissement=etab, annee=annee_active,
                defaults={'id_pays': etab.pays_id}
            )
            active_cycle_ids = set(
                EtablissementAnneeClasse.objects.filter(
                    etablissement_annee=etab_annee
                ).select_related('classe').values_list('classe__cycle_id', flat=True).distinct()
            )
            if active_cycle_ids:
                allowed_type_ids = set(
                    RepartitionConfigCycle.objects.filter(
                        cycle_id__in=active_cycle_ids,
                        is_active=True,
                        id_pays=etab.pays_id
                    ).values_list('type_racine_id', flat=True)
                )
                # Toujours inclure le type Période
                periode_type = RepartitionType.objects.filter(code='P', id_pays=etab.pays_id).values_list('id_type', flat=True).first()
                if periode_type:
                    allowed_type_ids.add(periode_type)

        if not is_synched:
            # Auto-provision: copier les RepartitionInstance nationales dans RepartitionConfigEtabAnnee
            if annee_active:
                national_instances = RepartitionInstance.objects.filter(
                    annee=annee_active, pays=etab.pays, is_active=True
                ).select_related('type').order_by('type__nom', 'ordre')
                if allowed_type_ids:
                    national_instances = national_instances.filter(type_id__in=allowed_type_ids)

                existing_ids = set(
                    RepartitionConfigEtabAnnee.objects.filter(
                        etablissement_annee=etab_annee
                    ).values_list('repartition_id', flat=True)
                )

                for ri in national_instances:
                    if ri.pk not in existing_ids:
                        RepartitionConfigEtabAnnee.objects.create(
                            etablissement_annee=etab_annee,
                            repartition=ri,
                            parent=None, has_parent=False,
                            debut=ri.date_debut, fin=ri.date_fin,
                            is_open=True, is_national=False
                        )

                # Return all configs for this etab (filtered by allowed types)
                configs = RepartitionConfigEtabAnnee.objects.filter(
                    etablissement_annee=etab_annee
                ).select_related('repartition__type').order_by('repartition__type__nom', 'repartition__ordre')
                if allowed_type_ids:
                    configs = configs.filter(repartition__type_id__in=allowed_type_ids)
                for c in configs:
                    repartitions.append({
                        'id': c.id,
                        'id_instance': c.repartition_id,
                        'nom': c.repartition.nom,
                        'code': c.repartition.code,
                        'type': c.repartition.type.nom if c.repartition.type else '',
                        'type_code': c.repartition.type.code if c.repartition.type else '',
                        'ordre': c.repartition.ordre,
                        'is_open': c.is_open,
                        'debut': str(c.debut or ''),
                        'fin': str(c.fin or ''),
                    })
        else:
            # Return national instances (filtered by allowed types)
            if annee_active:
                ri_qs = RepartitionInstance.objects.filter(
                    annee=annee_active, pays=etab.pays, is_active=True
                ).select_related('type').order_by('type__nom', 'ordre')
                if allowed_type_ids:
                    ri_qs = ri_qs.filter(type_id__in=allowed_type_ids)
                for ri in ri_qs:
                    repartitions.append({
                        'id': ri.pk,
                        'id_instance': ri.id_instance,
                        'nom': ri.nom, 'code': ri.code,
                        'type': ri.type.nom if ri.type else '',
                        'type_code': ri.type.code if ri.type else '',
                        'ordre': ri.ordre, 'is_open': ri.is_active,
                        'debut': str(ri.date_debut or ''),
                        'fin': str(ri.date_fin or ''),
                    })

        return JsonResponse({
            'success': True,
            'is_calendar_synched': etab.is_calendar_synched,
            'repartitions': repartitions,
            'message': 'Calendrier synchronisé avec le Hub.' if is_synched
                       else 'Calendrier en mode personnalisé.'
        })
    except AdminUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Utilisateur introuvable.'}, status=404)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def update_calendar_config(request):
    """
    Sauvegarder les modifications de dates/statut sur RepartitionConfigEtabAnnee.
    Body JSON: { "id": config_id, "field": "debut"|"fin"|"is_open", "value": "..." }
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        etab, err = _get_tenant_etab(request)
        if err: return err
        if not etab:
            return JsonResponse({'success': False, 'error': 'Pas un admin établissement.'}, status=403)
        if etab.is_calendar_synched:
            return JsonResponse({'success': False, 'error': 'Calendrier synchronisé — modification interdite.'}, status=403)

        data = json.loads(request.body)
        config_id = data.get('id')
        field = data.get('field')
        value = data.get('value')

        if not config_id or field not in ('debut', 'fin', 'is_open'):
            return JsonResponse({'success': False, 'error': 'Paramètres invalides.'}, status=400)

        config = RepartitionConfigEtabAnnee.objects.select_related(
            'etablissement_annee__etablissement'
        ).get(id=config_id)

        if config.etablissement_annee.etablissement_id != etab.id_etablissement:
            return JsonResponse({'success': False, 'error': 'Accès refusé.'}, status=403)

        if field == 'is_open':
            config.is_open = bool(value)
        elif field == 'debut':
            from datetime import date as dt_date
            config.debut = dt_date.fromisoformat(value) if value else None
        elif field == 'fin':
            from datetime import date as dt_date
            config.fin = dt_date.fromisoformat(value) if value else None

        config.is_national = False  # Mark as customized
        config.save()

        return JsonResponse({'success': True, 'message': 'Mis à jour.'})
    except RepartitionConfigEtabAnnee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Config introuvable.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def upload_etab_logo(request):
    """Upload the establishment logo. Saves as media/Logos/LogoEtab_x.ext"""
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        etab, err = _get_tenant_etab(request)
        if err: return err
        if not etab:
            return JsonResponse({'success': False, 'error': 'Pas un admin établissement.'}, status=403)

        logo_file = request.FILES.get('logo')
        if not logo_file:
            return JsonResponse({'success': False, 'error': 'Fichier logo requis.'}, status=400)

        import os
        from django.conf import settings

        logos_dir = os.path.join(settings.MEDIA_ROOT, 'Logos')
        os.makedirs(logos_dir, exist_ok=True)
        try:
            os.chmod(logos_dir, 0o755)
        except Exception:
            pass

        ext = os.path.splitext(logo_file.name)[1].lower() or '.png'
        filename = f"LogoEtab_{etab.id_etablissement}{ext}"
        filepath = os.path.join(logos_dir, filename)

        # Remove old logos with different extensions
        for old_ext in ['.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg']:
            old_file = os.path.join(logos_dir, f"LogoEtab_{etab.id_etablissement}{old_ext}")
            if os.path.exists(old_file) and old_file != filepath:
                try:
                    os.remove(old_file)
                except Exception:
                    pass

        with open(filepath, 'wb+') as f:
            for chunk in logo_file.chunks():
                f.write(chunk)

        try:
            os.chmod(filepath, 0o644)
        except Exception:
            pass

        logo_url = f"/media/Logos/{filename}"
        etab.logo_ecole = logo_url
        etab.save(update_fields=['logo_ecole'])

        return JsonResponse({'success': True, 'logo_url': logo_url})
    except AdminUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Utilisateur introuvable.'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def upload_etab_document(request):
    """Upload a document for the establishment (e.g. agrément ref document).
    Saves to media/Documents/Agrements/Agrement_EtabID_x.ext
    The 'doc_type' field determines the type of document (currently: 'agrement').
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        etab, err = _get_tenant_etab(request)
        if err: return err
        if not etab:
            return JsonResponse({'success': False, 'error': 'Pas un admin établissement.'}, status=403)

        doc_file = request.FILES.get('document')
        doc_type = request.POST.get('doc_type', 'agrement')

        if not doc_file:
            return JsonResponse({'success': False, 'error': 'Fichier document requis.'}, status=400)

        import os
        from django.conf import settings

        docs_dir = os.path.join(settings.MEDIA_ROOT, 'Documents', 'Agrements')
        os.makedirs(docs_dir, exist_ok=True)
        try:
            os.chmod(docs_dir, 0o755)
        except Exception:
            pass

        ext = os.path.splitext(doc_file.name)[1].lower() or '.pdf'
        filename = f"Agrement_EtabID_{etab.id_etablissement}{ext}"
        filepath = os.path.join(docs_dir, filename)

        # Remove old documents with different extensions
        for old_ext in ['.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx']:
            old_file = os.path.join(docs_dir, f"Agrement_EtabID_{etab.id_etablissement}{old_ext}")
            if os.path.exists(old_file) and old_file != filepath:
                try:
                    os.remove(old_file)
                except Exception:
                    pass

        with open(filepath, 'wb+') as f:
            for chunk in doc_file.chunks():
                f.write(chunk)

        try:
            os.chmod(filepath, 0o644)
        except Exception:
            pass

        doc_url = f"/media/Documents/Agrements/{filename}"
        etab.document_agrement = doc_url
        etab.save(update_fields=['document_agrement'])

        return JsonResponse({'success': True, 'document_url': doc_url})
    except AdminUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Utilisateur introuvable.'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_mon_etablissement(request):
    """API pour récupérer les données de l'établissement de l'admin courant."""
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        etab, err = _get_tenant_etab(request)
        if err: return err

        regime_nom = '-'
        if etab.id_regime:
            try:
                regime_nom = Regime.objects.get(id_regime=etab.id_regime).regime
            except Regime.DoesNotExist:
                pass

        return JsonResponse({
            'success': True,
            'etablissement': {
                'id_etablissement': etab.id_etablissement,
                'nom': etab.nom,
                'sigle': etab.sigle or '',
                'code_ecole': etab.code_ecole or '',
                'matricule': etab.matricule or '',
                'no_dinacope': etab.no_dinacope or '',
                'reference_agrement': etab.reference_agrement or '',
                'adresse': etab.adresse or '',
                'email': etab.email or '',
                'telephone': etab.telephone or '',
                'fax': etab.fax or '',
                'boite_postale': etab.boite_postale or '',
                'representant': etab.representant or '',
                'emplacement': etab.emplacement or '',
                'url': etab.url or '',
                'logo_ecole': etab.logo_ecole or '',
                'regime_nom': regime_nom,
                'id_regime': etab.id_regime,
                'pays_nom': etab.pays.nom,
                'id_pays': etab.pays.id_pays,
                'annee_creation': etab.annee_creation or '',
                'annee_agrement': etab.annee_agrement or '',
                'document_agrement': etab.document_agrement or '',
            }
        })
    except AdminUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Utilisateur introuvable.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# RÉPARTITIONS TEMPORELLES — API CRUD
# ============================================================

@require_http_methods(["GET"])
def get_repartition_types(request):
    """Retourne tous les types de répartition."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        types = list(RepartitionType.objects.filter(id_pays=etab.pays_id).values(
            'id_type', 'nom', 'code', 'description', 'is_active'
        ))
        return JsonResponse({'success': True, 'types': types})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_repartition_type(request):
    """Créer ou modifier un type de répartition."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        id_type = data.get('id_type')
        nom = data.get('nom', '').strip()
        code = data.get('code', '').strip().upper()
        description = data.get('description', '').strip()
        is_active = data.get('is_active', True)

        if not nom or not code:
            return JsonResponse({'success': False, 'error': 'Nom et code requis.'}, status=400)

        if id_type:
            obj = get_object_or_404(RepartitionType, pk=id_type, id_pays=etab.pays_id)
            obj.nom = nom
            obj.code = code
            obj.description = description
            obj.is_active = is_active
            obj.save()
        else:
            obj = RepartitionType.objects.create(
                nom=nom, code=code, description=description, is_active=is_active,
                id_pays=etab.pays_id
            )
        return JsonResponse({'success': True, 'id_type': obj.id_type})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_repartition_type(request):
    """Supprimer un type de répartition."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        obj = get_object_or_404(RepartitionType, pk=data.get('id_type'), id_pays=etab.pays_id)
        # Vérifier qu'il n'y a pas d'instances liées
        if obj.instances.exists():
            return JsonResponse({
                'success': False,
                'error': f'Impossible de supprimer : {obj.instances.count()} instance(s) liée(s).'
            }, status=400)
        obj.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_repartition_instances(request):
    """Retourne les instances de répartition, filtrable par type et/ou année."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        type_id = request.GET.get('type_id')
        annee_id = request.GET.get('annee_id')
        qs = RepartitionInstance.objects.select_related('type', 'annee').filter(pays_id=etab.pays_id)
        if type_id:
            qs = qs.filter(type_id=type_id)
        if annee_id:
            qs = qs.filter(annee_id=annee_id)

        instances = []
        for inst in qs:
            instances.append({
                'id_instance': inst.id_instance,
                'type_id': inst.type_id,
                'type_nom': inst.type.nom,
                'type_code': inst.type.code,
                'annee_id': inst.annee_id,
                'annee_nom': inst.annee.annee if inst.annee else '',
                'nom': inst.nom,
                'code': inst.code,
                'ordre': inst.ordre,
                'date_debut': inst.date_debut.isoformat() if inst.date_debut else '',
                'date_fin': inst.date_fin.isoformat() if inst.date_fin else '',
                'is_active': inst.is_active,
            })
        return JsonResponse({'success': True, 'instances': instances})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_repartition_instance(request):
    """Créer ou modifier une instance de répartition."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        id_instance = data.get('id_instance')
        type_id = data.get('type_id')
        annee_id = data.get('annee_id')
        nom = data.get('nom', '').strip()
        code = data.get('code', '').strip().upper()
        ordre = int(data.get('ordre', 1))
        date_debut = data.get('date_debut') or None
        date_fin = data.get('date_fin') or None
        is_active = data.get('is_active', True)

        if not type_id or not nom or not code:
            return JsonResponse({'success': False, 'error': 'Type, nom et code requis.'}, status=400)

        rtype = get_object_or_404(RepartitionType, pk=type_id, id_pays=etab.pays_id)
        annee = Annee.objects.filter(pk=annee_id, pays_id=etab.pays_id).first() if annee_id else None

        if id_instance:
            obj = get_object_or_404(RepartitionInstance, pk=id_instance, pays_id=etab.pays_id)
            obj.type = rtype
            obj.annee = annee
            obj.nom = nom
            obj.code = code
            obj.ordre = ordre
            obj.date_debut = date_debut
            obj.date_fin = date_fin
            obj.is_active = is_active
            obj.save()
        else:
            obj = RepartitionInstance.objects.create(
                type=rtype, annee=annee, nom=nom, code=code,
                ordre=ordre, date_debut=date_debut, date_fin=date_fin,
                is_active=is_active, pays_id=etab.pays_id
            )
        return JsonResponse({'success': True, 'id_instance': obj.id_instance})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_repartition_instance(request):
    """Supprimer une instance de répartition."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        obj = get_object_or_404(RepartitionInstance, pk=data.get('id_instance'), pays_id=etab.pays_id)
        obj.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_repartition_hierarchies(request):
    """Retourne les hiérarchies entre types."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        qs = RepartitionHierarchie.objects.select_related('type_parent', 'type_enfant').filter(id_pays=etab.pays_id)
        hierarchies = []
        for h in qs:
            hierarchies.append({
                'id_hierarchie': h.id_hierarchie,
                'type_parent_id': h.type_parent_id,
                'type_parent_nom': h.type_parent.nom,
                'type_parent_code': h.type_parent.code,
                'type_enfant_id': h.type_enfant_id,
                'type_enfant_nom': h.type_enfant.nom,
                'type_enfant_code': h.type_enfant.code,
                'nombre_enfants': h.nombre_enfants,
                'is_active': h.is_active,
            })
        return JsonResponse({'success': True, 'hierarchies': hierarchies})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_repartition_hierarchie(request):
    """Créer ou modifier une hiérarchie entre types."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        id_hierarchie = data.get('id_hierarchie')
        type_parent_id = data.get('type_parent_id')
        type_enfant_id = data.get('type_enfant_id')
        nombre_enfants = int(data.get('nombre_enfants', 1))
        is_active = data.get('is_active', True)

        if not type_parent_id or not type_enfant_id:
            return JsonResponse({'success': False, 'error': 'Types parent et enfant requis.'}, status=400)

        if type_parent_id == type_enfant_id:
            return JsonResponse({'success': False, 'error': 'Un type ne peut pas être son propre enfant.'}, status=400)

        tp = get_object_or_404(RepartitionType, pk=type_parent_id, id_pays=etab.pays_id)
        te = get_object_or_404(RepartitionType, pk=type_enfant_id, id_pays=etab.pays_id)

        if id_hierarchie:
            obj = get_object_or_404(RepartitionHierarchie, id_hierarchie=id_hierarchie, id_pays=etab.pays_id)
            obj.type_parent = tp
            obj.type_enfant = te
            obj.nombre_enfants = nombre_enfants
            obj.is_active = is_active
            obj.save()
        else:
            obj = RepartitionHierarchie.objects.create(
                type_parent=tp, type_enfant=te,
                nombre_enfants=nombre_enfants, is_active=is_active,
                id_pays=etab.pays_id
            )
        return JsonResponse({'success': True, 'id_hierarchie': obj.id_hierarchie})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_repartition_hierarchie(request):
    """Supprimer une hiérarchie."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        obj = get_object_or_404(RepartitionHierarchie, id_hierarchie=data.get('id_hierarchie'), id_pays=etab.pays_id)
        obj.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_repartition_configs_cycle(request):
    """Retourne les configurations répartition par cycle."""
    try:
        id_pays = request.GET.get('id_pays')
        qs = RepartitionConfigCycle.objects.select_related('cycle', 'type_racine', 'cycle__pays').all()
        if id_pays:
            qs = qs.filter(cycle__pays__id_pays=id_pays)

        configs = []
        for c in qs:
            configs.append({
                'id': c.id,
                'cycle_id': c.cycle_id,
                'cycle_nom': c.cycle.nom,
                'type_racine_id': c.type_racine_id,
                'type_racine_nom': c.type_racine.nom,
                'type_racine_code': c.type_racine.code,
                'nombre_au_niveau_racine': c.nombre_au_niveau_racine,
                'is_active': c.is_active,
            })
        return JsonResponse({'success': True, 'configs': configs})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_repartition_config_cycle(request):
    """Créer ou modifier une config de répartition par cycle."""
    try:
        data = json.loads(request.body)
        config_id = data.get('id')
        cycle_id = data.get('cycle_id')
        type_racine_id = data.get('type_racine_id')
        nombre = int(data.get('nombre_au_niveau_racine', 1))
        is_active = data.get('is_active', True)

        if not cycle_id or not type_racine_id:
            return JsonResponse({'success': False, 'error': 'Cycle et type racine requis.'}, status=400)

        cycle = get_object_or_404(Cycle, id_cycle=cycle_id)
        tr = get_object_or_404(RepartitionType, pk=type_racine_id)

        if config_id:
            obj = get_object_or_404(RepartitionConfigCycle, id=config_id)
            obj.cycle = cycle
            obj.type_racine = tr
            obj.nombre_au_niveau_racine = nombre
            obj.is_active = is_active
            obj.save()
        else:
            obj = RepartitionConfigCycle.objects.create(
                cycle=cycle, type_racine=tr,
                nombre_au_niveau_racine=nombre, is_active=is_active
            )
        return JsonResponse({'success': True, 'id': obj.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def delete_repartition_config_cycle(request):
    """Supprimer une config de répartition par cycle."""
    try:
        data = json.loads(request.body)
        obj = get_object_or_404(RepartitionConfigCycle, id=data.get('id'))
        obj.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# RÉPARTITIONS — AUTO-GÉNÉRATION INSTANCES
# ============================================================

@require_http_methods(["POST"])
def auto_generate_instances(request):
    """
    Auto-génère les RepartitionInstance à partir des RepartitionConfigCycle
    et RepartitionHierarchie existants.
    
    Ex: Config = Maternel → 3 Trimestres  +  Hiérarchie = 1 Trimestre → 2 Périodes
    → Crée: Premier Trimestre, Deuxième Trimestre, Troisième Trimestre
    → Crée: 1ère Période, 2ème Période, ..., 6ème Période
    
    Si pas de hiérarchie pour un type (ex: Semestre sans subdivision), 
    seuls les semestres sont créés.
    
    Payload: { "annee_id": int }
    """
    try:
        data = json.loads(request.body)
        annee_id = data.get('annee_id')
        
        if not annee_id:
            return JsonResponse({'success': False, 'error': 'annee_id requis.'}, status=400)
        
        annee = get_object_or_404(Annee, id_annee=annee_id)
        pays = Pays.objects.filter(id_pays=annee.pays_id).first()
        if not pays:
            return JsonResponse({'success': False, 'error': 'Pays introuvable pour cette année.'}, status=400)
        
        # French ordinal names
        ORDINALS = {
            1: "Premier", 2: "Deuxième", 3: "Troisième",
            4: "Quatrième", 5: "Cinquième", 6: "Sixième",
            7: "Septième", 8: "Huitième", 9: "Neuvième", 10: "Dixième"
        }
        ORDINALS_FEM = {
            1: "Première", 2: "Deuxième", 3: "Troisième",
            4: "Quatrième", 5: "Cinquième", 6: "Sixième",
            7: "Septième", 8: "Huitième", 9: "Neuvième", 10: "Dixième"
        }
        
        def get_ordinal_name(n, type_nom):
            """Retourne le nom ordinal français adapté au genre du type."""
            is_feminine = type_nom.lower() in ('période',)
            ords = ORDINALS_FEM if is_feminine else ORDINALS
            ord_name = ords.get(n, f"{n}e")
            # Special: "Second Semestre" instead of "Deuxième"
            if type_nom.lower() == 'semestre' and n == 2:
                ord_name = "Second"
            return f"{ord_name} {type_nom}"
        
        # Get all active cycle configs
        configs = RepartitionConfigCycle.objects.filter(
            is_active=True, id_pays=pays.id_pays
        ).select_related('type_racine')
        
        if not configs.exists():
            return JsonResponse({
                'success': True, 'created_root': 0, 'created_child': 0,
                'message': 'Aucune configuration de cycle trouvée.'
            })
        
        # Deduplicate: type_racine → max count needed
        type_max = {}
        for cc in configs:
            tid = cc.type_racine_id
            if tid not in type_max or cc.nombre_au_niveau_racine > type_max[tid]:
                type_max[tid] = cc.nombre_au_niveau_racine
        
        # Get hierarchies (parent_type → list of children)
        hierarchies = {}
        for h in RepartitionHierarchie.objects.filter(is_active=True, id_pays=pays.id_pays).select_related('type_parent', 'type_enfant'):
            hierarchies.setdefault(h.type_parent_id, []).append(h)
        
        created_root = 0
        created_child = 0
        
        for type_id, count in type_max.items():
            rtype = RepartitionType.objects.get(pk=type_id)
            
            # Check existing instances for this type + année + pays
            existing = list(
                RepartitionInstance.objects.filter(
                    type=rtype, annee=annee, pays=pays
                ).order_by('ordre')
            )
            # Fallback: also check without pays (legacy)
            if not existing:
                existing = list(
                    RepartitionInstance.objects.filter(
                        type=rtype, annee=annee
                    ).order_by('ordre')
                )
                for inst in existing:
                    if not inst.pays_id:
                        inst.pays = pays
                        inst.save(update_fields=['pays'])
            
            # Create missing root instances
            while len(existing) < count:
                n = len(existing) + 1
                instance_name = get_ordinal_name(n, rtype.nom)
                instance_code = f"{rtype.code}{n}"
                
                new_inst = RepartitionInstance.objects.create(
                    type=rtype, annee=annee, pays=pays,
                    nom=instance_name, code=instance_code,
                    ordre=n, is_active=True
                )
                existing.append(new_inst)
                created_root += 1
            
            # Handle children via hierarchies (only if hierarchy exists!)
            if type_id in hierarchies:
                for hier in hierarchies[type_id]:
                    child_type = hier.type_enfant
                    nb_per_parent = hier.nombre_enfants
                    
                    # Get existing children
                    existing_children = list(
                        RepartitionInstance.objects.filter(
                            type=child_type, annee=annee, pays=pays
                        ).order_by('ordre')
                    )
                    if not existing_children:
                        existing_children = list(
                            RepartitionInstance.objects.filter(
                                type=child_type, annee=annee
                            ).order_by('ordre')
                        )
                        for ci in existing_children:
                            if not ci.pays_id:
                                ci.pays = pays
                                ci.save(update_fields=['pays'])
                    
                    total_children_needed = count * nb_per_parent
                    
                    while len(existing_children) < total_children_needed:
                        n = len(existing_children) + 1
                        child_name = get_ordinal_name(n, child_type.nom)
                        child_code = f"{child_type.code}{n}"
                        
                        new_child = RepartitionInstance.objects.create(
                            type=child_type, annee=annee, pays=pays,
                            nom=child_name, code=child_code,
                            ordre=n, is_active=True
                        )
                        existing_children.append(new_child)
                        created_child += 1
        
        return JsonResponse({
            'success': True,
            'created_root': created_root,
            'created_child': created_child,
            'message': f'{created_root} instances principales + {created_child} subdivisions créées.'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# RÉPARTITIONS — PROPAGATION CALENDRIER NATIONAL
# ============================================================

@require_http_methods(["POST"])
def propagate_national_calendar(request):
    """
    Propage les modifications du calendrier national à tous les établissements.
    
    Quand un admin modifie les dates/état d'un RepartitionInstance national,
    cette API met à jour toutes les RepartitionConfigEtabAnnee marquées is_national=True
    qui utilisent cette instance.
    
    Payload: { "instance_id": int, "date_debut": str|null, "date_fin": str|null, "is_open": bool }
    """
    try:
        data = json.loads(request.body)
        instance_id = data.get('instance_id')
        
        if not instance_id:
            return JsonResponse({'success': False, 'error': 'instance_id requis.'}, status=400)
        
        instance = get_object_or_404(RepartitionInstance, pk=instance_id)
        
        # Update instance dates/state if provided
        date_debut = data.get('date_debut')
        date_fin = data.get('date_fin')
        is_open = data.get('is_open')
        
        if date_debut is not None:
            instance.date_debut = date_debut or None
        if date_fin is not None:
            instance.date_fin = date_fin or None
        if is_open is not None:
            instance.is_active = is_open
        instance.save()
        
        # Propagate to all national configs using this instance
        updated = RepartitionConfigEtabAnnee.objects.filter(
            repartition=instance,
            is_national=True  # Only propagate to national (non-custom) configs
        ).update(
            debut=instance.date_debut,
            fin=instance.date_fin,
            is_open=instance.is_active
        )
        
        return JsonResponse({
            'success': True,
            'instance_id': instance.id_instance,
            'establishments_updated': updated,
            'message': f'Calendrier propagé à {updated} établissement(s).'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def provision_repartitions_for_etab(request):
    """
    Provisionne manuellement les répartitions pour un établissement donné.
    Utile pour re-provisioner ou pour les établissements déjà existants
    qui n'avaient pas encore de répartitions.
    
    Payload: { "etablissement_id": int, "annee_id": int }
    """
    try:
        data = json.loads(request.body)
        etab_id = data.get('etablissement_id')
        annee_id = data.get('annee_id')
        
        if not etab_id or not annee_id:
            return JsonResponse({'success': False, 'error': 'etablissement_id et annee_id requis.'}, status=400)
        
        etablissement = get_object_or_404(Etablissement, id_etablissement=etab_id)
        annee = get_object_or_404(Annee, id_annee=annee_id)
        etab_pays = etablissement.pays
        
        # Get or create EtablissementAnnee
        etab_annee, _ = EtablissementAnnee.objects.get_or_create(
            etablissement=etablissement, annee=annee,
            defaults={'id_pays': etablissement.pays_id}
        )
        
        # Collect cycle IDs from activated classes
        activated_cycle_ids = set(
            EtablissementAnneeClasse.objects.filter(
                etablissement_annee=etab_annee
            ).values_list('classe__cycle_id', flat=True).distinct()
        )
        
        if not activated_cycle_ids:
            return JsonResponse({
                'success': False, 
                'error': 'Aucune classe activée pour cet établissement/année.'
            }, status=400)
        
        repartitions_created = 0
        
        # Find RepartitionConfigCycle for these cycles
        cycle_configs = RepartitionConfigCycle.objects.filter(
            cycle_id__in=activated_cycle_ids, is_active=True, id_pays=etab.pays_id
        ).select_related('type_racine')
        
        hierarchies = {}
        for h in RepartitionHierarchie.objects.filter(is_active=True, id_pays=etab.pays_id).select_related('type_parent', 'type_enfant'):
            hierarchies.setdefault(h.type_parent_id, []).append(h)
        
        existing_instance_ids = set(
            RepartitionConfigEtabAnnee.objects.filter(
                etablissement_annee=etab_annee
            ).values_list('repartition_id', flat=True)
        )
        
        type_combos = {}
        for cc in cycle_configs:
            tid = cc.type_racine_id
            if tid not in type_combos or cc.nombre_au_niveau_racine > type_combos[tid]:
                type_combos[tid] = cc.nombre_au_niveau_racine
        
        for type_id, nombre_racine in type_combos.items():
            rtype = RepartitionType.objects.get(pk=type_id)
            
            # Find or create root instances
            root_instances = list(
                RepartitionInstance.objects.filter(
                    type=rtype, annee=annee, pays=etab_pays
                ).order_by('ordre')
            )
            if not root_instances:
                root_instances = list(
                    RepartitionInstance.objects.filter(type=rtype, annee=annee).order_by('ordre')
                )
                for inst in root_instances:
                    if not inst.pays_id:
                        inst.pays = etab_pays
                        inst.save(update_fields=['pays'])
            
            while len(root_instances) < nombre_racine:
                n = len(root_instances) + 1
                root_instances.append(RepartitionInstance.objects.create(
                    type=rtype, annee=annee, pays=etab_pays,
                    nom=f"{rtype.nom} {n}", code=f"{rtype.code}{n}",
                    ordre=n, is_active=True
                ))
            
            for inst in root_instances[:nombre_racine]:
                if inst.id_instance not in existing_instance_ids:
                    RepartitionConfigEtabAnnee.objects.create(
                        etablissement_annee=etab_annee, repartition=inst,
                        parent=None, has_parent=False,
                        debut=inst.date_debut, fin=inst.date_fin,
                        is_open=True, is_national=True
                    )
                    existing_instance_ids.add(inst.id_instance)
                    repartitions_created += 1
            
            if type_id in hierarchies:
                for hier in hierarchies[type_id]:
                    child_type = hier.type_enfant
                    nb_children = hier.nombre_enfants
                    
                    root_configs = list(
                        RepartitionConfigEtabAnnee.objects.filter(
                            etablissement_annee=etab_annee,
                            repartition__type=rtype, has_parent=False
                        ).select_related('repartition').order_by('repartition__ordre')
                    )
                    
                    for parent_config in root_configs[:nombre_racine]:
                        child_instances = list(
                            RepartitionInstance.objects.filter(
                                type=child_type, annee=annee, pays=etab_pays
                            ).order_by('ordre')
                        )
                        if not child_instances:
                            child_instances = list(
                                RepartitionInstance.objects.filter(
                                    type=child_type, annee=annee
                                ).order_by('ordre')
                            )
                            for ci in child_instances:
                                if not ci.pays_id:
                                    ci.pays = etab_pays
                                    ci.save(update_fields=['pays'])
                        
                        parent_idx = parent_config.repartition.ordre - 1
                        start_child = parent_idx * nb_children
                        total_needed = start_child + nb_children
                        
                        while len(child_instances) < total_needed:
                            n = len(child_instances) + 1
                            child_instances.append(RepartitionInstance.objects.create(
                                type=child_type, annee=annee, pays=etab_pays,
                                nom=f"{child_type.nom} {n}", code=f"{child_type.code}{n}",
                                ordre=n, is_active=True
                            ))
                        
                        for child_inst in child_instances[start_child:start_child + nb_children]:
                            if child_inst.id_instance not in existing_instance_ids:
                                RepartitionConfigEtabAnnee.objects.create(
                                    etablissement_annee=etab_annee, repartition=child_inst,
                                    parent=parent_config, has_parent=True,
                                    debut=child_inst.date_debut, fin=child_inst.date_fin,
                                    is_open=True, is_national=True
                                )
                                existing_instance_ids.add(child_inst.id_instance)
                                repartitions_created += 1
        
        return JsonResponse({
            'success': True,
            'repartitions_created': repartitions_created,
            'message': f'{repartitions_created} répartition(s) provisionnée(s) pour {etablissement.nom}.'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# ÉVALUATIONS — API CRUD
# ============================================================

@require_http_methods(["GET"])
def get_evaluation_types(request):
    """Retourne les types d'évaluations depuis le Hub."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        types = list(EvaluationType.objects.filter(is_active=True, id_pays=etab.pays_id).values(
            'id_type_eval', 'nom', 'sigle', 'description', 'is_active'
        ))
        return JsonResponse({'success': True, 'types': types})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def save_evaluation_type(request):
    """Créer ou modifier un type d'évaluation."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        id_type = data.get('id_type_eval')
        nom = (data.get('nom') or '').strip()
        sigle = (data.get('sigle') or '').strip().upper()
        description = (data.get('description') or '').strip()

        if not nom or not sigle:
            return JsonResponse({'success': False, 'error': 'Nom et sigle requis.'}, status=400)

        if id_type:
            obj = EvaluationType.objects.get(id_type_eval=id_type, id_pays=etab.pays_id)
            obj.nom = nom
            obj.sigle = sigle
            obj.description = description
            obj.save()
        else:
            obj = EvaluationType.objects.create(nom=nom, sigle=sigle, description=description, id_pays=etab.pays_id)

        return JsonResponse({'success': True, 'id_type_eval': obj.id_type_eval})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def delete_evaluation_type(request):
    """Supprimer un type d'évaluation."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        data = json.loads(request.body)
        obj = EvaluationType.objects.get(id_type_eval=data.get('id_type_eval'), id_pays=etab.pays_id)
        obj.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _get_spoke_connection():
    """Helper: ouvre une connexion pymysql vers la spoke database."""
    from django.db import connections
    import pymysql
    db_settings = connections['default'].settings_dict
    return pymysql.connect(
        host=db_settings.get('HOST', 'localhost') or 'localhost',
        user=db_settings['USER'],
        password=db_settings['PASSWORD'],
        port=int(db_settings.get('PORT', 3306) or 3306),
        database=db_settings['NAME'],
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=5,
        read_timeout=10,
    )


@require_http_methods(["GET"])
def get_evaluation_cours(request):
    """Retourne les cours pour une classe via le Hub (CoursAnnee/Cours)."""
    try:
        eac_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        if not eac_id:
            return JsonResponse({'success': False, 'error': 'classe_id requis.'}, status=400)

        # already imported at top
        # Resolve EAC to actual classe + annee
        eac = EtablissementAnneeClasse.objects.select_related(
            'classe', 'etablissement_annee', 'etablissement_annee__annee',
            'etablissement_annee__etablissement'
        ).get(id=int(eac_id))

        classe_id = eac.classe_id
        annee_id = eac.etablissement_annee.annee_id
        etab_id = eac.etablissement_annee.etablissement_id
        id_pays = eac.etablissement_annee.id_pays or eac.etablissement_annee.etablissement.pays_id

        # Get CoursAnnee for this classe + annee (national OR etab-specific)
        from django.db.models import Q
        cours_annee_qs = CoursAnnee.objects.filter(
            cours__classe_id=classe_id,
            annee_id=annee_id,
            id_pays=id_pays
        ).filter(
            Q(etablissement__isnull=True) | Q(etablissement_id=etab_id)
        ).select_related('cours').order_by('cours__cours')

        cours = [{
            'id_cours_classe': ca.id_cours_annee,
            'nom': ca.cours.cours,
            'code': ca.cours.code_cours,
            'maxima_exam': ca.maxima_exam,
            'maxima': ca.maxima_exam,
        } for ca in cours_annee_qs]

        return JsonResponse({'success': True, 'cours': cours})
    except EtablissementAnneeClasse.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Classe non trouvée.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_evaluations_list(request):
    """Retourne les évaluations pour une classe (optionnellement filtrées par cours)."""
    try:
        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        cours_id = request.GET.get('cours_id')  # Optional filter
        if not classe_id:
            return JsonResponse({'success': False, 'error': 'classe_id requis.'}, status=400)

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        # Resolve EAC.id → business keys
        bk = _resolve_eac_orm(classe_id)
        if not bk:
            return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Build optional cours filter
                cours_filter = ""
                params = [bk['classe_id'], bk['groupe'], bk['section_id'], etab_id]
                if cours_id:
                    cours_filter = "AND e.id_cours_classe_id = %s"
                    params.append(cours_id)

                cur.execute(f"""
                    SELECT e.id_evaluation, e.title, e.id_type_eval,
                           e.ponderer_eval, e.date_eval, e.date_soumission,
                           e.contenu_evaluation, e.document_url,
                           e.id_repartition_instance, e.id_cours_classe_id,
                           et.sigle AS type_sigle, et.nom AS type_nom,
                           ri.nom AS repartition_nom,
                           ca.cours AS cours_nom, ca.code_cours AS cours_code,
                           (SELECT COUNT(*) FROM evaluation_repartition er WHERE er.id_evaluation = e.id_evaluation) AS assign_count
                    FROM evaluation e
                    LEFT JOIN countryStructure.evaluation_types et ON et.id = e.id_type_eval
                    LEFT JOIN countryStructure.repartition_instances ri ON ri.id = e.id_repartition_instance
                    LEFT JOIN countryStructure.cours_annee cann ON cann.id_cours_annee = e.id_cours_classe_id
                    LEFT JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    WHERE e.classe_id = %s AND e.groupe <=> %s AND e.section_id <=> %s
                          AND e.id_etablissement = %s
                          {cours_filter}
                    ORDER BY ca.cours ASC, e.date_eval DESC, e.id_evaluation DESC
                """, params)
                evals = []
                for r in cur.fetchall():
                    evals.append({
                        'id_evaluation': r['id_evaluation'],
                        'title': r['title'],
                        'id_type_eval': r['id_type_eval'],
                        'type_sigle': r['type_sigle'],
                        'type_nom': r['type_nom'],
                        'ponderer_eval': r['ponderer_eval'],
                        'date_eval': r['date_eval'].isoformat() if r['date_eval'] else None,
                        'date_soumission': r['date_soumission'].isoformat() if r['date_soumission'] else None,
                        'contenu': r['contenu_evaluation'] or '',
                        'document_url': r['document_url'],
                        'is_assigned': r['assign_count'] > 0,
                        'id_repartition_instance': r['id_repartition_instance'],
                        'repartition_nom': r['repartition_nom'] or '',
                        'cours_nom': r['cours_nom'] or '',
                        'cours_code': r['cours_code'] or '',
                        'id_cours_classe': r['id_cours_classe_id'],
                    })
            return JsonResponse({'success': True, 'evaluations': evals})
        finally:
            conn.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def save_evaluation(request):
    """Créer ou modifier une évaluation (avec upload PDF optionnel)."""
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        # FormData (not JSON)
        eval_id = request.POST.get('id_evaluation', '').strip()
        title = request.POST.get('title', '').strip()
        id_type_eval = request.POST.get('id_type_eval', '').strip() or None
        date_eval = request.POST.get('date_eval', '').strip()
        date_soumission = request.POST.get('date_soumission', '').strip() or None
        ponderer_eval = request.POST.get('ponderer_eval', '').strip()
        contenu = request.POST.get('contenu_evaluation', '').strip()
        classe_id = request.POST.get('id_classe_id', '').strip()
        cours_id = request.POST.get('id_cours_classe_id', '').strip()
        doc_file = request.FILES.get('document')
        repartition_id = request.POST.get('id_repartition_instance', '').strip() or None

        if not title or not date_eval or not ponderer_eval:
            return JsonResponse({'success': False, 'error': 'Titre, date et maximum requis.'}, status=400)

        # Handle PDF upload
        document_url = None
        if doc_file:
            import os
            from django.conf import settings
            eval_dir = os.path.join(settings.MEDIA_ROOT, 'Evaluations', f'EtabID_{etab_id}')
            os.makedirs(eval_dir, exist_ok=True)
            try:
                os.chmod(eval_dir, 0o755)
            except Exception:
                pass

            from datetime import datetime
            ts = datetime.now().strftime('%Y%m%d%H%M%S')
            ext = os.path.splitext(doc_file.name)[1].lower() or '.pdf'
            # Use eval_id if editing, else use timestamp
            fname = f"Eval_{eval_id or 'new'}_{ts}{ext}"
            filepath = os.path.join(eval_dir, fname)
            with open(filepath, 'wb+') as f:
                for chunk in doc_file.chunks():
                    f.write(chunk)
            try:
                os.chmod(filepath, 0o644)
            except Exception:
                pass
            document_url = f"/media/Evaluations/EtabID_{etab_id}/{fname}"

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Get annee_id from the session context
                # Resolve EAC → business keys
                bk = _resolve_eac_keys(cur, classe_id)

                # Get campus_id
                cur.execute("""
                    SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1
                """, [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                if eval_id:
                    # UPDATE
                    update_parts = [
                        "title=%s", "id_type_eval=%s", "date_eval=%s",
                        "date_soumission=%s", "ponderer_eval=%s", "contenu_evaluation=%s",
                        "id_repartition_instance=%s"
                    ]
                    params = [title, id_type_eval, date_eval, date_soumission, int(ponderer_eval), contenu, repartition_id]
                    if document_url:
                        update_parts.append("document_url=%s")
                        params.append(document_url)
                    params.extend([int(eval_id), etab_id])
                    cur.execute(f"UPDATE evaluation SET {', '.join(update_parts)} WHERE id_evaluation=%s AND id_etablissement=%s", params)
                    new_id = int(eval_id)

                    # Sync evaluation_repartition when repartition changes
                    if repartition_id:
                        config = _get_or_create_repartition_config(int(repartition_id), etab_id)
                        if config:
                            cur.execute("DELETE FROM evaluation_repartition WHERE id_evaluation = %s", [new_id])
                            cur.execute("""
                                INSERT INTO evaluation_repartition (id_evaluation, id_repartition_config)
                                VALUES (%s, %s)
                            """, [new_id, config.id])
                else:
                    # INSERT — id_classe_id is the legacy EAC FK (NOT NULL)
                    cur.execute("""
                        INSERT INTO evaluation (title, id_type_eval, ponderer_eval, date_eval, date_soumission,
                            contenu_evaluation, document_url, id_annee_id, idCampus_id, id_classe_id, classe_id,
                            groupe, section_id, id_cours_classe_id, id_repartition_instance, id_etablissement, id_pays, date_creation)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    """, [title, id_type_eval, int(ponderer_eval), date_eval, date_soumission,
                          contenu, document_url, bk['annee_id'] if bk else None, campus_id,
                          int(classe_id),
                          bk['classe_id'] if bk else int(classe_id),
                          bk['groupe'] if bk else None, bk['section_id'] if bk else None,
                          int(cours_id), repartition_id, etab_id, etab.pays_id])
                    new_id = cur.lastrowid

                    # Auto-create evaluation_repartition entry
                    if repartition_id:
                        config = _get_or_create_repartition_config(int(repartition_id), etab_id)
                        if config:
                            cur.execute("""
                                INSERT INTO evaluation_repartition (id_evaluation, id_repartition_config)
                                VALUES (%s, %s)
                            """, [new_id, config.id])

                    # Rename the uploaded file with actual ID
                    if document_url and 'new' in document_url:
                        import os
                        from django.conf import settings
                        old_path = os.path.join(settings.MEDIA_ROOT, document_url.lstrip('/media/'))
                        from datetime import datetime
                        ts = datetime.now().strftime('%Y%m%d%H%M%S')
                        new_fname = f"Eval_{new_id}_{ts}{ext}"
                        new_path = os.path.join(os.path.dirname(old_path), new_fname)
                        if os.path.exists(old_path):
                            os.rename(old_path, new_path)
                        document_url = f"/media/Evaluations/EtabID_{etab_id}/{new_fname}"
                        cur.execute("UPDATE evaluation SET document_url=%s WHERE id_evaluation=%s",
                                    [document_url, new_id])

            conn.commit()
            return JsonResponse({'success': True, 'id_evaluation': new_id, 'document_url': document_url})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def delete_evaluation(request):
    """Supprimer une évaluation."""
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        data = json.loads(request.body)
        eval_id = data.get('id_evaluation')
        if not eval_id:
            return JsonResponse({'success': False, 'error': 'id_evaluation requis.'}, status=400)

        conn = _get_spoke_connection()
        try:
            etab, err = _get_tenant_etab(request)
            if err: return err
            etab_id = etab.id_etablissement
            with conn.cursor() as cur:
                # Get document_url before deleting to remove file
                cur.execute("SELECT document_url FROM evaluation WHERE id_evaluation=%s AND id_etablissement=%s", [eval_id, etab_id])
                row = cur.fetchone()
                if row and row['document_url']:
                    import os
                    from django.conf import settings
                    filepath = os.path.join(settings.MEDIA_ROOT, row['document_url'].lstrip('/media/'))
                    if os.path.exists(filepath):
                        try:
                            os.remove(filepath)
                        except Exception:
                            pass

                cur.execute("DELETE FROM evaluation WHERE id_evaluation=%s AND id_etablissement=%s", [eval_id, etab_id])
            conn.commit()
            return JsonResponse({'success': True})
        finally:
            conn.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_evaluation_candidates(request):
    """Retourne les évaluations candidates pour une répartition (filtrées par intervalle de dates)."""
    try:
        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        cours_id = request.GET.get('cours_id')
        repartition_id = request.GET.get('repartition_id')

        if not classe_id or not cours_id or not repartition_id:
            return JsonResponse({'success': False, 'error': 'Paramètres requis.'}, status=400)

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        # Resolve EAC.id → business keys
        bk = _resolve_eac_orm(classe_id)
        if not bk:
            return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

        # Get repartition config dates
        config = RepartitionConfigEtabAnnee.objects.filter(
            repartition_id=repartition_id
        ).first()

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Candidates: evaluations matching by date range OR by repartition_instance
                if config and config.debut and config.fin:
                    cur.execute("""
                        SELECT e.id_evaluation, e.title, e.id_type_eval,
                               e.ponderer_eval, e.date_eval,
                               et.sigle AS type_sigle
                        FROM evaluation e
                        LEFT JOIN countryStructure.evaluation_types et ON et.id = e.id_type_eval
                        WHERE e.classe_id = %s AND e.groupe <=> %s AND e.section_id <=> %s
                              AND e.id_cours_classe_id = %s
                              AND e.id_etablissement = %s
                              AND (e.date_eval BETWEEN %s AND %s OR e.id_repartition_instance = %s)
                        ORDER BY e.date_eval ASC
                    """, [bk['classe_id'], bk['groupe'], bk['section_id'], cours_id, etab_id,
                          config.debut, config.fin, int(repartition_id)])
                else:
                    cur.execute("""
                        SELECT e.id_evaluation, e.title, e.id_type_eval,
                               e.ponderer_eval, e.date_eval,
                               et.sigle AS type_sigle
                        FROM evaluation e
                        LEFT JOIN countryStructure.evaluation_types et ON et.id = e.id_type_eval
                        WHERE e.classe_id = %s AND e.groupe <=> %s AND e.section_id <=> %s
                              AND e.id_cours_classe_id = %s
                              AND e.id_etablissement = %s
                              AND (e.id_repartition_instance = %s OR 1=1)
                        ORDER BY e.date_eval ASC
                    """, [bk['classe_id'], bk['groupe'], bk['section_id'], cours_id, etab_id,
                          int(repartition_id)])

                candidates = [{
                    'id_evaluation': r['id_evaluation'],
                    'title': r['title'],
                    'id_type_eval': r['id_type_eval'],
                    'type_sigle': r['type_sigle'],
                    'ponderer_eval': r['ponderer_eval'],
                    'date_eval': r['date_eval'].isoformat() if r['date_eval'] else None,
                } for r in cur.fetchall()]

                # Already assigned
                if config:
                    cur.execute("""
                        SELECT er.id_evaluation, er.pourcentage, er.id_note_type
                        FROM evaluation_repartition er
                        WHERE er.id_repartition_config = %s
                            AND er.id_evaluation IN (
                                SELECT e.id_evaluation FROM evaluation e
                                WHERE e.id_cours_classe_id = %s AND e.classe_id = %s
                                  AND e.groupe <=> %s AND e.section_id <=> %s
                            )
                    """, [config.id, cours_id, bk['classe_id'], bk['groupe'], bk['section_id']])
                    assigned = [{
                        'id_evaluation': r['id_evaluation'],
                        'pourcentage': float(r['pourcentage']) if r['pourcentage'] else None,
                        'id_note_type': r['id_note_type'],
                    } for r in cur.fetchall()]
                else:
                    assigned = []

            return JsonResponse({'success': True, 'candidates': candidates, 'assigned': assigned})
        finally:
            conn.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def assign_evaluations(request):
    """Assigner/mettre à jour les évaluations sélectionnées pour une répartition."""
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        data = json.loads(request.body)
        repartition_id = data.get('repartition_id')
        cours_id = data.get('cours_id')
        classe_id = data.get('classe_id') or data.get('id_classe_id')
        assignments = data.get('assignments', [])

        if not repartition_id or not cours_id:
            return JsonResponse({'success': False, 'error': 'repartition_id et cours_id requis.'}, status=400)

        # Resolve EAC.id → business keys
        bk = _resolve_eac_orm(classe_id)
        if not bk:
            return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

        config = RepartitionConfigEtabAnnee.objects.filter(
            repartition_id=repartition_id
        ).first()
        if not config:
            return JsonResponse({'success': False, 'error': 'Configuration de répartition introuvable.'}, status=404)

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Delete existing assignments for this config + cours + classe
                cur.execute("""
                    DELETE er FROM evaluation_repartition er
                    JOIN evaluation e ON e.id_evaluation = er.id_evaluation
                    WHERE er.id_repartition_config = %s
                      AND e.id_cours_classe_id = %s
                      AND e.classe_id = %s AND e.groupe <=> %s AND e.section_id <=> %s
                """, [config.id, cours_id, bk['classe_id'], bk['groupe'], bk['section_id']])

                # Insert new assignments
                for a in assignments:
                    eval_id = a.get('id_evaluation')
                    pct = a.get('pourcentage')
                    if eval_id:
                        cur.execute("""
                            INSERT INTO evaluation_repartition (id_evaluation, id_repartition_config, pourcentage)
                            VALUES (%s, %s, %s)
                        """, [eval_id, config.id, pct])

            conn.commit()
            return JsonResponse({'success': True, 'count': len(assignments)})
        finally:
            conn.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# NOTES — HELPER: AUTO-PROVISION REPARTITION CONFIG
# ============================================================

def _get_or_create_repartition_config(repartition_id, etab_id):
    """
    Ensures a RepartitionConfigEtabAnnee exists for the given repartition instance
    AND the current establishment. Without the etab filter, multi-tenant setups
    could return a config from a different establishment, causing config_id mismatches
    when reading notes back on the bulletin.
    """
    # Find the EtablissementAnnee for this etab first
    try:
        ri = RepartitionInstance.objects.select_related('type').get(pk=repartition_id)
    except RepartitionInstance.DoesNotExist:
        # Fallback: try without etab filter (single-tenant)
        config = RepartitionConfigEtabAnnee.objects.filter(
            repartition_id=repartition_id
        ).select_related('repartition', 'repartition__type').first()
        return config

    etab_annee = EtablissementAnnee.objects.filter(
        etablissement_id=etab_id,
        annee=ri.annee
    ).first()

    if etab_annee:
        # Filter by BOTH repartition AND establishment to get the correct config_id
        config = RepartitionConfigEtabAnnee.objects.filter(
            repartition_id=repartition_id,
            etablissement_annee_id=etab_annee.id
        ).select_related('repartition', 'repartition__type').first()

        if config:
            return config

    # Config doesn't exist — auto-provision
    if not etab_annee:
        return None

    config = RepartitionConfigEtabAnnee.objects.create(
        etablissement_annee=etab_annee,
        repartition=ri,
        parent=None,
        has_parent=False,
        debut=ri.date_debut,
        fin=ri.date_fin,
        is_open=True,
        is_national=True,
    )
    return RepartitionConfigEtabAnnee.objects.filter(
        id=config.id
    ).select_related('repartition', 'repartition__type').first()


# ============================================================
# NOTES — API SAISIE & IMPORT
# ============================================================

@require_http_methods(["GET"])
def get_notes_grid(request):
    """
    Charge la grille de saisie des notes.
    Params: classe_id (EAC id), repartition_id, note_type (TJ|EXAM)
    Retourne: élèves inscrits, cours avec évaluations assignées, notes existantes.
    """
    try:
        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        repartition_id = request.GET.get('repartition_id')
        note_type = request.GET.get('note_type', 'TJ')  # TJ ou EXAM

        if not classe_id or not repartition_id:
            return JsonResponse({'success': False, 'error': 'classe_id et repartition_id requis.'}, status=400)

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        # Get repartition config for dates (auto-creates if synched mode)
        config = _get_or_create_repartition_config(repartition_id, etab_id)

        repartition_name = config.repartition.nom if config else ''

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # 1. Get annee + campus + business keys from the EAC
                cur.execute("""
                    SELECT ea.annee_id AS id_annee, ea.etablissement_id AS id_etab,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()
                if not ctx:
                    return JsonResponse({'success': False, 'error': 'Classe non trouvée.'}, status=404)

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # 2. Get enrolled students (filtered by business keys)
                cur.execute("""
                    SELECT DISTINCT e.id_eleve, e.nom, e.prenom
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                      AND ei.status = 1
                    ORDER BY e.nom, e.prenom
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleves = cur.fetchall()

                # 3. Get evaluations for this repartition (period)
                #    Loads ALL evaluations whose id_repartition_instance matches,
                #    no need to be formally assigned via evaluation_repartition.
                cur.execute("""
                    SELECT ev.id_evaluation, ev.title, ev.ponderer_eval,
                           ev.id_cours_classe_id, ev.date_eval,
                           et.sigle AS type_sigle,
                           ca.cours AS cours_nom, ca.code_cours AS cours_code
                    FROM evaluation ev
                    LEFT JOIN countryStructure.evaluation_types et ON et.id = ev.id_type_eval
                    LEFT JOIN countryStructure.cours_annee cann ON cann.id_cours_annee = ev.id_cours_classe_id
                    LEFT JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    WHERE ev.id_repartition_instance = %s
                      AND ev.classe_id = %s AND ev.groupe <=> %s AND ev.section_id <=> %s
                      AND ev.id_etablissement = %s
                    ORDER BY ev.id_cours_classe_id, ev.date_eval
                """, [repartition_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section'], etab_id])
                evaluations = cur.fetchall()

                # 4. Get existing notes for these evaluations + students
                eval_ids = [e['id_evaluation'] for e in evaluations]
                notes = {}
                if eval_ids:
                    placeholders = ','.join(['%s'] * len(eval_ids))
                    cur.execute(f"""
                        SELECT en.id_eleve_id, en.id_evaluation_id, en.note, en.id_note
                        FROM eleve_note en
                        WHERE en.id_evaluation_id IN ({placeholders})
                    """, eval_ids)
                    for n in cur.fetchall():
                        key = f"{n['id_eleve_id']}_{n['id_evaluation_id']}"
                        notes[key] = {'note': float(n['note']) if n['note'] is not None else None, 'id_note': n['id_note']}

                # 5. Group evaluations by cours
                cours_evals = {}
                for ev in evaluations:
                    cid = ev['id_cours_classe_id']
                    if cid not in cours_evals:
                        cours_evals[cid] = {
                            'cours_id': cid,
                            'cours_nom': ev['cours_nom'] or f'Cours #{cid}',
                            'cours_code': ev['cours_code'] or '',
                            'evaluations': []
                        }
                    cours_evals[cid]['evaluations'].append({
                        'id_evaluation': ev['id_evaluation'],
                        'title': ev['title'],
                        'max': ev['ponderer_eval'],
                        'type_sigle': ev['type_sigle'],
                        'date_eval': ev['date_eval'].isoformat() if ev['date_eval'] else None,
                    })

            # Get bulletin maxima (TJ expected max for this period type)
            bulletin_maxima = None
            if config and config.repartition:
                try:
                    from django.db import connections as _conns
                    _hub_cur = _conns['countryStructure'].cursor()
                    try:
                        _hub_cur.execute("""
                            SELECT rtn.ponderation_max
                            FROM repartition_type_notes rtn
                            JOIN note_types nt ON nt.id = rtn.note_type_id
                            WHERE rtn.repartition_type_id = %s AND nt.sigle = 'TJ'
                              AND rtn.source_type = 'EVALUATIONS' AND rtn.is_active = 1
                            LIMIT 1
                        """, [config.repartition.type_id])
                        _tj_row = _hub_cur.fetchone()
                        if _tj_row:
                            bulletin_maxima = float(_tj_row[0]) if _tj_row[0] else 20
                    finally:
                        _hub_cur.close()
                except Exception:
                    bulletin_maxima = 20

            return JsonResponse({
                'success': True,
                'repartition_name': repartition_name,
                'eleves': [{'id': e['id_eleve'], 'nom': e['nom'], 'prenom': e['prenom']} for e in eleves],
                'cours': list(cours_evals.values()),
                'notes': notes,
                'bulletin_maxima': bulletin_maxima,
                'context': {
                    'annee_id': ctx['id_annee'],
                    'campus_id': campus_id,
                    'etab_id': etab_id,
                    'config_id': config.id if config else None,
                }
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def save_notes(request):
    """
    Enregistre les notes en batch.
    Payload: { notes: [{eleve_id, evaluation_id, note, id_note?}], context: {annee_id, campus_id, ...} }
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        data = json.loads(request.body)
        notes_data = data.get('notes', [])
        ctx = data.get('context', {})

        if not notes_data:
            return JsonResponse({'success': False, 'error': 'Aucune note à enregistrer.'}, status=400)

        conn = _get_spoke_connection()
        try:
            saved = 0
            updated = 0
            with conn.cursor() as cur:
                for nd in notes_data:
                    eleve_id = nd.get('eleve_id')
                    eval_id = nd.get('evaluation_id')
                    note_val = nd.get('note')
                    id_note = nd.get('id_note')

                    # Skip empty notes
                    if note_val is None or str(note_val).strip() == '':
                        # If existing note, delete it
                        if id_note:
                            cur.execute("DELETE FROM eleve_note WHERE id_note = %s", [id_note])
                        continue

                    try:
                        note_float = float(note_val)
                    except (ValueError, TypeError):
                        continue

                    if id_note:
                        # UPDATE existing
                        cur.execute("UPDATE eleve_note SET note = %s WHERE id_note = %s", [note_float, id_note])
                        updated += 1
                    else:
                        # Check if exists already
                        cur.execute("""
                            SELECT id_note FROM eleve_note
                            WHERE id_eleve_id = %s AND id_evaluation_id = %s
                            LIMIT 1
                        """, [eleve_id, eval_id])
                        existing = cur.fetchone()
                        if existing:
                            cur.execute("UPDATE eleve_note SET note = %s WHERE id_note = %s",
                                        [note_float, existing['id_note']])
                            updated += 1
                        else:
                            # Get evaluation context for FK fields
                            cur.execute("""
                                SELECT id_annee_id, idCampus_id, classe_id, groupe, section_id,
                                       id_cours_classe_id, id_classe_id, id_cycle_id,
                                       id_repartition_instance, id_session_id, id_type_eval
                                FROM evaluation WHERE id_evaluation = %s
                            """, [eval_id])
                            ev_ctx = cur.fetchone()
                            if not ev_ctx:
                                continue

                            # Map evaluation type to note type: default TJ=1
                            id_type_note = 1  # TJ by default

                            cur.execute("""
                                INSERT INTO eleve_note
                                    (id_eleve_id, id_evaluation_id, note, id_annee_id, idCampus_id,
                                     id_classe_id, classe_id, groupe, section_id,
                                     id_cycle_id, id_repartition_instance,
                                     id_cours_id, id_type_note_id, id_session_id,
                                     id_etablissement, id_pays, date_saisie)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                            """, [eleve_id, eval_id, note_float,
                                  ev_ctx['id_annee_id'], ev_ctx['idCampus_id'],
                                  ev_ctx['id_classe_id'] or 0, ev_ctx['classe_id'], ev_ctx['groupe'], ev_ctx['section_id'],
                                  ev_ctx['id_cycle_id'] or 0,
                                  ev_ctx['id_repartition_instance'] or 0,
                                  ev_ctx['id_cours_classe_id'], id_type_note,
                                  ev_ctx['id_session_id'] or 0, etab_id, etab.pays_id])
                            saved += 1

            conn.commit()
            return JsonResponse({
                'success': True,
                'saved': saved,
                'updated': updated,
                'message': f'{saved} note(s) créée(s), {updated} mise(s) à jour.'
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def download_notes_template(request):
    """Génère un modèle Excel pour la saisie des notes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Protection, PatternFill, Alignment, Border, Side
        from io import BytesIO

        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        repartition_id = request.GET.get('repartition_id')
        cours_id_filter = request.GET.get('cours_id', '')
        eval_id_filter = request.GET.get('eval_id', '')

        if not classe_id or not repartition_id:
            return JsonResponse({'success': False, 'error': 'Paramètres requis.'}, status=400)

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        config = _get_or_create_repartition_config(repartition_id, etab_id)

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT ea.annee_id AS id_annee,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # Get students (filtered by business keys)
                cur.execute("""
                    SELECT DISTINCT e.id_eleve, e.nom, e.prenom
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                      AND ei.status = 1
                    ORDER BY e.nom, e.prenom
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleves = cur.fetchall()

                # Get evaluations for this period
                eval_sql = """
                    SELECT ev.id_evaluation, ev.title, ev.ponderer_eval,
                           ca.cours AS cours_nom, ev.id_cours_classe_id
                    FROM evaluation ev
                    LEFT JOIN countryStructure.cours_annee cann ON cann.id_cours_annee = ev.id_cours_classe_id
                    LEFT JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    WHERE ev.id_repartition_instance = %s
                      AND ev.classe_id = %s AND ev.groupe <=> %s AND ev.section_id <=> %s
                      AND ev.id_etablissement = %s
                """
                eval_params = [repartition_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section'], etab_id]

                # Apply optional cours filter
                if cours_id_filter:
                    eval_sql += " AND ev.id_cours_classe_id = %s"
                    eval_params.append(cours_id_filter)  

                # Apply optional single-eval filter
                if eval_id_filter:
                    eval_sql += " AND ev.id_evaluation = %s"
                    eval_params.append(eval_id_filter)

                eval_sql += " ORDER BY ev.id_cours_classe_id, ev.date_eval"
                cur.execute(eval_sql, eval_params)
                evals = cur.fetchall()

                # ---- Load existing notes for these evaluations + students ----
                existing_notes = {}
                eval_ids = [e['id_evaluation'] for e in evals]
                eleve_ids = [e['id_eleve'] for e in eleves]
                if eval_ids and eleve_ids:
                    placeholders_ev = ','.join(['%s'] * len(eval_ids))
                    placeholders_el = ','.join(['%s'] * len(eleve_ids))
                    cur.execute(f"""
                        SELECT en.id_eleve_id, en.id_evaluation_id, en.note
                        FROM eleve_note en
                        WHERE en.id_evaluation_id IN ({placeholders_ev})
                          AND en.id_eleve_id IN ({placeholders_el})
                    """, eval_ids + eleve_ids)
                    for n in cur.fetchall():
                        key = f"{n['id_eleve_id']}_{n['id_evaluation_id']}"
                        existing_notes[key] = float(n['note']) if n['note'] is not None else None
        finally:
            conn.close()

        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Saisie Notes"

        header_fill = PatternFill('solid', fgColor='4F46E5')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        locked = Protection(locked=True)
        unlocked = Protection(locked=False)
        note_fill = PatternFill('solid', fgColor='F0FDF4')  # Light green for pre-filled notes
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = ['ID Élève', 'Nom', 'Prénom']
        for ev in evals:
            cname = ev['cours_nom'] or '?'
            headers.append(f"{cname[:12]} - {ev['title'][:15]} (/{ev['ponderer_eval']})")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Data rows — pre-fill with existing notes
        for row, el in enumerate(eleves, 2):
            ws.cell(row=row, column=1, value=el['id_eleve']).protection = locked
            ws.cell(row=row, column=2, value=el['nom']).protection = locked
            ws.cell(row=row, column=3, value=el['prenom']).protection = locked
            for col_offset, ev in enumerate(evals):
                key = f"{el['id_eleve']}_{ev['id_evaluation']}"
                note_val = existing_notes.get(key)
                cell = ws.cell(row=row, column=4 + col_offset, value=note_val if note_val is not None else '')
                cell.protection = unlocked
                cell.border = thin_border
                if note_val is not None:
                    cell.fill = note_fill

        # Hidden metadata sheet
        ws2 = wb.create_sheet('_meta')
        ws2.cell(row=1, column=1, value='classe_id')
        ws2.cell(row=1, column=2, value=int(classe_id))
        ws2.cell(row=2, column=1, value='repartition_id')
        ws2.cell(row=2, column=2, value=int(repartition_id))
        ws2.cell(row=3, column=1, value='etab_id')
        ws2.cell(row=3, column=2, value=etab_id)
        for i, ev in enumerate(evals):
            ws2.cell(row=4 + i, column=1, value=f'eval_{i}')
            ws2.cell(row=4 + i, column=2, value=ev['id_evaluation'])
            ws2.cell(row=4 + i, column=3, value=ev['ponderer_eval'])
        ws2.sheet_state = 'hidden'

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        for i in range(len(evals)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(4 + i)].width = 22

        ws.protection.enable()

        rep_name = config.repartition.nom if config else 'Notes'
        filename = f"Modele_Notes_{rep_name.replace(' ', '_')}.xlsx"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            content=buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def import_notes_excel(request):
    """Importe les notes depuis un fichier Excel (modèle généré par download_notes_template)."""
    try:
        import openpyxl
        from io import BytesIO

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({'success': False, 'error': 'Aucun fichier.'}, status=400)

        file_bytes = uploaded_file.read()
        # Load with data_only=True to read computed values (not formulas)
        wb_data = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        # Load without data_only for metadata (formulas aren't used there)
        wb_meta = openpyxl.load_workbook(BytesIO(file_bytes))

        # Read metadata from wb_meta (reliable for non-formula cells)
        if '_meta' not in wb_meta.sheetnames:
            return JsonResponse({'success': False, 'error': 'Fichier invalide: pas de métadonnées.'}, status=400)

        meta = wb_meta['_meta']
        classe_id = int(meta.cell(row=1, column=2).value)
        repartition_id = int(meta.cell(row=2, column=2).value)

        # Read eval IDs + maxima from meta
        eval_map = []  # list of {id_evaluation, max}
        row = 4
        while True:
            key = meta.cell(row=row, column=1).value
            if not key or not str(key).startswith('eval_'):
                break
            eval_map.append({
                'id_evaluation': int(meta.cell(row=row, column=2).value),
                'max': float(meta.cell(row=row, column=3).value or 0)
            })
            row += 1

        if not eval_map:
            return JsonResponse({'success': False, 'error': 'Aucune évaluation dans le modèle.'}, status=400)

        # Read notes from main sheet (using data_only workbook for formula results)
        ws = wb_data.active
        notes_to_save = []
        errors = []

        for row_idx in range(2, ws.max_row + 1):
            eleve_id = ws.cell(row=row_idx, column=1).value
            nom = ws.cell(row=row_idx, column=2).value
            if not eleve_id:
                continue

            for col_idx, em in enumerate(eval_map):
                cell_val = ws.cell(row=row_idx, column=4 + col_idx).value
                if cell_val is None or str(cell_val).strip() == '':
                    continue
                try:
                    note_val = float(cell_val)
                except (ValueError, TypeError):
                    errors.append(f"Note invalide pour {nom} col {col_idx + 1}: '{cell_val}'")
                    continue

                if note_val < 0:
                    errors.append(f"Note négative pour {nom}: {note_val}")
                    continue
                if note_val > em['max'] and em['max'] > 0:
                    errors.append(f"Note {note_val} > max {em['max']} pour {nom}")
                    continue

                notes_to_save.append({
                    'eleve_id': int(eleve_id),
                    'evaluation_id': em['id_evaluation'],
                    'note': note_val
                })

        if errors and not notes_to_save:
            return JsonResponse({'success': False, 'error': 'Erreurs: ' + '; '.join(errors[:5])}, status=400)

        # Save notes
        conn = _get_spoke_connection()
        try:
            saved = 0
            with conn.cursor() as cur:
                for nd in notes_to_save:
                    cur.execute("""
                        SELECT id_note FROM eleve_note
                        WHERE id_eleve_id = %s AND id_evaluation_id = %s LIMIT 1
                    """, [nd['eleve_id'], nd['evaluation_id']])
                    existing = cur.fetchone()

                    if existing:
                        cur.execute("UPDATE eleve_note SET note = %s WHERE id_note = %s",
                                    [nd['note'], existing['id_note']])
                    else:
                        cur.execute("""
                            SELECT id_annee_id, idCampus_id, classe_id, groupe, section_id,
                                   id_cours_classe_id, id_classe_id, id_cycle_id,
                                   id_repartition_instance, id_session_id, id_type_eval
                            FROM evaluation WHERE id_evaluation = %s
                        """, [nd['evaluation_id']])
                        ev_ctx = cur.fetchone()
                        if not ev_ctx:
                            continue

                        # Map evaluation type to note type: default TJ=1
                        id_type_note = 1  # TJ by default

                        cur.execute("""
                            INSERT INTO eleve_note
                                (id_eleve_id, id_evaluation_id, note, id_annee_id, idCampus_id,
                                 id_classe_id, classe_id, groupe, section_id,
                                 id_cycle_id, id_repartition_instance,
                                 id_cours_id, id_type_note_id, id_session_id,
                                 id_etablissement, id_pays, date_saisie)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        """, [nd['eleve_id'], nd['evaluation_id'], nd['note'],
                              ev_ctx['id_annee_id'], ev_ctx['idCampus_id'],
                              ev_ctx['id_classe_id'] or 0, ev_ctx['classe_id'], ev_ctx['groupe'], ev_ctx['section_id'],
                              ev_ctx['id_cycle_id'] or 0,
                              ev_ctx['id_repartition_instance'] or 0,
                              ev_ctx['id_cours_classe_id'], id_type_note,
                              ev_ctx['id_session_id'] or 0, etab_id, etab.pays_id])
                    saved += 1

            conn.commit()
            result = {'success': True, 'saved': saved, 'message': f'{saved} note(s) importée(s).'}
            if errors:
                result['warnings'] = errors[:5]
            return JsonResponse(result)
        finally:
            conn.close()

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# IMPORT EXAM NOTES FROM EXCEL
# ============================================================

@csrf_exempt
@require_http_methods(["POST"])
def import_exam_notes_excel(request):
    """Importe les notes d'examen depuis un fichier Excel (template d'examen avec cours_*)."""
    try:
        import openpyxl
        from io import BytesIO

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({'success': False, 'error': 'Aucun fichier.'}, status=400)

        file_bytes = uploaded_file.read()
        wb_data = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        wb_meta = openpyxl.load_workbook(BytesIO(file_bytes))

        if '_meta' not in wb_meta.sheetnames:
            return JsonResponse({'success': False, 'error': 'Fichier invalide: pas de métadonnées.'}, status=400)

        meta = wb_meta['_meta']
        classe_id = int(meta.cell(row=1, column=2).value)
        repartition_id = int(meta.cell(row=2, column=2).value)

        # Verify it's an exam template
        file_type = meta.cell(row=4, column=2).value
        if file_type != 'EXAM':
            return JsonResponse({'success': False, 'error': "Ce fichier n'est pas un modèle d'examen."}, status=400)

        # Read cours IDs + maxima from meta (cours_0, cours_1, ...)
        cours_map = []
        row = 5
        while True:
            key = meta.cell(row=row, column=1).value
            if not key or not str(key).startswith('cours_'):
                break
            cours_map.append({
                'id_cours_annee': int(meta.cell(row=row, column=2).value),
                'max': float(meta.cell(row=row, column=3).value or 0)
            })
            row += 1

        if not cours_map:
            return JsonResponse({'success': False, 'error': 'Aucun cours dans le modèle.'}, status=400)

        # Get config
        config = _get_or_create_repartition_config(repartition_id, etab_id)
        if not config:
            return JsonResponse({'success': False, 'error': 'Configuration de répartition introuvable.'}, status=404)

        from django.db import connections
        rep_type_id = config.repartition.type_id

        # Find EX note_type via raw SQL
        conn_hub = connections['countryStructure'].cursor()
        try:
            conn_hub.execute("""
                SELECT rtn.ponderation_max, nt.id_type_note
                FROM repartition_type_notes rtn
                JOIN note_types nt ON nt.id = rtn.note_type_id
                WHERE rtn.repartition_type_id = %s AND nt.sigle = 'EX' AND rtn.is_active = 1
                LIMIT 1
            """, [rep_type_id])
            ex_row = conn_hub.fetchone()
        finally:
            conn_hub.close()

        if not ex_row:
            return JsonResponse({'success': False, 'error': 'Type de note Examen non configuré.'}, status=404)

        ex_nt_id = ex_row[1]

        # Read notes from main sheet
        ws = wb_data.active
        notes_to_save = []
        errors = []

        for row_idx in range(2, ws.max_row + 1):
            eleve_id = ws.cell(row=row_idx, column=1).value
            nom = ws.cell(row=row_idx, column=2).value
            if not eleve_id:
                continue
            for col_idx, cm in enumerate(cours_map):
                cell_val = ws.cell(row=row_idx, column=4 + col_idx).value
                if cell_val is None or str(cell_val).strip() == '':
                    continue
                try:
                    note_val = float(cell_val)
                except (ValueError, TypeError):
                    errors.append(f"Note invalide pour {nom} col {col_idx + 1}: '{cell_val}'")
                    continue
                if note_val < 0:
                    errors.append(f"Note négative pour {nom}: {note_val}")
                    continue
                if note_val > cm['max'] and cm['max'] > 0:
                    errors.append(f"Note {note_val} > max {cm['max']} pour {nom}")
                    continue
                notes_to_save.append({
                    'eleve_id': int(eleve_id),
                    'cours_annee_id': cm['id_cours_annee'],
                    'note': note_val,
                    'maxima': int(cm['max'])
                })

        if errors and not notes_to_save:
            return JsonResponse({'success': False, 'error': 'Erreurs: ' + '; '.join(errors[:5])}, status=400)

        # Save exam notes into note_bulletin
        conn = _get_spoke_connection()
        try:
            saved = 0
            with conn.cursor() as cur:
                for nd in notes_to_save:
                    cur.execute("""
                        INSERT INTO note_bulletin
                            (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                             note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                        VALUES (%s, %s, %s, %s, %s, %s, 'SAISIE_DIRECTE', NOW(), %s, %s)
                        ON DUPLICATE KEY UPDATE
                            note = VALUES(note), maxima = VALUES(maxima),
                            source_type = 'SAISIE_DIRECTE',
                            date_calcul = NOW(), updated_at = NOW()
                    """, [nd['eleve_id'], nd['cours_annee_id'], config.id, ex_nt_id,
                          nd['note'], nd['maxima'], etab_id, etab.pays_id])
                    saved += 1
            conn.commit()
            result = {'success': True, 'saved': saved, 'message': f"{saved} note(s) d'examen importée(s)."}
            if errors:
                result['warnings'] = errors[:5]
            return JsonResponse(result)
        finally:
            conn.close()

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# NOTES BULLETIN — CALCUL & AFFICHAGE
# ============================================================

@csrf_exempt
@require_http_methods(["POST"])
def calculate_period_batch(request):
    """
    Calcule les TJ de TOUTES les périodes enfants d'un parent (semestre)
    pour UN cours OU tous les cours de la classe.
    Body JSON:
      - classe_id: EAC id
      - parent_config_id: repartition_config id du parent (semestre/trimestre)
      - cours_id (optional): cours_annee id — if provided, only this course
      - periods_config (optional): [{config_id, evaluations: [{id, included, weight}]}]
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        data = json.loads(request.body)
        classe_id = data.get('classe_id')
        parent_config_id = data.get('parent_config_id')

        if not classe_id or not parent_config_id:
            return JsonResponse({'success': False, 'error': 'classe_id et parent_config_id requis.'}, status=400)

        from django.db import connections

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Get business keys
                cur.execute("""
                    SELECT ea.annee_id AS id_annee,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()
                if not ctx:
                    return JsonResponse({'success': False, 'error': 'Classe non trouvée.'}, status=404)

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # Get enrolled students
                cur.execute("""
                    SELECT DISTINCT e.id_eleve
                    FROM eleve_inscription ei JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s AND ei.status = 1
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleve_ids = [r['id_eleve'] for r in cur.fetchall()]
                if not eleve_ids:
                    return JsonResponse({'success': False, 'error': 'Aucun élève.'}, status=404)

                # Get parent config info (type, repartition)
                conn_hub = connections['countryStructure'].cursor()
                try:
                    conn_hub.execute("""
                        SELECT rc.repartition_id, rc.etablissement_annee_id,
                               r.type_id AS parent_type_id
                        FROM repartition_configs_etab_annee rc
                        JOIN repartition_instances r ON r.id = rc.repartition_id
                        WHERE rc.id = %s
                    """, [parent_config_id])
                    parent_info = conn_hub.fetchone()
                    if not parent_info:
                        return JsonResponse({'success': False, 'error': 'Config parent introuvable.'}, status=404)
                    parent_type_id = parent_info[2]
                    etab_annee_id = parent_info[1]

                    # Get child type
                    conn_hub.execute("""
                        SELECT rh.type_enfant_id
                        FROM repartition_hierarchies rh
                        WHERE rh.type_parent_id = %s AND rh.is_active = 1 LIMIT 1
                    """, [parent_type_id])
                    child_type_row = conn_hub.fetchone()
                    if not child_type_row:
                        return JsonResponse({'success': False, 'error': 'Hiérarchie non trouvée.'}, status=404)
                    child_type_id = child_type_row[0]

                    # Use frontend-provided config_ids if available (correct parent-child mapping)
                    periods_cfg_raw = data.get('periods_config') or []
                    if periods_cfg_raw:
                        frontend_config_ids = [pc['config_id'] for pc in periods_cfg_raw]
                        ph_fc = ','.join(['%s'] * len(frontend_config_ids))
                        conn_hub.execute(f"""
                            SELECT rc.id AS config_id, rc.repartition_id, r.taux_participation, r.nom
                            FROM repartition_configs_etab_annee rc
                            JOIN repartition_instances r ON r.id = rc.repartition_id
                            WHERE rc.id IN ({ph_fc})
                            ORDER BY r.ordre
                        """, frontend_config_ids)
                    else:
                        # Fallback: discover children, but properly filter by parent ordering
                        conn_hub.execute("""
                            SELECT rc.id AS config_id, rc.repartition_id, r.taux_participation, r.nom
                            FROM repartition_configs_etab_annee rc
                            JOIN repartition_instances r ON r.id = rc.repartition_id
                            WHERE rc.etablissement_annee_id = %s AND r.type_id = %s AND rc.is_open = 1
                            ORDER BY r.ordre
                        """, [etab_annee_id, child_type_id])

                    child_configs = [{'config_id': r[0], 'repartition_id': r[1],
                                      'taux': float(r[2]) if r[2] else 100.0, 'nom': r[3]}
                                     for r in conn_hub.fetchall()]

                    # Fallback filtering: if no periods_config, filter by parent ordering
                    if not periods_cfg_raw and child_configs:
                        # Get all parent configs to determine child-parent mapping
                        conn_hub.execute("""
                            SELECT rc.id AS config_id
                            FROM repartition_configs_etab_annee rc
                            JOIN repartition_instances r ON r.id = rc.repartition_id
                            WHERE rc.etablissement_annee_id = %s AND r.type_id = %s AND rc.is_open = 1
                            ORDER BY r.ordre
                        """, [etab_annee_id, parent_type_id])
                        all_parent_ids = [r[0] for r in conn_hub.fetchall()]
                        # Determine child_count_per_root
                        if all_parent_ids:
                            child_count_per_root = len(child_configs) // len(all_parent_ids)
                            if child_count_per_root > 0:
                                # Find index of our parent in the parent list
                                try:
                                    parent_idx = all_parent_ids.index(parent_config_id)
                                except ValueError:
                                    parent_idx = 0
                                start = parent_idx * child_count_per_root
                                end = start + child_count_per_root
                                child_configs = child_configs[start:end]

                    if not child_configs:
                        return JsonResponse({'success': False, 'error': 'Aucune période enfant trouvée.'}, status=404)

                    total_taux = sum(c['taux'] for c in child_configs) or 100.0

                    # Get TJ note_type for child periods
                    conn_hub.execute("""
                        SELECT nt.id_type_note, rtn.ponderation_max
                        FROM repartition_type_notes rtn
                        JOIN note_types nt ON nt.id = rtn.note_type_id
                        WHERE rtn.repartition_type_id = %s AND nt.sigle = 'TJ' AND rtn.is_active = 1
                        LIMIT 1
                    """, [child_type_id])
                    child_tj_row = conn_hub.fetchone()
                    if not child_tj_row:
                        return JsonResponse({'success': False, 'error': 'Pas de TJ configuré pour enfants.'}, status=404)
                    child_tj_nt_id = child_tj_row[0]
                    child_tj_max_global = child_tj_row[1] or 20

                    # Get parent note types (TJ, EX, TOTAL)
                    conn_hub.execute("""
                        SELECT nt.id_type_note, nt.sigle, rtn.ponderation_max
                        FROM repartition_type_notes rtn
                        JOIN note_types nt ON nt.id = rtn.note_type_id
                        WHERE rtn.repartition_type_id = %s AND rtn.is_active = 1
                        ORDER BY rtn.ordre
                    """, [parent_type_id])
                    parent_nts = {}
                    for prow in conn_hub.fetchall():
                        parent_nts[prow[1]] = {'nt_id': prow[0], 'max': prow[2]}

                finally:
                    conn_hub.close()

                # Get courses — single or all
                single_cours_id = data.get('cours_id')
                # Build lookup: config_id -> {eval_id -> {included, weight}}
                pcfg_map = {}
                for pcc in periods_cfg_raw:
                    ev_map = {}
                    for evc in (pcc.get('evaluations') or []):
                        ev_map[evc['id']] = {'included': evc.get('included', True), 'weight': float(evc.get('weight', 0))}
                    pcfg_map[pcc['config_id']] = ev_map

                if single_cours_id:
                    cur.execute("""
                        SELECT cann.id_cours_annee, cann.maxima_tj, cann.maxima_exam
                        FROM countryStructure.cours_annee cann
                        WHERE cann.id_cours_annee = %s
                    """, [single_cours_id])
                else:
                    cur.execute("""
                        SELECT cann.id_cours_annee, cann.maxima_tj, cann.maxima_exam
                        FROM countryStructure.cours_annee cann
                        JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                        JOIN countryStructure.etablissements_annees ea3 ON ea3.annee_id = cann.annee_id
                        JOIN countryStructure.etablissements_annees_classes eac3 ON eac3.etablissement_annee_id = ea3.id
                        WHERE eac3.id = %s AND ca.classe_id = eac3.classe_id
                        GROUP BY cann.id_cours_annee, cann.maxima_tj, cann.maxima_exam
                    """, [classe_id])
                all_cours = cur.fetchall()

                total_calculated = 0
                total_errors = 0
                period_details = []

                # === PHASE 1: Calculate TJ for each child period × each course ===
                for child_cfg in child_configs:
                    cfg_id = child_cfg['config_id']
                    this_taux = child_cfg['taux']
                    period_calc = 0

                    for cr in all_cours:
                        cours_id = cr['id_cours_annee']
                        cours_maxima_tj = int(cr['maxima_tj']) if cr['maxima_tj'] else None

                        # Period max = maxima_tj × (this_taux / total_taux)
                        if cours_maxima_tj and total_taux > 0:
                            tj_max = round(cours_maxima_tj * (this_taux / total_taux), 2)
                        else:
                            tj_max = child_tj_max_global

                        # Get evaluations for this config + cours
                        cur.execute("""
                            SELECT ev.id_evaluation, ev.ponderer_eval
                            FROM evaluation ev
                            JOIN evaluation_repartition er ON er.id_evaluation = ev.id_evaluation
                            WHERE er.id_repartition_config = %s
                              AND ev.id_cours_classe_id = %s
                              AND ev.id_etablissement = %s
                            ORDER BY ev.id_evaluation
                        """, [cfg_id, cours_id, etab_id])
                        db_evals = cur.fetchall()
                        if not db_evals:
                            continue

                        # Use user-provided weights if available
                        ev_cfg = pcfg_map.get(cfg_id, {})
                        if ev_cfg:
                            included_evals = []
                            for ev in db_evals:
                                ec = ev_cfg.get(ev['id_evaluation'])
                                if ec and ec['included']:
                                    included_evals.append({
                                        'id': ev['id_evaluation'],
                                        'max': ev['ponderer_eval'] or 0,
                                        'weight': ec.get('weight', 0)
                                    })
                        else:
                            # Auto equal weight
                            n_evals = len(db_evals)
                            included_evals = [{
                                'id': ev['id_evaluation'],
                                'max': ev['ponderer_eval'] or 0,
                                'weight': round(100.0 / n_evals, 2) if n_evals > 0 else 0
                            } for ev in db_evals]

                        eval_ids = [e['id'] for e in included_evals]
                        placeholders = ','.join(['%s'] * len(eval_ids))

                        for eleve_id in eleve_ids:
                            cur.execute(f"""
                                SELECT en.id_evaluation_id, en.note
                                FROM eleve_note en
                                WHERE en.id_eleve_id = %s AND en.id_evaluation_id IN ({placeholders})
                            """, [eleve_id] + eval_ids)
                            raw_notes = {r['id_evaluation_id']: float(r['note']) if r['note'] is not None else None
                                         for r in cur.fetchall()}

                            weighted_sum = 0
                            weight_used = 0
                            for ev in included_evals:
                                note = raw_notes.get(ev['id'])
                                if note is not None and ev['max'] > 0:
                                    normalized = note / ev['max']
                                    weighted_sum += normalized * ev['weight']
                                    weight_used += ev['weight']

                            if weight_used > 0:
                                scaled = round((weighted_sum / weight_used) * tj_max, 2)
                            else:
                                scaled = None

                            if scaled is not None:
                                # Build calc_details JSON for audit trail
                                details_json = json.dumps({
                                    'taux_participation': this_taux,
                                    'total_taux': total_taux,
                                    'tj_max': tj_max,
                                    'evaluations': [
                                        {
                                            'id': ev['id'],
                                            'max': ev['max'],
                                            'weight': ev['weight'],
                                            'note': raw_notes.get(ev['id'])
                                        } for ev in included_evals
                                    ]
                                })
                                cur.execute("""
                                    INSERT INTO note_bulletin
                                        (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                         note, maxima, source_type, calc_details, date_calcul, id_etablissement, id_pays)
                                    VALUES (%s, %s, %s, %s, %s, %s, 'EVALUATIONS', %s, NOW(), %s, %s)
                                    ON DUPLICATE KEY UPDATE
                                        note = VALUES(note), maxima = VALUES(maxima),
                                        calc_details = VALUES(calc_details),
                                        date_calcul = NOW(), updated_at = NOW()
                                """, [eleve_id, cours_id, cfg_id, child_tj_nt_id, scaled, tj_max, details_json, etab_id, etab.pays_id])
                                period_calc += 1

                    total_calculated += period_calc
                    period_details.append({'nom': child_cfg['nom'], 'calculated': period_calc})

                # === PHASE 2: Cascade to parent (HERITAGE TJ + TOTAL) ===
                parent_tj = parent_nts.get('TJ')
                parent_tot = parent_nts.get('TOTAL')
                child_config_ids = [c['config_id'] for c in child_configs]
                ch_ph = ','.join(['%s'] * len(child_config_ids))

                if parent_tj:
                    for cr in all_cours:
                        cours_id = cr['id_cours_annee']
                        cours_maxima_tj = int(cr['maxima_tj']) if cr['maxima_tj'] else None
                        p_max = cours_maxima_tj if cours_maxima_tj else parent_tj['max']

                        for eleve_id in eleve_ids:
                            cur.execute(f"""
                                SELECT COALESCE(SUM(nb.note), 0) AS total,
                                       COALESCE(SUM(nb.maxima), 0) AS total_max
                                FROM note_bulletin nb
                                WHERE nb.id_eleve_id = %s AND nb.id_cours_annee = %s
                                  AND nb.id_note_type = %s
                                  AND nb.id_repartition_config IN ({ch_ph})
                            """, [eleve_id, cours_id, child_tj_nt_id] + child_config_ids)
                            row = cur.fetchone()
                            raw_total = float(row['total']) if row and row['total'] else None
                            raw_max = float(row['total_max']) if row and row['total_max'] else None

                            if raw_total is not None and raw_max and raw_max > 0:
                                p_note = round((raw_total / raw_max) * p_max, 2)
                                cur.execute("""
                                    INSERT INTO note_bulletin
                                        (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                         note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                    VALUES (%s, %s, %s, %s, %s, %s, 'HERITAGE', NOW(), %s, %s)
                                    ON DUPLICATE KEY UPDATE
                                        note = VALUES(note), maxima = VALUES(maxima),
                                        date_calcul = NOW(), updated_at = NOW()
                                """, [eleve_id, cours_id, parent_config_id, parent_tj['nt_id'], p_note, p_max, etab_id, etab.pays_id])

                            # TOTAL = TJ + EX
                            if parent_tot:
                                cur.execute("""
                                    SELECT COALESCE(SUM(nb.note), 0) AS total,
                                           COALESCE(SUM(nb.maxima), 0) AS total_max
                                    FROM note_bulletin nb
                                    WHERE nb.id_eleve_id = %s AND nb.id_cours_annee = %s
                                      AND nb.id_repartition_config = %s
                                      AND nb.id_note_type != %s
                                """, [eleve_id, cours_id, parent_config_id, parent_tot['nt_id']])
                                trow = cur.fetchone()
                                t_note = round(float(trow['total']), 2) if trow and trow['total'] else None
                                t_max = int(trow['total_max']) if trow and trow['total_max'] else parent_tot['max']
                                if t_note is not None:
                                    cur.execute("""
                                        INSERT INTO note_bulletin
                                            (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                             note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                        VALUES (%s, %s, %s, %s, %s, %s, 'FORMULE', NOW(), %s, %s)
                                        ON DUPLICATE KEY UPDATE
                                            note = VALUES(note), maxima = VALUES(maxima),
                                            date_calcul = NOW(), updated_at = NOW()
                                    """, [eleve_id, cours_id, parent_config_id, parent_tot['nt_id'], t_note, t_max, etab_id, etab.pays_id])

                conn.commit()
                return JsonResponse({
                    'success': True,
                    'calculated': total_calculated,
                    'periods': period_details,
                })
        finally:
            conn.close()
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def calculate_period_notes(request):
    """
    Calcule la note TJ d'une SEULE période pour UN cours.
    Body JSON:
      - classe_id: EAC id
      - config_id: repartition_config id (période)
      - cours_id: id_cours_annee
      - evaluations: [{id, included, weight}] — weight in % (sum=100 for included)
    Si evaluations absent: auto-equal pour toutes.
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        data = json.loads(request.body)
        classe_id = data.get('classe_id')
        config_id = data.get('config_id')
        cours_id = data.get('cours_id')
        eval_config = data.get('evaluations')  # [{id, included, weight}] or None

        if not classe_id or not config_id or not cours_id:
            return JsonResponse({'success': False, 'error': 'classe_id, config_id et cours_id requis.'}, status=400)

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Get business keys
                cur.execute("""
                    SELECT ea.annee_id AS id_annee,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()
                if not ctx:
                    return JsonResponse({'success': False, 'error': 'Classe non trouvée.'}, status=404)

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # Get enrolled students
                cur.execute("""
                    SELECT DISTINCT e.id_eleve
                    FROM eleve_inscription ei JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s AND ei.status = 1
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleve_ids = [r['id_eleve'] for r in cur.fetchall()]
                if not eleve_ids:
                    return JsonResponse({'success': False, 'error': 'Aucun élève.'}, status=404)

                # Get evaluations for this config + cours
                cur.execute("""
                    SELECT ev.id_evaluation, ev.title, ev.ponderer_eval
                    FROM evaluation ev
                    JOIN evaluation_repartition er ON er.id_evaluation = ev.id_evaluation
                    WHERE er.id_repartition_config = %s
                      AND ev.id_cours_classe_id = %s
                      AND ev.id_etablissement = %s
                    ORDER BY ev.id_evaluation
                """, [config_id, cours_id, etab_id])
                db_evals = cur.fetchall()
                if not db_evals:
                    return JsonResponse({'success': False, 'error': 'Aucune évaluation trouvée.'}, status=404)

                # Build weight map
                if eval_config:
                    # User provided config
                    cfg_map = {int(e['id']): e for e in eval_config}
                    included_evals = []
                    for ev in db_evals:
                        ec = cfg_map.get(ev['id_evaluation'])
                        if ec and ec.get('included', True):
                            included_evals.append({
                                'id': ev['id_evaluation'],
                                'max': ev['ponderer_eval'] or 0,
                                'weight': float(ec.get('weight', 0))
                            })
                else:
                    # Auto: all included, equal weight
                    n = len(db_evals)
                    included_evals = [{
                        'id': ev['id_evaluation'],
                        'max': ev['ponderer_eval'] or 0,
                        'weight': round(100.0 / n, 2) if n > 0 else 0
                    } for ev in db_evals]

                if not included_evals:
                    return JsonResponse({'success': False, 'error': 'Aucune évaluation incluse.'}, status=400)

                # Normalize weights to sum=100
                total_weight = sum(e['weight'] for e in included_evals)
                if total_weight <= 0:
                    total_weight = len(included_evals)
                    for e in included_evals:
                        e['weight'] = 100.0 / total_weight

                # Find the TJ note_type for this period's repartition type
                from django.db import connections
                cur.execute("SELECT repartition_id FROM countryStructure.repartition_configs_etab_annee WHERE id = %s", [config_id])
                cfg_row = cur.fetchone()
                if not cfg_row:
                    return JsonResponse({'success': False, 'error': 'Config introuvable.'}, status=404)

                conn_hub = connections['countryStructure'].cursor()
                try:
                    conn_hub.execute("""
                        SELECT r.type_id FROM repartition_instances r WHERE r.id_instance = %s
                    """, [cfg_row['repartition_id']])
                    rt_row = conn_hub.fetchone()
                    if not rt_row:
                        return JsonResponse({'success': False, 'error': 'Type répartition introuvable.'}, status=404)
                    rep_type_id = rt_row[0]

                    conn_hub.execute("""
                        SELECT rtn.ponderation_max, nt.id_type_note, nt.sigle
                        FROM repartition_type_notes rtn
                        JOIN note_types nt ON nt.id = rtn.note_type_id
                        WHERE rtn.repartition_type_id = %s AND nt.sigle = 'TJ' AND rtn.is_active = 1
                        LIMIT 1
                    """, [rep_type_id])
                    tj_row = conn_hub.fetchone()

                    # Also find the parent repartition type for auto-cascade
                    conn_hub.execute("""
                        SELECT rh.type_parent_id FROM repartition_hierarchies rh
                        WHERE rh.type_enfant_id = %s LIMIT 1
                    """, [rep_type_id])
                    parent_type_row = conn_hub.fetchone()
                    parent_type_id = parent_type_row[0] if parent_type_row else None

                    # Get parent note types (TJ, TOTAL) for auto-cascade
                    parent_tj_info = None
                    parent_tot_info = None
                    if parent_type_id:
                        conn_hub.execute("""
                            SELECT rtn.ponderation_max, nt.id_type_note, nt.sigle
                            FROM repartition_type_notes rtn
                            JOIN note_types nt ON nt.id = rtn.note_type_id
                            WHERE rtn.repartition_type_id = %s AND rtn.is_active = 1
                            ORDER BY rtn.ordre
                        """, [parent_type_id])
                        for pr in conn_hub.fetchall():
                            if pr[2] == 'TJ':
                                parent_tj_info = {'max': pr[0], 'nt_id': pr[1]}
                            elif pr[2] == 'TOTAL':
                                parent_tot_info = {'max': pr[0], 'nt_id': pr[1]}
                finally:
                    conn_hub.close()

                if not tj_row:
                    return JsonResponse({'success': False, 'error': "Pas de note TJ configurée pour cette période."}, status=404)

                tj_max_global = tj_row[0] or 20
                tj_nt_id = tj_row[1]

                # === Fetch course-level maxima_tj ===
                cur.execute("""
                    SELECT maxima_tj
                    FROM countryStructure.cours_annee
                    WHERE id_cours_annee = %s LIMIT 1
                """, [cours_id])
                cours_row = cur.fetchone()
                cours_maxima_tj = int(cours_row['maxima_tj']) if cours_row and cours_row['maxima_tj'] else None

                # === Fetch taux_participation for this period + sum of all sibling periods ===
                # Get this period's taux_participation
                cur.execute("""
                    SELECT r.taux_participation
                    FROM countryStructure.repartition_instances r
                    WHERE r.id_instance = %s
                """, [cfg_row['repartition_id']])
                tp_row = cur.fetchone()
                this_taux = float(tp_row['taux_participation']) if tp_row and tp_row['taux_participation'] else 100.0

                # Get sum of taux for all sibling periods (same type, same parent)
                total_taux = this_taux  # fallback
                if parent_type_id:
                    cur.execute("""
                        SELECT COALESCE(SUM(r.taux_participation), 0) AS total_taux
                        FROM countryStructure.repartition_configs_etab_annee rc
                        JOIN countryStructure.repartition_instances r ON r.id = rc.repartition_id
                        WHERE rc.etablissement_annee_id = (
                            SELECT etablissement_annee_id FROM countryStructure.repartition_configs_etab_annee WHERE id = %s
                        ) AND r.type_id = %s AND rc.is_open = 1
                    """, [config_id, rep_type_id])
                    sum_row = cur.fetchone()
                    total_taux = float(sum_row['total_taux']) if sum_row and sum_row['total_taux'] else this_taux

                # Determine effective TJ max for this period:
                # Period max = maxima_tj × (this_taux / total_taux_of_all_siblings)
                if cours_maxima_tj and total_taux > 0:
                    tj_max = round(cours_maxima_tj * (this_taux / total_taux), 2)
                else:
                    tj_max = tj_max_global

                # Calculate for each student
                eval_ids = [e['id'] for e in included_evals]
                placeholders = ','.join(['%s'] * len(eval_ids))
                calculated = 0

                for eleve_id in eleve_ids:
                    # Get raw notes
                    cur.execute(f"""
                        SELECT en.id_evaluation_id, en.note
                        FROM eleve_note en
                        WHERE en.id_eleve_id = %s AND en.id_evaluation_id IN ({placeholders})
                    """, [eleve_id] + eval_ids)
                    raw_notes = {r['id_evaluation_id']: float(r['note']) if r['note'] is not None else None for r in cur.fetchall()}

                    # Weighted calculation
                    weighted_sum = 0
                    weight_used = 0
                    for ev in included_evals:
                        note = raw_notes.get(ev['id'])
                        if note is not None and ev['max'] > 0:
                            # Note normalized to [0,1] then weighted
                            normalized = note / ev['max']
                            weighted_sum += normalized * ev['weight']
                            weight_used += ev['weight']

                    if weight_used > 0:
                        # Scale to TJ max: (weighted_sum / weight_used) * tj_max
                        scaled = round((weighted_sum / weight_used) * tj_max, 2)
                    else:
                        scaled = None

                    if scaled is not None:
                        cur.execute("""
                            INSERT INTO note_bulletin
                                (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                 note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                            VALUES (%s, %s, %s, %s, %s, %s, 'EVALUATIONS', NOW(), %s, %s)
                            ON DUPLICATE KEY UPDATE
                                note = VALUES(note), maxima = VALUES(maxima),
                                date_calcul = NOW(), updated_at = NOW()
                        """, [eleve_id, cours_id, config_id, tj_nt_id, scaled, tj_max, etab_id, etab.pays_id])
                        calculated += 1

                # === AUTO-CASCADE: update parent TJ (heritage) and TOTAL ===
                if parent_tj_info and parent_type_id:
                    # Find parent config(s) for this cours
                    cur.execute("""
                        SELECT rc.id AS parent_config_id
                        FROM countryStructure.repartition_configs_etab_annee rc
                        JOIN countryStructure.repartition_instances r ON r.id = rc.repartition_id
                        WHERE rc.etablissement_annee_id = (
                            SELECT etablissement_annee_id FROM countryStructure.repartition_configs_etab_annee WHERE id = %s
                        ) AND r.type_id = %s AND rc.is_open = 1
                    """, [config_id, parent_type_id])
                    parent_configs = [r['parent_config_id'] for r in cur.fetchall()]

                    # Find all child configs (siblings of current config)
                    child_configs = []
                    for pc_id in parent_configs:
                        cur.execute("""
                            SELECT rc.id FROM countryStructure.repartition_configs_etab_annee rc
                            JOIN countryStructure.repartition_instances r ON r.id = rc.repartition_id
                            WHERE rc.etablissement_annee_id = (
                                SELECT etablissement_annee_id FROM countryStructure.repartition_configs_etab_annee WHERE id = %s
                            ) AND r.type_id = %s AND rc.is_open = 1
                        """, [config_id, rep_type_id])
                        child_configs = [r['id'] for r in cur.fetchall()]

                    if parent_configs and child_configs:
                        ch_ph = ','.join(['%s'] * len(child_configs))
                        # Parent TJ maxima = course's maxima_tj (authoritative source)
                        p_max = cours_maxima_tj if cours_maxima_tj else parent_tj_info['max']

                        for pc_id in parent_configs:
                            for eleve_id in eleve_ids:
                                # Sum child TJs for this cours
                                cur.execute(f"""
                                    SELECT COALESCE(SUM(nb.note), 0) AS total,
                                           COALESCE(SUM(nb.maxima), 0) AS total_max
                                    FROM note_bulletin nb
                                    WHERE nb.id_eleve_id = %s AND nb.id_cours_annee = %s
                                      AND nb.id_note_type = %s
                                      AND nb.id_repartition_config IN ({ch_ph})
                                """, [eleve_id, cours_id, tj_nt_id] + child_configs)
                                row = cur.fetchone()
                                raw_total = float(row['total']) if row and row['total'] else None
                                raw_max = float(row['total_max']) if row and row['total_max'] else None

                                # Scale note to course maxima_tj
                                if raw_total is not None and raw_max and raw_max > 0:
                                    p_note = round((raw_total / raw_max) * p_max, 2)
                                else:
                                    p_note = None

                                if p_note is not None:
                                    cur.execute("""
                                        INSERT INTO note_bulletin
                                            (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                             note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                        VALUES (%s, %s, %s, %s, %s, %s, 'HERITAGE', NOW(), %s, %s)
                                        ON DUPLICATE KEY UPDATE
                                            note = VALUES(note), maxima = VALUES(maxima),
                                            date_calcul = NOW(), updated_at = NOW()
                                    """, [eleve_id, cours_id, pc_id, parent_tj_info['nt_id'], p_note, p_max, etab_id, etab.pays_id])

                                # Also compute TOTAL = TJ + EX at parent level
                                if parent_tot_info:
                                    cur.execute("""
                                        SELECT COALESCE(SUM(nb.note), 0) AS total,
                                               COALESCE(SUM(nb.maxima), 0) AS total_max
                                        FROM note_bulletin nb
                                        WHERE nb.id_eleve_id = %s AND nb.id_cours_annee = %s
                                          AND nb.id_repartition_config = %s
                                          AND nb.id_note_type != %s
                                    """, [eleve_id, cours_id, pc_id, parent_tot_info['nt_id']])
                                    trow = cur.fetchone()
                                    t_note = round(float(trow['total']), 2) if trow and trow['total'] else None
                                    t_max = int(trow['total_max']) if trow and trow['total_max'] else parent_tot_info['max']
                                    if t_note is not None:
                                        cur.execute("""
                                            INSERT INTO note_bulletin
                                                (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                                 note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                            VALUES (%s, %s, %s, %s, %s, %s, 'FORMULE', NOW(), %s, %s)
                                            ON DUPLICATE KEY UPDATE
                                                note = VALUES(note), maxima = VALUES(maxima),
                                                date_calcul = NOW(), updated_at = NOW()
                                        """, [eleve_id, cours_id, pc_id, parent_tot_info['nt_id'], t_note, t_max, etab_id, etab.pays_id])

                conn.commit()
                return JsonResponse({
                    'success': True,
                    'calculated': calculated,
                    'evals_used': len(included_evals),
                    'tj_max': tj_max
                })
        finally:
            conn.close()
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def calculate_notes_bulletin(request):
    """
    Calcule les notes du bulletin pour une classe/répartition.
    Gère les 3 sources:
      - EVALUATIONS: agrège les notes d'évaluations assignées
      - HERITAGE: somme/moyenne depuis les sous-répartitions
      - FORMULE: somme d'autres note_types au même niveau
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        data = json.loads(request.body)
        classe_id = data.get('classe_id') or data.get('id_classe_id')
        repartition_id = data.get('repartition_id')

        if not classe_id or not repartition_id:
            return JsonResponse({'success': False, 'error': 'classe_id et repartition_id requis.'}, status=400)

        # Get the repartition config (auto-creates if synched mode)
        config = _get_or_create_repartition_config(repartition_id, etab_id)

        if not config:
            return JsonResponse({'success': False, 'error': 'Configuration de répartition introuvable.'}, status=404)

        rep_type = config.repartition.type
        rep_type_id = config.repartition.type_id

        # Get expected note types for this repartition type via raw SQL
        from django.db import connections
        conn_hub = connections['countryStructure'].cursor()
        try:
            conn_hub.execute("""
                SELECT rtn.id, rtn.ponderation_max, rtn.source_type, rtn.mode_calcul, rtn.ordre,
                       nt.id_type_note, nt.sigle, nt.nom
                FROM repartition_type_notes rtn
                JOIN note_types nt ON nt.id = rtn.note_type_id
                WHERE rtn.repartition_type_id = %s AND rtn.is_active = 1
                ORDER BY rtn.ordre
            """, [rep_type_id])
            columns = [col[0] for col in conn_hub.description]
            expected_notes_raw = [dict(zip(columns, row)) for row in conn_hub.fetchall()]
        finally:
            conn_hub.close()

        if not expected_notes_raw:
            return JsonResponse({'success': False, 'error': 'Aucune note attendue configurée pour ce type de répartition.'}, status=404)

        # Build lightweight wrapper objects for compatibility with the rest of the function
        class _NoteTypeProxy:
            def __init__(self, row):
                self.id_type_note = row['id_type_note']
                self.sigle = row['sigle']
                self.nom = row['nom']
        class _ExpectedNote:
            def __init__(self, row):
                self.note_type = _NoteTypeProxy(row)
                self.ponderation_max = row['ponderation_max']
                self.source_type = row['source_type']
                self.mode_calcul = row.get('mode_calcul', 'SOMME')
                self.ordre = row['ordre']
        expected_notes = [_ExpectedNote(r) for r in expected_notes_raw]

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Get annee + business keys
                cur.execute("""
                    SELECT ea.annee_id AS id_annee,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()
                if not ctx:
                    return JsonResponse({'success': False, 'error': 'Classe non trouvée.'}, status=404)

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # Get enrolled students for this class (business keys)
                cur.execute("""
                    SELECT DISTINCT e.id_eleve
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                      AND ei.status = 1
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleve_ids = [r['id_eleve'] for r in cur.fetchall()]

                if not eleve_ids:
                    return JsonResponse({'success': False, 'error': 'Aucun élève inscrit.'}, status=404)

                # Get cours for this class (with maxima_exam + maxima_tj) — filtered by classe
                cur.execute("""
                    SELECT MIN(cann.id_cours_annee) AS id_cours_annee,
                           MAX(cann.maxima_exam) AS maxima_exam,
                           MAX(cann.maxima_tj) AS maxima_tj
                    FROM countryStructure.cours_annee cann
                    JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    JOIN countryStructure.etablissements_annees ea ON ea.annee_id = cann.annee_id
                    JOIN countryStructure.etablissements_annees_classes eac ON eac.etablissement_annee_id = ea.id
                    WHERE eac.id = %s AND ca.classe_id = eac.classe_id
                    GROUP BY cann.cours_id
                """, [classe_id])
                cours_rows = cur.fetchall()
                # Dict: cours_id → {maxima_exam, maxima_tj}
                cours_maximas = {r['id_cours_annee']: r['maxima_exam'] for r in cours_rows}
                cours_maximas_tj = {r['id_cours_annee']: r['maxima_tj'] for r in cours_rows}
                cours_ids = list(cours_maximas.keys())

                calculated = 0

                for en in expected_notes:
                    nt_id = en.note_type.id_type_note
                    default_max = en.ponderation_max or 20

                    if en.source_type == 'EVALUATIONS':
                        # Calculate from evaluation notes assigned to this repartition
                        for cours_id in cours_ids:
                            # Use the course's own maxima_exam, fallback to default
                            cours_maxima = cours_maximas.get(cours_id) or default_max

                            # Get evaluations assigned to this repartition for this cours
                            cur.execute("""
                                SELECT ev.id_evaluation, ev.ponderer_eval
                                FROM evaluation ev
                                JOIN evaluation_repartition er ON er.id_evaluation = ev.id_evaluation
                                WHERE er.id_repartition_config = %s
                                  AND ev.id_cours_classe_id = %s
                                  AND ev.classe_id = %s AND ev.groupe <=> %s AND ev.section_id <=> %s
                                  AND ev.id_etablissement = %s
                            """, [config.id, cours_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section'], etab_id])
                            evals = cur.fetchall()

                            if not evals:
                                continue

                            eval_ids = [e['id_evaluation'] for e in evals]
                            total_max = sum(e['ponderer_eval'] or 0 for e in evals)

                            if total_max == 0:
                                continue

                            # Get notes for each student
                            placeholders = ','.join(['%s'] * len(eval_ids))
                            for eleve_id in eleve_ids:
                                # PROTECTION: skip if a SAISIE_DIRECTE note already exists
                                # (user manually entered exam notes — don't overwrite)
                                cur.execute("""
                                    SELECT source_type FROM note_bulletin
                                    WHERE id_eleve_id = %s AND id_cours_annee = %s
                                      AND id_repartition_config = %s AND id_note_type = %s
                                      AND source_type = 'SAISIE_DIRECTE'
                                """, [eleve_id, cours_id, config.id, nt_id])
                                if cur.fetchone():
                                    continue  # Manually entered — preserve it

                                cur.execute(f"""
                                    SELECT COALESCE(SUM(en.note), 0) AS total_note
                                    FROM eleve_note en
                                    WHERE en.id_eleve_id = %s
                                      AND en.id_evaluation_id IN ({placeholders})
                                """, [eleve_id] + eval_ids)
                                row = cur.fetchone()
                                raw_total = float(row['total_note']) if row else 0

                                # Scale: (raw_total / total_max_evals) * cours_maxima
                                scaled_note = round((raw_total / total_max) * cours_maxima, 2) if total_max > 0 else 0

                                # Upsert into note_bulletin
                                cur.execute("""
                                    INSERT INTO note_bulletin
                                        (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                         note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                    VALUES (%s, %s, %s, %s, %s, %s, 'EVALUATIONS', NOW(), %s, %s)
                                    ON DUPLICATE KEY UPDATE
                                        note = VALUES(note), maxima = VALUES(maxima),
                                        date_calcul = NOW(), updated_at = NOW()
                                """, [eleve_id, cours_id, config.id, nt_id, scaled_note, cours_maxima, etab_id, etab.pays_id])
                                calculated += 1

                    elif en.source_type == 'HERITAGE':
                        # Sum/Average from child repartitions
                        # Find child configs via hierarchy
                        # Find child repartition type via hierarchy (raw SQL)
                        conn_hub2 = connections['countryStructure'].cursor()
                        try:
                            conn_hub2.execute("""
                                SELECT type_enfant_id FROM repartition_hierarchies
                                WHERE type_parent_id = %s LIMIT 1
                            """, [rep_type_id])
                            hier_row = conn_hub2.fetchone()
                        finally:
                            conn_hub2.close()

                        if not hier_row:
                            continue

                        child_type_id = hier_row[0]

                        # Find child repartition configs
                        child_configs = list(RepartitionConfigEtabAnnee.objects.filter(
                            etablissement_annee=config.etablissement_annee,
                            repartition__type_id=child_type_id,
                            is_open=True
                        ).values_list('id', flat=True))

                        if not child_configs:
                            continue

                        child_placeholders = ','.join(['%s'] * len(child_configs))

                        for cours_id in cours_ids:
                            # Authoritative TJ max from the course itself
                            c_maxima_tj = cours_maximas_tj.get(cours_id)
                            heritage_max = int(c_maxima_tj) if c_maxima_tj else default_max

                            for eleve_id in eleve_ids:
                                cur.execute(f"""
                                    SELECT COALESCE(SUM(nb.note), 0) AS total,
                                           COALESCE(SUM(nb.maxima), 0) AS total_max
                                    FROM note_bulletin nb
                                    WHERE nb.id_eleve_id = %s AND nb.id_cours_annee = %s
                                      AND nb.id_note_type = %s
                                      AND nb.id_repartition_config IN ({child_placeholders})
                                """, [eleve_id, cours_id, nt_id] + child_configs)
                                row = cur.fetchone()
                                raw_total = float(row['total']) if row and row['total'] else None
                                raw_max = float(row['total_max']) if row and row['total_max'] else None

                                # Scale to course's maxima_tj
                                if raw_total is not None and raw_max and raw_max > 0:
                                    note_val = round((raw_total / raw_max) * heritage_max, 2)
                                else:
                                    note_val = None

                                if note_val is not None:
                                    cur.execute("""
                                        INSERT INTO note_bulletin
                                            (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                             note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                        VALUES (%s, %s, %s, %s, %s, %s, 'HERITAGE', NOW(), %s, %s)
                                        ON DUPLICATE KEY UPDATE
                                            note = VALUES(note), maxima = VALUES(maxima),
                                            date_calcul = NOW(), updated_at = NOW()
                                    """, [eleve_id, cours_id, config.id, nt_id, note_val, heritage_max, etab_id, etab.pays_id])
                                    calculated += 1

                    elif en.source_type == 'FORMULE':
                        # Sum other note_types at same repartition level
                        other_nt_ids = [
                            e.note_type.id_type_note for e in expected_notes
                            if e.source_type != 'FORMULE' and e.note_type.id_type_note != nt_id
                        ]
                        if not other_nt_ids:
                            continue

                        other_placeholders = ','.join(['%s'] * len(other_nt_ids))

                        for cours_id in cours_ids:
                            for eleve_id in eleve_ids:
                                cur.execute(f"""
                                    SELECT COALESCE(SUM(nb.note), 0) AS total,
                                           COALESCE(SUM(nb.maxima), 0) AS total_max
                                    FROM note_bulletin nb
                                    WHERE nb.id_eleve_id = %s AND nb.id_cours_annee = %s
                                      AND nb.id_repartition_config = %s
                                      AND nb.id_note_type IN ({other_placeholders})
                                """, [eleve_id, cours_id, config.id] + other_nt_ids)
                                row = cur.fetchone()
                                note_val = round(float(row['total']), 2) if row and row['total'] else None
                                total_max_val = int(row['total_max']) if row and row['total_max'] else default_max

                                if note_val is not None:
                                    cur.execute("""
                                        INSERT INTO note_bulletin
                                            (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                             note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                        VALUES (%s, %s, %s, %s, %s, %s, 'FORMULE', NOW(), %s, %s)
                                        ON DUPLICATE KEY UPDATE
                                            note = VALUES(note), maxima = VALUES(maxima),
                                            date_calcul = NOW(), updated_at = NOW()
                                    """, [eleve_id, cours_id, config.id, nt_id, note_val, total_max_val, etab_id, etab.pays_id])
                                    calculated += 1

            conn.commit()
            return JsonResponse({
                'success': True,
                'calculated': calculated,
                'message': f'{calculated} note(s) de bulletin calculée(s).'
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# SYNC ALL NOTES → note_bulletin (GLOBAL)
# ============================================================

@csrf_exempt
@require_http_methods(["POST"])
def sync_all_notes_bulletin(request):
    """
    Synchronise TOUTES les notes d'une classe vers note_bulletin.
    Parcourt toutes les répartitions (périodes enfants d'abord, puis parents).
    Pour chaque période+cours+élève :
      - Calcule TJ pondéré depuis eleve_note (évaluations brutes)
      - Écrit dans note_bulletin
    Pour chaque parent (trimestre/semestre) :
      - Cascade TJ des enfants → TJ parent (HERITAGE)
      - Conserve les EX existants (SAISIE_DIRECTE)
      - Calcule TOTAL = TJ + EX (FORMULE)
    Body JSON: { classe_id: EAC id }
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        data = json.loads(request.body)
        classe_id = data.get('classe_id')
        if not classe_id:
            return JsonResponse({'success': False, 'error': 'classe_id requis.'}, status=400)

        from django.db import connections

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # 1. Get business keys from EAC
                cur.execute("""
                    SELECT ea.annee_id AS id_annee, ea.id AS etab_annee_id,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()
                if not ctx:
                    return JsonResponse({'success': False, 'error': 'Classe non trouvée.'}, status=404)

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # 2. Get enrolled students
                cur.execute("""
                    SELECT DISTINCT e.id_eleve
                    FROM eleve_inscription ei JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s AND ei.status = 1
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleve_ids = [r['id_eleve'] for r in cur.fetchall()]
                if not eleve_ids:
                    return JsonResponse({'success': False, 'error': 'Aucun élève inscrit.'}, status=404)

                # 3. Get all cours_annee for this class
                cur.execute("""
                    SELECT cann.id_cours_annee, cann.maxima_exam, cann.maxima_tj, cann.maxima_periode
                    FROM countryStructure.cours_annee cann
                    JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    WHERE ca.classe_id = %s
                """, [ctx['bk_classe']])
                cours_rows = cur.fetchall()
                cours_ids = [r['id_cours_annee'] for r in cours_rows]
                cours_maximas_tj = {r['id_cours_annee']: r['maxima_tj'] for r in cours_rows}
                cours_maximas_exam = {r['id_cours_annee']: r['maxima_exam'] for r in cours_rows}

                if not cours_ids:
                    return JsonResponse({'success': False, 'error': 'Aucun cours trouvé.'}, status=404)

                # 4. Get ALL repartition configs for this establishment/year
                #    Split into children (has_parent=1, i.e. periods) and parents (has_parent=0, i.e. trimesters/semesters)
                cur.execute("""
                    SELECT rc.id AS config_id, rc.repartition_id, rc.parent_id, rc.has_parent,
                           r.type_id, r.code, r.nom, r.taux_participation
                    FROM countryStructure.repartition_configs_etab_annee rc
                    JOIN countryStructure.repartition_instances r ON r.id = rc.repartition_id
                    WHERE rc.etablissement_annee_id = %s AND rc.is_open = 1
                    ORDER BY rc.has_parent ASC, r.ordre ASC
                """, [ctx['etab_annee_id']])
                all_configs = cur.fetchall()

                child_configs = [c for c in all_configs if c['has_parent']]
                parent_configs = [c for c in all_configs if not c['has_parent']]

                # 5. For each repartition type, get note_types (TJ, EX, TOTAL)
                type_ids = set(c['type_id'] for c in all_configs)
                note_types_by_rep_type = {}
                conn_hub = connections['countryStructure'].cursor()
                try:
                    for tid in type_ids:
                        conn_hub.execute("""
                            SELECT rtn.ponderation_max, rtn.source_type, nt.id_type_note, nt.sigle
                            FROM repartition_type_notes rtn
                            JOIN note_types nt ON nt.id = rtn.note_type_id
                            WHERE rtn.repartition_type_id = %s AND rtn.is_active = 1
                            ORDER BY rtn.ordre
                        """, [tid])
                        cols = [c[0] for c in conn_hub.description]
                        note_types_by_rep_type[tid] = [dict(zip(cols, r)) for r in conn_hub.fetchall()]
                finally:
                    conn_hub.close()

                calculated = 0

                # ============================================
                # PHASE 1: Process CHILDREN (periods → TJ)
                # ============================================
                for cfg in child_configs:
                    config_id = cfg['config_id']
                    rep_type_id = cfg['type_id']
                    nts = note_types_by_rep_type.get(rep_type_id, [])
                    tj_nt = next((n for n in nts if n['sigle'] == 'TJ'), None)
                    if not tj_nt:
                        continue

                    tj_nt_id = tj_nt['id_type_note']
                    taux = float(cfg['taux_participation']) if cfg['taux_participation'] else 100.0

                    # Get total taux for all siblings (same type, same etab_annee)
                    cur.execute("""
                        SELECT COALESCE(SUM(r.taux_participation), 0) AS total_taux
                        FROM countryStructure.repartition_configs_etab_annee rc
                        JOIN countryStructure.repartition_instances r ON r.id = rc.repartition_id
                        WHERE rc.etablissement_annee_id = %s AND r.type_id = %s AND rc.is_open = 1
                    """, [ctx['etab_annee_id'], rep_type_id])
                    sum_row = cur.fetchone()
                    total_taux = float(sum_row['total_taux']) if sum_row and sum_row['total_taux'] else taux

                    for cours_id in cours_ids:
                        c_maxima_tj = cours_maximas_tj.get(cours_id)
                        if c_maxima_tj and total_taux > 0:
                            period_max = round(float(c_maxima_tj) * (taux / total_taux), 2)
                        else:
                            period_max = tj_nt['ponderation_max'] or 20

                        # Get evaluations assigned to this config+cours
                        cur.execute("""
                            SELECT ev.id_evaluation, ev.ponderer_eval
                            FROM evaluation ev
                            JOIN evaluation_repartition er ON er.id_evaluation = ev.id_evaluation
                            WHERE er.id_repartition_config = %s
                              AND ev.id_cours_classe_id = %s
                              AND ev.id_etablissement = %s
                        """, [config_id, cours_id, etab_id])
                        evals = cur.fetchall()

                        if not evals:
                            continue

                        eval_ids = [e['id_evaluation'] for e in evals]
                        total_max_evals = sum(e['ponderer_eval'] or 0 for e in evals)
                        if total_max_evals == 0:
                            continue

                        placeholders = ','.join(['%s'] * len(eval_ids))

                        for eleve_id in eleve_ids:
                            # Skip if SAISIE_DIRECTE already exists
                            cur.execute("""
                                SELECT source_type FROM note_bulletin
                                WHERE id_eleve_id = %s AND id_cours_annee = %s
                                  AND id_repartition_config = %s AND id_note_type = %s
                                  AND source_type = 'SAISIE_DIRECTE'
                            """, [eleve_id, cours_id, config_id, tj_nt_id])
                            if cur.fetchone():
                                continue

                            # Sum raw grades from eleve_note
                            cur.execute(f"""
                                SELECT COALESCE(SUM(en.note), 0) AS total_note
                                FROM eleve_note en
                                WHERE en.id_eleve_id = %s
                                  AND en.id_evaluation_id IN ({placeholders})
                            """, [eleve_id] + eval_ids)
                            row = cur.fetchone()
                            raw_total = float(row['total_note']) if row else 0

                            # Scale to period max
                            scaled = round((raw_total / total_max_evals) * period_max, 2)

                            cur.execute("""
                                INSERT INTO note_bulletin
                                    (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                     note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                VALUES (%s, %s, %s, %s, %s, %s, 'EVALUATIONS', NOW(), %s, %s)
                                ON DUPLICATE KEY UPDATE
                                    note = VALUES(note), maxima = VALUES(maxima),
                                    date_calcul = NOW(), updated_at = NOW()
                            """, [eleve_id, cours_id, config_id, tj_nt_id, scaled, period_max, etab_id, etab.pays_id])
                            calculated += 1

                # ============================================
                # PHASE 2: Process PARENTS (trimestres/semestres → HERITAGE + FORMULE)
                # ============================================
                for cfg in parent_configs:
                    config_id = cfg['config_id']
                    rep_type_id = cfg['type_id']
                    nts = note_types_by_rep_type.get(rep_type_id, [])

                    tj_nt = next((n for n in nts if n['sigle'] == 'TJ'), None)
                    ex_nt = next((n for n in nts if n['sigle'] == 'EX'), None)
                    tot_nt = next((n for n in nts if n['sigle'] in ('TOTAL', 'TOT')), None)

                    # Find child config IDs for this parent
                    child_cfg_ids = [c['config_id'] for c in child_configs if c['parent_id'] == config_id]
                    if not child_cfg_ids:
                        # Try matching by type hierarchy
                        conn_hub2 = connections['countryStructure'].cursor()
                        try:
                            conn_hub2.execute("""
                                SELECT type_enfant_id FROM repartition_hierarchies
                                WHERE type_parent_id = %s LIMIT 1
                            """, [rep_type_id])
                            hier_row = conn_hub2.fetchone()
                        finally:
                            conn_hub2.close()
                        if hier_row:
                            child_type_id = hier_row[0]
                            child_cfg_ids = [c['config_id'] for c in child_configs if c['type_id'] == child_type_id]

                    # --- TJ HERITAGE: sum child TJs → parent TJ ---
                    if tj_nt and child_cfg_ids:
                        tj_nt_id = tj_nt['id_type_note']
                        # Find the child TJ note_type_id (might be different type)
                        child_type_ids = set(c['type_id'] for c in child_configs if c['config_id'] in child_cfg_ids)
                        child_tj_nt_id = tj_nt_id  # default same
                        for ct_id in child_type_ids:
                            child_nts = note_types_by_rep_type.get(ct_id, [])
                            child_tj = next((n for n in child_nts if n['sigle'] == 'TJ'), None)
                            if child_tj:
                                child_tj_nt_id = child_tj['id_type_note']
                                break

                        ch_ph = ','.join(['%s'] * len(child_cfg_ids))

                        for cours_id in cours_ids:
                            c_maxima_tj = cours_maximas_tj.get(cours_id)
                            heritage_max = int(c_maxima_tj) if c_maxima_tj else (tj_nt['ponderation_max'] or 20)

                            for eleve_id in eleve_ids:
                                cur.execute(f"""
                                    SELECT COALESCE(SUM(nb.note), 0) AS total,
                                           COALESCE(SUM(nb.maxima), 0) AS total_max
                                    FROM note_bulletin nb
                                    WHERE nb.id_eleve_id = %s AND nb.id_cours_annee = %s
                                      AND nb.id_note_type = %s
                                      AND nb.id_repartition_config IN ({ch_ph})
                                """, [eleve_id, cours_id, child_tj_nt_id] + child_cfg_ids)
                                row = cur.fetchone()
                                raw_total = float(row['total']) if row and row['total'] else None
                                raw_max = float(row['total_max']) if row and row['total_max'] else None

                                if raw_total is not None and raw_max and raw_max > 0:
                                    note_val = round((raw_total / raw_max) * heritage_max, 2)
                                else:
                                    note_val = None

                                if note_val is not None:
                                    cur.execute("""
                                        INSERT INTO note_bulletin
                                            (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                             note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                        VALUES (%s, %s, %s, %s, %s, %s, 'HERITAGE', NOW(), %s, %s)
                                        ON DUPLICATE KEY UPDATE
                                            note = VALUES(note), maxima = VALUES(maxima),
                                            date_calcul = NOW(), updated_at = NOW()
                                    """, [eleve_id, cours_id, config_id, tj_nt_id, note_val, heritage_max, etab_id, etab.pays_id])
                                    calculated += 1

                    # --- TOTAL FORMULE: TJ + EX = TOTAL ---
                    if tot_nt:
                        tot_nt_id = tot_nt['id_type_note']
                        other_nt_ids = [n['id_type_note'] for n in nts
                                        if n['sigle'] != 'TOTAL' and n['sigle'] != 'TOT'
                                        and n['id_type_note'] != tot_nt_id]
                        if other_nt_ids:
                            other_ph = ','.join(['%s'] * len(other_nt_ids))
                            for cours_id in cours_ids:
                                for eleve_id in eleve_ids:
                                    cur.execute(f"""
                                        SELECT COALESCE(SUM(nb.note), 0) AS total,
                                               COALESCE(SUM(nb.maxima), 0) AS total_max
                                        FROM note_bulletin nb
                                        WHERE nb.id_eleve_id = %s AND nb.id_cours_annee = %s
                                          AND nb.id_repartition_config = %s
                                          AND nb.id_note_type IN ({other_ph})
                                    """, [eleve_id, cours_id, config_id] + other_nt_ids)
                                    row = cur.fetchone()
                                    note_val = round(float(row['total']), 2) if row and row['total'] else None
                                    total_max_val = int(row['total_max']) if row and row['total_max'] else (tot_nt['ponderation_max'] or 40)

                                    if note_val is not None and note_val > 0:
                                        cur.execute("""
                                            INSERT INTO note_bulletin
                                                (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                                                 note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                                            VALUES (%s, %s, %s, %s, %s, %s, 'FORMULE', NOW(), %s, %s)
                                            ON DUPLICATE KEY UPDATE
                                                note = VALUES(note), maxima = VALUES(maxima),
                                                date_calcul = NOW(), updated_at = NOW()
                                        """, [eleve_id, cours_id, config_id, tot_nt_id, note_val, total_max_val, etab_id, etab.pays_id])
                                        calculated += 1

            conn.commit()
            return JsonResponse({
                'success': True,
                'calculated': calculated,
                'students': len(eleve_ids),
                'courses': len(cours_ids),
                'child_configs': len(child_configs),
                'parent_configs': len(parent_configs),
                'message': f'{calculated} note(s) synchronisées dans note_bulletin.'
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# EXAM NOTES — SAISIE DIRECTE PAR RÉPARTITION PRINCIPALE
# ============================================================

@require_http_methods(["GET"])
def get_exam_grid(request):
    """
    Charge la grille de saisie des notes d'examen pour une répartition principale.
    Params: classe_id (EAC id), repartition_id
    Retourne: élèves, cours avec maxima_exam, notes d'examen existantes.
    """
    try:
        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        repartition_id = request.GET.get('repartition_id')

        if not classe_id or not repartition_id:
            return JsonResponse({'success': False, 'error': 'classe_id et repartition_id requis.'}, status=400)

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        # Get repartition config (auto-creates if synched mode)
        config = _get_or_create_repartition_config(repartition_id, etab_id)
        repartition_name = config.repartition.nom if config else ''
        rep_type = config.repartition.type if config else None

        # Find the EX note_type for this repartition type via raw SQL
        from django.db import connections
        rep_type_id = config.repartition.type_id if config else None
        conn_hub = connections['countryStructure'].cursor()
        try:
            conn_hub.execute("""
                SELECT rtn.id, rtn.ponderation_max, nt.id_type_note
                FROM repartition_type_notes rtn
                JOIN note_types nt ON nt.id = rtn.note_type_id
                WHERE rtn.repartition_type_id = %s AND nt.sigle = 'EX' AND rtn.is_active = 1
                LIMIT 1
            """, [rep_type_id])
            ex_row = conn_hub.fetchone()
        finally:
            conn_hub.close()

        if not ex_row:
            return JsonResponse({'success': False, 'error': 'Aucun type de note "Examen" configuré.'}, status=404)

        ex_note_type_id = ex_row[2]
        default_max = ex_row[1] or 20

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # 1. Get annee + business keys from EAC
                cur.execute("""
                    SELECT ea.annee_id AS id_annee, ea.etablissement_id AS id_etab,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()
                if not ctx:
                    return JsonResponse({'success': False, 'error': 'Classe non trouvée.'}, status=404)

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # 2. Get enrolled students
                cur.execute("""
                    SELECT DISTINCT e.id_eleve, e.nom, e.prenom
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                      AND ei.status = 1
                    ORDER BY e.nom, e.prenom
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleves = cur.fetchall()

                # 3. Get cours for this class with maxima_exam
                cur.execute("""
                    SELECT cann.id_cours_annee, ca.cours AS cours_nom, ca.code_cours,
                           cann.maxima_exam, cann.ordre
                    FROM countryStructure.cours_annee cann
                    JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    JOIN countryStructure.etablissements_annees ea ON ea.annee_id = cann.annee_id
                    JOIN countryStructure.etablissements_annees_classes eac ON eac.etablissement_annee_id = ea.id
                    WHERE eac.id = %s AND ca.classe_id = eac.classe_id
                    GROUP BY cann.cours_id
                    ORDER BY cann.ordre, ca.cours
                """, [classe_id])
                cours_rows = cur.fetchall()

                # 4. Get existing exam notes from note_bulletin
                cours_list = []
                notes = {}
                if config and cours_rows:
                    cours_ids = [r['id_cours_annee'] for r in cours_rows]
                    placeholders = ','.join(['%s'] * len(cours_ids))
                    cur.execute(f"""
                        SELECT nb.id_eleve_id, nb.id_cours_annee, nb.note, nb.maxima
                        FROM note_bulletin nb
                        WHERE nb.id_repartition_config = %s
                          AND nb.id_note_type = %s
                          AND nb.id_cours_annee IN ({placeholders})
                          AND nb.id_etablissement = %s
                    """, [config.id, ex_note_type_id] + cours_ids + [etab_id])
                    for n in cur.fetchall():
                        key = f"{n['id_eleve_id']}_{n['id_cours_annee']}"
                        notes[key] = float(n['note']) if n['note'] is not None else None

                for r in cours_rows:
                    cours_list.append({
                        'id_cours_annee': r['id_cours_annee'],
                        'cours_nom': r['cours_nom'] or f'Cours #{r["id_cours_annee"]}',
                        'code_cours': r['code_cours'] or '',
                        'maxima_exam': r['maxima_exam'] or default_max,
                    })

            return JsonResponse({
                'success': True,
                'repartition_name': repartition_name,
                'note_type_id': ex_note_type_id,
                'eleves': [{'id': e['id_eleve'], 'nom': e['nom'], 'prenom': e['prenom']} for e in eleves],
                'cours': cours_list,
                'notes': notes,
                'context': {
                    'annee_id': ctx['id_annee'],
                    'campus_id': campus_id,
                    'etab_id': etab_id,
                    'config_id': config.id if config else None,
                }
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def save_exam_notes(request):
    """
    Enregistre les notes d'examen directement dans note_bulletin.
    Payload: { classe_id, repartition_id, notes: [{eleve_id, cours_annee_id, note}] }
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        data = json.loads(request.body)
        classe_id = data.get('classe_id')
        repartition_id = data.get('repartition_id')
        notes_data = data.get('notes', [])

        if not classe_id or not repartition_id:
            return JsonResponse({'success': False, 'error': 'classe_id et repartition_id requis.'}, status=400)
        if not notes_data:
            return JsonResponse({'success': False, 'error': 'Aucune note à enregistrer.'}, status=400)

        # Get config
        config = _get_or_create_repartition_config(repartition_id, etab_id)
        if not config:
            return JsonResponse({'success': False, 'error': 'Configuration de répartition introuvable.'}, status=404)

        from django.db import connections
        rep_type_id = config.repartition.type_id

        # Find EX note_type via raw SQL
        conn_hub = connections['countryStructure'].cursor()
        try:
            conn_hub.execute("""
                SELECT rtn.ponderation_max, nt.id_type_note
                FROM repartition_type_notes rtn
                JOIN note_types nt ON nt.id = rtn.note_type_id
                WHERE rtn.repartition_type_id = %s AND nt.sigle = 'EX' AND rtn.is_active = 1
                LIMIT 1
            """, [rep_type_id])
            ex_row = conn_hub.fetchone()
        finally:
            conn_hub.close()

        if not ex_row:
            return JsonResponse({'success': False, 'error': 'Type de note Examen non configuré.'}, status=404)

        ex_nt_id = ex_row[1]

        conn = _get_spoke_connection()
        try:
            saved = 0
            with conn.cursor() as cur:
                for nd in notes_data:
                    eleve_id = nd.get('eleve_id')
                    cours_annee_id = nd.get('cours_annee_id')
                    note_val = nd.get('note')
                    maxima = nd.get('maxima')

                    if not eleve_id or not cours_annee_id:
                        continue

                    # Handle empty/null notes: delete existing
                    if note_val is None or str(note_val).strip() == '':
                        cur.execute("""
                            DELETE FROM note_bulletin
                            WHERE id_eleve_id = %s AND id_cours_annee = %s
                              AND id_repartition_config = %s AND id_note_type = %s
                        """, [eleve_id, cours_annee_id, config.id, ex_nt_id])
                        continue

                    try:
                        note_float = round(float(note_val), 2)
                    except (ValueError, TypeError):
                        continue

                    maxima_val = int(maxima) if maxima else (ex_row[0] or 20)

                    # Upsert
                    cur.execute("""
                        INSERT INTO note_bulletin
                            (id_eleve_id, id_cours_annee, id_repartition_config, id_note_type,
                             note, maxima, source_type, date_calcul, id_etablissement, id_pays)
                        VALUES (%s, %s, %s, %s, %s, %s, 'SAISIE_DIRECTE', NOW(), %s, %s)
                        ON DUPLICATE KEY UPDATE
                            note = VALUES(note), maxima = VALUES(maxima),
                            source_type = 'SAISIE_DIRECTE',
                            date_calcul = NOW(), updated_at = NOW()
                    """, [eleve_id, cours_annee_id, config.id, ex_nt_id, note_float, maxima_val, etab_id, etab.pays_id])
                    saved += 1

            conn.commit()
            return JsonResponse({
                'success': True,
                'saved': saved,
                'message': f'{saved} note(s) d\'examen enregistrée(s).'
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def download_exam_template(request):
    """
    Génère un modèle Excel pour les notes d'examen (un cours par colonne),
    pré-rempli avec les notes existantes.
    Params: classe_id (EAC id), repartition_id (parent instance id)
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Border, Side, Protection, Alignment
        from io import BytesIO

        classe_id = request.GET.get('classe_id')
        repartition_id = request.GET.get('repartition_id')

        if not classe_id or not repartition_id:
            return JsonResponse({'success': False, 'error': 'classe_id et repartition_id requis.'}, status=400)

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        config = _get_or_create_repartition_config(repartition_id, etab_id)
        from django.db import connections
        rep_type_id = config.repartition.type_id if config else None

        # Find EX note type via raw SQL
        conn_hub = connections['countryStructure'].cursor()
        try:
            conn_hub.execute("""
                SELECT rtn.ponderation_max, nt.id_type_note
                FROM repartition_type_notes rtn
                JOIN note_types nt ON nt.id = rtn.note_type_id
                WHERE rtn.repartition_type_id = %s AND nt.sigle = 'EX' AND rtn.is_active = 1
                LIMIT 1
            """, [rep_type_id])
            ex_row = conn_hub.fetchone()
        finally:
            conn_hub.close()

        if not ex_row:
            return JsonResponse({'success': False, 'error': 'Type de note Examen non configuré.'}, status=404)

        ex_note_type_id = ex_row[1]
        default_max = ex_row[0] or 20

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Get context from EAC
                cur.execute("""
                    SELECT ea.annee_id AS id_annee,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # Students
                cur.execute("""
                    SELECT DISTINCT e.id_eleve, e.nom, e.prenom
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                      AND ei.status = 1
                    ORDER BY e.nom, e.prenom
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleves = cur.fetchall()

                # Courses with maxima_exam
                cur.execute("""
                    SELECT cann.id_cours_annee, ca.cours AS cours_nom, ca.code_cours,
                           cann.maxima_exam, cann.ordre
                    FROM countryStructure.cours_annee cann
                    JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    JOIN countryStructure.etablissements_annees ea ON ea.annee_id = cann.annee_id
                    JOIN countryStructure.etablissements_annees_classes eac ON eac.etablissement_annee_id = ea.id
                    WHERE eac.id = %s AND ca.classe_id = eac.classe_id
                    GROUP BY cann.cours_id
                    ORDER BY cann.ordre, ca.cours
                """, [classe_id])
                cours_rows = cur.fetchall()

                # Existing exam notes
                existing_notes = {}
                if config and cours_rows:
                    cours_ids = [r['id_cours_annee'] for r in cours_rows]
                    eleve_ids = [e['id_eleve'] for e in eleves]
                    if cours_ids and eleve_ids:
                        ph_c = ','.join(['%s'] * len(cours_ids))
                        ph_e = ','.join(['%s'] * len(eleve_ids))
                        cur.execute(f"""
                            SELECT nb.id_eleve_id, nb.id_cours_annee, nb.note
                            FROM note_bulletin nb
                            WHERE nb.id_repartition_config = %s
                              AND nb.id_note_type = %s
                              AND nb.id_cours_annee IN ({ph_c})
                              AND nb.id_eleve_id IN ({ph_e})
                              AND nb.id_etablissement = %s
                        """, [config.id, ex_note_type_id] + cours_ids + eleve_ids + [etab_id])
                        for n in cur.fetchall():
                            key = f"{n['id_eleve_id']}_{n['id_cours_annee']}"
                            existing_notes[key] = float(n['note']) if n['note'] is not None else None
        finally:
            conn.close()

        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Examen Notes"

        header_fill = PatternFill('solid', fgColor='7C3AED')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        locked = Protection(locked=True)
        unlocked = Protection(locked=False)
        note_fill = PatternFill('solid', fgColor='F5F3FF')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = ['ID Élève', 'Nom', 'Prénom']
        for c in cours_rows:
            cname = c['cours_nom'] or '?'
            max_ex = c['maxima_exam'] or default_max
            headers.append(f"{cname[:20]} (/{max_ex})")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Data rows with pre-fill
        for row, el in enumerate(eleves, 2):
            ws.cell(row=row, column=1, value=el['id_eleve']).protection = locked
            ws.cell(row=row, column=2, value=el['nom']).protection = locked
            ws.cell(row=row, column=3, value=el['prenom']).protection = locked
            for col_offset, c in enumerate(cours_rows):
                key = f"{el['id_eleve']}_{c['id_cours_annee']}"
                note_val = existing_notes.get(key)
                cell = ws.cell(row=row, column=4 + col_offset, value=note_val if note_val is not None else '')
                cell.protection = unlocked
                cell.border = thin_border
                if note_val is not None:
                    cell.fill = note_fill

        # Meta sheet
        ws2 = wb.create_sheet('_meta')
        ws2.cell(row=1, column=1, value='classe_id')
        ws2.cell(row=1, column=2, value=int(classe_id))
        ws2.cell(row=2, column=1, value='repartition_id')
        ws2.cell(row=2, column=2, value=int(repartition_id))
        ws2.cell(row=3, column=1, value='etab_id')
        ws2.cell(row=3, column=2, value=etab_id)
        ws2.cell(row=4, column=1, value='type')
        ws2.cell(row=4, column=2, value='EXAM')
        for i, c in enumerate(cours_rows):
            ws2.cell(row=5 + i, column=1, value=f'cours_{i}')
            ws2.cell(row=5 + i, column=2, value=c['id_cours_annee'])
            ws2.cell(row=5 + i, column=3, value=c['maxima_exam'] or default_max)
        ws2.sheet_state = 'hidden'

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        for i in range(len(cours_rows)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(4 + i)].width = 22

        ws.protection.enable()

        rep_name = config.repartition.nom if config else 'Examen'
        filename = f"Modele_Examen_{rep_name.replace(' ', '_')}.xlsx"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            content=buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_bulletin_overview(request):
    """
    Retourne TOUTES les notes de bulletin pour une classe, toutes répartitions confondues.
    Utilisé par la vue accordéon de Notes Pondérées.
    Retourne: élèves, cours, répartitions hiérarchiques, note_types par type, notes.
    """
    try:
        classe_id = request.GET.get('classe_id')
        if not classe_id:
            return JsonResponse({'success': False, 'error': 'classe_id requis.'}, status=400)

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        annee = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        if not annee:
            return JsonResponse({'success': False, 'error': 'Pas d\'année ouverte.'})

        ea = EtablissementAnnee.objects.filter(
            etablissement_id=etab_id, annee=annee, id_pays=etab.pays_id
        ).first()
        if not ea:
            return JsonResponse({'success': False, 'error': 'Config étab-année introuvable.'})

        # Resolve cycle for the class
        eac = EtablissementAnneeClasse.objects.filter(id=classe_id).select_related('classe').first()
        if not eac:
            return JsonResponse({'success': False, 'error': 'Classe introuvable.'})

        cycle_id = getattr(eac.classe, 'cycle_id', None) or getattr(eac, 'cycle_id', None)

        # Get root type + hierarchy for this cycle
        root_type_id = None
        child_type_id = None
        root_count = 0
        child_count_per_root = 0
        allowed_type_ids = set()

        if cycle_id:
            cycle_config = RepartitionConfigCycle.objects.filter(
                cycle_id=cycle_id, is_active=True, id_pays=etab.pays_id
            ).first()
            if cycle_config:
                root_type_id = cycle_config.type_racine_id
                root_count = cycle_config.nombre_au_niveau_racine
                allowed_type_ids.add(root_type_id)

                hierarchies = RepartitionHierarchie.objects.filter(
                    type_parent_id=root_type_id, is_active=True, id_pays=etab.pays_id
                )
                for h in hierarchies:
                    child_type_id = h.type_enfant_id
                    child_count_per_root = h.nombre_enfants
                    allowed_type_ids.add(child_type_id)

        # Get all repartition configs for this etab/annee filtered by allowed types
        configs = RepartitionConfigEtabAnnee.objects.filter(
            etablissement_annee=ea
        ).select_related('repartition', 'repartition__type')

        if allowed_type_ids:
            configs = configs.filter(repartition__type_id__in=allowed_type_ids)

        # Build repartitions list with parent-child mapping
        repartitions = []
        parent_reps = []
        child_reps = []

        for cfg in configs.order_by('repartition__type_id', 'repartition__ordre'):
            rep = cfg.repartition
            type_id = rep.type_id
            is_parent = (type_id == root_type_id)

            rep_data = {
                'config_id': cfg.id,
                'repartition_id': rep.id_instance,
                'nom': rep.nom,
                'code': rep.code,
                'type_id': type_id,
                'type_code': rep.type.code if rep.type else '',
                'type_nom': rep.type.nom if rep.type else '',
                'ordre': rep.ordre,
                'is_parent': is_parent,
                'is_open': cfg.is_open,
                'parent_config_id': None,
                'taux_participation': float(rep.taux_participation) if rep.taux_participation else 100.0,
            }

            if is_parent:
                parent_reps.append(rep_data)
            else:
                child_reps.append(rep_data)
            repartitions.append(rep_data)

        # Assign children to parents by order
        if parent_reps and child_reps and child_count_per_root > 0:
            for ci, child in enumerate(child_reps):
                pi = ci // child_count_per_root
                if pi < len(parent_reps):
                    child['parent_config_id'] = parent_reps[pi]['config_id']

        # Limit by expected count
        if root_count > 0:
            parent_reps = parent_reps[:root_count]
            total_children = root_count * child_count_per_root
            child_reps = child_reps[:total_children]
            allowed_config_ids = set(r['config_id'] for r in parent_reps + child_reps)
            repartitions = [r for r in repartitions if r['config_id'] in allowed_config_ids]

        # Get note_types per repartition type
        from django.db import connections
        note_types_by_type = {}
        if allowed_type_ids:
            conn_hub = connections['countryStructure'].cursor()
            try:
                type_placeholders = ','.join(['%s'] * len(allowed_type_ids))
                conn_hub.execute(f"""
                    SELECT rtn.repartition_type_id, rtn.ponderation_max, rtn.source_type,
                           rtn.mode_calcul, rtn.ordre,
                           nt.id_type_note, nt.sigle, nt.nom
                    FROM repartition_type_notes rtn
                    JOIN note_types nt ON nt.id = rtn.note_type_id
                    WHERE rtn.repartition_type_id IN ({type_placeholders})
                      AND rtn.is_active = 1
                    ORDER BY rtn.repartition_type_id, rtn.ordre
                """, list(allowed_type_ids))
                columns = [col[0] for col in conn_hub.description]
                for row in conn_hub.fetchall():
                    r = dict(zip(columns, row))
                    tid = r['repartition_type_id']
                    note_types_by_type.setdefault(tid, []).append({
                        'id': r['id_type_note'],
                        'sigle': r['sigle'],
                        'nom': r['nom'],
                        'max': r['ponderation_max'],
                        'source': r['source_type'],
                        'ordre': r['ordre'],
                    })
            finally:
                conn_hub.close()

        # Get students + cours via spoke
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT ea2.annee_id AS id_annee,
                           eac2.classe_id AS bk_classe, eac2.groupe AS bk_groupe, eac2.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac2
                    JOIN countryStructure.etablissements_annees ea2 ON ea2.id = eac2.etablissement_annee_id
                    WHERE eac2.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()
                if not ctx:
                    return JsonResponse({'success': False, 'error': 'Contexte classe introuvable.'})

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                # Students
                cur.execute("""
                    SELECT DISTINCT e.id_eleve, e.nom, e.prenom
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                      AND ei.status = 1
                    ORDER BY e.nom, e.prenom
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleves = [dict(r) for r in cur.fetchall()]

                # Courses
                cur.execute("""
                    SELECT MIN(cann.id_cours_annee) AS id_cours_annee, ca.cours AS cours_nom, ca.code_cours,
                           MAX(cann.maxima_exam) AS maxima_exam, MAX(cann.maxima_tj) AS maxima_tj,
                           MAX(cann.maxima_periode) AS maxima_periode
                    FROM countryStructure.cours_annee cann
                    JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    JOIN countryStructure.etablissements_annees ea3 ON ea3.annee_id = cann.annee_id
                    JOIN countryStructure.etablissements_annees_classes eac3 ON eac3.etablissement_annee_id = ea3.id
                    WHERE eac3.id = %s AND ca.classe_id = eac3.classe_id
                    GROUP BY cann.cours_id, ca.cours, ca.code_cours
                    ORDER BY ca.cours
                """, [classe_id])
                cours = [dict(c) for c in cur.fetchall()]

                # Get ALL note_bulletin for all configs of this class
                config_ids = [r['config_id'] for r in repartitions]
                notes = {}
                if config_ids:
                    ph = ','.join(['%s'] * len(config_ids))
                    cur.execute(f"""
                        SELECT nb.id_eleve_id, nb.id_cours_annee, nb.id_repartition_config,
                               nb.id_note_type, nb.note, nb.maxima
                        FROM note_bulletin nb
                        WHERE nb.id_repartition_config IN ({ph}) AND nb.id_etablissement = %s
                    """, config_ids + [etab_id])
                    for n in cur.fetchall():
                        key = f"{n['id_eleve_id']}_{n['id_cours_annee']}_{n['id_repartition_config']}_{n['id_note_type']}"
                        notes[key] = {
                            'note': float(n['note']) if n['note'] is not None else None,
                            'maxima': n['maxima'],
                        }

                # Get evaluations linked to these configs (filtered by this class's courses)
                evaluations = []
                eval_ids_all = []
                cours_ids_list = [c['id_cours_annee'] for c in cours]
                if config_ids and cours_ids_list:
                    ph2 = ','.join(['%s'] * len(config_ids))
                    ph_cours = ','.join(['%s'] * len(cours_ids_list))
                    cur.execute(f"""
                        SELECT ev.id_evaluation, ev.title, ev.ponderer_eval,
                               ev.id_cours_classe_id AS cours_annee_id,
                               er.id_repartition_config AS config_id
                        FROM evaluation ev
                        JOIN evaluation_repartition er ON er.id_evaluation = ev.id_evaluation
                        WHERE er.id_repartition_config IN ({ph2})
                          AND ev.id_etablissement = %s
                          AND ev.id_cours_classe_id IN ({ph_cours})
                        ORDER BY ev.id_evaluation
                    """, config_ids + [etab_id] + cours_ids_list)
                    for ev in cur.fetchall():
                        evaluations.append({
                            'id': ev['id_evaluation'],
                            'nom': ev['title'],
                            'max': ev['ponderer_eval'],
                            'cours_id': ev['cours_annee_id'],
                            'config_id': ev['config_id'],
                        })
                        eval_ids_all.append(ev['id_evaluation'])

                # Get raw eleve_note for those evaluations
                raw_notes = {}
                if eval_ids_all:
                    ph3 = ','.join(['%s'] * len(eval_ids_all))
                    cur.execute(f"""
                        SELECT en.id_eleve_id, en.id_evaluation_id, en.note
                        FROM eleve_note en
                        WHERE en.id_evaluation_id IN ({ph3})
                    """, eval_ids_all)
                    for rn in cur.fetchall():
                        key = f"{rn['id_eleve_id']}_{rn['id_evaluation_id']}"
                        raw_notes[key] = float(rn['note']) if rn['note'] is not None else None
        finally:
            conn.close()

        return JsonResponse({
            'success': True,
            'classe_nom': str(eac.classe) if eac.classe else '',
            'eleves': [{'id': e['id_eleve'], 'nom': e['nom'], 'prenom': e['prenom']} for e in eleves],
            'cours': [{'id': c['id_cours_annee'], 'nom': c['cours_nom'], 'code': c['code_cours'],
                       'maxima_exam': c['maxima_exam'], 'maxima_tj': c['maxima_tj'],
                       'maxima_periode': c['maxima_periode']} for c in cours],
            'repartitions': repartitions,
            'note_types_by_type': note_types_by_type,
            'notes': notes,
            'evaluations': evaluations,
            'raw_notes': raw_notes,
            'root_type_id': root_type_id,
            'child_type_id': child_type_id,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_notes_bulletin(request):
    """
    Récupère les notes de bulletin calculées pour affichage.
    Params: classe_id, repartition_id
    """
    try:
        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        repartition_id = request.GET.get('repartition_id')

        if not classe_id or not repartition_id:
            return JsonResponse({'success': False, 'error': 'Paramètres requis.'}, status=400)

        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)
        etab, err = _get_tenant_etab(request)
        if err: return err
        etab_id = etab.id_etablissement

        config = _get_or_create_repartition_config(repartition_id, etab_id)

        if not config:
            return JsonResponse({'success': False, 'error': 'Config introuvable.'}, status=404)

        rep_type = config.repartition.type
        rep_type_id = config.repartition.type_id

        # Get expected note types via raw SQL
        from django.db import connections
        conn_hub = connections['countryStructure'].cursor()
        try:
            conn_hub.execute("""
                SELECT rtn.id, rtn.ponderation_max, rtn.source_type, rtn.ordre,
                       nt.id_type_note, nt.sigle, nt.nom
                FROM repartition_type_notes rtn
                JOIN note_types nt ON nt.id = rtn.note_type_id
                WHERE rtn.repartition_type_id = %s AND rtn.is_active = 1
                  AND rtn.is_visible_bulletin = 1
                ORDER BY rtn.ordre
            """, [rep_type_id])
            columns = [col[0] for col in conn_hub.description]
            expected_raw = [dict(zip(columns, row)) for row in conn_hub.fetchall()]
        finally:
            conn_hub.close()

        class _NT:
            def __init__(self, r):
                self.id_type_note = r['id_type_note']
                self.sigle = r['sigle']
                self.nom = r['nom']
        class _RTN:
            def __init__(self, r):
                self.note_type = _NT(r)
                self.ponderation_max = r['ponderation_max']
                self.source_type = r['source_type']
                self.ordre = r['ordre']
        expected = [_RTN(r) for r in expected_raw]

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Get students + business keys
                cur.execute("""
                    SELECT ea.annee_id AS id_annee,
                           eac.classe_id AS bk_classe, eac.groupe AS bk_groupe, eac.section_id AS bk_section
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.etablissements_annees ea ON ea.id = eac.etablissement_annee_id
                    WHERE eac.id = %s LIMIT 1
                """, [classe_id])
                ctx = cur.fetchone()

                cur.execute("SELECT idCampus FROM campus WHERE id_etablissement = %s AND is_active=1 LIMIT 1", [etab_id])
                campus_row = cur.fetchone()
                campus_id = campus_row['idCampus'] if campus_row else None

                cur.execute("""
                    SELECT DISTINCT e.id_eleve, e.nom, e.prenom
                    FROM eleve_inscription ei
                    JOIN eleve e ON e.id_eleve = ei.id_eleve_id
                    WHERE ei.id_annee_id = %s AND ei.idCampus_id = %s
                      AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                      AND ei.status = 1
                    ORDER BY e.nom, e.prenom
                """, [ctx['id_annee'], campus_id, ctx['bk_classe'], ctx['bk_groupe'], ctx['bk_section']])
                eleves = cur.fetchall()

                # Get cours (with maxima_exam) — filtered by classe
                cur.execute("""
                    SELECT MIN(cann.id_cours_annee) AS id_cours_annee, ca.cours AS cours_nom, ca.code_cours,
                           MAX(cann.maxima_exam) AS maxima_exam
                    FROM countryStructure.cours_annee cann
                    JOIN countryStructure.cours ca ON ca.id = cann.cours_id
                    JOIN countryStructure.etablissements_annees ea ON ea.annee_id = cann.annee_id
                    JOIN countryStructure.etablissements_annees_classes eac ON eac.etablissement_annee_id = ea.id
                    WHERE eac.id = %s AND ca.classe_id = eac.classe_id
                    GROUP BY cann.cours_id, ca.cours, ca.code_cours
                    ORDER BY ca.cours
                """, [classe_id])
                cours = cur.fetchall()

                # Get all note_bulletin entries for this config
                # ORDER BY id_note_bulletin so latest entry (highest id) overwrites older duplicates in dict
                cur.execute("""
                    SELECT nb.id_eleve_id, nb.id_cours_annee, nb.id_note_type,
                           nb.note, nb.maxima, nb.source_type, nb.date_calcul
                    FROM note_bulletin nb
                    WHERE nb.id_repartition_config = %s AND nb.id_etablissement = %s
                    ORDER BY nb.id_note_bulletin ASC
                """, [config.id, etab_id])
                notes_raw = cur.fetchall()

                # Build notes dict: key = "eleve_cours_notetype"
                notes = {}
                for n in notes_raw:
                    key = f"{n['id_eleve_id']}_{n['id_cours_annee']}_{n['id_note_type']}"
                    notes[key] = {
                        'note': float(n['note']) if n['note'] is not None else None,
                        'maxima': n['maxima'],
                        'source': n['source_type'],
                    }

            return JsonResponse({
                'success': True,
                'repartition_name': config.repartition.nom,
                'repartition_type': rep_type.nom,
                'eleves': [{'id': e['id_eleve'], 'nom': e['nom'], 'prenom': e['prenom']} for e in eleves],
                'cours': [{'id': c['id_cours_annee'], 'nom': c['cours_nom'], 'code': c['code_cours'],
                           'maxima_exam': c['maxima_exam']} for c in cours],
                'note_types': [{'id': e.note_type.id_type_note, 'sigle': e.note_type.sigle,
                                'nom': e.note_type.nom, 'max': e.ponderation_max,
                                'source': e.source_type, 'ordre': e.ordre} for e in expected],
                'notes': notes,
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# --- CRUD DOMAINES ---
# ============================================================

@require_http_methods(["GET"])
def get_domaines_data(request):
    """Liste les domaines pour un pays donné."""
    try:
        id_pays = request.GET.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})
        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'domaines': []})
        if Domaine is None:
            return JsonResponse({'success': True, 'domaines': []})
        domaines = list(Domaine.objects.filter(pays=pays).values(
            'id_domaine', 'nom', 'code'
        ))
        return JsonResponse({'success': True, 'domaines': domaines})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def save_domaine(request):
    """Crée ou modifie un domaine."""
    try:
        data = json.loads(request.body)
        id_domaine = data.get('id_domaine')
        nom = data.get('nom', '').strip()
        code = data.get('code', '').strip()
        sigle = data.get('sigle', '').strip()
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')

        if not nom:
            return JsonResponse({'success': False, 'error': 'Le nom du domaine est requis.'}, status=400)
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant.'}, status=400)
        if Domaine is None:
            return JsonResponse({'success': False, 'error': 'Modèle Domaine non disponible.'}, status=500)

        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': False, 'error': 'Pays introuvable.'}, status=400)

        if id_domaine:
            d = Domaine.objects.filter(id_domaine=id_domaine, pays=pays).first()
            if not d:
                return JsonResponse({'success': False, 'error': 'Domaine introuvable.'}, status=404)
            d.nom = nom
            d.code = code
            d.save()
        else:
            if Domaine.objects.filter(pays=pays, nom=nom).exists():
                return JsonResponse({'success': False, 'error': f'Le domaine "{nom}" existe déjà.'}, status=400)
            Domaine.objects.create(nom=nom, code=code, pays=pays)

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def delete_domaine(request):
    """Supprime un domaine s'il n'est pas référencé."""
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        if Domaine is None:
            return JsonResponse({'success': False, 'error': 'Modèle Domaine non disponible.'}, status=500)
        pays = Pays.objects.filter(id_pays=id_pays).first() if id_pays else None
        domaine = Domaine.objects.filter(id_domaine=data.get('id_domaine')).first()
        if not domaine:
            return JsonResponse({'success': False, 'error': 'Domaine introuvable.'}, status=404)

        if Cours and pays:
            cours_count = Cours.objects.filter(domaine=domaine, id_pays=pays.id_pays).count()
        else:
            cours_count = 0
        if pays:
            ca_count = CoursAnnee.objects.filter(domaine=domaine, id_pays=pays.id_pays).count()
        else:
            ca_count = 0
        if cours_count > 0 or ca_count > 0:
            return JsonResponse({
                'success': False,
                'error': f'Ce domaine est utilisé par {cours_count} cours et {ca_count} config(s) annuelle(s).'
            }, status=400)

        domaine.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# --- CRUD COURS (Catalogue) ---
# ============================================================

@require_http_methods(["GET"])
def get_cours_data(request):
    """Liste les cours, filtré par pays (id_pays) et optionnellement par classe+section."""
    try:
        id_pays = request.GET.get('id_pays')
        id_classe = request.GET.get('id_classe')
        id_section = request.GET.get('id_section')

        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})

        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'cours': [], 'classes': []})
        if Cours is None:
            return JsonResponse({'success': True, 'cours': [], 'classes': []})

        all_sections = []
        if Section:
            all_sections = list(Section.objects.filter(pays_id=pays.id_pays).order_by('nom').values('id_section', 'nom', 'code'))

        # Récupérer les classes activées pour cet établissement
        id_etablissement = getattr(request, 'id_etablissement', None) or request.session.get('id_etablissement')
        activated_classe_ids = set()
        if id_etablissement:
            annee_active = Annee.objects.filter(pays_id=pays.id_pays, isOpen=True).order_by('-annee').first()
            if annee_active:
                etab_annee = EtablissementAnnee.objects.filter(
                    etablissement_id=id_etablissement, annee=annee_active, id_pays=pays.id_pays
                ).first()
                if etab_annee:
                    activated_classe_ids = set(
                        EtablissementAnneeClasse.objects.filter(
                            etablissement_annee=etab_annee
                        ).values_list('classe_id', flat=True)
                    )

        if activated_classe_ids:
            activated_cycle_ids = set(
                Classe.objects.filter(id__in=activated_classe_ids).values_list('cycle_id', flat=True)
            )
            cycles = Cycle.objects.filter(id__in=activated_cycle_ids).order_by('ordre')
        else:
            cycles = pays.cycles.all().order_by('ordre')

        classes_data = []
        for cycle in cycles:
            if activated_classe_ids:
                cycle_classes = Classe.objects.filter(
                    id__in=activated_classe_ids, cycle_id=cycle.id
                ).order_by('ordre')
            else:
                cycle_classes = cycle.classes.order_by('ordre')
            if cycle.duree and cycle.duree > 0:
                cycle_classes = cycle_classes[:cycle.duree]

            if cycle.hasSections and all_sections:
                for section in all_sections:
                    for cls in cycle_classes:
                        classes_data.append({
                            'id_classe': cls.id_classe, 'id_section': section['id_section'],
                            'nom': f"{cls.nom} — {section['nom']}",
                            'cycle_nom': f"{cycle.nom} / {section['nom']}",
                            'cycle_id': cycle.id, 'has_sections': True,
                        })
            else:
                for cls in cycle_classes:
                    classes_data.append({
                        'id_classe': cls.id_classe, 'id_section': None,
                        'nom': cls.nom, 'cycle_nom': cycle.nom,
                        'cycle_id': cycle.id, 'has_sections': False,
                    })

        # Build lookup dicts (Cours has IntegerField domaine_id/section_id, NOT FKs)
        domaine_map = {}
        if Domaine:
            domaine_map = {d['id_domaine']: d['nom'] for d in Domaine.objects.filter(pays=pays).values('id_domaine', 'nom')}
        section_map = {}
        if Section:
            section_map = {s['id_section']: s['nom'] for s in Section.objects.filter(pays_id=pays.id_pays).values('id_section', 'nom')}

        cours_qs = Cours.objects.filter(classe__cycle__pays=pays)
        if id_classe:
            cours_qs = cours_qs.filter(classe__id_classe=id_classe)
            if id_section:
                cours_qs = cours_qs.filter(section_id=id_section)
            else:
                cours_qs = cours_qs.filter(section_id__isnull=True)

        cours_data = []
        for c in cours_qs:
            cours_data.append({
                'id_cours': c.id_cours, 'cours': c.cours, 'code_cours': c.code_cours,
                'domaine_id': c.domaine_id,
                'domaine_nom': domaine_map.get(c.domaine_id, ''),
                'id_classe_id': c.classe_id,
                'section_id': c.section_id,
                'section_nom': section_map.get(c.section_id, ''),
            })

        domaines = list(Domaine.objects.filter(pays=pays).values('id_domaine', 'nom', 'code')) if Domaine else []

        return JsonResponse({
            'success': True, 'cours': cours_data,
            'classes': classes_data, 'domaines': domaines,
            'sections': all_sections,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def save_cours(request):
    """Crée ou modifie un cours."""
    try:
        data = json.loads(request.body)
        id_cours = data.get('id_cours')
        code_cours = data.get('code_cours', '').strip()
        nom_cours = data.get('cours', '').strip()
        domaine_id = data.get('domaine_id')
        id_classe = data.get('id_classe')
        id_section = data.get('id_section')
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')

        if not code_cours or not nom_cours or not id_classe:
            return JsonResponse({'success': False, 'error': 'Code, nom du cours et classe sont requis.'}, status=400)
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant.'}, status=400)
        if Cours is None:
            return JsonResponse({'success': False, 'error': 'Modèle Cours non disponible.'}, status=500)

        classe = Classe.objects.filter(id_classe=id_classe).first()
        if not classe:
            return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)
        # domaine_id and section_id are IntegerFields on Cours, not FKs
        dom_id = int(domaine_id) if domaine_id else None
        sec_id = int(id_section) if id_section else None

        if id_cours:
            c = Cours.objects.filter(id_cours=id_cours, id_pays=id_pays).first()
            if not c:
                return JsonResponse({'success': False, 'error': 'Cours introuvable.'}, status=404)
            c.cours = nom_cours
            c.code_cours = code_cours
            c.domaine_id = dom_id
            c.classe = classe
            c.section_id = sec_id
            c.save()
        else:
            if Cours.objects.filter(classe=classe, section_id=sec_id, code_cours=code_cours, id_pays=id_pays).exists():
                return JsonResponse({'success': False, 'error': f'Le code "{code_cours}" existe déjà pour cette classe.'}, status=400)
            Cours.objects.create(cours=nom_cours, code_cours=code_cours, domaine_id=dom_id, classe=classe, section_id=sec_id, id_pays=id_pays)

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def delete_cours(request):
    """Supprime un cours."""
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        if Cours is None:
            return JsonResponse({'success': False, 'error': 'Modèle Cours non disponible.'}, status=500)
        c = Cours.objects.filter(id_cours=data.get('id_cours'), id_pays=id_pays).first()
        if c:
            c.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# --- CRUD COURS_ANNEE (Configuration Annuelle des Cours) ---
# ============================================================

@require_http_methods(["GET"])
def get_cours_annee_data(request):
    """Liste la configuration annuelle des cours, filtrée par classe (+section) et année."""
    try:
        id_classe = request.GET.get('id_classe')
        id_annee = request.GET.get('id_annee')
        id_section = request.GET.get('id_section')
        id_pays = request.GET.get('id_pays')

        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant'})

        pays = Pays.objects.filter(id_pays=id_pays).first()
        if not pays:
            return JsonResponse({'success': True, 'cours_annee': [], 'annees': []})

        annees = list(Annee.objects.filter(pays_id=pays.id_pays).order_by('-annee').values('id_annee', 'annee', 'isOpen'))
        domaines = list(Domaine.objects.filter(pays=pays).values('id_domaine', 'nom', 'code')) if Domaine else []

        if not id_classe or not id_annee:
            return JsonResponse({'success': True, 'cours_annee': [], 'annees': annees, 'domaines': domaines})

        if Cours is None:
            return JsonResponse({'success': True, 'cours_annee': [], 'annees': annees, 'domaines': domaines})

        # Tous les cours du catalogue pour cette classe (+section)
        # NOTE: Cours.domaine_id is IntegerField, NOT FK — cannot use select_related/order_by FK
        cours_catalogue = Cours.objects.filter(classe__id_classe=id_classe, id_pays=id_pays).order_by('cours')
        if id_section:
            cours_catalogue = cours_catalogue.filter(section_id=id_section)
        else:
            cours_catalogue = cours_catalogue.filter(section_id__isnull=True)

        # Build domaine/section lookup maps
        domaine_map = {}
        if Domaine:
            domaine_map = {d['id_domaine']: d['nom'] for d in Domaine.objects.filter(pays=pays).values('id_domaine', 'nom')}
        section_map = {}
        if Section:
            section_map = {s['id_section']: s['nom'] for s in Section.objects.filter(pays_id=pays.id_pays).values('id_section', 'nom')}

        # Get tenant etab for per-establishment configs
        etab_id = getattr(request, 'id_etablissement', None) or request.session.get('id_etablissement')

        # Les configs existantes (check for etablissement-specific then national)
        configs_map = {}
        configs_qs = CoursAnnee.objects.filter(id_pays=id_pays,
            cours__classe__id_classe=id_classe, annee__id_annee=id_annee
        ).select_related('cours')
        if id_section:
            configs_qs = configs_qs.filter(cours__section_id=id_section)
        else:
            configs_qs = configs_qs.filter(cours__section_id__isnull=True)
        for ca in configs_qs:
            configs_map[ca.cours_id] = ca

        # Construire la réponse
        cours_data = []
        for c in cours_catalogue:
            ca = configs_map.get(c.pk)
            # Domaine: priority to CoursAnnee.domaine_id, else fallback to Cours.domaine_id (both IntegerFields)
            dom_id = (ca.domaine_id if ca and ca.domaine_id else c.domaine_id)
            dom_nom = domaine_map.get(dom_id, '') if dom_id else ''
            entry = {
                'id_cours': c.id_cours, 'code_cours': c.code_cours, 'cours': c.cours,
                'domaine_id': dom_id,
                'domaine_nom': dom_nom,
                'is_configured': ca is not None,
                'section_id': c.section_id,
                'section_nom': section_map.get(c.section_id, ''),
            }
            if ca:
                entry.update({
                    'id_cours_annee': ca.id_cours_annee,
                    'maxima_exam': ca.maxima_exam,
                    'maxima_tj': ca.maxima_tj, 'maxima_periode': ca.maxima_periode,
                    'credits': ca.credits, 'heure_semaine': ca.heure_semaine,
                    'is_obligatory': ca.is_obligatory, 'ordre': ca.ordre,
                    'compte_au_nombre_echec': ca.compte_au_nombre_echec,
                    'total_considerable_trimestre': ca.total_considerable_trimestre,
                    'est_considerer_echec_lorsque_pourcentage_est': ca.est_considerer_echec_lorsque_pourcentage_est,
                    'is_second_semester': ca.is_second_semester,
                })
            cours_data.append(entry)

        return JsonResponse({'success': True, 'cours_annee': cours_data, 'annees': annees, 'domaines': domaines})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def save_cours_annee(request):
    """Crée ou modifie une configuration annuelle de cours."""
    try:
        data = json.loads(request.body)
        id_cours_annee = data.get('id_cours_annee')
        cours_id = data.get('cours_id')
        annee_id = data.get('annee_id')
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')

        int_fields = ['maxima_exam', 'maxima_tj', 'maxima_periode', 'credits',
                      'heure_semaine', 'ordre', 'est_considerer_echec_lorsque_pourcentage_est']
        bool_fields = ['is_obligatory', 'compte_au_nombre_echec', 'total_considerable_trimestre', 'is_second_semester']

        if id_cours_annee:
            ca = CoursAnnee.objects.filter(id_cours_annee=id_cours_annee, id_pays=id_pays).first()
            if not ca:
                return JsonResponse({'success': False, 'error': 'Config introuvable.'}, status=404)
            for f in int_fields:
                if f in data:
                    v = data[f]
                    try:
                        setattr(ca, f, int(v) if v is not None and v != '' else None)
                    except (ValueError, TypeError):
                        setattr(ca, f, None)
            for f in bool_fields:
                if f in data:
                    setattr(ca, f, bool(data[f]))
            if 'domaine_id' in data and Domaine:
                ca.domaine_id = int(data['domaine_id']) if data['domaine_id'] else None
            ca.save()
        else:
            if not cours_id or not annee_id:
                return JsonResponse({'success': False, 'error': 'Cours et année requis.'}, status=400)
            if Cours is None:
                return JsonResponse({'success': False, 'error': 'Modèle Cours non disponible.'}, status=500)

            fields = {}
            for f in int_fields:
                v = data.get(f)
                try:
                    fields[f] = int(v) if v is not None and v != '' else None
                except (ValueError, TypeError):
                    fields[f] = None
            for f in bool_fields:
                fields[f] = bool(data.get(f, False))

            cours = Cours.objects.filter(id_cours=cours_id, id_pays=id_pays).first()
            annee = Annee.objects.filter(pk=annee_id, pays_id=id_pays).first()
            if not cours or not annee:
                return JsonResponse({'success': False, 'error': 'Cours ou année introuvable.'}, status=404)

            etab_id = data.get('etablissement_id') or getattr(request, 'id_etablissement', None) or request.session.get('id_etablissement')
            etab = Etablissement.objects.filter(id_etablissement=etab_id).first() if etab_id else None

            if CoursAnnee.objects.filter(cours=cours, annee=annee, etablissement=etab, id_pays=cours.id_pays).exists():
                return JsonResponse({'success': False, 'error': 'Ce cours est déjà configuré pour cette année.'}, status=400)

            dom_id = data.get('domaine_id') or (cours.domaine_id if cours.domaine_id else None)
            if dom_id:
                dom_id = int(dom_id)
            CoursAnnee.objects.create(cours=cours, annee=annee, etablissement=etab, domaine_id=dom_id, id_pays=cours.id_pays, **fields)

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def delete_cours_annee(request):
    """Supprime une configuration annuelle de cours."""
    try:
        data = json.loads(request.body)
        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        ca = CoursAnnee.objects.filter(id_cours_annee=data.get('id_cours_annee'), id_pays=id_pays).first()
        if ca:
            ca.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def bulk_activate_cours_annee(request):
    """Active en masse les cours du catalogue pour une classe/année (+section)."""
    try:
        data = json.loads(request.body)
        id_classe = data.get('id_classe')
        id_annee = data.get('id_annee')
        id_section = data.get('id_section')

        if not id_classe or not id_annee:
            return JsonResponse({'success': False, 'error': 'Classe et année requis.'}, status=400)
        if Cours is None:
            return JsonResponse({'success': False, 'error': 'Modèle Cours non disponible.'}, status=500)

        id_pays = data.get('id_pays') or getattr(request, 'id_pays', None) or request.session.get('id_pays')
        if not id_pays:
            return JsonResponse({'success': False, 'error': 'id_pays manquant.'}, status=400)

        cours_catalogue = Cours.objects.filter(classe__id_classe=id_classe, id_pays=id_pays)
        if id_section:
            cours_catalogue = cours_catalogue.filter(section_id=id_section)
        else:
            cours_catalogue = cours_catalogue.filter(section_id__isnull=True)
        annee = Annee.objects.filter(pk=id_annee, pays_id=id_pays).first()
        if not annee:
            return JsonResponse({'success': False, 'error': 'Année introuvable.'}, status=404)

        created = 0
        for c in cours_catalogue:
            _, was_created = CoursAnnee.objects.get_or_create(
                cours=c, annee=annee, defaults={'ordre': c.id_cours, 'id_pays': id_pays}
            )
            if was_created:
                created += 1

        return JsonResponse({'success': True, 'count': created})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# EMPLOI DU TEMPS / HORAIRE
# ============================================================

@csrf_exempt
@require_http_methods(["POST"])
def dashboard_horaire(request):
    """API pour la gestion de l'emploi du temps."""
    try:
        data = json.loads(request.body)
        action = data.get('action')
        etab_id = data.get('id_etablissement')

        if not etab_id:
            return JsonResponse({'success': False, 'error': 'id_etablissement requis'}, status=400)
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')

        if action == 'get-periods':
            annee_id = data.get('id_annee')
            if not annee_id:
                return JsonResponse({'success': False, 'error': 'id_annee requis'})

            etab_annee = EtablissementAnnee.objects.filter(
                etablissement_id=etab_id, annee_id=annee_id
            ).first()
            if not etab_annee:
                return JsonResponse({'success': True, 'periods': []})

            configs = RepartitionConfigEtabAnnee.objects.filter(
                etablissement_annee=etab_annee
            ).select_related('repartition', 'repartition__type')

            periods = []
            for cfg in configs:
                rep = cfg.repartition
                if not rep:
                    continue
                periods.append({
                    'id': cfg.id,
                    'repartition_id': rep.id_instance,
                    'nom': rep.nom,
                    'code': rep.code,
                    'type_nom': rep.type.nom if rep.type else '',
                    'type_code': rep.type.code if rep.type else '',
                    'debut': str(cfg.debut) if cfg.debut else str(rep.date_debut) if rep.date_debut else '',
                    'fin': str(cfg.fin) if cfg.fin else str(rep.date_fin) if rep.date_fin else '',
                    'ordre': rep.ordre,
                })
            periods.sort(key=lambda x: (x['type_code'], x['ordre']))
            return JsonResponse({'success': True, 'periods': periods})

        elif action == 'get-dates':
            date_debut = data.get('date_debut')
            date_fin = data.get('date_fin')
            if not date_debut or not date_fin:
                return JsonResponse({'success': False, 'error': 'date_debut et date_fin requis'})

            from datetime import datetime, timedelta
            d_start = datetime.strptime(date_debut, '%Y-%m-%d').date()
            d_end = datetime.strptime(date_fin, '%Y-%m-%d').date()

            jour_noms = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']

            weeks = []
            current = d_start
            while current.weekday() != 0:
                current -= timedelta(days=1)

            while current <= d_end:
                week_start = current
                week_dates = []
                for i in range(7):  # 7 jours (Lun-Dim)
                    day = week_start + timedelta(days=i)
                    if d_start <= day <= d_end:
                        week_dates.append({
                            'date': str(day),
                            'jour': jour_noms[day.weekday()],
                            'is_weekend': day.weekday() >= 5,
                        })
                if week_dates:
                    weeks.append({
                        'week_label': f"Sem. {week_start.isocalendar()[1]} ({week_dates[0]['date']} au {week_dates[-1]['date']})",
                        'dates': week_dates,
                    })
                current += timedelta(days=7)

            return JsonResponse({'success': True, 'weeks': weeks})

        elif action == 'list':
            id_classe = data.get('id_classe')
            date_debut = data.get('date_debut')
            date_fin = data.get('date_fin')

            if not id_classe:
                return JsonResponse({'success': False, 'error': 'id_classe requis'})

            # Resolve EAC.id → business keys
            bk = _resolve_eac_orm(id_classe)
            if not bk:
                return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

            from MonEcole_app.models.horaire import Horaire, Horaire_type
            from MonEcole_app.models.enseignmnts.matiere import Cours_par_classe

            qs = Horaire.objects.filter(
                id_etablissement=etab_id,
                id_classe_id=bk['classe_id'],
                groupe=bk['groupe'],
                section=bk['section_id'],
                id_pays=id_pays,
            )
            if date_debut and date_fin:
                qs = qs.filter(date__gte=date_debut, date__lte=date_fin)

            # Get all horaires WITHOUT cross-DB joins
            horaire_list = list(qs.order_by('date', 'debut').values(
                'id_horaire', 'date', 'debut', 'fin', 'id_cours_id', 'id_horaire_type_id'
            ))

            # Resolve course names from Hub
            cours_ids = set(h['id_cours_id'] for h in horaire_list if h['id_cours_id'])
            cours_map = {}
            if cours_ids:
                for ca in Cours_par_classe.objects.filter(id_cours_classe__in=cours_ids).select_related('id_cours'):
                    try:
                        cours_map[ca.id_cours_classe] = ca.id_cours.cours if ca.id_cours else str(ca.id_cours_classe)
                    except:
                        cours_map[ca.id_cours_classe] = str(ca.id_cours_classe)

            horaires = []
            for h in horaire_list:
                horaires.append({
                    'id': h['id_horaire'],
                    'date': str(h['date']),
                    'debut': h['debut'],
                    'fin': h['fin'],
                    'cours_id': h['id_cours_id'],
                    'cours_nom': cours_map.get(h['id_cours_id'], f"Cours #{h['id_cours_id']}"),
                    'type_id': h['id_horaire_type_id'],
                })

            etab, err = _get_tenant_etab(request)
            if err: return err
            types = list(Horaire_type.objects.filter(id_pays=etab.pays_id).values('id_horaire_type', 'horaire_type'))
            return JsonResponse({'success': True, 'horaires': horaires, 'types': types})

        elif action == 'save':
            from MonEcole_app.models.horaire import Horaire, Horaire_type
            from MonEcole_app.models.enseignmnts.matiere import Attribution_cours

            id_horaire = data.get('id_horaire')
            id_classe = data.get('id_classe')
            id_cours = data.get('id_cours')
            date_val = data.get('date')
            debut = data.get('debut')
            fin = data.get('fin')
            type_id = data.get('type_id', 1)

            if not all([id_classe, id_cours, date_val, debut, fin]):
                return JsonResponse({'success': False, 'error': 'Tous les champs sont requis'})

            # Resolve EAC.id → business keys (Spoke only)
            bk = _resolve_eac_orm(id_classe)
            if not bk:
                return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

            overlap = Horaire.objects.filter(
                id_etablissement=etab_id,
                id_classe_id=bk['classe_id'],
                groupe=bk['groupe'],
                section=bk['section_id'],
                date=date_val,
            ).exclude(id_horaire=id_horaire if id_horaire else 0)

            for existing in overlap:
                if not (fin <= existing.debut or debut >= existing.fin):
                    return JsonResponse({
                        'success': False,
                        'error': f'Chevauchement avec un créneau existant ({existing.debut}-{existing.fin})'
                    })

            attr = Attribution_cours.objects.filter(
                id_cours_id=id_cours,
                id_classe_id=bk['classe_id'],
                groupe=bk['groupe'],
                section_id=bk['section_id'],
                id_etablissement=etab_id,
            ).first()

            if not attr:
                return JsonResponse({'success': False, 'error': "Pas d'attribution trouvée pour ce cours/classe"})

            if id_horaire:
                h = Horaire.objects.get(id_horaire=id_horaire, id_etablissement=etab_id)
                h.date = date_val
                h.debut = debut
                h.fin = fin
                h.id_cours_id = id_cours
                h.id_horaire_type_id = type_id
                h.save()
            else:
                Horaire.objects.create(
                    id_etablissement=etab_id,
                    id_classe_id=bk['classe_id'],
                    groupe=bk['groupe'],
                    section=bk['section_id'],
                    id_cours_id=id_cours,
                    id_annee_id=attr.id_annee_id,
                    idCampus_id=attr.idCampus_id,
                    id_cycle_id=attr.id_cycle_id,
                    id_horaire_type_id=type_id,
                    date=date_val,
                    debut=debut,
                    fin=fin,
                    id_pays=id_pays,
                )

            return JsonResponse({'success': True})

        elif action == 'delete':
            from MonEcole_app.models.horaire import Horaire

            id_horaire = data.get('id_horaire')
            if not id_horaire:
                return JsonResponse({'success': False, 'error': 'id_horaire requis'})

            Horaire.objects.filter(id_horaire=id_horaire, id_etablissement=etab_id).delete()
            return JsonResponse({'success': True})

        elif action == 'get-cours':
            from MonEcole_app.models.enseignmnts.matiere import Attribution_cours, Cours_par_classe, Cours
            from MonEcole_app.models.personnel import Personnel

            id_classe = data.get('id_classe')
            if not id_classe:
                return JsonResponse({'success': False, 'error': 'id_classe requis'})

            # Resolve EAC.id → business keys (Spoke only)
            bk = _resolve_eac_orm(id_classe)
            if not bk:
                return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

            # Ne PAS faire select_related cross-database (spoke→hub)
            attrs = Attribution_cours.objects.filter(
                id_classe_id=bk['classe_id'],
                groupe=bk['groupe'],
                section_id=bk['section_id'],
            ).values('id_cours_id', 'id_personnel_id')

            # Collecter les IDs
            cours_annee_ids = set()
            personnel_ids = set()
            for a in attrs:
                cours_annee_ids.add(a['id_cours_id'])
                personnel_ids.add(a['id_personnel_id'])

            # Charger les noms depuis le Hub (cours_annee → cours) — Hub inchangé
            cours_map = {}
            if cours_annee_ids:
                for ca in Cours_par_classe.objects.filter(id_cours_classe__in=cours_annee_ids).select_related('id_cours'):
                    try:
                        cours_map[ca.id_cours_classe] = ca.id_cours.cours if ca.id_cours else str(ca.id_cours_classe)
                    except:
                        cours_map[ca.id_cours_classe] = str(ca.id_cours_classe)

            # Charger les noms depuis le Spoke (personnel)
            pers_map = {}
            if personnel_ids:
                for p in Personnel.objects.filter(id_personnel__in=personnel_ids):
                    try:
                        pers_map[p.id_personnel] = f"{p.prenom or ''} {p.nom or ''} ({p.matricule})"
                    except:
                        pers_map[p.id_personnel] = f"Personnel #{p.id_personnel}"

            # Construire la liste
            cours_list = []
            for a in Attribution_cours.objects.filter(
                id_classe_id=bk['classe_id'],
                groupe=bk['groupe'],
                section_id=bk['section_id'],
            ).values('id_cours_id', 'id_personnel_id'):
                cours_list.append({
                    'cours_annee_id': a['id_cours_id'],
                    'cours_nom': cours_map.get(a['id_cours_id'], f"Cours #{a['id_cours_id']}"),
                    'enseignant': pers_map.get(a['id_personnel_id'], ''),
                })

            return JsonResponse({'success': True, 'cours': cours_list})

        elif action == 'save-bulk':
            # Sauvegarder un créneau sur plusieurs dates (toute la semaine)
            from MonEcole_app.models.horaire import Horaire
            from MonEcole_app.models.enseignmnts.matiere import Attribution_cours

            id_classe = data.get('id_classe')
            id_cours = data.get('id_cours')
            dates = data.get('dates', [])
            debut = data.get('debut')
            fin = data.get('fin')
            type_id = data.get('type_id', 1)

            if not all([id_classe, id_cours, dates, debut, fin]):
                return JsonResponse({'success': False, 'error': 'Tous les champs sont requis'})

            # Resolve EAC.id → business keys (Spoke only)
            bk = _resolve_eac_orm(id_classe)
            if not bk:
                return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

            attr = Attribution_cours.objects.filter(
                id_cours_id=id_cours,
                id_classe_id=bk['classe_id'],
                groupe=bk['groupe'],
                section_id=bk['section_id'],
            ).first()
            if not attr:
                return JsonResponse({'success': False, 'error': "Pas d'attribution trouvée"})

            created = 0
            skipped = 0
            for date_val in dates:
                # Vérifier chevauchement
                overlap = Horaire.objects.filter(
                    id_etablissement=etab_id,
                    id_classe_id=bk['classe_id'],
                    groupe=bk['groupe'],
                    section=bk['section_id'],
                    date=date_val,
                )
                has_overlap = False
                for existing in overlap:
                    if not (fin <= existing.debut or debut >= existing.fin):
                        has_overlap = True
                        break
                if has_overlap:
                    skipped += 1
                    continue

                Horaire.objects.create(
                    id_etablissement=etab_id,
                    id_classe_id=bk['classe_id'],
                    groupe=bk['groupe'],
                    section=bk['section_id'],
                    id_cours_id=id_cours,
                    id_annee_id=attr.id_annee_id,
                    idCampus_id=attr.idCampus_id,
                    id_cycle_id=attr.id_cycle_id,
                    id_horaire_type_id=type_id,
                    date=date_val,
                    debut=debut,
                    fin=fin,
                )
                created += 1

            msg = f'{created} créneau(x) créé(s)'
            if skipped:
                msg += f', {skipped} ignoré(s) (chevauchement)'
            return JsonResponse({'success': True, 'message': msg, 'created': created, 'skipped': skipped})

        elif action == 'copy-week':
            # Copier les horaires d'une semaine source vers une semaine cible
            from MonEcole_app.models.horaire import Horaire
            from datetime import datetime, timedelta

            id_classe = data.get('id_classe')
            source_dates = data.get('source_dates', [])
            target_dates = data.get('target_dates', [])

            if not id_classe or not source_dates or not target_dates:
                return JsonResponse({'success': False, 'error': 'Données insuffisantes'})

            if len(source_dates) != len(target_dates):
                return JsonResponse({'success': False, 'error': 'Nombre de jours incompatible'})

            # Resolve EAC.id → business keys (Spoke only)
            bk = _resolve_eac_orm(id_classe)
            if not bk:
                return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

            # Map source day -> target day
            day_map = {}
            for i, sd in enumerate(source_dates):
                if i < len(target_dates):
                    day_map[sd] = target_dates[i]

            source_horaires = Horaire.objects.filter(
                id_etablissement=etab_id,
                id_classe_id=bk['classe_id'],
                groupe=bk['groupe'],
                section=bk['section_id'],
                date__in=source_dates,
            )

            created = 0
            skipped = 0
            for h in source_horaires:
                target_date = day_map.get(str(h.date))
                if not target_date:
                    continue

                # Vérifier si déjà existant
                exists = Horaire.objects.filter(
                    id_etablissement=etab_id,
                    id_classe_id=bk['classe_id'],
                    groupe=bk['groupe'],
                    section=bk['section_id'],
                    date=target_date,
                    debut=h.debut,
                    fin=h.fin,
                ).exists()
                if exists:
                    skipped += 1
                    continue

                Horaire.objects.create(
                    id_etablissement=etab_id,
                    id_classe_id=bk['classe_id'],
                    groupe=bk['groupe'],
                    section=bk['section_id'],
                    id_cours_id=h.id_cours_id,
                    id_annee_id=h.id_annee_id,
                    idCampus_id=h.idCampus_id,
                    id_cycle_id=h.id_cycle_id,
                    id_horaire_type_id=h.id_horaire_type_id,
                    date=target_date,
                    debut=h.debut,
                    fin=h.fin,
                )
                created += 1

            msg = f'{created} créneau(x) copié(s)'
            if skipped:
                msg += f', {skipped} ignoré(s) (déjà existant)'
            return JsonResponse({'success': True, 'message': msg, 'created': created, 'skipped': skipped})

        else:
            return JsonResponse({'success': False, 'error': f'Action inconnue: {action}'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# TRANSFER ELEVES BETWEEN CLASSES
# ============================================================
@require_http_methods(["POST"])
def dashboard_transfer_eleves(request):
    """Transfer students from one class to another by updating their inscription."""
    try:
        data = json.loads(request.body)
        eleves = data.get('eleves', [])
        dest_classe_id = data.get('dest_classe_id')
        if not eleves or not dest_classe_id:
            return JsonResponse({'success': False, 'error': 'Paramètres manquants'}, status=400)

        conn = _get_spoke_connection()
        try:
            transferred = 0
            with conn.cursor() as cur:
                # Resolve destination EAC.id → business keys
                cur.execute("""
                    SELECT eac.classe_id, eac.groupe, eac.section_id, cl.cycle_id
                    FROM countryStructure.etablissements_annees_classes eac
                    JOIN countryStructure.classes cl ON cl.id = eac.classe_id
                    WHERE eac.id = %s
                """, [dest_classe_id])
                bk = cur.fetchone()
                if not bk:
                    return JsonResponse({'success': False, 'error': 'Classe destination introuvable'}, status=404)

                for e in eleves:
                    insc_id = e.get('id_inscription')
                    if insc_id:
                        cur.execute("""
                            UPDATE eleve_inscription
                            SET classe_id=%s, groupe=%s, section_id=%s, id_cycle_id=%s
                            WHERE id_inscription=%s
                        """, [bk['classe_id'], bk['groupe'], bk['section_id'], bk['cycle_id'], insc_id])
                        transferred += cur.rowcount
                conn.commit()
            return JsonResponse({'success': True, 'transferred': transferred})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# DOSSIER ADMINISTRATIF — Document Types CRUD
# ============================================================
@require_http_methods(["GET", "POST", "DELETE"])
def document_types_api(request):
    """CRUD for document types (per establishment)."""
    try:
        conn = _get_spoke_connection()
        try:
            if request.method == 'GET':
                id_etablissement = request.GET.get('id_etablissement')
                with conn.cursor() as cur:
                    cur.execute("SELECT * FROM document_type WHERE id_etablissement=%s ORDER BY nom", [id_etablissement])
                    rows = cur.fetchall()
                return JsonResponse({'success': True, 'types': [
                    {'id': r['id'], 'nom': r['nom'], 'description': r.get('description') or '',
                     'isObligatoire': bool(r.get('isObligatoire', 0))}
                    for r in rows
                ]})

            elif request.method == 'POST':
                data = json.loads(request.body)
                nom = data.get('nom', '').strip()
                if not nom:
                    return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
                with conn.cursor() as cur:
                    cur.execute(
                        "INSERT INTO document_type (nom, description, isObligatoire, id_etablissement) VALUES (%s,%s,%s,%s)",
                        [nom, data.get('description', ''), int(data.get('isObligatoire', 0)), data.get('id_etablissement')]
                    )
                    conn.commit()
                return JsonResponse({'success': True, 'id': cur.lastrowid})

            elif request.method == 'DELETE':
                data = json.loads(request.body)
                doc_id = data.get('id')
                with conn.cursor() as cur:
                    # Also delete associated student documents
                    cur.execute("DELETE FROM document_eleve WHERE id_document_type=%s", [doc_id])
                    cur.execute("DELETE FROM document_type WHERE id=%s", [doc_id])
                    conn.commit()
                return JsonResponse({'success': True})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# DOSSIER ADMINISTRATIF — Eleve Documents CRUD
# ============================================================
@require_http_methods(["GET", "POST", "DELETE"])
def eleve_documents_api(request):
    """Upload/list/delete student documents."""
    try:
        conn = _get_spoke_connection()
        try:
            if request.method == 'GET':
                id_eleve = request.GET.get('id_eleve')
                id_etablissement = request.GET.get('id_etablissement')
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT * FROM document_eleve WHERE id_eleve=%s AND id_etablissement=%s",
                        [id_eleve, id_etablissement]
                    )
                    rows = cur.fetchall()
                return JsonResponse({'success': True, 'documents': [
                    {'id': r['id'], 'id_eleve': r['id_eleve'], 'id_document_type': r['id_document_type'],
                     'file_url': r['file_url'], 'date_upload': str(r['date_upload']) if r.get('date_upload') else ''}
                    for r in rows
                ]})

            elif request.method == 'POST':
                file = request.FILES.get('file')
                id_eleve = request.POST.get('id_eleve')
                id_document_type = request.POST.get('id_document_type')
                id_etablissement = request.POST.get('id_etablissement')
                if not file or not id_eleve or not id_document_type or not id_etablissement:
                    return JsonResponse({'success': False, 'error': 'Paramètres manquants'}, status=400)

                import os
                from django.conf import settings
                import datetime

                # Directory: media/Documents/EtabID_x/Eleve_y/
                doc_dir = os.path.join(settings.MEDIA_ROOT, 'Documents', f'EtabID_{id_etablissement}', f'Eleve_{id_eleve}')
                os.makedirs(doc_dir, exist_ok=True)

                ext = os.path.splitext(file.name)[1].lower() or '.pdf'
                filename = f'Doc_{id_document_type}_{id_eleve}{ext}'
                filepath = os.path.join(doc_dir, filename)

                # Remove old file if exists with different extension
                for old_ext in ['.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx', '.webp']:
                    old_file = os.path.join(doc_dir, f'Doc_{id_document_type}_{id_eleve}{old_ext}')
                    if os.path.exists(old_file) and old_file != filepath:
                        try:
                            os.remove(old_file)
                        except Exception:
                            pass

                with open(filepath, 'wb+') as f:
                    for chunk in file.chunks():
                        f.write(chunk)

                file_url = f'/media/Documents/EtabID_{id_etablissement}/Eleve_{id_eleve}/{filename}'
                today = datetime.date.today().strftime('%Y-%m-%d')

                with conn.cursor() as cur:
                    # Upsert: delete old then insert
                    cur.execute("DELETE FROM document_eleve WHERE id_eleve=%s AND id_document_type=%s AND id_etablissement=%s",
                                [id_eleve, id_document_type, id_etablissement])
                    cur.execute(
                        "INSERT INTO document_eleve (id_eleve, id_document_type, file_url, date_upload, id_etablissement) VALUES (%s,%s,%s,%s,%s)",
                        [id_eleve, id_document_type, file_url, today, id_etablissement]
                    )
                    conn.commit()
                return JsonResponse({'success': True, 'file_url': file_url})

            elif request.method == 'DELETE':
                data = json.loads(request.body)
                doc_id = data.get('id')
                with conn.cursor() as cur:
                    # Get file path to delete physical file
                    cur.execute("SELECT file_url FROM document_eleve WHERE id=%s", [doc_id])
                    row = cur.fetchone()
                    if row and row.get('file_url'):
                        import os
                        from django.conf import settings
                        filepath = os.path.join(settings.MEDIA_ROOT, row['file_url'].lstrip('/media/'))
                        if os.path.exists(filepath):
                            try:
                                os.remove(filepath)
                            except Exception:
                                pass
                    cur.execute("DELETE FROM document_eleve WHERE id=%s", [doc_id])
                    conn.commit()
                return JsonResponse({'success': True})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# GESTION DES UTILISATEURS & DROITS
# ============================================================

def _get_admin_email_for_etab(etab_id, pays_id=None):
    """Récupère l'admin_email depuis le Hub pour l'établissement (scoped par pays_id)."""
    try:
        with connections['countryStructure'].cursor() as cur:
            sql = "SELECT admin_email FROM etablissements WHERE id_etablissement=%s"
            params = [etab_id]
            if pays_id:
                sql += " AND pays_id=%s"
                params.append(pays_id)
            cur.execute(sql, params)
            row = cur.fetchone()
            return row[0] if row else None
    except Exception:
        return None


@csrf_exempt
@require_http_methods(["GET"])
def dashboard_users_list(request):
    """
    Liste des utilisateurs (personnel) avec leurs modules assignés.
    GET /api/dashboard/users/
    """
    try:
        etab_id = getattr(request, 'id_etablissement', None) or request.session.get('id_etablissement')
        if not etab_id:
            return JsonResponse({'success': False, 'error': 'Pas de tenant'}, status=400)
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')

        admin_email = _get_admin_email_for_etab(etab_id, pays_id=id_pays)

        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                # Tous les personnels en fonction pour cet établissement (filtré par pays)
                cur.execute("""
                    SELECT p.id_personnel, p.email, p.prenom as first_name, p.nom as last_name,
                           p.matricule, p.isUser, p.is_verified, p.en_fonction,
                           p.telephone, p.id_personnel as personnel_id
                    FROM personnel p
                    WHERE p.id_etablissement = %s AND p.id_pays = %s
                    ORDER BY p.nom, p.prenom
                """, [etab_id, id_pays])
                personnels = cur.fetchall()

                # Tous les modules
                cur.execute("SELECT id_module, module FROM module ORDER BY id_module")
                modules = cur.fetchall()

                # User_module actifs pour cet établissement
                cur.execute("""
                    SELECT um.user_id, um.module_id, um.is_active
                    FROM user_module um
                    WHERE um.id_etablissement = %s
                """, [etab_id])
                user_modules_raw = cur.fetchall()

            # Indexer les modules par personnel
            um_map = {}  # {personnel_id: {module_id: is_active}}
            for um in user_modules_raw:
                pid = um['user_id']
                mid = um['module_id']
                if pid not in um_map:
                    um_map[pid] = {}
                um_map[pid][mid] = bool(um['is_active'])

            users = []
            for p in personnels:
                email = p['email'] or ''
                is_super = admin_email and email.lower() == admin_email.lower()
                pid = p['id_personnel']
                user_mods = um_map.get(pid, {})
                users.append({
                    'id_personnel': pid,
                    'personnel_id': pid,
                    'email': email,
                    'first_name': p['first_name'] or '',
                    'last_name': p['last_name'] or '',
                    'matricule': p['matricule'] or '',
                    'telephone': str(p['telephone'] or ''),
                    'isUser': bool(p['isUser']),
                    'is_verified': bool(p['is_verified']),
                    'en_fonction': bool(p['en_fonction']),
                    'is_super_admin': is_super,
                    'modules': user_mods,
                })

            modules_list = [{'id': m['id_module'], 'name': m['module']} for m in modules]

            return JsonResponse({
                'success': True,
                'users': users,
                'modules': modules_list,
                'admin_email': admin_email or '',
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def dashboard_users_toggle(request):
    """
    Active/désactive un utilisateur (isUser, is_verified).
    POST /api/dashboard/users/toggle/
    Body: {id_personnel, isUser:bool, is_verified:bool}
    """
    try:
        etab_id = getattr(request, 'id_etablissement', None) or request.session.get('id_etablissement')
        if not etab_id:
            return JsonResponse({'success': False, 'error': 'Pas de tenant'}, status=400)
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')

        data = json.loads(request.body)
        pid = data.get('id_personnel')
        is_user = data.get('isUser')
        is_verified = data.get('is_verified')

        if not pid:
            return JsonResponse({'success': False, 'error': 'id_personnel requis'}, status=400)

        # Vérifier que ce n'est pas le super admin
        admin_email = _get_admin_email_for_etab(etab_id, pays_id=id_pays)
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT p.email FROM personnel p
                    WHERE p.id_personnel = %s AND p.id_pays = %s
                """, [pid, id_pays])
                row = cur.fetchone()
                if row and admin_email and row['email'].lower() == admin_email.lower():
                    return JsonResponse({
                        'success': False,
                        'error': 'Impossible de modifier le super administrateur'
                    }, status=403)

                updates = []
                params = []
                if is_user is not None:
                    updates.append("isUser = %s")
                    params.append(1 if is_user else 0)
                if is_verified is not None:
                    updates.append("is_verified = %s")
                    params.append(1 if is_verified else 0)

                if updates:
                    params.extend([pid, id_pays])
                    cur.execute(f"UPDATE personnel SET {', '.join(updates)} WHERE id_personnel = %s AND id_pays = %s", params)
                    conn.commit()

            return JsonResponse({'success': True})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def dashboard_users_modules(request):
    """
    Assigne les modules à un utilisateur.
    POST /api/dashboard/users/modules/
    Body: {id_personnel, modules: {module_id: bool, ...}}
    """
    try:
        etab_id = getattr(request, 'id_etablissement', None) or request.session.get('id_etablissement')
        if not etab_id:
            return JsonResponse({'success': False, 'error': 'Pas de tenant'}, status=400)
        id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')

        data = json.loads(request.body)
        pid = data.get('id_personnel')
        modules_map = data.get('modules', {})

        if not pid:
            return JsonResponse({'success': False, 'error': 'id_personnel requis'}, status=400)

        # Vérifier que ce n'est pas le super admin
        admin_email = _get_admin_email_for_etab(etab_id, pays_id=id_pays)
        conn = _get_spoke_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT p.email FROM personnel p
                    WHERE p.id_personnel = %s AND p.id_pays = %s
                """, [pid, id_pays])
                row = cur.fetchone()
                if row and admin_email and row['email'].lower() == admin_email.lower():
                    return JsonResponse({
                        'success': False,
                        'error': 'Les droits du super administrateur ne peuvent pas être modifiés'
                    }, status=403)

                # Récupérer l'année active depuis le Hub (table annees dans countryStructure)
                from django.db import connections as db_connections
                with db_connections['countryStructure'].cursor() as hub_cur:
                    hub_cur.execute("""
                        SELECT id_annee FROM annees
                        WHERE isOpen = 1 AND pays_id = %s
                        ORDER BY annee DESC LIMIT 1
                    """, [id_pays])
                    hub_row = hub_cur.fetchone()
                    annee_id = hub_row[0] if hub_row else 1

                for mod_id_str, is_active in modules_map.items():
                    mod_id = int(mod_id_str)
                    # Check if user_module exists
                    cur.execute("""
                        SELECT id_user_module FROM user_module
                        WHERE user_id=%s AND module_id=%s AND id_etablissement=%s
                    """, [pid, mod_id, etab_id])
                    existing = cur.fetchone()

                    if existing:
                        cur.execute("""
                            UPDATE user_module SET is_active=%s
                            WHERE id_user_module=%s
                        """, [1 if is_active else 0, existing['id_user_module']])
                    elif is_active:
                        cur.execute("""
                            INSERT INTO user_module (id_annee_id, user_id, module_id, is_active, id_etablissement, date_creation)
                            VALUES (%s, %s, %s, 1, %s, CURDATE())
                        """, [annee_id, pid, mod_id, etab_id])

                conn.commit()

                # Activer automatiquement isUser + is_verified si au moins un module actif
                cur.execute("""
                    SELECT COUNT(*) as cnt FROM user_module
                    WHERE user_id=%s AND id_etablissement=%s AND is_active=1
                """, [pid, etab_id])
                cnt = cur.fetchone()['cnt']
                if cnt > 0:
                    cur.execute("""
                        UPDATE personnel SET isUser=1, is_verified=1
                        WHERE id_personnel=%s AND id_pays=%s
                    """, [pid, id_pays])
                    conn.commit()

            return JsonResponse({'success': True})
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# DÉLIBÉRATIONS API
# =============================================================================

@require_http_methods(["GET"])
def get_evaluations_sessions(request):
    """Retourne la liste des sessions pour le dropdown délibérations."""
    try:
        etab, err = _get_tenant_etab(request)
        if err: return err
        sessions = list(Session.objects.filter(is_active=True, id_pays=etab.pays_id).values('id_session', 'session'))
        return JsonResponse({'success': True, 'sessions': sessions})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_evaluations_repartitions(request):
    """Retourne les répartitions configurées pour une classe, filtrées par cycle.
    
    Logique de filtrage:
      1. Trouve le cycle de la classe sélectionnée (via EAC → classe → cycle)
      2. Trouve le type racine du cycle (Trimestre ou Semestre) via RepartitionConfigCycle
      3. Trouve les types enfants via RepartitionHierarchie (généralement Période)
      4. Filtre les configs pour ne garder que les répartitions dont le type
         correspond au type racine ou ses enfants
      5. Limite le nombre de répartitions au nombre prévu par la config cycle × hiérarchie
    """
    try:
        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        if not classe_id:
            return JsonResponse({'success': False, 'error': 'classe_id requis'}, status=400)

        etab, err = _get_tenant_etab(request)
        if err:
            return err

        annee = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        if not annee:
            return JsonResponse({'success': True, 'repartitions': []})

        ea = EtablissementAnnee.objects.filter(
            etablissement_id=etab.id_etablissement,
            annee=annee
        ).first()
        if not ea:
            return JsonResponse({'success': True, 'repartitions': []})

        # Résoudre le cycle de la classe sélectionnée
        eac = EtablissementAnneeClasse.objects.filter(id=classe_id).select_related('classe').first()
        cycle_id = None
        if eac and hasattr(eac.classe, 'cycle_id'):
            cycle_id = eac.classe.cycle_id
        elif eac and hasattr(eac, 'cycle_id'):
            cycle_id = eac.cycle_id

        # Déterminer les types de répartition autorisés pour ce cycle
        allowed_type_ids = set()
        root_count = 0  # nombre de racines (ex: 3 trimestres, 2 semestres)
        child_count_per_root = 0  # nombre d'enfants par racine (ex: 2 périodes/trimestre)

        if cycle_id:
            # Config cycle → type racine (ex: Cycle "Ecole de Base" → Semestre, nombre=2)
            cycle_config = RepartitionConfigCycle.objects.filter(
                cycle_id=cycle_id, is_active=True, id_pays=etab.pays_id
            ).select_related('type_racine').first()

            if cycle_config:
                root_type_id = cycle_config.type_racine_id
                root_count = cycle_config.nombre_au_niveau_racine
                allowed_type_ids.add(root_type_id)

                # Hiérarchie → types enfants (ex: Semestre → Période, nombre=2)
                hierarchies = RepartitionHierarchie.objects.filter(
                    type_parent_id=root_type_id, is_active=True, id_pays=etab.pays_id
                )
                for h in hierarchies:
                    allowed_type_ids.add(h.type_enfant_id)
                    child_count_per_root = h.nombre_enfants

        # Charger toutes les configs pour cet établissement/année
        configs = RepartitionConfigEtabAnnee.objects.filter(
            etablissement_annee=ea
        ).select_related('repartition', 'repartition__type')

        # Filtrer par types autorisés et limiter le nombre
        repartitions = []
        count_by_type = {}  # {type_id: count} pour limiter

        for cfg in configs:
            rep = cfg.repartition
            type_id = rep.type_id if rep.type else None

            # Si on a un cycle défini, filtrer par types autorisés
            if allowed_type_ids and type_id not in allowed_type_ids:
                continue

            # Limiter le nombre par type
            if type_id not in count_by_type:
                count_by_type[type_id] = 0

            # Calculer la limite pour ce type
            if cycle_id and allowed_type_ids:
                is_root_type = (type_id == cycle_config.type_racine_id) if cycle_config else False
                max_count = root_count if is_root_type else (root_count * child_count_per_root)
                if count_by_type[type_id] >= max_count:
                    continue

            count_by_type[type_id] = count_by_type.get(type_id, 0) + 1

            repartitions.append({
                'id': cfg.id,
                'repartition_id': rep.id_instance,
                'nom': rep.nom,
                'code': rep.code,
                'type_code': rep.type.code if rep.type else '',
                'type_nom': rep.type.nom if rep.type else '',
                'is_open': cfg.is_open,
                'ordre': rep.ordre,
            })

        # Trier par type puis par ordre
        repartitions.sort(key=lambda r: (r['type_code'], r['ordre']))

        return JsonResponse({'success': True, 'repartitions': repartitions})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_deliberation_conditions(request):
    """
    Retourne les conditions de délibération annuelle depuis le Hub
    pour une classe donnée (EtablissementAnneeClasse).
    """
    try:
        from MonEcole_app.models.evaluations.note import (
            Deliberation_annuelle_condition, Deliberation_annuelle_finalite
        )
        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        if not classe_id:
            return JsonResponse({'success': False, 'error': 'classe_id requis'}, status=400)

        etab, err = _get_tenant_etab(request)
        if err:
            return err

        annee = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        if not annee:
            return JsonResponse({'success': True, 'conditions': []})

        # Get the EtablissementAnneeClasse to resolve classe and cycle
        eac = EtablissementAnneeClasse.objects.filter(id=classe_id).first()
        if not eac:
            return JsonResponse({'success': True, 'conditions': []})

        # Query conditions from Hub — nationales, filtrées par pays
        conditions = Deliberation_annuelle_condition.objects.filter(
            id_annee=annee, id_pays=etab.pays_id
        )

        # Fetch mentions and finalites for display
        mention_map = {m.id_mention: str(m) for m in Mention.objects.filter(id_pays=etab.pays_id)}
        finalite_map = {}
        try:
            finalite_map = {f.id_finalite: f.finalite for f in Deliberation_annuelle_finalite.objects.filter(id_pays=etab.pays_id)}
        except Exception:
            pass

        result = []
        for c in conditions:
            result.append({
                'id_decision': c.id_decision,
                'mention': mention_map.get(c.id_mention_id, f'Mention {c.id_mention_id}') if hasattr(c, 'id_mention') and c.id_mention_id else '—',
                'max_echecs': c.max_echecs_acceptable,
                'seuil_profondeur': c.seuil_profondeur_echec,
                'sanction': c.sanction_disciplinaire or '',
                'finalite': finalite_map.get(c.id_finalite_id, f'Finalité {c.id_finalite_id}') if hasattr(c, 'id_finalite') and c.id_finalite_id else '—',
            })

        return JsonResponse({'success': True, 'conditions': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def execute_deliberation(request):
    """
    Exécute la délibération pour une classe et un type donné.
    Types: 'periode', 'trimestre', 'annee', 'repechage'
    
    Algorithme:
      1. Récupère les élèves inscrits dans la CLASSE.
      2. Pour chaque élève, calcule total obtenu / total maxima
         sur la période/trimestre/année sélectionnée.
      3. Calcule le pourcentage, détermine la mention.
      4. Classe les élèves par pourcentage décroissant.
      5. Sauvegarde les résultats dans la table appropriée.
    """
    try:
        from MonEcole_app.models.evaluations.note import (
            Deliberation_annuelle_condition,
            Deliberation_annuelle_resultat,
            Deliberation_periodique_resultat,
            Deliberation_trimistrielle_resultat,
            Deliberation_examen_resultat,
            Deliberation_repechage_resultat,
            Deliberation_annuelle_finalite,
        )
        from MonEcole_app.models.eleves.eleve import Eleve, Eleve_inscription
        from MonEcole_app.models.campus import Campus

        data = json.loads(request.body)
        classe_id = data.get('classe_id') or data.get('id_classe_id')
        delib_type = data.get('type')  # 'periode', 'trimestre', 'annee', 'repechage'
        repartition_id = data.get('repartition_id')
        session_id = data.get('session_id')

        if not classe_id or not delib_type:
            return JsonResponse({'success': False, 'error': 'classe_id et type requis'}, status=400)

        etab, err = _get_tenant_etab(request)
        if err:
            return err

        annee = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        if not annee:
            return JsonResponse({'success': False, 'error': 'Aucune année en cours.'}, status=400)

        # Resolve EAC
        eac = EtablissementAnneeClasse.objects.filter(id=classe_id).first()
        if not eac:
            return JsonResponse({'success': False, 'error': 'Classe non trouvée.'}, status=404)

        # Get enrolled students for THIS CLASS (via business keys)
        from django.db import connection
        with connection.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT e.id_eleve, e.nom, e.prenom, e.genre
                FROM eleve_inscription ei
                JOIN eleve e ON ei.id_eleve_id = e.id_eleve
                WHERE ei.id_annee_id = %s
                  AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                  AND ei.id_etablissement = %s
                  AND ei.status = 1
                ORDER BY e.nom, e.prenom
            """, [annee.pk, eac.classe_id, eac.groupe, eac.section_id, etab.id_etablissement])
            columns = [col[0] for col in cur.description]
            eleves = [dict(zip(columns, row)) for row in cur.fetchall()]

        if not eleves:
            return JsonResponse({'success': False, 'error': 'Aucun élève inscrit dans cette classe.'}, status=400)

        # ─── Build the list of repartition_config IDs and note_type based on delib_type ───
        from django.db import connections
        etab_annee_id = eac.etablissement_annee_id

        # Determine which config_ids and note_types to query based on delib_type
        config_ids = []
        note_types = []

        if delib_type == 'periode' and repartition_id:
            # Period deliberation: sum TJ (type=1) for this single period config
            config_ids = [int(repartition_id)]
            note_types = [1]  # TJ only

        elif delib_type == 'trimestre' and repartition_id:
            # Trimester/Semester deliberation: 
            # Find ALL child period configs under this parent config
            # Sum TJ (type=1) of all child periods + EX (type=2) on parent
            with connections['countryStructure'].cursor() as hub_cur:
                # Get parent type
                hub_cur.execute("""
                    SELECT r.type_id FROM repartition_configs_etab_annee rc
                    JOIN repartition_instances r ON r.id = rc.repartition_id
                    WHERE rc.id = %s
                """, [int(repartition_id)])
                parent_row = hub_cur.fetchone()
                if parent_row:
                    parent_type_id = parent_row[0]
                    # Get child type
                    hub_cur.execute("""
                        SELECT type_enfant_id FROM repartition_hierarchies
                        WHERE type_parent_id = %s AND is_active = 1 LIMIT 1
                    """, [parent_type_id])
                    child_type_row = hub_cur.fetchone()
                    if child_type_row:
                        child_type_id = child_type_row[0]
                        # Get ALL child configs for this etab_annee
                        hub_cur.execute("""
                            SELECT rc.id FROM repartition_configs_etab_annee rc
                            JOIN repartition_instances r ON r.id = rc.repartition_id
                            WHERE rc.etablissement_annee_id = %s AND r.type_id = %s
                            ORDER BY r.ordre
                        """, [etab_annee_id, child_type_id])
                        all_child_ids = [r[0] for r in hub_cur.fetchall()]

                        # Get all parent configs to determine which children belong to this parent
                        hub_cur.execute("""
                            SELECT rc.id FROM repartition_configs_etab_annee rc
                            JOIN repartition_instances r ON r.id = rc.repartition_id
                            WHERE rc.etablissement_annee_id = %s AND r.type_id = %s
                            ORDER BY r.ordre
                        """, [etab_annee_id, parent_type_id])
                        all_parent_ids = [r[0] for r in hub_cur.fetchall()]

                        if all_parent_ids and all_child_ids:
                            kids_per_parent = len(all_child_ids) // len(all_parent_ids)
                            try:
                                pidx = all_parent_ids.index(int(repartition_id))
                            except ValueError:
                                pidx = 0
                            start = pidx * kids_per_parent
                            end = start + kids_per_parent
                            my_child_ids = all_child_ids[start:end]
                        else:
                            my_child_ids = all_child_ids

                        # Config IDs = child periods (TJ type=1) + parent (EX type=2)
                        config_ids = my_child_ids + [int(repartition_id)]
                        note_types = [1, 2]  # TJ from children + EX from parent

            if not config_ids:
                config_ids = [int(repartition_id)]
                note_types = [1, 2, 5]

        elif delib_type == 'examen' and repartition_id:
            # Exam deliberation: EX (type=2) on parent config
            config_ids = [int(repartition_id)]
            note_types = [2]

        elif delib_type == 'annee':
            # Annual deliberation: get ALL configs, sum all TJ + EX
            with connections['countryStructure'].cursor() as hub_cur:
                hub_cur.execute("""
                    SELECT rc.id FROM repartition_configs_etab_annee rc
                    WHERE rc.etablissement_annee_id = %s
                """, [etab_annee_id])
                config_ids = [r[0] for r in hub_cur.fetchall()]
            note_types = [1, 2]  # All TJ + EX

        else:
            return JsonResponse({'success': False, 'error': f'Type de délibération non supporté: {delib_type}'}, status=400)

        import sys
        print(f"[DELIB DEBUG] type={delib_type}, config_ids={config_ids}, note_types={note_types}", file=sys.stderr)

        # Calculate percentages for each student from note_bulletin
        # note_bulletin is already scoped by id_etablissement, no need for cours_annee filter
        cfg_placeholders = ','.join(['%s'] * len(config_ids))
        nt_placeholders = ','.join(['%s'] * len(note_types))

        resultats = []
        for eleve in eleves:
            id_eleve = eleve['id_eleve']

            with connection.cursor() as cur:
                cur.execute(f"""
                    SELECT 
                        COALESCE(SUM(nb.note), 0) as total_note,
                        COALESCE(SUM(nb.maxima), 0) as total_max
                    FROM note_bulletin nb
                    WHERE nb.id_eleve_id = %s
                      AND nb.id_etablissement = %s
                      AND nb.id_repartition_config IN ({cfg_placeholders})
                      AND nb.id_note_type IN ({nt_placeholders})
                """, [id_eleve, etab.id_etablissement] + config_ids + note_types)
                row = cur.fetchone()
                total_note = float(row[0]) if row and row[0] else 0
                total_max = float(row[1]) if row and row[1] else 0

            pourcentage = (total_note * 100 / total_max) if total_max > 0 else 0

            resultats.append({
                'id_eleve': id_eleve,
                'nom': eleve['nom'],
                'prenom': eleve['prenom'],
                'genre': eleve.get('genre', 'M'),
                'pourcentage': round(pourcentage, 2),
                'total_note': round(total_note, 1),
                'total_max': round(total_max, 1),
            })

        # Sort by percentage descending for ranking
        resultats.sort(key=lambda x: x['pourcentage'], reverse=True)

        # Assign places
        for i, r in enumerate(resultats):
            rank = i + 1
            total_eleves = len(resultats)
            r['place'] = f'{rank}/{total_eleves}'

        # Determine mention for each student
        mentions = list(Mention.objects.filter(id_pays=etab.pays_id))
        for r in resultats:
            pct = r['pourcentage']
            r['mention'] = '—'
            for m in mentions:
                if m.min <= pct <= m.max:
                    r['mention'] = str(m)
                    r['mention_id'] = m.id_mention
                    break

        # Determine decision from Hub conditions (nationales, filtrées par pays)
        conditions = list(Deliberation_annuelle_condition.objects.filter(
            id_annee=annee, id_pays=etab.pays_id
        ))

        for r in resultats:
            r['decision'] = '—'
            mention_id = r.get('mention_id')
            if mention_id and conditions:
                for cond in conditions:
                    if cond.id_mention_id == mention_id:
                        try:
                            finalite = Deliberation_annuelle_finalite.objects.get(
                                id_finalite=cond.id_finalite_id, id_pays=etab.pays_id
                            )
                            r['decision'] = finalite.finalite
                            r['finalite_id'] = finalite.id_finalite
                        except Deliberation_annuelle_finalite.DoesNotExist:
                            r['decision'] = cond.sanction_disciplinaire or '—'
                        break

        # Get campus for the establishment
        campus = Campus.objects.filter(id_etablissement=etab.id_etablissement).first()
        campus_id = campus.idCampus if campus else 1

        # Resolve cycle from EAC
        cycle_id = eac.cycle_id if hasattr(eac, 'cycle_id') and eac.cycle_id else 3

        # Save results to Spoke DB (depending on type)
        saved_count = 0

        if delib_type == 'periode' and repartition_id:
            # Save to deliberation_periodique_resultats
            # We need both id_trimestre and id_periode
            # For now, use repartition_id as both (they map to same repartition_configs)
            for r in resultats:
                try:
                    with connection.cursor() as cur:
                        cur.execute("""
                            INSERT INTO deliberation_periodique_resultats
                            (id_eleve_id, idCampus_id, id_annee_id, id_cycle_id, classe_id,
                             groupe, section_id,
                             id_trimestre_id, id_periode_id, pourcentage, place, date_creation, id_etablissement, id_pays)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURDATE(), %s, %s)
                            ON DUPLICATE KEY UPDATE
                                pourcentage=VALUES(pourcentage), place=VALUES(place)
                        """, [r['id_eleve'], campus_id, annee.pk, cycle_id,
                              eac.classe_id, eac.groupe, eac.section_id,
                              int(repartition_id), int(repartition_id),
                              r['pourcentage'], r['place'], etab.id_etablissement, etab.pays_id])
                    saved_count += 1
                except Exception as save_err:
                    logging.getLogger(__name__).error(f"Save error periode eleve {r['id_eleve']}: {save_err}")

        elif delib_type == 'trimestre' and repartition_id:
            # Save to deliberation_trimistrielle_resultats
            for r in resultats:
                try:
                    with connection.cursor() as cur:
                        cur.execute("""
                            INSERT INTO deliberation_trimistrielle_resultats
                            (id_eleve_id, idCampus_id, id_annee_id, id_cycle_id, classe_id,
                             groupe, section_id,
                             id_trimestre_id, pourcentage, place, date_creation, id_etablissement, id_pays)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURDATE(), %s, %s)
                            ON DUPLICATE KEY UPDATE
                                pourcentage=VALUES(pourcentage), place=VALUES(place)
                        """, [r['id_eleve'], campus_id, annee.pk, cycle_id,
                              eac.classe_id, eac.groupe, eac.section_id,
                              int(repartition_id),
                              r['pourcentage'], r['place'], etab.id_etablissement, etab.pays_id])
                    saved_count += 1
                except Exception as save_err:
                    logging.getLogger(__name__).error(f"Save error trimestre eleve {r['id_eleve']}: {save_err}")

        elif delib_type == 'annee' and session_id:
            for r in resultats:
                try:
                    Deliberation_annuelle_resultat.objects.update_or_create(
                        id_eleve_id=r['id_eleve'],
                        id_annee=annee,
                        id_etablissement=etab.id_etablissement,
                        defaults={
                            'idCampus_id': campus_id,
                            'id_cycle_id': cycle_id,
                            'id_classe_id': eac.classe_id,
                            'groupe': eac.groupe,
                            'section_id': eac.section_id,
                            'id_session_id': int(session_id),
                            'id_mention_id': r.get('mention_id', 1),
                            'id_decision_id': r.get('finalite_id', 1),
                            'pourcentage': r['pourcentage'],
                            'place': r['place'],
                            'id_pays': etab.pays_id,
                        }
                    )
                    saved_count += 1
                except Exception as save_err:
                    logging.getLogger(__name__).error(f"Save error annee eleve {r['id_eleve']}: {save_err}")

        return JsonResponse({
            'success': True,
            'message': f'Délibération {delib_type} effectuée pour {len(resultats)} élèves ({saved_count} résultats sauvegardés).',
            'resultats': resultats,
            'debug_repartition_id': repartition_id,
            'debug_type': delib_type,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def cancel_deliberation(request):
    """
    Annule (supprime) les résultats d'une délibération pour une classe.
    Cascade : annuler une période P annule aussi toutes les périodes supérieures,
    les trimestres/semestres concernés, et l'annuelle.
    Si preview=true dans le body, retourne la liste de ce qui sera annulé sans exécuter.
    """
    try:
        from MonEcole_app.models.evaluations.note import (
            Deliberation_annuelle_resultat,
            Deliberation_periodique_resultat,
            Deliberation_trimistrielle_resultat,
            Deliberation_examen_resultat,
        )
        from django.db import connections

        data = json.loads(request.body)
        classe_id = data.get('classe_id') or data.get('id_classe_id')
        delib_type = data.get('type')
        repartition_id = data.get('repartition_id')
        preview_mode = data.get('preview', False)

        etab, err = _get_tenant_etab(request)
        if err:
            return err

        annee = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        if not annee:
            return JsonResponse({'success': False, 'error': 'Aucune année en cours.'}, status=400)

        eac = EtablissementAnneeClasse.objects.filter(id=classe_id).first() if classe_id else None
        base_filter = {
            'id_annee': annee,
            'id_etablissement': etab.id_etablissement,
            'id_pays': etab.pays_id,
        }
        if eac:
            base_filter['id_classe_id'] = eac.classe_id
            base_filter['groupe'] = eac.groupe
            base_filter['section_id'] = eac.section_id

        # ── Ordre structurel des codes ──
        CODE_POSITION = {
            'P1': 1, 'P2': 2, 'S1': 5, 'T1': 5,
            'P3': 6, 'P4': 7, 'S2': 10, 'T2': 10,
            'P5': 11, 'P6': 12, 'S3': 15, 'T3': 15,
        }
        ANNUAL_POS = 100

        # Résoudre config_id → code + nom + type
        config_to_info = {}  # config_id → {'code': 'P1', 'nom': '...', 'pos': N, 'type_code': 'P'}
        if eac:
            try:
                with connections['countryStructure'].cursor() as cur:
                    cur.execute("""
                        SELECT rc.id, ri.code, ri.nom, rc.parent_id, rc.has_parent, rt.code as type_code
                        FROM repartition_configs_etab_annee rc
                        JOIN repartition_instances ri ON ri.id = rc.repartition_id
                        LEFT JOIN repartition_types rt ON rt.id = ri.type_id
                        WHERE rc.etablissement_annee_id = %s
                    """, [eac.etablissement_annee_id])
                    for row in cur.fetchall():
                        cfg_id, code, nom, parent_id, has_parent, type_code = row
                        pos = CODE_POSITION.get(code, 0)
                        # Déterminer si c'est une période via le type_code ('P') ou le code (commence par 'P')
                        is_period = (type_code == 'P') or (code and code.startswith('P'))
                        is_sem_tri = (type_code in ('S', 'T')) or (code and code[:1] in ('S', 'T'))
                        config_to_info[cfg_id] = {
                            'code': code, 'nom': nom, 'pos': pos,
                            'parent_id': parent_id, 'has_parent': bool(has_parent),
                            'type_code': type_code or '',
                            'is_period': is_period,
                            'is_sem_tri': is_sem_tri,
                        }
            except Exception as e:
                import traceback; traceback.print_exc()

        import sys
        print(f"[CANCEL DEBUG] config_to_info = { {k: (v['code'], v['pos'], v['is_period'], v['is_sem_tri']) for k,v in config_to_info.items()} }", file=sys.stderr)

        # Déterminer la position de la délibération qu'on annule
        cancel_pos = 0
        cancel_label = ''
        if delib_type == 'annee':
            cancel_pos = ANNUAL_POS
            cancel_label = 'Annuelle'
        elif delib_type == 'trimestre' and repartition_id:
            info = config_to_info.get(int(repartition_id), {})
            cancel_pos = info.get('pos', 0)
            cancel_label = info.get('nom', f'Trimestre {repartition_id}')
        elif delib_type == 'periode' and repartition_id:
            info = config_to_info.get(int(repartition_id), {})
            cancel_pos = info.get('pos', 0)
            cancel_label = info.get('nom', f'Période {repartition_id}')

        print(f"[CANCEL DEBUG] cancel_pos={cancel_pos}, cancel_label={cancel_label}, delib_type={delib_type}", file=sys.stderr)

        # Construire la liste de TOUT ce qui doit être annulé (position >= cancel_pos)
        to_cancel = []

        # Périodes dont la position est >= cancel_pos (identifiées par type_code='P' ou code startswith 'P')
        for cfg_id, info in config_to_info.items():
            if info['is_period'] and info['pos'] >= cancel_pos:
                f = dict(base_filter)
                f['id_periode_id'] = cfg_id
                exists = Deliberation_periodique_resultat.objects.filter(**f).exists()
                print(f"[CANCEL DEBUG]   Period {info['code']} cfg={cfg_id} pos={info['pos']} exists={exists}", file=sys.stderr)
                if exists:
                    to_cancel.append({
                        'type': 'periode', 'config_id': cfg_id,
                        'label': info['nom'], 'code': info['code'], 'pos': info['pos']
                    })

        # Semestres/Trimestres dont la position est >= cancel_pos
        for cfg_id, info in config_to_info.items():
            if info['is_sem_tri'] and info['pos'] >= cancel_pos:
                f = dict(base_filter)
                f['id_trimestre_id'] = cfg_id
                if Deliberation_trimistrielle_resultat.objects.filter(**f).exists():
                    to_cancel.append({
                        'type': 'trimestre', 'config_id': cfg_id,
                        'label': info['nom'], 'code': info['code'], 'pos': info['pos']
                    })
                # Aussi les examens du semestre/trimestre
                fe = dict(base_filter)
                fe['id_trimestre_id'] = cfg_id
                if Deliberation_examen_resultat.objects.filter(**fe).exists():
                    to_cancel.append({
                        'type': 'examen', 'config_id': cfg_id,
                        'label': f"Examen {info['nom']}", 'code': info['code'], 'pos': info['pos']
                    })

        # Annuelle (pos >= cancel_pos, toujours vrai puisque ANNUAL=100)
        if ANNUAL_POS >= cancel_pos:
            f_ann = dict(base_filter)
            if Deliberation_annuelle_resultat.objects.filter(**f_ann).exists():
                to_cancel.append({
                    'type': 'annee', 'config_id': None,
                    'label': 'Délibération Annuelle', 'code': 'ANN', 'pos': ANNUAL_POS
                })

        # Trier par position structurelle
        to_cancel.sort(key=lambda x: x['pos'])
        print(f"[CANCEL DEBUG] to_cancel = {[(c['type'], c['code'], c['pos']) for c in to_cancel]}", file=sys.stderr)

        # ── Mode preview ──
        if preview_mode:
            labels = [item['label'] for item in to_cancel]
            return JsonResponse({
                'success': True,
                'preview': True,
                'cancel_label': cancel_label,
                'affected': labels,
                'count': len(to_cancel),
            })

        # ── Mode exécution : supprimer en cascade ──
        total_deleted = 0
        for item in to_cancel:
            f = dict(base_filter)
            if item['type'] == 'periode':
                f['id_periode_id'] = item['config_id']
                d, _ = Deliberation_periodique_resultat.objects.filter(**f).delete()
                total_deleted += d
            elif item['type'] == 'trimestre':
                f['id_trimestre_id'] = item['config_id']
                d, _ = Deliberation_trimistrielle_resultat.objects.filter(**f).delete()
                total_deleted += d
            elif item['type'] == 'examen':
                f['id_trimestre_id'] = item['config_id']
                d, _ = Deliberation_examen_resultat.objects.filter(**f).delete()
                total_deleted += d
            elif item['type'] == 'annee':
                d, _ = Deliberation_annuelle_resultat.objects.filter(**f).delete()
                total_deleted += d

        cancelled_labels = [item['label'] for item in to_cancel]
        return JsonResponse({
            'success': True,
            'message': f'Délibérations annulées en cascade : {", ".join(cancelled_labels)}. {total_deleted} résultat(s) supprimé(s).'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_http_methods(["GET"])
def get_deliberation_results(request):
    """
    Charge les résultats de délibération existants.
    GET /api/deliberations/results/?classe_id=X&type=periode&repartition_id=Y
    """
    try:
        from MonEcole_app.models.evaluations.note import (
            Deliberation_annuelle_resultat,
            Deliberation_periodique_resultat,
            Deliberation_trimistrielle_resultat,
            Deliberation_annuelle_condition,
            Deliberation_annuelle_finalite,
        )

        classe_id = request.GET.get('classe_id')
        delib_type = request.GET.get('type', 'periode')
        repartition_id = request.GET.get('repartition_id')

        etab, err = _get_tenant_etab(request)
        if err:
            return err

        annee = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        if not annee:
            return JsonResponse({'success': True, 'resultats': []})

        eac = EtablissementAnneeClasse.objects.filter(id=classe_id).first() if classe_id else None
        if not eac:
            return JsonResponse({'success': True, 'resultats': []})

        base_filter = {
            'id_annee': annee,
            'id_etablissement': etab.id_etablissement,
            'id_pays': etab.pays_id,
            'id_classe_id': eac.classe_id,
            'groupe': eac.groupe,
            'section_id': eac.section_id,
        }

        results_qs = None

        if delib_type == 'periode' and repartition_id:
            f = dict(base_filter)
            f['id_periode_id'] = int(repartition_id)
            results_qs = Deliberation_periodique_resultat.objects.filter(**f).select_related('id_eleve').order_by('pourcentage').reverse()
        elif delib_type == 'trimestre' and repartition_id:
            f = dict(base_filter)
            f['id_trimestre_id'] = int(repartition_id)
            results_qs = Deliberation_trimistrielle_resultat.objects.filter(**f).select_related('id_eleve').order_by('pourcentage').reverse()
        elif delib_type == 'annee':
            results_qs = Deliberation_annuelle_resultat.objects.filter(**base_filter).select_related('id_eleve').order_by('pourcentage').reverse()
        else:
            return JsonResponse({'success': True, 'resultats': []})

        if results_qs is None or not results_qs.exists():
            return JsonResponse({'success': True, 'resultats': []})

        # Build response
        resultats = []
        mentions = {m.id_mention: str(m) for m in Mention.objects.filter(id_pays=etab.pays_id)}

        for r in results_qs:
            eleve = r.id_eleve
            mention_str = '—'
            decision_str = '—'

            # Get mention from percentage
            for m in Mention.objects.filter(id_pays=etab.pays_id):
                if m.min <= r.pourcentage <= m.max:
                    mention_str = str(m)
                    # Get decision from conditions
                    cond = Deliberation_annuelle_condition.objects.filter(
                        id_annee=annee, id_mention_id=m.id_mention, id_pays=etab.pays_id
                    ).first()
                    if cond:
                        try:
                            fin = Deliberation_annuelle_finalite.objects.get(id_finalite=cond.id_finalite_id, id_pays=etab.pays_id)
                            decision_str = fin.finalite
                        except:
                            decision_str = cond.sanction_disciplinaire or '—'
                    break

            resultats.append({
                'id_eleve': eleve.id_eleve,
                'nom': eleve.nom or '',
                'prenom': eleve.prenom or '',
                'pourcentage': round(r.pourcentage, 2),
                'total_note': '',
                'total_max': '',
                'place': r.place,
                'mention': mention_str,
                'decision': decision_str,
            })

        return JsonResponse({'success': True, 'resultats': resultats, 'exists': True})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# BULLETINS API
# =============================================================================

@require_http_methods(["GET"])
def get_deliberated_classes(request):
    """
    Retourne les classes qui ont été délibérées (au moins un résultat de délibération),
    avec le modèle de bulletin associé (depuis le Hub).
    GET /api/bulletins/classes/?session_id=X
    """
    try:
        etab, err = _get_tenant_etab(request)
        if err:
            return err

        annee = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        if not annee:
            return JsonResponse({'success': True, 'classes': []})

        ea = EtablissementAnnee.objects.filter(
            etablissement_id=etab.id_etablissement,
            annee=annee
        ).first()
        if not ea:
            return JsonResponse({'success': True, 'classes': []})

        # Récupérer toutes les EAC pour cet établissement/année
        eacs = EtablissementAnneeClasse.objects.filter(
            etablissement_annee=ea
        ).select_related('classe', 'classe__cycle', 'section')

        # Lire les modèles de bulletin depuis le Hub (filtré par pays)
        from MonEcole_app.models.evaluations.bulletin_model import BulletinClasseModel, BulletinModel
        bcm_map = {}  # {classe_id_hub: model_info}
        for bcm in BulletinClasseModel.objects.select_related('id_model').filter(id_pays=etab.pays_id):
            bcm_map[bcm.id_classe_id] = {
                'model_id': bcm.id_model_id,
                'model_name': bcm.id_model.model_name,
                'roundedValues': bcm.roundedValues,
            }

        # Checker quelles classes ont des résultats de délibération
        from MonEcole_app.models.evaluations.note import (
            Deliberation_trimistrielle_resultat,
            Deliberation_annuelle_resultat,
            Deliberation_periodique_resultat,
        )

        # Classes avec résultats (n'importe quel type de délibération)
        # On utilise des tuples (classe_id, groupe, section_id) pour distinguer les groupes
        delib_classes = set()

        # Trimestrielle
        for vals in Deliberation_trimistrielle_resultat.objects.filter(
            id_annee=annee, id_etablissement=etab.id_etablissement
        ).values_list('id_classe_id', 'groupe', 'section_id').distinct():
            delib_classes.add(vals)

        # Annuelle
        for vals in Deliberation_annuelle_resultat.objects.filter(
            id_annee=annee, id_etablissement=etab.id_etablissement
        ).values_list('id_classe_id', 'groupe', 'section_id').distinct():
            delib_classes.add(vals)

        # Périodique
        for vals in Deliberation_periodique_resultat.objects.filter(
            id_annee=annee, id_etablissement=etab.id_etablissement
        ).values_list('id_classe_id', 'groupe', 'section_id').distinct():
            delib_classes.add(vals)

        classes = []
        for eac in eacs:
            hub_classe_id = eac.classe_id
            model_info = bcm_map.get(hub_classe_id, None)
            is_deliberated = (hub_classe_id, eac.groupe, eac.section_id) in delib_classes

            # Compter les élèves inscrits
            from django.db import connection
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT COUNT(DISTINCT id_eleve_id) FROM eleve_inscription
                    WHERE id_annee_id=%s AND classe_id=%s AND groupe <=> %s AND section_id <=> %s
                      AND id_etablissement=%s AND status=1
                """, [annee.pk, eac.classe_id, eac.groupe, eac.section_id, etab.id_etablissement])
                nb_eleves = cur.fetchone()[0]

            cycle_name = eac.classe.cycle.cycle if eac.classe and hasattr(eac.classe, 'cycle') and eac.classe.cycle else '—'
            section_name = eac.section.nom if eac.section else None

            classes.append({
                'eac_id': eac.id,
                'classe_name': eac.classe.classe if eac.classe else '—',
                'groupe': eac.groupe or '',
                'cycle_name': cycle_name,
                'section_name': section_name,
                'nb_eleves': nb_eleves,
                'is_deliberated': is_deliberated,
                'model_name': model_info['model_name'] if model_info else None,
                'model_id': model_info['model_id'] if model_info else None,
                'roundedValues': model_info['roundedValues'] if model_info else False,
            })

        # Trier : délibérées en premier, puis par cycle/classe
        classes.sort(key=lambda c: (not c['is_deliberated'], c['cycle_name'], c['classe_name']))

        return JsonResponse({'success': True, 'classes': classes})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_bulletin_eleves(request):
    """
    Retourne les élèves d'une classe pour la génération de bulletin.
    GET /api/bulletins/eleves/?classe_id=X
    """
    try:
        classe_id = request.GET.get('classe_id') or request.GET.get('id_classe_id')
        if not classe_id:
            return JsonResponse({'success': False, 'error': 'classe_id requis'}, status=400)

        etab, err = _get_tenant_etab(request)
        if err:
            return err

        annee = Annee.objects.filter(isOpen=True, pays_id=etab.pays_id).first()
        if not annee:
            return JsonResponse({'success': True, 'eleves': []})

        # Resolve EAC.id → business keys
        eac = EtablissementAnneeClasse.objects.filter(id=classe_id).first()
        if not eac:
            return JsonResponse({'success': False, 'error': 'Classe non trouvée'}, status=404)

        from django.db import connection
        with connection.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT e.id_eleve, e.nom, e.prenom, e.genre, e.date_naissance
                FROM eleve_inscription ei
                JOIN eleve e ON ei.id_eleve_id = e.id_eleve
                WHERE ei.id_annee_id = %s
                  AND ei.classe_id = %s AND ei.groupe <=> %s AND ei.section_id <=> %s
                  AND ei.id_etablissement = %s
                  AND ei.status = 1
                ORDER BY e.nom, e.prenom
            """, [annee.pk, eac.classe_id, eac.groupe, eac.section_id, etab.id_etablissement])
            columns = [col[0] for col in cur.description]
            rows = cur.fetchall()

        eleves = []
        for row in rows:
            d = dict(zip(columns, row))
            eleves.append({
                'id_eleve': d['id_eleve'],
                'nom': d['nom'] or '',
                'prenom': d['prenom'] or '',
                'genre': d.get('genre', 'M'),
                'date_naissance': str(d['date_naissance']) if d.get('date_naissance') else '',
            })

        return JsonResponse({'success': True, 'eleves': eleves})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def delete_inscriptions(request):
    """
    Supprime les inscriptions d'élèves sélectionnés UNIQUEMENT s'ils ne sont
    référencés dans aucune autre table (notes, presences, bulletins, deliberations, etc.)
    pour cette classe et cette année.
    """
    try:
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Non authentifié.'}, status=401)

        etab, err = _get_tenant_etab(request)
        if err:
            return err
        etab_id = etab.id_etablissement

        data = json.loads(request.body)
        eleve_ids = data.get('eleve_ids', [])
        classe_par_annee_id = data.get('classe_par_annee_id')

        if not eleve_ids or not classe_par_annee_id:
            return JsonResponse({'success': False, 'error': 'eleve_ids et classe_par_annee_id requis.'}, status=400)

        # Resolve EAC → business keys
        bk = _resolve_eac_orm(classe_par_annee_id)
        if not bk:
            return JsonResponse({'success': False, 'error': 'Classe introuvable.'}, status=404)

        conn = _get_spoke_connection()
        try:
            deleted = []
            blocked = []

            with conn.cursor() as cur:
                for eid in eleve_ids:
                    eid = int(eid)
                    reasons = []

                    # Check all related tables
                    dependency_checks = [
                        ("eleve_note", "id_eleve_id", "Notes d'évaluation"),
                        ("note_bulletin", "id_eleve_id", "Notes de bulletin"),
                        ("horaire_presence", "id_eleve_id", "Présences"),
                        ("eleve_conduite", "id_eleve_id", "Conduite"),
                        ("deliberation_annuelle_resultats", "id_eleve_id", "Délibération annuelle"),
                        ("deliberation_periodique_resultats", "id_eleve_id", "Délibération périodique"),
                        ("deliberation_trimistrielle_resultats", "id_eleve_id", "Délibération trimestrielle"),
                        ("deliberation_examen_resultats", "id_eleve_id", "Délibération examen"),
                        ("biblio_emprunt", "id_eleve_id", "Emprunts bibliothèque"),
                        ("document_eleve", "id_eleve", "Documents administratifs"),
                        ("recouvrment_paiement", "id_eleve_id", "Paiements"),
                    ]

                    for table, col, label in dependency_checks:
                        try:
                            cur.execute(f"SELECT COUNT(*) AS cnt FROM {table} WHERE {col} = %s", [eid])
                            row = cur.fetchone()
                            cnt = row['cnt'] if isinstance(row, dict) else row[0]
                            if cnt > 0:
                                reasons.append(f"{label} ({cnt})")
                        except Exception:
                            # Table might not exist in this spoke
                            pass

                    if reasons:
                        # Get student name for feedback
                        cur.execute("SELECT nom, prenom FROM eleve WHERE id_eleve = %s", [eid])
                        stu = cur.fetchone()
                        nom = f"{stu['nom'] or ''} {stu['prenom'] or ''}".strip() if stu else f"ID {eid}"
                        blocked.append({
                            'id_eleve': eid,
                            'nom': nom,
                            'reasons': reasons
                        })
                    else:
                        # Safe to delete — remove inscription first, then eleve
                        cur.execute("""
                            DELETE FROM eleve_inscription
                            WHERE id_eleve_id = %s
                              AND classe_id = %s AND groupe <=> %s AND section_id <=> %s
                              AND id_etablissement = %s
                        """, [eid, bk['classe_id'], bk['groupe'], bk['section_id'], etab_id])

                        # Check if eleve has any other inscriptions remaining
                        cur.execute("""
                            SELECT COUNT(*) AS cnt FROM eleve_inscription WHERE id_eleve_id = %s
                        """, [eid])
                        remaining = cur.fetchone()
                        remaining_cnt = remaining['cnt'] if isinstance(remaining, dict) else remaining[0]

                        if remaining_cnt == 0:
                            # No more inscriptions — safe to delete the eleve record
                            id_pays = getattr(request, 'id_pays', None) or request.session.get('id_pays')
                            cur.execute("DELETE FROM eleve WHERE id_eleve = %s AND id_pays = %s", [eid, int(id_pays)])

                        deleted.append(eid)

            conn.commit()

            return JsonResponse({
                'success': True,
                'deleted_count': len(deleted),
                'deleted_ids': deleted,
                'blocked_count': len(blocked),
                'blocked': blocked,
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
